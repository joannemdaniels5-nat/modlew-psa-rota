import io
import re
from dataclasses import dataclass
from datetime import datetime, date, time, timedelta
from typing import Dict, List, Tuple, Optional, Set

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# =========================================================
# Rota Generator Engine — v14+++ Full Refactor (rules-first)
# =========================================================

# ---------- Time grid ----------
DAY_START = time(8, 0)
DAY_END = time(18, 30)
SLOT_MIN = 30

def timeslots() -> List[time]:
    cur = datetime(2000, 1, 1, DAY_START.hour, DAY_START.minute)
    end = datetime(2000, 1, 1, DAY_END.hour, DAY_END.minute)
    out: List[time] = []
    while cur < end:
        out.append(cur.time())
        cur += timedelta(minutes=SLOT_MIN)
    return out

def normalize(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s).strip().lower())

def find_sheet(xls: pd.ExcelFile, candidates: List[str]) -> Optional[str]:
    names = {normalize(n): n for n in xls.sheet_names}
    for c in candidates:
        k = normalize(c)
        if k in names:
            return names[k]
    for n in xls.sheet_names:
        nn = normalize(n)
        for c in candidates:
            if normalize(c) in nn:
                return n
    return None

def pick_col(df: pd.DataFrame, candidates: List[str], required: bool = True) -> Optional[str]:
    cols = {normalize(c): c for c in df.columns}
    for cand in candidates:
        k = normalize(cand)
        if k in cols:
            return cols[k]
    for c in df.columns:
        nc = normalize(c)
        for cand in candidates:
            if normalize(cand) in nc:
                return c
    if required:
        raise KeyError(f"Missing required column among {candidates}. Available: {list(df.columns)}")
    return None

def to_time(x):
    if pd.isna(x) or x == "":
        return None
    if isinstance(x, time):
        return x
    if isinstance(x, datetime):
        return x.time()
    if isinstance(x, (float, int)):
        seconds = int(round(float(x) * 86400))
        return (datetime(2000, 1, 1) + timedelta(seconds=seconds)).time()
    return pd.to_datetime(str(x)).time()

def to_date(x):
    if pd.isna(x) or x == "":
        return None
    if isinstance(x, (datetime, pd.Timestamp)):
        return x.date()
    if isinstance(x, date):
        return x
    return pd.to_datetime(x, dayfirst=True).date()

def dt_of(d: date, t: time) -> datetime:
    return datetime(d.year, d.month, d.day, t.hour, t.minute)

def add_minutes(t: time, mins: int) -> time:
    return (datetime(2000, 1, 1, t.hour, t.minute) + timedelta(minutes=mins)).time()

def ensure_monday(d: date) -> date:
    return d - timedelta(days=d.weekday())

def day_name(d: date) -> str:
    return ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"][d.weekday()]

def t_in_range(t: time, a: time, b: time) -> bool:
    return (t >= a) and (t < b)

# ---------- Rules / bands ----------
SITES = ["SLGP", "JEN", "BGS"]

FD_BANDS = [
    (time(8, 0), time(11, 0)),
    (time(11, 0), time(13, 30)),
    (time(13, 30), time(16, 0)),
    (time(16, 0), time(18, 30)),
]

TRIAGE_BANDS = [
    (time(8, 0), time(10, 30)),
    (time(10, 30), time(13, 0)),
    (time(13, 30), time(16, 0)),
]

BREAK_WINDOW = (time(12, 0), time(14, 0))
BREAK_CANDIDATES = [time(12, 0), time(12, 30), time(13, 0), time(13, 30)]
BREAK_THRESHOLD_HOURS = 6.0

# Variable task block rules (30-min slots)
# Variety: aim 2.5–4.5h stints where possible
MIN_2P5H = 5   # 2.5h
MAX_4H   = 8   # 4h
MAX_4P5H = 9   # 4.5h
MIN_DOCMAN = 6 # 3h
# --- Hard defaults for EMIS / Docman (requested) ---
EMIS_TARGET_HOURS = 20.0
DOCMAN_TARGET_HOURS = 14.0
TARGET_TOLERANCE = 0.15  # ±15%
EMIS_MIN_HOURS = EMIS_TARGET_HOURS * (1 - TARGET_TOLERANCE)
EMIS_MAX_HOURS = EMIS_TARGET_HOURS * (1 + TARGET_TOLERANCE)
DOCMAN_MIN_HOURS = DOCMAN_TARGET_HOURS * (1 - TARGET_TOLERANCE)
DOCMAN_MAX_HOURS = DOCMAN_TARGET_HOURS * (1 + TARGET_TOLERANCE)

# hard slot caps
EMIS_SLOT_CAP = 1
DOCMAN_SLOT_CAP = 1


# Priorities (strict) — quotas are soft
# 1) FrontDesk  2) Triage + Email  3) Phones + Awaiting/PSA Admin  4) Bookings  5) EMIS/Docman fill
PRIORITY_ORDER = [
    "FrontDesk", "Triage", "Email", "Phones", "Awaiting", "Bookings", "EMIS", "Docman"
]

# Phones steal order (strict), never steal from FD/Triage
PHONES_STEAL_ORDER = ["Bookings", "Awaiting", "Email_after_16"]

# ---------- Data models ----------
@dataclass
class Staff:
    name: str
    home: str
    can_frontdesk: bool
    can_triage: bool
    can_email: bool
    can_phones: bool
    can_bookings: bool
    can_emis: bool
    can_docman: bool
    weights: Dict[str, int]  # tie-break only

@dataclass
class TemplateData:
    staff: List[Staff]
    hours_map: Dict[str, Dict[str, Optional[time]]]
    hols: List[Tuple[str, date, date, str]]
    call_handlers: pd.DataFrame  # hours rows
    handler_leave: pd.DataFrame  # leave ranges
    phones_targets: Dict[Tuple[str, time], int]  # (Mon, 08:00)->5
    bookings_targets: Dict[Tuple[str, time], int]  # optional
    weekly_targets: Dict[str, float]  # Bookings/EMIS/Docman weekly hours
    swaps: List[Tuple[date, str, str, Optional[time], Optional[time]]]
    buddies: Dict[str, str]

# ---------- Template parsing ----------
def yn(v) -> bool:
    if pd.isna(v):
        return False
    s = str(v).strip().lower()
    return s in {"y", "yes", "true", "1", "t"}

def read_template(uploaded_bytes: bytes) -> TemplateData:
    xls = pd.ExcelFile(io.BytesIO(uploaded_bytes))

    staff_sheet = find_sheet(xls, ["Staff"])
    hours_sheet = find_sheet(xls, ["WorkingHours", "Hours"])
    hols_sheet = find_sheet(xls, ["Holidays", "Leave", "Absence"])

    callh_sheet = find_sheet(xls, ["CallHandlers", "Call Handlers"])
    hleave_sheet = find_sheet(xls, ["Handler_Leave", "Handler Leave", "CallHandler_Leave"])

    tph_sheet = find_sheet(xls, ["Targets_Phones_Hourly", "PhonesTargets"])
    tbk_sheet = find_sheet(xls, ["Targets_Bookings_Hourly", "BookingsTargets"])
    tweek_sheet = find_sheet(xls, ["Targets_Weekly"])

    swaps_sheet = find_sheet(xls, ["Swaps"])
    new_sheet = find_sheet(xls, ["NewStarters", "New Starters"])

    if not staff_sheet or not hours_sheet:
        raise ValueError(f"Missing Staff/WorkingHours sheets. Found: {xls.sheet_names}")

    staff_df = pd.read_excel(xls, sheet_name=staff_sheet)
    hours_df = pd.read_excel(xls, sheet_name=hours_sheet)
    hols_df = pd.read_excel(xls, sheet_name=hols_sheet) if hols_sheet else pd.DataFrame()

    callh_df = pd.read_excel(xls, sheet_name=callh_sheet) if callh_sheet else pd.DataFrame()
    hleave_df = pd.read_excel(xls, sheet_name=hleave_sheet) if hleave_sheet else pd.DataFrame()

    tph_df = pd.read_excel(xls, sheet_name=tph_sheet) if tph_sheet else pd.DataFrame()
    tbk_df = pd.read_excel(xls, sheet_name=tbk_sheet) if tbk_sheet else pd.DataFrame()
    tweek_df = pd.read_excel(xls, sheet_name=tweek_sheet) if tweek_sheet else pd.DataFrame()

    swaps_df = pd.read_excel(xls, sheet_name=swaps_sheet) if swaps_sheet else pd.DataFrame()
    new_df = pd.read_excel(xls, sheet_name=new_sheet) if new_sheet else pd.DataFrame()

    # Staff
    name_c = pick_col(staff_df, ["Name", "StaffName"])
    home_c = pick_col(staff_df, ["HomeSite", "Site", "BaseSite"], required=False)

    staff_df = staff_df.copy()
    staff_df["Name"] = staff_df[name_c].astype(str).str.strip()
    staff_df["HomeSite"] = staff_df[home_c].astype(str).str.strip().str.upper() if home_c else ""

    def bool_col(df, cands):
        c = pick_col(df, cands, required=False)
        if not c:
            return pd.Series([False] * len(df))
        return df[c].apply(yn)

    staff_df["CanFrontDesk"] = bool_col(staff_df, ["CanFrontDesk"])
    staff_df["CanTriage"] = bool_col(staff_df, ["CanTriage"])
    staff_df["CanEmail"] = bool_col(staff_df, ["CanEmail"])
    staff_df["CanPhones"] = bool_col(staff_df, ["CanPhones"])
    staff_df["CanBookings"] = bool_col(staff_df, ["CanBookings"])
    staff_df["CanEMIS"] = bool_col(staff_df, ["CanEMIS"])
    staff_df["CanDocman"] = bool_col(staff_df, ["CanDocman_PSA"]) | bool_col(staff_df, ["CanDocman_AWAIT"]) | bool_col(staff_df, ["CanDocman"])

    weight_map = {
        "Phones": "PhonesWeight",
        "Bookings": "BookingsWeight",
        "EMIS": "EmisWeight",
        "Docman": "DocmanWeight",
        "Awaiting": "AwaitingWeight",
        "Email": "EmailWeight",
        "Triage": "TriageWeight",
        "FrontDesk": "FrontDeskWeight",
    }

    staff_list: List[Staff] = []
    for _, r in staff_df.iterrows():
        weights = {}
        for k, col in weight_map.items():
            v = r.get(col, 3)
            try:
                if pd.isna(v):
                    v = 3
                v = int(float(v))
                v = max(0, min(5, v))
            except Exception:
                v = 3
            weights[k] = v

        staff_list.append(
            Staff(
                name=str(r["Name"]).strip(),
                home=str(r.get("HomeSite", "")).strip().upper(),
                can_frontdesk=bool(r.get("CanFrontDesk", False)),
                can_triage=bool(r.get("CanTriage", False)),
                can_email=bool(r.get("CanEmail", False)),
                can_phones=bool(r.get("CanPhones", False)),
                can_bookings=bool(r.get("CanBookings", False)),
                can_emis=bool(r.get("CanEMIS", False)),
                can_docman=bool(r.get("CanDocman", False)),
                weights=weights,
            )
        )

    # Working hours
    hours_df = hours_df.copy()
    hn = pick_col(hours_df, ["Name", "StaffName"])
    hs = pick_col(hours_df, ["HomeSite", "Site", "BaseSite"], required=False)
    hours_df["Name"] = hours_df[hn].astype(str).str.strip()
    if hs:
        hours_df["HomeSite"] = hours_df[hs].astype(str).str.strip().str.upper()
    else:
        hours_df["HomeSite"] = ""

    for dn in ["Mon", "Tue", "Wed", "Thu", "Fri"]:
        sc = pick_col(hours_df, [f"{dn}Start", f"{dn} Start", f"{dn}_Start"], required=False)
        ec = pick_col(hours_df, [f"{dn}End", f"{dn} End", f"{dn}_End"], required=False)
        hours_df[f"{dn}Start"] = hours_df[sc].apply(to_time) if sc else None
        hours_df[f"{dn}End"] = hours_df[ec].apply(to_time) if ec else None

    hours_map = {}
    for _, r in hours_df.iterrows():
        hours_map[r["Name"]] = {k: r.get(k) for k in hours_df.columns}

    # Holidays (ranges)
    hols: List[Tuple[str, date, date, str]] = []
    if not hols_df.empty:
        ncol = pick_col(hols_df, ["Name", "StaffName"], required=False) or hols_df.columns[0]
        sdcol = pick_col(hols_df, ["StartDate", "Start"], required=False) or hols_df.columns[1]
        edcol = pick_col(hols_df, ["EndDate", "End"], required=False) or hols_df.columns[2]
        notes_c = pick_col(hols_df, ["Notes", "Note", "Reason"], required=False)

        for _, r in hols_df.iterrows():
            nm = str(r.get(ncol, "")).strip()
            sd = to_date(r.get(sdcol))
            ed = to_date(r.get(edcol))
            note = "" if (not notes_c or pd.isna(r.get(notes_c))) else str(r.get(notes_c)).strip().lower()
            kind = "Holiday"
            if "sick" in note or "sickness" in note:
                kind = "Sick"
            elif "bank" in note:
                kind = "Bank Holiday"
            if nm and sd and ed:
                hols.append((nm, sd, ed, kind))

    # Targets
    phones_targets: Dict[Tuple[str, time], int] = {}
    bookings_targets: Dict[Tuple[str, time], int] = {}

    def parse_hourly(df: pd.DataFrame) -> Dict[Tuple[str, time], int]:
        out = {}
        if df is None or df.empty:
            return out
        time_col = pick_col(df, ["Time"], required=False) or df.columns[0]
        ddf = df.copy()
        ddf["Time"] = ddf[time_col].apply(to_time)
        for dn in ["Mon", "Tue", "Wed", "Thu", "Fri"]:
            if dn not in ddf.columns:
                continue
            for _, r in ddf.iterrows():
                hh = r.get("Time")
                if not hh:
                    continue
                val = r.get(dn)
                if pd.isna(val) or val == "":
                    continue
                out[(dn, time(hh.hour, 0))] = int(float(val))
        return out

    phones_targets = parse_hourly(tph_df)
    bookings_targets = parse_hourly(tbk_df)

    weekly_targets = {"Bookings": 0.0, "EMIS": 0.0, "Docman": 0.0}
    if tweek_df is not None and not tweek_df.empty:
        task_c = pick_col(tweek_df, ["Task"], required=False) or tweek_df.columns[0]
        val_c = pick_col(tweek_df, ["WeekHoursTarget", "Target", "Hours"], required=False) or tweek_df.columns[1]
        for _, r in tweek_df.iterrows():
            tsk = str(r.get(task_c, "")).strip()
            val = r.get(val_c)
            if pd.isna(val) or val == "":
                continue
            if tsk in weekly_targets:
                weekly_targets[tsk] = float(val)

    # Swaps
    swaps: List[Tuple[date, str, str, Optional[time], Optional[time]]] = []
    if swaps_df is not None and not swaps_df.empty:
        dcol = pick_col(swaps_df, ["Date"], required=False) or swaps_df.columns[0]
        ncol = pick_col(swaps_df, ["Name"], required=False) or swaps_df.columns[1]
        swcol = pick_col(swaps_df, ["SwapWith", "SwapWith (OPTION A)", "Swap With"], required=False)
        nscol = pick_col(swaps_df, ["NewStart", "NewStart (OPTION B)", "New Start"], required=False)
        necol = pick_col(swaps_df, ["NewEnd", "NewEnd (OPTION B)", "New End"], required=False)
        for _, r in swaps_df.iterrows():
            dd = to_date(r.get(dcol))
            if not dd:
                continue
            nm = str(r.get(ncol, "")).strip()
            sw = str(r.get(swcol, "")).strip() if swcol else ""
            ns = to_time(r.get(nscol)) if nscol else None
            ne = to_time(r.get(necol)) if necol else None
            if nm:
                swaps.append((dd, nm, sw, ns, ne))

    # Buddy map
    buddies: Dict[str, str] = {}
    if new_df is not None and not new_df.empty:
        nc = pick_col(new_df, ["NewStarterName", "NewStarter", "Starter"], required=False) or new_df.columns[0]
        bc = pick_col(new_df, ["BuddyName", "Buddy"], required=False) or new_df.columns[1]
        for _, r in new_df.iterrows():
            n = str(r.get(nc, "")).strip()
            b = str(r.get(bc, "")).strip()
            if n and b:
                buddies[n] = b

    return TemplateData(
        staff=staff_list,
        hours_map=hours_map,
        hols=hols,
        call_handlers=callh_df if callh_df is not None else pd.DataFrame(),
        handler_leave=hleave_df if hleave_df is not None else pd.DataFrame(),
        phones_targets=phones_targets,
        bookings_targets=bookings_targets,
        weekly_targets=weekly_targets,
        swaps=swaps,
        buddies=buddies,
    )

# ---------- Availability ----------
def holiday_kind(name: str, d: date, hols: List[Tuple[str, date, date, str]]) -> Optional[str]:
    for n, s, e, k in hols:
        if n.strip().lower() == name.strip().lower() and s and e and s <= d <= e:
            return k
    return None

def shift_window(hours_map: Dict[str, Dict[str, Optional[time]]], d: date, name: str) -> Tuple[Optional[time], Optional[time]]:
    dn = day_name(d)
    hr = hours_map.get(name)
    if hr is None:
        return None, None
    return hr.get(f"{dn}Start"), hr.get(f"{dn}End")

def is_working(hours_map: Dict[str, Dict[str, Optional[time]]], d: date, t: time, name: str) -> bool:
    stt, end = shift_window(hours_map, d, name)
    return bool(stt and end and (t >= stt) and (t < end))

def staff_home_from_hours(hours_map: Dict[str, Dict[str, Optional[time]]], name: str) -> str:
    hr = hours_map.get(name, {})
    hs = hr.get("HomeSite")
    return str(hs).strip().upper() if hs else ""

# ---------- Call handler leave impacts (phones) ----------
def parse_handler_leave(df: pd.DataFrame) -> List[Tuple[str, date, date]]:
    if df is None or df.empty:
        return []
    ncol = pick_col(df, ["Name", "HandlerName", "CallHandler"], required=False) or df.columns[0]
    sdcol = pick_col(df, ["LeaveStartDate", "LeaveStart", "StartDate"], required=False) or df.columns[1]
    edcol = pick_col(df, ["LeaveEndDate", "LeaveEnd", "EndDate"], required=False) or df.columns[2]
    out = []
    for _, r in df.iterrows():
        nm = str(r.get(ncol, "")).strip()
        sd = to_date(r.get(sdcol))
        ed = to_date(r.get(edcol))
        if nm and sd and ed:
            out.append((nm, sd, ed))
    return out

def handler_working(callh_row: pd.Series, d: date, t: time) -> bool:
    dn = day_name(d)
    stt = to_time(callh_row.get(f"{dn}Start"))
    end = to_time(callh_row.get(f"{dn}End"))
    return bool(stt and end and (t >= stt) and (t < end))

def phones_required(tpl: TemplateData, d: date, t: time) -> int:
    dn = day_name(d)
    hour_key = time(t.hour, 0)
    base = int(tpl.phones_targets.get((dn, hour_key), 0))

    # Add 1 for each call handler who would otherwise be working but is on leave that day
    leave_ranges = parse_handler_leave(tpl.handler_leave)
    off = 0
    if tpl.call_handlers is not None and not tpl.call_handlers.empty:
        for _, r in tpl.call_handlers.iterrows():
            nm = str(r.get("Name", "")).strip()
            if not nm:
                continue
            if not handler_working(r, d, t):
                continue
            for ln, sd, ed in leave_ranges:
                if ln.strip().lower() == nm.strip().lower() and sd <= d <= ed:
                    off += 1
                    break
    return base + off

def bookings_required(tpl: TemplateData, d: date, t: time) -> int:
    if not tpl.bookings_targets:
        return 0
    dn = day_name(d)
    hour_key = time(t.hour, 0)
    return int(tpl.bookings_targets.get((dn, hour_key), 0) or 0)

# ---------- Site-of-day rules ----------
def awaiting_site_for_day(d: date) -> str:
    wd = d.weekday()
    if wd in (0, 4):  # Mon/Fri
        return "SLGP"
    if wd in (1, 3):  # Tue/Thu
        return "JEN"
    return "BGS"  # Wed

def email_site_allowed(staff: Staff, d: date) -> bool:
    # Per your standard rule: Email covered by JEN/BGS; move only when no suitable on that site.
    # We'll keep to JEN/BGS by default. SLGP only used as last resort (gap note).
    return staff.home in ("JEN", "BGS") and staff.can_email

# ---------- Break placement (site-balanced, avoid <1h fragments) ----------
def pick_breaks_site_balanced(staff_list: List[Staff], hours_map: Dict[str, Dict[str, Optional[time]]], hols, week_dates: List[date]) -> Dict[Tuple[date, time], Set[str]]:
    breaks: Dict[Tuple[date, time], Set[str]] = {}
    break_load: Dict[Tuple[date, str, time], int] = {}

    for d in week_dates:
        for st in staff_list:
            if holiday_kind(st.name, d, hols):
                continue
            stt, end = shift_window(hours_map, d, st.name)
            if not stt or not end:
                continue
            dur = (dt_of(d, end) - dt_of(d, stt)).total_seconds() / 3600.0
            if dur <= BREAK_THRESHOLD_HOURS:
                continue

            # Prefer midpoint but balance within site
            midpoint = dt_of(d, stt) + (dt_of(d, end) - dt_of(d, stt)) / 2
            best = None
            for bt in BREAK_CANDIDATES:
                if bt < stt or add_minutes(bt, 30) > end:
                    continue
                if not t_in_range(bt, BREAK_WINDOW[0], BREAK_WINDOW[1]):
                    continue

                # Avoid leaving <1h on either side if possible
                before = (dt_of(d, bt) - dt_of(d, stt)).total_seconds() / 3600.0
                after = (dt_of(d, end) - dt_of(d, add_minutes(bt, 30))).total_seconds() / 3600.0
                frag_penalty = 0
                if before < 1.0:
                    frag_penalty += 10_000
                if after < 1.0:
                    frag_penalty += 10_000

                load = break_load.get((d, st.home, bt), 0)
                dist = abs((dt_of(d, bt) - midpoint).total_seconds())
                score = frag_penalty + (load * 3600) + dist
                if best is None or score < best[0]:
                    best = (score, bt)

            if best:
                bt = best[1]
                breaks.setdefault((d, bt), set()).add(st.name)
                break_load[(d, st.home, bt)] = break_load.get((d, st.home, bt), 0) + 1

    return breaks

# ---------- Scheduling core ----------
def apply_swaps(hours_map: Dict[str, Dict[str, Optional[time]]], swaps: List[Tuple[date, str, str, Optional[time], Optional[time]]], week_dates: List[date]) -> Dict[str, Dict[str, Optional[time]]]:
    """Apply swap rules to hours_map.

    Supports:
      A) SwapWith another staff member on a specific date (swap that day's start/end).
      B) NewStart/NewEnd override on a specific date.
      C) Self day-swap: two rows for the same staff in the same week with SwapWith blank and NewStart/NewEnd blank.
         Those two dates' start/end are swapped.
    """
    out = {k: dict(v) for k, v in hours_map.items()}
    in_week = set(week_dates)

    # First pass: apply explicit swap-with and explicit overrides
    self_swap_rows: Dict[str, List[date]] = {}
    for dd, nm, sw, ns, ne in swaps:
        if dd not in in_week:
            continue
        if nm not in out:
            continue
        dn = day_name(dd)

        sw = (sw or "").strip()
        # Overrides win
        if ns and ne:
            out[nm][f"{dn}Start"], out[nm][f"{dn}End"] = ns, ne
            continue

        # Swap with another person
        if sw:
            if sw not in out:
                continue
            a1, a2 = out[nm].get(f"{dn}Start"), out[nm].get(f"{dn}End")
            b1, b2 = out[sw].get(f"{dn}Start"), out[sw].get(f"{dn}End")
            out[nm][f"{dn}Start"], out[nm][f"{dn}End"] = b1, b2
            out[sw][f"{dn}Start"], out[sw][f"{dn}End"] = a1, a2
            continue

        # Candidate for self day-swap pairing
        if (not sw) and (not ns) and (not ne):
            self_swap_rows.setdefault(nm, []).append(dd)

    # Second pass: self day-swap pairs (exactly 2 dates in the same week)
    for nm, dts in self_swap_rows.items():
        # only pair within the provided week_dates
        dts = sorted([d for d in dts if d in in_week])
        if len(dts) != 2:
            continue
        d1, d2 = dts
        dn1, dn2 = day_name(d1), day_name(d2)
        a1s, a1e = out[nm].get(f"{dn1}Start"), out[nm].get(f"{dn1}End")
        a2s, a2e = out[nm].get(f"{dn2}Start"), out[nm].get(f"{dn2}End")
        out[nm][f"{dn1}Start"], out[nm][f"{dn1}End"] = a2s, a2e
        out[nm][f"{dn2}Start"], out[nm][f"{dn2}End"] = a1s, a1e

    return out


def schedule_week(tpl: TemplateData, week_start: date) -> Tuple[Dict[Tuple[date, time, str], str], Dict[Tuple[date, time], Set[str]], List[Tuple[date, time, str, str]], List[date], List[time], Dict[str, Dict[str, Optional[time]]]]:
    slots = timeslots()
    dates = [week_start + timedelta(days=i) for i in range(5)]
    staff_by_name = {s.name: s for s in tpl.staff}

    # apply swaps to hours
    hours_map = apply_swaps(tpl.hours_map, tpl.swaps, dates)

    # breaks
    breaks = pick_breaks_site_balanced(tpl.staff, hours_map, tpl.hols, dates)

    a: Dict[Tuple[date, time, str], str] = {}
    gaps: List[Tuple[date, time, str, str]] = []

    # active blocks for variable tasks: (d,name)->(task,end_idx)
    active: Dict[Tuple[date, str], Tuple[str, int]] = {}

    # tracking minutes
    mins_task: Dict[Tuple[str, str], int] = {}  # (name, task)->minutes
    mins_task_day: Dict[Tuple[date, str, str], int] = {}  # (d,name,task)->minutes

    def add_minutes_track(d: date, name: str, task: str, minutes: int):
        mins_task[(name, task)] = mins_task.get((name, task), 0) + minutes
        mins_task_day[(d, name, task)] = mins_task_day.get((d, name, task), 0) + minutes

    def on_break(name: str, d: date, t: time) -> bool:
        return name in breaks.get((d, t), set())

    def is_free(name: str, d: date, t: time) -> bool:
        return (d, t, name) not in a

    def assigned_task(name: str, d: date, t: time) -> Optional[str]:
        return a.get((d, t, name))

    def staff_on(task: str, d: date, t: time) -> List[str]:
        return [nm for (dd, tt, nm), rr in a.items() if dd == d and tt == t and rr == task]

    def can_do(name: str, task: str, d: date, t: time) -> bool:
        st = staff_by_name[name]
        if holiday_kind(name, d, tpl.hols):
            return False
        if not is_working(hours_map, d, t, name):
            return False
        if on_break(name, d, t):
            return False

        # site rules + skills
        if task.startswith("FrontDesk_"):
            site = task.split("_", 1)[1]
            return st.can_frontdesk and st.home == site
        if task.startswith("Triage_Admin_"):
            site = task.split("_")[-1]
            return st.can_triage and st.home == site
        if task == "Email_Box":
            return email_site_allowed(st, d)
        if task == "Phones":
            return st.can_phones
        if task == "Awaiting_PSA_Admin":
            return st.can_docman and st.home == awaiting_site_for_day(d)
        if task == "Bookings":
            return st.can_bookings and st.home == "SLGP"
        if task == "EMIS":
            return st.can_emis
        if task == "Docman":
            return st.can_docman
        if task == "Misc_Tasks":
            return True
        return True

    def block_limits(task: str) -> Tuple[int, int]:
        if task == "Phones":
            return MIN_2P5H, MAX_4H
        if task == "Bookings":
            return MIN_2P5H, MAX_4H
        if task == "Email_Box":
            return MIN_2P5H, MAX_4H
        if task == "Awaiting_PSA_Admin":
            return MIN_2P5H, MAX_4H
        if task == "Docman":
            return 4, 8  # 2h–4h
        if task == "EMIS":
            return 4, 8  # 2h–4h
        return 4, 8

    def start_block(name: str, task: str, d: date, start_idx: int) -> bool:
        mn, mx = block_limits(task)
        stt, end = shift_window(hours_map, d, name)
        if not stt or not end:
            return False

        # end idx based on shift end
        end_idx = start_idx
        while end_idx < len(slots) and slots[end_idx] < end:
            end_idx += 1

        # cut at break
        for k in range(start_idx, end_idx):
            if name in breaks.get((d, slots[k]), set()):
                end_idx = k
                break

        remaining = end_idx - start_idx
        if remaining <= 0:
            return False

        L = remaining if remaining < mn else min(mx, remaining)
        if remaining >= mn and L < mn:
            L = mn

        # variety: avoid exceeding 4.5h/day on same variable task
        day_used = mins_task_day.get((d, name, task), 0)
        if task in {"Phones", "Bookings", "Email_Box", "Awaiting_PSA_Admin", "EMIS", "Docman"} and day_used >= 270:
            # already 4.5h
            return False

        active[(d, name)] = (task, start_idx + L)
        return True

    def apply_active(name: str, d: date, idx: int) -> bool:
        b = active.get((d, name))
        if not b:
            return False
        task, end_idx = b
        if idx >= end_idx:
            del active[(d, name)]
            return False
        t = slots[idx]
        a[(d, t, name)] = task
        add_minutes_track(d, name, task, SLOT_MIN)
        return True

    def stop_block(name: str, d: date):
        if (d, name) in active:
            del active[(d, name)]

    # -------- Phase 1: lock Front Desk bands (exactly 1 per site) --------
    def lock_band(d: date, band_start: time, band_end: time, task: str, candidates: List[str]):
        # Choose who can cover whole band
        ok: List[str] = []
        for nm in candidates:
            stt, end = shift_window(hours_map, d, nm)
            if not stt or not end:
                continue
            # must be working across whole band
            if stt > band_start or end < band_end:
                continue
            # no break inside band
            has_break = any(nm in breaks.get((d, tt), set()) for tt in slots if tt >= band_start and tt < band_end)
            if has_break:
                continue
            ok.append(nm)

        if not ok:
            gaps.append((d, band_start, task, "No suitable staff available for full band"))
            return

        # tie-break: least minutes on that task this week + weight
        def score(nm: str):
            st = staff_by_name[nm]
            used = mins_task.get((nm, task), 0)
            wkey = "Phones" if task == "Phones" else ("Bookings" if task == "Bookings" else ("Email" if task == "Email_Box" else ("Awaiting" if task == "Awaiting_PSA_Admin" else task)))
            w = st.weights.get(wkey, 3)
            # weight first (highest), then least used, then name
            return (-w, used, nm.lower())

        ok.sort(key=score)
        chosen = ok[0]

        for tt in slots:
            if tt < band_start or tt >= band_end:
                continue
            a[(d, tt, chosen)] = task
            add_minutes_track(d, chosen, task, SLOT_MIN)

    for d in dates:
        for site in SITES:
            task = f"FrontDesk_{site}"
            cands = [s.name for s in tpl.staff if s.home == site and s.can_frontdesk]
            for bs, be in FD_BANDS:
                lock_band(d, bs, be, task, cands)

    # -------- Phase 2: lock Triage bands (SLGP/JEN) --------
    for d in dates:
        for site in ("SLGP", "JEN"):
            task = f"Triage_Admin_{site}"
            cands = [s.name for s in tpl.staff if s.home == site and s.can_triage]
            for bs, be in TRIAGE_BANDS:
                lock_band(d, bs, be, task, cands)

    # -------- Phase 3: time-step fill by priority with blocks --------
    def pick_free_candidates(task: str, d: date, t: time) -> List[str]:
        out = []
        for nm in staff_by_name.keys():
            if not is_free(nm, d, t):
                continue
            if not can_do(nm, task, d, t):
                continue
            out.append(nm)
        # tie-break by least time already on task, weight, then name
        def score(nm: str):
            st = staff_by_name[nm]
            used = mins_task.get((nm, task), 0)
            wkey = "Phones" if task == "Phones" else ("Bookings" if task == "Bookings" else ("Email" if task == "Email_Box" else ("Awaiting" if task == "Awaiting_PSA_Admin" else task)))
            w = st.weights.get(wkey, 3)
            # weight first (highest), then least used, then name
            return (-w, used, nm.lower())
        out.sort(key=score)
        return out

    def assign_block_or_slot(nm: str, task: str, d: date, idx: int):
        # If already active same task, just apply
        b = active.get((d, nm))
        if b and b[0] == task:
            apply_active(nm, d, idx)
            return
        # If active different task, stop it (cut) to respect priority
        if b and b[0] != task:
            stop_block(nm, d)
        # Start new block if possible
        started = start_block(nm, task, d, idx)
        if not started:
            # Do not create 30-min floaters; park as Misc and let smoothing handle any edge cases.
            t = slots[idx]
            a[(d, t, nm)] = "Misc_Tasks"
            add_minutes_track(d, nm, "Misc_Tasks", SLOT_MIN)
            return
        else:
            apply_active(nm, d, idx)
            return

    def enforce(task: str, need: int, d: date, idx: int, allow_cross_site: bool = False, after_16_email_optional: bool = False):
        t = slots[idx]
        current = len(staff_on(task, d, t))

        while current < need:
            cands = pick_free_candidates(task, d, t)
            if cands:
                nm = cands[0]
                assign_block_or_slot(nm, task, d, idx)
                current = len(staff_on(task, d, t))
                continue

            # No free staff: steal if allowed for task
            if task == "Phones":
                if not steal_for_phones(d, idx):
                    gaps.append((d, t, "Phones", f"Short by {need-current}"))
                    break
                current = len(staff_on(task, d, t))
                continue

            if task in {"Email_Box", "Awaiting_PSA_Admin", "Bookings"}:
                if not steal_generic(task, d, idx):
                    # For optional email after 16, don't log as gap unless it's within mandatory window
                    if not (task == "Email_Box" and t >= time(16, 0)):
                        gaps.append((d, t, task, f"Short by {need-current}"))
                    break
                current = len(staff_on(task, d, t))
                continue

            gaps.append((d, t, task, f"Short by {need-current}"))
            break

    def steal_for_phones(d: date, idx: int) -> bool:
        t = slots[idx]
        # steal order: Bookings -> Awaiting -> Email(after 16 only)
        # never steal from FD/Triage bands (they are hard locked and those people will already be assigned)
        for rule in PHONES_STEAL_ORDER:
            if rule == "Bookings":
                donors = [nm for nm in staff_by_name.keys() if assigned_task(nm, d, t) == "Bookings"]
            elif rule == "Awaiting":
                donors = [nm for nm in staff_by_name.keys() if assigned_task(nm, d, t) == "Awaiting_PSA_Admin"]
            else:  # Email_after_16
                if t < time(16, 0):
                    donors = []
                else:
                    donors = [nm for nm in staff_by_name.keys() if assigned_task(nm, d, t) == "Email_Box"]

            for nm in donors:
                if not can_do(nm, "Phones", d, t):
                    continue
                # reassign this slot and cut their block
                prev = assigned_task(nm, d, t)
                a[(d, t, nm)] = "Phones"
                add_minutes_track(d, nm, "Phones", SLOT_MIN)
                # adjust prev minutes (we can't "un-add" cleanly without a full recompute; log note instead)
                gaps.append((d, t, "Phones", f"Stole {nm} from {prev}"))
                stop_block(nm, d)
                return True
        return False

    def steal_generic(task: str, d: date, idx: int) -> bool:
        t = slots[idx]
        # Generic: steal from EMIS/Docman first, then from Bookings if needed (but not from FD/Triage)
        donor_tasks = ["EMIS", "Docman", "Bookings"]
        if task == "Bookings":
            donor_tasks = ["EMIS", "Docman"]  # don't steal from awaiting/email to do bookings
        if task == "Email_Box":
            donor_tasks = ["Bookings", "EMIS", "Docman"]  # bookings can drop for email before 16
        if task == "Awaiting_PSA_Admin":
            donor_tasks = ["Bookings", "EMIS", "Docman"]

        for dtask in donor_tasks:
            donors = [nm for nm in staff_by_name.keys() if assigned_task(nm, d, t) == dtask]
            for nm in donors:
                if not can_do(nm, task, d, t):
                    continue
                prev = assigned_task(nm, d, t)
                a[(d, t, nm)] = task
                add_minutes_track(d, nm, task, SLOT_MIN)
                gaps.append((d, t, task, f"Stole {nm} from {prev}"))
                stop_block(nm, d)
                return True
        return False

    def fill_filler(d: date, idx: int):
        t = slots[idx]

        # Targets: use template if provided (>0), else hard defaults (with ±15% tolerance)
        emis_t_h = float(tpl.weekly_targets.get("EMIS", 0.0) or 0.0)
        doc_t_h  = float(tpl.weekly_targets.get("Docman", 0.0) or 0.0)
        emis_target = int(round(((emis_t_h if emis_t_h > 0 else EMIS_TARGET_HOURS) * 60)))
        doc_target  = int(round(((doc_t_h  if doc_t_h  > 0 else DOCMAN_TARGET_HOURS) * 60)))

        emis_min = int(round((emis_target * (1 - TARGET_TOLERANCE))))
        doc_min  = int(round((doc_target  * (1 - TARGET_TOLERANCE))))

        emis_done = sum(v for (nm, task), v in mins_task.items() if task == "EMIS")
        doc_done  = sum(v for (nm, task), v in mins_task.items() if task == "Docman")

        # Slot caps (requested): at most 1 EMIS + 1 Docman at a time
        cur_emis = len(staff_on("EMIS", d, t))
        cur_doc  = len(staff_on("Docman", d, t))

        for nm, st in staff_by_name.items():
            if not is_working(hours_map, d, t, nm):
                continue
            if holiday_kind(nm, d, tpl.hols):
                continue
            if on_break(nm, d, t):
                continue
            if not is_free(nm, d, t):
                continue

            chosen = None

            # Prioritise hitting minimum targets before Misc
            if doc_done < doc_min and cur_doc < DOCMAN_SLOT_CAP and can_do(nm, "Docman", d, t):
                chosen = "Docman"
            elif emis_done < emis_min and cur_emis < EMIS_SLOT_CAP and can_do(nm, "EMIS", d, t):
                chosen = "EMIS"
            else:
                # If minimums met, still top up towards target (soft) but never at the expense of mandatory tasks
                if doc_done < doc_target and cur_doc < DOCMAN_SLOT_CAP and can_do(nm, "Docman", d, t):
                    chosen = "Docman"
                elif emis_done < emis_target and cur_emis < EMIS_SLOT_CAP and can_do(nm, "EMIS", d, t):
                    chosen = "EMIS"
                else:
                    chosen = "Misc_Tasks"

            if chosen in ("EMIS", "Docman"):
                assign_block_or_slot(nm, chosen, d, idx)
                # refresh counters
                emis_done = sum(v for (nn, task), v in mins_task.items() if task == "EMIS")
                doc_done  = sum(v for (nn, task), v in mins_task.items() if task == "Docman")
                cur_emis = len(staff_on("EMIS", d, t))
                cur_doc  = len(staff_on("Docman", d, t))
            else:
                a[(d, t, nm)] = "Misc_Tasks"
                add_minutes_track(d, nm, "Misc_Tasks", SLOT_MIN)


    # Pre-apply active blocks each slot (so people don't jump)
    for d in dates:
        for idx, t in enumerate(slots):
            # extend existing blocks first
            for nm in staff_by_name.keys():
                if (d, t, nm) in a:  # fixed tasks already set
                    continue
                if on_break(nm, d, t):
                    continue
                apply_active(nm, d, idx)

            # Mandatory windows
            # Email mandatory 10:30–16:00; optional after 16 only if capacity
            if t_in_range(t, time(10, 30), time(16, 0)):
                enforce("Email_Box", 1, d, idx)
            # Awaiting mandatory 10:00–16:00 on site-of-day
            if t_in_range(t, time(10, 0), time(16, 0)):
                enforce("Awaiting_PSA_Admin", 1, d, idx)
            # Phones required all day by hourly matrix (applies to both half-hours)
            req_phones = phones_required(tpl, d, t)
            if req_phones > 0:
                enforce("Phones", req_phones, d, idx)

            # Bookings requirement (optional matrix); only from 10:30 onwards; lower priority so do it after phones
            if t >= time(10, 30):
                req_b = bookings_required(tpl, d, t)
                if req_b > 0:
                    enforce("Bookings", req_b, d, idx)

            # After enforcing, fill remaining with EMIS/Docman
            fill_filler(d, idx)

    # Buddy system (post-pass per slot): try align buddy with new starter's task (no FD/Triage steals)
    training_tasks = {"Phones", "Bookings", "EMIS", "Docman", "Email_Box", "Awaiting_PSA_Admin"}
    for d in dates:
        for t in slots:
            for trainee, buddy in tpl.buddies.items():
                if (d, t, trainee) not in a:
                    continue
                task = a[(d, t, trainee)]
                if task not in training_tasks:
                    continue
                # If buddy already on same task, fine
                if a.get((d, t, buddy)) == task:
                    continue
                # Buddy must be working, not on break, not fixed FD/Triage
                if not is_working(hours_map, d, t, buddy):
                    continue
                if on_break(buddy, d, t):
                    continue
                bt = a.get((d, t, buddy), "")
                if bt.startswith("FrontDesk_") or bt.startswith("Triage_Admin_"):
                    continue
                # Can buddy do the task?
                if buddy in staff_by_name and can_do(buddy, task, d, t):
                    a[(d, t, buddy)] = task
                    gaps.append((d, t, "Buddy", f"Aligned {buddy} with {trainee} on {task} (was {bt or 'free'})"))

    
    # --- Post-pass: remove any single-slot (30-min) fragments for non-fixed tasks ---
    FIXED_PREFIXES = ("FrontDesk_", "Triage_Admin_")
    SPECIAL = ("Break", "Holiday", "Bank Holiday", "Sick", "")
    for d in dates:
        for nm in staff_by_name.keys():
            seq = [a.get((d, tt, nm), "") for tt in slots]
            i = 0
            while i < len(slots):
                task = seq[i]
                j = i + 1
                while j < len(slots) and seq[j] == task:
                    j += 1
                block_len = j - i
                if block_len == 1:
                    tsk = str(task or "")
                    if tsk.startswith(FIXED_PREFIXES) or tsk in SPECIAL:
                        i = j
                        continue
                    prev = seq[i-1] if i-1 >= 0 else ""
                    nxt  = seq[j] if j < len(slots) else ""
                    chosen = ""
                    # Prefer extending previous real task, else next
                    if prev and (not str(prev).startswith(FIXED_PREFIXES)) and str(prev) not in SPECIAL:
                        chosen = prev
                    elif nxt and (not str(nxt).startswith(FIXED_PREFIXES)) and str(nxt) not in SPECIAL:
                        chosen = nxt
                    if chosen:
                        a[(d, slots[i], nm)] = chosen
                        seq[i] = chosen
                i = j
    return a, breaks, gaps, dates, slots, hours_map

# ---------- Excel output (master + formula-linked site sheets + counts) ----------
ROLE_COLORS = {
    "FrontDesk_SLGP": "FFF2CC",
    "FrontDesk_JEN":  "FFF2CC",
    "FrontDesk_BGS":  "FFF2CC",
    "Triage_Admin_SLGP": "D9EAD3",
    "Triage_Admin_JEN":  "D9EAD3",
    "Email_Box": "CFE2F3",
    "Phones": "C9DAF8",
    "Bookings": "FCE5CD",
    "EMIS": "EAD1DC",
    "Docman": "D0E0E3",
    "Awaiting_PSA_Admin": "D0E0E3",
    "Unassigned": "FFFFFF",
    "Break": "DDDDDD",
    "Holiday": "FFF2CC",
    "Bank Holiday": "FFE599",
    "Sick": "F4CCCC",
    "": "DDDDDD",
}

def fill_for(value: str) -> PatternFill:
    return PatternFill("solid", fgColor=ROLE_COLORS.get(value, "FFFFFF"))

THICK = Side(style="thick")
THIN = Side(style="thin")
DAY_BORDER = Border(top=THICK)
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def build_workbook(tpl: TemplateData, start_monday: date, weeks: int) -> Workbook:
    """
    Clean Excel builder (single source of truth = SITE timelines)

    - Creates ONLY site timelines (WeekX_SLGP_Timeline / WeekX_JEN_Timeline / WeekX_BGS_Timeline)
      containing ONLY that site's staff.
    - Applies colour fills directly (PatternFill) and thick borders between days.
    - Creates WeekX_Totals with formulas that COUNTIF against the relevant site timeline column
      so manual edits to a site timeline update totals automatically.
    - Creates WeekX_Coverage_By_Slot_By_Site (static, derived from engine output) with coloured columns.
    - Creates WeekX_Dashboard and WeekX_NotesAndGaps.
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule

    wb = Workbook()
    wb.remove(wb.active)

    all_staff = [s.name for s in tpl.staff]
    staff_by_name = {s.name: s for s in tpl.staff}

    # Pre-calc row count for slot area
    def slot_rows_count(dates, slots):
        return len(dates) * len(slots)

    # Border helper: thick line at the start of each day (on the first row of that day)
    def apply_day_separators(ws, date_col=1, start_row=2):
        last_date = None
        for r in range(start_row, ws.max_row + 1):
            cur_date = ws.cell(r, date_col).value
            if cur_date != last_date:
                # new day -> thick top border
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(r, c)
                    cell.border = Border(
                        left=cell.border.left or THIN,
                        right=cell.border.right or THIN,
                        top=THICK,
                        bottom=cell.border.bottom or THIN,
                    )
                last_date = cur_date

    # Consistent widths
    DATE_W, TIME_W, STAFF_W = 14, 8, 18

    def style_timeline_sheet(ws):
        ws.freeze_panes = "C2"
        ws.print_title_rows = "1:1"
        ws.column_dimensions["A"].width = DATE_W
        ws.column_dimensions["B"].width = TIME_W
        # staff cols set later after we know count

        # Header styling
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")

        # Cell borders + wrap + fills
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                cell.border = CELL_BORDER
                if c >= 3:
                    v = str(cell.value or "")
                    cell.fill = fill_for(v)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                else:
                    cell.alignment = Alignment(vertical="top")

        apply_day_separators(ws)

    def make_site_timeline(week_num: int, site: str, dates, slots, a, breaks, hours_map):
        site_staff = [s.name for s in tpl.staff if (str(s.home).upper() == site)]
        if not site_staff:
            return None, []

        ws = wb.create_sheet(f"Week{week_num}_{site}_Timeline")
        ws.append(["Date", "Time"] + site_staff)

        # widths
        for i, _nm in enumerate(site_staff):
            ws.column_dimensions[get_column_letter(3 + i)].width = STAFF_W

        # populate
        for d in dates:
            for t in slots:
                row = [d.strftime("%a %d-%b"), t.strftime("%H:%M")]
                for nm in site_staff:
                    hk = holiday_kind(nm, d, tpl.hols)
                    if hk:
                        val = hk
                    elif not is_working(hours_map, d, t, nm):
                        val = ""
                    elif nm in breaks.get((d, t), set()):
                        val = "Break"
                    else:
                        val = a.get((d, t, nm), "Misc_Tasks")
                        if not val or val == "Unassigned":
                            val = "Misc_Tasks"
                    row.append(val)
                ws.append(row)

        style_timeline_sheet(ws)
        return ws, site_staff

    def build_dynamic_totals(week_num: int, site_map: dict, max_row: int):
        """
        site_map: name -> (sheet_name, col_letter)
        """
        ws = wb.create_sheet(f"Week{week_num}_Totals")

        task_cols = [
            ("FrontDesk", 'FrontDesk_*'),
            ("Triage", 'Triage_Admin_*'),
            ("Email", 'Email_Box'),
            ("Awaiting", 'Awaiting_PSA_Admin'),
            ("Phones", 'Phones'),
            ("Bookings", 'Bookings'),
            ("EMIS", 'EMIS'),
            ("Docman", 'Docman'),
            ("Misc", 'Misc_Tasks'),
            ("Break", 'Break'),
        ]

        ws.append(["Name"] + [t for t, _ in task_cols] + ["WeeklyTotal"])
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "B2"
        ws.column_dimensions["A"].width = 22
        for i in range(2, 2 + len(task_cols) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 12

        for r_i, nm in enumerate(all_staff, start=2):
            row = [nm]
            if nm not in site_map:
                # not present in any site sheet -> zeros
                for _ in task_cols:
                    row.append(0)
                row.append(0)
                ws.append(row)
                continue

            sheet_name, col_letter = site_map[nm]
            rng = f"{col_letter}$2:{col_letter}${max_row}"

            for _task, crit in task_cols:
                # 0.5 hours per slot
                row.append(f"=0.5*COUNTIF('{sheet_name}'!{rng},\"{crit}\")")

            # WeeklyTotal = sum of the task columns in this row
            start_letter = get_column_letter(2)
            end_letter = get_column_letter(1 + len(task_cols))
            row.append(f"=SUM({start_letter}{r_i}:{end_letter}{r_i})")
            ws.append(row)

        # Progress bars for weekly total (optional visual)
        last_row = ws.max_row
        ws.conditional_formatting.add(
            f"{get_column_letter(2+len(task_cols))}2:{get_column_letter(2+len(task_cols))}{last_row}",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=45, color="63C384")
        )
        return ws


    def build_coverage_by_slot_by_site(week_num: int, dates, slots):
        """Coverage sheet (names per slot) that updates when Site Timelines are edited.

        Uses TEXTJOIN/IF array formulas referencing the three site timeline sheets.
        Requires Excel 365 / modern Excel dynamic array behaviour.
        """
        ws = wb.create_sheet(f"Week{week_num}_Coverage_By_Slot_By_Site")
        cov_cols = ["FD_SLGP","FD_JEN","FD_BGS","Triage_SLGP","Triage_JEN","Phones","Bookings","EMIS","Docman","Awaiting","Email","Misc"]
        ws.append(["Date","Time"] + cov_cols)
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "C2"
        ws.print_title_rows = "1:1"
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 8
        for i in range(len(cov_cols)):
            ws.column_dimensions[get_column_letter(3+i)].width = 26

        def names_on_sheet(sheet_name: str, row_num: int, crit: str) -> str:
            # crit is exact match, or a wildcard like "FrontDesk_*" where we match by SEARCH(prefix)
            if crit.endswith("*"):
                prefix = crit[:-1]
                return f'IFERROR(TEXTJOIN(", ",TRUE,IF(ISNUMBER(SEARCH("{prefix}",\'{sheet_name}\'!$C{row_num}:$ZZ{row_num})),\'{sheet_name}\'!$C$1:$ZZ$1,"")),"")'
            else:
                return f'IFERROR(TEXTJOIN(", ",TRUE,IF(\'{sheet_name}\'!$C{row_num}:$ZZ{row_num}="{crit}",\'{sheet_name}\'!$C$1:$ZZ$1,"")),"")'

        def names_all_sites(row_num: int, crit: str) -> str:
            a1 = names_on_sheet(f"Week{week_num}_SLGP_Timeline", row_num, crit)
            a2 = names_on_sheet(f"Week{week_num}_JEN_Timeline", row_num, crit)
            a3 = names_on_sheet(f"Week{week_num}_BGS_Timeline", row_num, crit)
            return f'=TEXTJOIN(", ",TRUE,{a1},{a2},{a3})'

        # Row alignment: Coverage row 2 corresponds to timeline row 2, etc.
        row_num = 2
        for d in dates:
            for t in slots:
                ws.append([d.strftime("%a %d-%b"), t.strftime("%H:%M")] + [""]*len(cov_cols))
                r = ws.max_row
                ws.cell(r, 3).value  = names_on_sheet(f"Week{week_num}_SLGP_Timeline", row_num, "FrontDesk_SLGP")
                ws.cell(r, 4).value  = names_on_sheet(f"Week{week_num}_JEN_Timeline",  row_num, "FrontDesk_JEN")
                ws.cell(r, 5).value  = names_on_sheet(f"Week{week_num}_BGS_Timeline",  row_num, "FrontDesk_BGS")
                ws.cell(r, 6).value  = names_on_sheet(f"Week{week_num}_SLGP_Timeline", row_num, "Triage_Admin_SLGP")
                ws.cell(r, 7).value  = names_on_sheet(f"Week{week_num}_JEN_Timeline",  row_num, "Triage_Admin_JEN")
                ws.cell(r, 8).value  = names_all_sites(row_num, "Phones")
                ws.cell(r, 9).value  = names_all_sites(row_num, "Bookings")
                ws.cell(r, 10).value = names_all_sites(row_num, "EMIS")
                ws.cell(r, 11).value = names_all_sites(row_num, "Docman")
                ws.cell(r, 12).value = names_all_sites(row_num, "Awaiting_PSA_Admin")
                ws.cell(r, 13).value = names_all_sites(row_num, "Email_Box")
                ws.cell(r, 14).value = names_all_sites(row_num, "Misc_Tasks")
                row_num += 1

        # simple colouring by column meaning (not dependent on formulas)
        col_fills = {
            "FD_": PatternFill("solid", fgColor=ROLE_COLORS.get("FrontDesk_SLGP","FFF2CC")),
            "Triage_": PatternFill("solid", fgColor=ROLE_COLORS.get("Triage_Admin_SLGP","D9EAD3")),
            "Phones": PatternFill("solid", fgColor=ROLE_COLORS.get("Phones","C9DAF8")),
            "Bookings": PatternFill("solid", fgColor=ROLE_COLORS.get("Bookings","FCE5CD")),
            "EMIS": PatternFill("solid", fgColor=ROLE_COLORS.get("EMIS","EAD1DC")),
            "Docman": PatternFill("solid", fgColor=ROLE_COLORS.get("Docman","D0E0E3")),
            "Awaiting": PatternFill("solid", fgColor=ROLE_COLORS.get("Awaiting_PSA_Admin","D0E0E3")),
            "Email": PatternFill("solid", fgColor=ROLE_COLORS.get("Email_Box","CFE2F3")),
            "Misc": PatternFill("solid", fgColor=ROLE_COLORS.get("Misc_Tasks","EFEFEF")),
        }
        for cc in range(3, ws.max_column+1):
            header = str(ws.cell(1,cc).value or "")
            fill = None
            for k,v in col_fills.items():
                if header.startswith(k) or header == k:
                    fill = v
                    break
            if fill:
                for rr in range(2, ws.max_row+1):
                    ws.cell(rr,cc).fill = fill
                    ws.cell(rr,cc).alignment = Alignment(wrap_text=True, vertical="top")

        try:
            apply_day_borders(ws)
        except Exception:
            pass
        return ws

        # Colour entire task columns to make it readable at a glance
        col_fills = {
            "FD_": PatternFill("solid", fgColor=ROLE_COLORS.get("FrontDesk_SLGP","FFF2CC")),
            "Triage_": PatternFill("solid", fgColor=ROLE_COLORS.get("Triage_Admin_SLGP","D9EAD3")),
            "Phones": PatternFill("solid", fgColor=ROLE_COLORS.get("Phones","C9DAF8")),
            "Bookings": PatternFill("solid", fgColor=ROLE_COLORS.get("Bookings","FCE5CD")),
            "EMIS": PatternFill("solid", fgColor=ROLE_COLORS.get("EMIS","EAD1DC")),
            "Docman": PatternFill("solid", fgColor=ROLE_COLORS.get("Docman","D0E0E3")),
            "Awaiting": PatternFill("solid", fgColor=ROLE_COLORS.get("Awaiting_PSA_Admin","D0E0E3")),
            "Email": PatternFill("solid", fgColor=ROLE_COLORS.get("Email_Box","CFE2F3")),
            "Misc": PatternFill("solid", fgColor="EFEFEF"),
        }
        for c in range(3, ws.max_column + 1):
            hdr = str(ws.cell(1,c).value or "")
            fill = None
            for k, v in col_fills.items():
                if hdr.startswith(k) or hdr == k:
                    fill = v
                    break
            if fill:
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(r,c)
                    cell.fill = fill
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.border = CELL_BORDER

        apply_day_separators(ws)
        return ws

    def build_dashboard(week_num: int, totals_ws, tpl_week_targets):
        ws = wb.create_sheet(f"Week{week_num}_Dashboard")
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 10

        ws["A1"] = "Coverage Dashboard"
        ws["A1"].font = Font(bold=True, size=14)

        # Locate columns in totals sheet
        header = {totals_ws.cell(1,c).value: c for c in range(1, totals_ws.max_column+1)}
        def col_letter(name):
            return get_column_letter(header[name])

        last = totals_ws.max_row
        # Achieved hours sums
        def sum_col(colname):
            col = col_letter(colname)
            return f"=SUM({totals_ws.title}!{col}2:{col}{last})"

        # Targets
        book_t = float(tpl_week_targets.get("Bookings", 0.0) or 0.0)
        emis_t = float(tpl_week_targets.get("EMIS", 0.0) or 0.0)
        doc_t  = float(tpl_week_targets.get("Docman", 0.0) or 0.0)

        ws.append(["Metric","Achieved","Target","%"])
        for c in ws[2]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")

        ws.append(["Bookings hours", sum_col("Bookings"), book_t, f"=IF(C3=0,1,B3/C3)"])
        ws.append(["EMIS hours", sum_col("EMIS"), emis_t, f"=IF(C4=0,1,B4/C4)"])
        ws.append(["Docman hours", sum_col("Docman"), doc_t, f"=IF(C5=0,1,B5/C5)"])

        # Data bars on %
        ws.conditional_formatting.add("D3:D5", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="63C384"))

        for r in range(2, ws.max_row+1):
            for c in range(1,5):
                ws.cell(r,c).border = CELL_BORDER
                ws.cell(r,c).alignment = Alignment(vertical="center")
        return ws

    for w in range(weeks):
        week_num = w + 1
        wk_start = start_monday + timedelta(days=7 * w)
        a, breaks, gaps, dates, slots, hours_map = schedule_week(tpl, wk_start)

        # Build site timelines (only site staff)
        site_map = {}  # name -> (sheet_name, col_letter)
        site_staff_map = {}
        for site in SITES:
            ws_site, site_staff = make_site_timeline(week_num, site, dates, slots, a, breaks, hours_map)
            if ws_site is None:
                continue
            site_staff_map[site] = site_staff
            for idx, nm in enumerate(site_staff):
                col = get_column_letter(3 + idx)
                site_map[nm] = (ws_site.title, col)

        # Dynamic totals based on site sheet columns
        max_row = 1 + slot_rows_count(dates, slots)
        totals_ws = build_dynamic_totals(week_num, site_map, max_row)

        # Coverage by slot by site (static)
        build_coverage_by_slot_by_site(week_num, dates, slots)

        # Dashboard
        build_dashboard(week_num, totals_ws, tpl.weekly_targets)

        # Notes / gaps
        ws_g = wb.create_sheet(f"Week{week_num}_NotesAndGaps")
        ws_g.append(["Date","Time","Task","Note"])
        for c in ws_g[1]:
            c.font = Font(bold=True)
        for d, t, task, note in gaps:
            ws_g.append([d.isoformat(), t.strftime("%H:%M") if t else "", task, note])

    return wb