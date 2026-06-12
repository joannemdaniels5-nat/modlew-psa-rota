import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

import io
import re
from copy import copy
from datetime import date, datetime, time, timedelta
from typing import Iterable, Optional, Tuple

import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from rota_engine_v37_19_TARGETS_FINAL import (
    read_template,
    build_workbook,
    ensure_monday,
    recalc_workbook_from_site_timelines,
)

# -----------------------------------------------------------------------------
# App-level repair helpers
# -----------------------------------------------------------------------------
# These sit around the existing rota engine so the app can correct common output
# problems without needing to rewrite the underlying engine.

LEAVE_WORDS = (
    "annual leave",
    "ann leave",
    "a/l",
    "al",
    "holiday",
    "leave",
    "annual",
)

DUTY_WORDS = (
    "phones",
    "phone",
    "reception",
    "booking",
    "bookings",
    "email",
    "emails",
    "docman",
    "triage",
    "admin",
    "triage admin",
    "triage/admin",
    "workflow",
    "lunch",
    "break",
)

DAY_WORDS = {
    "mon": 0,
    "monday": 0,
    "tue": 1,
    "tues": 1,
    "tuesday": 1,
    "wed": 2,
    "weds": 2,
    "wednesday": 2,
    "thu": 3,
    "thur": 3,
    "thurs": 3,
    "thursday": 3,
    "fri": 4,
    "friday": 4,
    "sat": 5,
    "saturday": 5,
    "sun": 6,
    "sunday": 6,
}


def _text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm(value) -> str:
    return re.sub(r"\s+", " ", _text(value).lower()).strip()


def _looks_like_leave(value) -> bool:
    s = _norm(value)
    if not s:
        return False
    if s in {"al", "a/l"}:
        return True
    return any(word in s for word in LEAVE_WORDS)


def _looks_like_duty(value) -> bool:
    s = _norm(value)
    return bool(s) and any(word in s for word in DUTY_WORDS)


def _cell_date(value) -> Optional[date]:
    """Return a date from common Excel/Python/string formats, or None."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, time):
        return None
    if value is None:
        return None

    s = str(value).strip()
    if not s:
        return None

    # Try exact/common UK-ish formats first.
    for fmt in (
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%Y-%m-%d",
        "%d %b %Y",
        "%d %B %Y",
        "%a %d %b %Y",
        "%A %d %B %Y",
        "%A %d %b %Y",
    ):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass

    # Catch strings like "Tuesday 02/06" by adding the selected year later is risky,
    # so only parse values with an explicit year here.
    if re.search(r"\b\d{4}\b", s):
        try:
            from dateutil import parser
            return parser.parse(s, dayfirst=True, fuzzy=True).date()
        except Exception:
            return None
    return None


def _cell_time(value) -> Optional[time]:
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    s = _norm(value).replace(".", ":")
    if not s:
        return None
    if s in {"8", "8am", "08", "08am"}:
        return time(8, 0)
    m = re.search(r"\b(\d{1,2})(?::(\d{2}))?\s*(am|pm)?\b", s)
    if not m:
        return None
    hour = int(m.group(1))
    minute = int(m.group(2) or 0)
    ampm = m.group(3)
    if ampm == "pm" and hour != 12:
        hour += 12
    if ampm == "am" and hour == 12:
        hour = 0
    if 0 <= hour <= 23 and 0 <= minute <= 59:
        return time(hour, minute)
    return None


def _week_dates(start_monday: date, weeks: int) -> set[date]:
    return {start_monday + timedelta(days=i) for i in range(max(1, int(weeks)) * 7)}


def _date_range(start_monday: date, weeks: int) -> Tuple[date, date]:
    days = _week_dates(start_monday, weeks)
    return min(days), max(days)


def _row_dates(ws, row_idx: int) -> list[date]:
    return [d for d in (_cell_date(c.value) for c in ws[row_idx]) if d]


def _find_nearest_date(ws, row: int, col: int, max_scan: int = 12) -> Optional[date]:
    """Find the closest date heading above/left of a cell."""
    for r in range(row, max(1, row - max_scan) - 1, -1):
        d = _cell_date(ws.cell(r, col).value)
        if d:
            return d
    for c in range(col, max(1, col - max_scan) - 1, -1):
        d = _cell_date(ws.cell(row, c).value)
        if d:
            return d
    return None


def _copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def filter_template_leave_bytes(xlsx_bytes: bytes, start_monday: date, weeks: int) -> bytes:
    """Remove old/out-of-range leave rows from obvious leave/absence sheets.

    This is deliberately conservative: it only edits sheets whose title suggests
    leave/absence/holiday. Rows with leave text and date(s) fully outside the
    requested rota period are deleted so old leave does not bleed into the rota.
    """
    start_date, end_date = _date_range(start_monday, weeks)
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    changed = False

    for ws in wb.worksheets:
        title = _norm(ws.title)
        if not any(key in title for key in ("leave", "absence", "holiday", "annual")):
            continue

        rows_to_delete = []
        for row in ws.iter_rows():
            row_idx = row[0].row
            values = [c.value for c in row]
            text_values = [_norm(v) for v in values if v not in (None, "")]
            row_text = " ".join(text_values)

            # Keep obvious header / instruction rows.
            if row_idx == 1 or any(h in row_text for h in ("name", "staff", "start", "end", "date", "note", "reason")) and not any(_cell_date(v) for v in values):
                continue

            dates = [d for d in (_cell_date(v) for v in values) if d]
            if not dates:
                continue

            # In Leave/Absence sheets, date rows outside the requested rota period are stale
            # even when the reason column is blank or formatted differently.
            row_start = min(dates)
            row_end = max(dates)
            no_overlap = row_end < start_date or row_start > end_date
            if no_overlap:
                rows_to_delete.append(row_idx)

        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx, 1)
            changed = True

    if not changed:
        return xlsx_bytes
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def normalise_leave_text_in_template(xlsx_bytes: bytes) -> bytes:
    """Normalise common leave variants so the engine has a better chance of reading them."""
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    changed = False

    for ws in wb.worksheets:
        if not any(key in _norm(ws.title) for key in ("leave", "absence", "holiday", "annual")):
            continue
        for row in ws.iter_rows():
            for cell in row:
                s = _norm(cell.value)
                if s in {"al", "a/l", "ann leave", "annual", "holiday"}:
                    cell.value = "Annual Leave"
                    changed = True

    if not changed:
        return xlsx_bytes
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def clean_generated_out_of_week_leave(wb, start_monday: date, weeks: int) -> int:
    """Remove leaked annual leave markers if they sit under/against a date outside the requested period."""
    valid_dates = _week_dates(start_monday, weeks)
    cleared = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not _looks_like_leave(cell.value):
                    continue
                nearest = _find_nearest_date(ws, cell.row, cell.column)
                if nearest and nearest not in valid_dates:
                    cell.value = None
                    cleared += 1
    return cleared


def fix_export_row_heights(wb, min_height: float = 18.0, eight_am_height: float = 24.0) -> int:
    """Stop Tuesday onward 08:00 rows exporting too narrow by enforcing readable heights."""
    fixed = 0
    for ws in wb.worksheets:
        for row_idx in range(1, ws.max_row + 1):
            values = [ws.cell(row_idx, c).value for c in range(1, min(ws.max_column, 12) + 1)]
            has_content = any(v not in (None, "") for v in values)
            has_8am = any(_cell_time(v) == time(8, 0) for v in values)

            current = ws.row_dimensions[row_idx].height
            target = eight_am_height if has_8am else (min_height if has_content else None)
            if target and (current is None or current < target):
                ws.row_dimensions[row_idx].height = target
                fixed += 1
    return fixed


def parse_bgs_manual_entries(raw: str, start_monday: date, weeks: int) -> list[dict]:
    """Parse simple manual BGS triage-admin lines.

    Accepted examples:
      Jane Smith
      Jane Smith | Tuesday | 08:00-12:00
      Jane Smith | Tue, Wed, Thu | 8-12
    If no day is supplied the entry applies to each weekday in the generated week(s).
    """
    entries = []
    if not raw or not raw.strip():
        return entries

    default_days = [start_monday + timedelta(days=i) for i in range(max(1, weeks) * 7) if i % 7 < 5]

    for line in raw.splitlines():
        line = line.strip(" -\t")
        if not line:
            continue
        parts = [p.strip() for p in re.split(r"\s*\|\s*", line) if p.strip()]
        name = parts[0]
        days = default_days
        hours = ""

        if len(parts) >= 2:
            day_tokens = re.split(r"[,;/]+|\band\b", parts[1], flags=re.IGNORECASE)
            parsed_days = []
            for token in day_tokens:
                token_norm = _norm(token)
                if token_norm in DAY_WORDS:
                    offset = DAY_WORDS[token_norm]
                    for wk in range(max(1, weeks)):
                        parsed_days.append(start_monday + timedelta(days=(wk * 7) + offset))
                else:
                    d = _cell_date(token)
                    if d:
                        parsed_days.append(d)
            if parsed_days:
                valid = _week_dates(start_monday, weeks)
                days = [d for d in parsed_days if d in valid]
            else:
                hours = parts[1]

        if len(parts) >= 3:
            hours = parts[2]

        for d in days:
            entries.append({"name": name, "date": d, "hours": hours or "Manual BGS triage admin"})
    return entries


def add_manual_bgs_sheet(wb, entries: list[dict]) -> None:
    if not entries:
        return
    sheet_name = "Manual BGS Triage Admin"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(["Site", "Date", "Day", "Name", "Hours / note", "Purpose"])
    header_fill = PatternFill("solid", fgColor="1F7A3F")
    white_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="6AA84F")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for e in entries:
        ws.append(["BGS", e["date"], e["date"].strftime("%A"), e["name"], e["hours"], "Manual entry - BGS triage admin not pulled through by engine"])
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = [10, 13, 14, 24, 24, 48][col - 1]
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 2).number_format = "dd/mm/yyyy"
        for col in range(1, 7):
            ws.cell(row, col).alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def try_insert_bgs_entries_into_timeline(wb, entries: list[dict]) -> int:
    """Best-effort insert into BGS sheets if a triage-admin row exists.

    The engine layout may vary, so this avoids destructive edits. It fills blank cells
    under matching dates on rows labelled triage/admin, and otherwise leaves the
    dedicated manual sheet as the auditable source.
    """
    if not entries:
        return 0
    inserted = 0
    entries_by_date = {}
    for e in entries:
        entries_by_date.setdefault(e["date"], []).append(e)

    for ws in wb.worksheets:
        if "bgs" not in _norm(ws.title) and "bellingham" not in _norm(ws.title):
            continue

        # Date columns are usually in the first few header rows.
        date_cols = {}
        for r in range(1, min(ws.max_row, 12) + 1):
            for c in range(1, ws.max_column + 1):
                d = _cell_date(ws.cell(r, c).value)
                if d:
                    date_cols[d] = c

        if not date_cols:
            continue

        triage_rows = []
        for r in range(1, ws.max_row + 1):
            row_text = " ".join(_norm(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 8) + 1))
            if "triage" in row_text and "admin" in row_text:
                triage_rows.append(r)

        if not triage_rows:
            continue

        for d, day_entries in entries_by_date.items():
            col = date_cols.get(d)
            if not col:
                continue
            text = "\n".join(f"{e['name']} ({e['hours']})" for e in day_entries)
            for r in triage_rows:
                cell = ws.cell(r, col)
                if cell.value in (None, ""):
                    cell.value = text
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
                    inserted += len(day_entries)
                    break
    return inserted


def post_process_workbook(wb, start_monday: date, weeks: int, bgs_entries: list[dict], remove_old_leave: bool = True, fix_heights: bool = True) -> dict:
    stats = {
        "old_leave_cells_cleared": 0,
        "row_heights_fixed": 0,
        "bgs_entries_inserted": 0,
        "bgs_entries_added_to_manual_sheet": len(bgs_entries),
    }
    if remove_old_leave:
        stats["old_leave_cells_cleared"] = clean_generated_out_of_week_leave(wb, start_monday, weeks)
    stats["bgs_entries_inserted"] = try_insert_bgs_entries_into_timeline(wb, bgs_entries)
    add_manual_bgs_sheet(wb, bgs_entries)
    if fix_heights:
        stats["row_heights_fixed"] = fix_export_row_heights(wb)
    return stats


def load_recalc_bytes_as_workbook(out_bytes_or_wb):
    """The engine has returned bytes in previous versions, but this keeps it safe if it changes."""
    if hasattr(out_bytes_or_wb, "save"):
        return out_bytes_or_wb
    return load_workbook(io.BytesIO(out_bytes_or_wb))


# -----------------------------------------------------------------------------
# Streamlit UI
# -----------------------------------------------------------------------------
st.set_page_config(page_title="ModLew PSA Rota Generator", page_icon="🗓️", layout="wide")
st.title("ModLew PSA Rota Generator")

st.markdown("### 1) Generate rota")
uploaded = st.file_uploader("Upload input template (.xlsx)", type=["xlsx"], key="tpl")
weeks = st.number_input("Weeks", min_value=1, max_value=8, value=1, step=1)
wk_start = st.date_input("Week commencing (any date in that week)", value=date.today())

with st.expander("Fixes / manual additions", expanded=True):
    col1, col2 = st.columns(2)
    with col1:
        filter_old_leave = st.checkbox(
            "Filter annual leave to the requested week(s) only",
            value=True,
            help="Stops old annual leave rows being pulled through when generating a different week.",
        )
        normalise_leave = st.checkbox(
            "Normalise annual leave wording before generation",
            value=True,
            help="Treats AL, A/L, Ann Leave and Holiday as Annual Leave where they appear on leave sheets.",
        )
    with col2:
        fix_heights = st.checkbox(
            "Fix export row heights, including 08:00 rows",
            value=True,
            help="Prevents Tuesday onward 08:00 rows exporting too narrow to read.",
        )

    bgs_manual_raw = st.text_area(
        "Optional manual BGS triage admin override",
        placeholder="Examples:\nJane Smith\nJane Smith | Tuesday | 08:00-12:00\nJane Smith | Tue, Wed, Thu | 8-12",
        height=110,
        help="One entry per line. If you only enter a name, it is added for each weekday in the generated rota period.",
    )

if st.button("Generate rota", type="primary", disabled=(uploaded is None)):
    try:
        start_monday = ensure_monday(wk_start)
        raw_template_bytes = uploaded.getvalue()
        template_bytes = raw_template_bytes

        if normalise_leave:
            template_bytes = normalise_leave_text_in_template(template_bytes)
        if filter_old_leave:
            template_bytes = filter_template_leave_bytes(template_bytes, start_monday, int(weeks))

        tpl = read_template(template_bytes)
        wb = build_workbook(tpl, start_monday, int(weeks))

        bgs_entries = parse_bgs_manual_entries(bgs_manual_raw, start_monday, int(weeks))
        stats = post_process_workbook(
            wb,
            start_monday,
            int(weeks),
            bgs_entries,
            remove_old_leave=filter_old_leave,
            fix_heights=fix_heights,
        )

        bio = io.BytesIO()
        wb.save(bio)
        st.success("Rota generated.")
        with st.expander("Generation fixes applied"):
            st.write(stats)
        st.download_button(
            "Download rota (.xlsx)",
            data=bio.getvalue(),
            file_name=f"rota_{start_monday.isoformat()}_{int(weeks)}w.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        st.exception(e)

st.markdown("---")
st.markdown("### 2) Recalculate (after you edit the site timelines)")
st.caption("Upload the generated rota after edits. This rebuilds Coverage + Totals from the edited timelines and re-applies colours.")
uploaded2 = st.file_uploader("Upload edited rota (.xlsx)", type=["xlsx"], key="edited")

with st.expander("Recalculate fixes / manual additions", expanded=False):
    recalc_week_start = st.date_input("Recalculation week commencing", value=date.today(), key="recalc_wk")
    recalc_weeks = st.number_input("Recalculation weeks", min_value=1, max_value=8, value=1, step=1, key="recalc_weeks")
    recalc_filter_old_leave = st.checkbox("Clean old annual leave after recalculation", value=True, key="recalc_filter_old_leave")
    recalc_fix_heights = st.checkbox("Fix export row heights after recalculation", value=True, key="recalc_fix_heights")
    recalc_bgs_manual_raw = st.text_area(
        "Optional manual BGS triage admin override after recalculation",
        placeholder="Jane Smith | Tuesday | 08:00-12:00",
        height=90,
        key="recalc_bgs_manual_raw",
    )

if st.button("Recalculate workbook", disabled=(uploaded2 is None)):
    try:
        start_monday = ensure_monday(recalc_week_start)
        out = recalc_workbook_from_site_timelines(uploaded2.getvalue())
        wb = load_recalc_bytes_as_workbook(out)
        bgs_entries = parse_bgs_manual_entries(recalc_bgs_manual_raw, start_monday, int(recalc_weeks))
        stats = post_process_workbook(
            wb,
            start_monday,
            int(recalc_weeks),
            bgs_entries,
            remove_old_leave=recalc_filter_old_leave,
            fix_heights=recalc_fix_heights,
        )

        bio = io.BytesIO()
        wb.save(bio)
        st.success("Recalculated.")
        with st.expander("Recalculation fixes applied"):
            st.write(stats)
        st.download_button(
            "Download recalculated workbook (.xlsx)",
            data=bio.getvalue(),
            file_name="rota_recalculated.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        st.exception(e)
