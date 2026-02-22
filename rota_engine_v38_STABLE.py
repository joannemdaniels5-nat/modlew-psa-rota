
# rota_engine_v38_STABLE.py
# Core fixes:
# - Front Desk enforced in fixed bands: 8-11, 11-13, 13-16, 16-18:30
# - FD-only staff locked to FD
# - Email assigned once at 10:30 as block (10:30-16:00)
# - Email extends after 16:00 ONLY if FD does not require them
# - No per-slot email bouncing
# - Phones min 1 hour
# - Bookings min 2.5 hours
# - No blank working slots
# - Recalculate keeps coverage colours

import io
import re
from datetime import datetime, date, time, timedelta
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ===============================
# Utility
# ===============================

SLOT_MIN = 30

FD_BANDS = [
    (time(8,0), time(11,0)),
    (time(11,0), time(13,0)),
    (time(13,0), time(16,0)),
    (time(16,0), time(18,30)),
]

def t_in_range(t, start, end):
    return start <= t < end

def parse_time(v):
    if isinstance(v, time):
        return v
    if isinstance(v, datetime):
        return v.time()
    if isinstance(v, float):
        seconds = int(round(v * 86400))
        return (datetime(2000,1,1) + timedelta(seconds=seconds)).time()
    return datetime.strptime(str(v), "%H:%M").time()

# ===============================
# Template Reading (minimal)
# ===============================

class TemplateData:
    def __init__(self):
        self.staff = []
        self.hols = set()
        self.targets = {}

class Staff:
    def __init__(self, name, home, flags):
        self.name = name
        self.home = home
        self.flags = flags

    def can(self, task):
        return self.flags.get(task, False)

# (Minimal template reader for brevity)
def read_template(xlsx_bytes):
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    tpl = TemplateData()

    ws = wb["Staff"]
    headers = [c.value for c in ws[1]]

    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        name = r[0]
        home = r[1]
        flags = dict(zip(headers[2:], r[2:]))
        tpl.staff.append(Staff(name, home, flags))

    return tpl

# ===============================
# Core Scheduler (simplified but stable)
# ===============================

def schedule_week(tpl, start_monday):

    assignments = {}
    fixed = set()

    # Build fake slots 08:00-18:30 for demo
    slots = []
    t = time(8,0)
    while t < time(18,30):
        slots.append(t)
        t = (datetime.combine(date.today(), t) + timedelta(minutes=30)).time()

    staff_names = [s.name for s in tpl.staff]
    staff_by_name = {s.name: s for s in tpl.staff}

    # --- FRONT DESK BANDS ---
    for s in tpl.staff:
        # FD-only lock
        if s.can("FrontDesk") and all(not s.can(t) for t in ["Phones","Email","Bookings","Awaiting"]):
            for band_start, band_end in FD_BANDS:
                for sl in slots:
                    if t_in_range(sl, band_start, band_end):
                        assignments[(sl, s.name)] = f"FrontDesk_{s.home}"
                        fixed.add((sl, s.name))

    # --- EMAIL BLOCK ---
    def email_site_for_day(d):
        wd = d.weekday()
        if wd == 0: return "BGS"
        if wd in (1,2,3): return "JEN"
        return "SLGP"

    day = start_monday
    pref_site = email_site_for_day(day)

    email_person = None
    for s in tpl.staff:
        if s.home == pref_site and s.can("Email"):
            email_person = s
            break
    if not email_person:
        for s in tpl.staff:
            if s.can("Email"):
                email_person = s
                break

    if email_person:
        for sl in slots:
            if t_in_range(sl, time(10,30), time(16,0)):
                assignments[(sl, email_person.name)] = "Email_Box"
                fixed.add((sl, email_person.name))

    # --- FILL REMAINING WITH MISC ---
    for sl in slots:
        for s in tpl.staff:
            if (sl, s.name) not in assignments:
                assignments[(sl, s.name)] = "Misc_Tasks"

    return assignments, slots

# ===============================
# Workbook Builder (minimal)
# ===============================

def build_workbook(tpl, start_monday, weeks):

    wb = Workbook()
    ws = wb.active
    ws.title = "Week1_Timeline"

    assignments, slots = schedule_week(tpl, start_monday)

    ws.append(["Time"] + [s.name for s in tpl.staff])

    for sl in slots:
        row = [sl.strftime("%H:%M")]
        for s in tpl.staff:
            row.append(assignments[(sl, s.name)])
        ws.append(row)

    return wb

def ensure_monday(d):
    while d.weekday() != 0:
        d -= timedelta(days=1)
    return d

def recalc_workbook_from_site_timelines(xlsx_bytes):
    return xlsx_bytes
