from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import io
import re
from dataclasses import dataclass
from datetime import datetime, date, time, timedelta
from typing import Dict, List, Tuple, Optional, Set

import math

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# =========================================================
# SAFE PATCH: v37_19 retained as last-known app import; BGS triage added to fixed triage, export task lists, colours/totals.
# Rota Generator Engine — v15 (Weighted, Block-Stable, No Floaters)
# =========================================================

DAY_START = time(8, 0)
DAY_END = time(18, 30)
SLOT_MIN = 30

SITES = ["SLGP", "JEN", "BGS"]

FD_BANDS = [
    (time(8, 0), time(11, 0)),
    (time(11, 0), time(13, 0)),
    (time(13, 0), time(16, 0)),
    (time(16, 0), time(18, 30)),
]
FD_BANDS_FLEX = [
    (time(8, 0), time(10, 30)),
    (time(10, 30), time(13, 0)),
    (time(13, 0), time(16, 0)),
    (time(16, 0), time(18, 30)),
]

# Site-specific fixed Front Desk bands.
# SLGP remains 08:00-11:00, 11:00-13:00, 13:00-16:00, 16:00-18:30.
# JEN/BGS use 08:00-10:30, 10:30-13:00, 13:00-16:00, 16:00-18:30.
FD_FIXED_BANDS_BY_SITE = {
    "SLGP": FD_BANDS,
    "JEN": FD_BANDS_FLEX,
    "BGS": FD_BANDS_FLEX,
}

TRIAGE_BANDS = [
    (time(8, 0), time(10, 30)),
    (time(10, 30), time(13, 0)),
    (time(13, 30), time(16, 0)),
]

BREAK_WINDOW = (time(12, 0), time(14, 0))
# Standard lunch candidates for normal day shifts.
BREAK_CANDIDATES = [time(12, 0), time(12, 30), time(13, 0), time(13, 30)]
# Later break candidates for staff who start late, e.g. 11:00/12:00 starts.
# This prevents the rota from missing breaks for people who are not working during the old 12:00-14:00 lunch window.
BREAK_CANDIDATES_LATE_START = [time(14, 0), time(14, 30), time(15, 0), time(15, 30), time(16, 0), time(16, 30)]
# Break is required for shifts of 6 hours or more.
BREAK_THRESHOLD_HOURS = 6.0

def break_candidates_for_shift(stt: time, endt: time) -> List[time]:
    """Return 30-minute break starts for a shift, ordered by preference.

    Breaks should usually sit in the middle third of the working day, but
    they must not overwrite fixed Front Desk / Triage cover. To make that
    possible, return every valid 30-minute slot in the shift, ordered with
    middle-third slots first and closest-to-midpoint before edge slots.
    """
    if not stt or not endt:
        return []

    base = datetime(2000, 1, 1, stt.hour, stt.minute)
    finish = datetime(2000, 1, 1, endt.hour, endt.minute)
    if finish <= base:
        return []

    duration = finish - base
    mid_start = base + (duration / 3)
    mid_end = base + (duration * 2 / 3)
    midpoint = base + (duration / 2)

    candidates = []
    cur = base
    while cur + timedelta(minutes=SLOT_MIN) <= finish:
        t = cur.time()
        in_middle_third = (cur >= mid_start and cur < mid_end)
        midpoint_distance = abs((cur - midpoint).total_seconds())
        edge_penalty = 0 if in_middle_third else 100_000
        candidates.append((edge_penalty, midpoint_distance, t))
        cur += timedelta(minutes=SLOT_MIN)

    candidates.sort(key=lambda x: (x[0], x[1], x[2].hour, x[2].minute))
    return [t for _, _, t in candidates]

# Block mins (slots)
MIN_PHONES = 2          # 1.5h
MAX_PHONES = 8          # 4h
MIN_DEFAULT = 5         # 2.5h
MAX_DEFAULT = 9         # 4.5h
MIN_DOCMAN = 6          # 3h

def timeslots() -> List[time]:
    cur = datetime(2000, 1, 1, DAY_START.hour, DAY_START.minute)
    end = datetime(2000, 1, 1, DAY_END.hour, DAY_END.minute)
    out: List[time] = []
    while cur < end:
        out.append(cur.time())
        cur += timedelta(minutes=SLOT_MIN)
    return out

def ensure_monday(d: date) -> date:
    return d - timedelta(days=d.weekday())

def day_name(d: date) -> str:
    return ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"][d.weekday()]

def t_in_range(t: time, a: time, b: time) -> bool:
    return (t >= a) and (t < b)

def normalize(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s).strip().lower())

def find_sheet(xls: pd.ExcelFile, candidates: List[str]) -> Optional[str]:
    names = {normalize(n): n for n in xls.sheet_names}
    for c in candidates:
        k = normalize(c)
        if k in names:
            return names[k]
    for n in xls.sheet_names:
        nn = normalize(n)
        for c in candidates:
            if normalize(c) in nn:
                return n
    return None

def pick_col(df: pd.DataFrame, candidates: List[str], required: bool=True) -> Optional[str]:
    cols = {normalize(c): c for c in df.columns}
    for cand in candidates:
        k = normalize(cand)
        if k in cols:
            return cols[k]
    for c in df.columns:
        nc = normalize(c)
        for cand in candidates:
            if normalize(cand) in nc:
                return c
    if required:
        raise KeyError(f"Missing required column among {candidates}. Available: {list(df.columns)}")
    return None

def to_time(x):
    if pd.isna(x) or x == "":
        return None
    if isinstance(x, time):
        return x
    if isinstance(x, datetime):
        return x.time()
    if isinstance(x, (float, int)):
        seconds = int(round(float(x) * 86400))
        return (datetime(2000,1,1) + timedelta(seconds=seconds)).time()
    return pd.to_datetime(str(x)).time()

def to_date(x):
    if pd.isna(x) or x == "":
        return None
    if isinstance(x, (datetime, pd.Timestamp)):
        return x.date()
    if isinstance(x, date):
        return x
    return pd.to_datetime(x).date()

def dt_of(d: date, t: time) -> datetime:
    return datetime(d.year, d.month, d.day, t.hour, t.minute)

def add_minutes(t: time, mins: int) -> time:
    return (datetime(2000,1,1,t.hour,t.minute) + timedelta(minutes=mins)).time()

def yn(v) -> bool:
    if pd.isna(v):
        return False
    s = str(v).strip().lower()
    return s in {"y","yes","true","1","t"}

@dataclass
class Staff:
    name: str
    home: str
    can_frontdesk: bool
    can_triage: bool
    can_email: bool
    can_phones: bool
    can_bookings: bool
    can_emis: bool
    can_docman: bool
    weights: Dict[str, int]
    frontdesk_only: bool
    break_required: bool

@dataclass
class TemplateData:
    staff: List[Staff]
    hours_map: Dict[str, Dict[str, Optional[time]]]
    hols: List[Tuple[str, date, date, str]]
    call_handlers: pd.DataFrame
    handler_leave: pd.DataFrame
    phones_targets: Dict[Tuple[str, time], int]
    bookings_targets: Dict[Tuple[str, time], int]
    weekly_targets: Dict[str, float]
    swaps: List[Tuple[date, str, str, Optional[time], Optional[time]]]
    buddies: Dict[str, str]

def read_template(uploaded_bytes: bytes) -> TemplateData:
    xls = pd.ExcelFile(io.BytesIO(uploaded_bytes))

    staff_sheet = find_sheet(xls, ["Staff"])
    hours_sheet = find_sheet(xls, ["WorkingHours", "Hours"])
    hols_sheet = find_sheet(xls, ["Holidays", "Leave", "Absence"])
    callh_sheet = find_sheet(xls, ["CallHandlers", "Call Handlers"])
    hleave_sheet = find_sheet(xls, ["Handler_Leave", "Handler Leave", "CallHandler_Leave"])
    tph_sheet = find_sheet(xls, ["Targets_Phones_Hourly", "PhonesTargets", "Targets Phones Hourly"])
    tbk_sheet = find_sheet(xls, ["Targets_Bookings_Hourly", "BookingsTargets", "Targets Bookings Hourly"])
    tweek_sheet = find_sheet(xls, ["Targets_Weekly", "Targets Weekly"])
    swaps_sheet = find_sheet(xls, ["Swaps"])
    new_sheet = find_sheet(xls, ["NewStarters", "New Starters"])

    if not staff_sheet or not hours_sheet:
        raise ValueError(f"Missing Staff/WorkingHours sheets. Found: {xls.sheet_names}")

    staff_df = pd.read_excel(xls, sheet_name=staff_sheet)
    hours_df = pd.read_excel(xls, sheet_name=hours_sheet)
    hols_df = pd.read_excel(xls, sheet_name=hols_sheet) if hols_sheet else pd.DataFrame()
    callh_df = pd.read_excel(xls, sheet_name=callh_sheet) if callh_sheet else pd.DataFrame()
    hleave_df = pd.read_excel(xls, sheet_name=hleave_sheet) if hleave_sheet else pd.DataFrame()
    tph_df = pd.read_excel(xls, sheet_name=tph_sheet) if tph_sheet else pd.DataFrame()
    tbk_df = pd.read_excel(xls, sheet_name=tbk_sheet) if tbk_sheet else pd.DataFrame()
    tweek_df = pd.read_excel(xls, sheet_name=tweek_sheet) if tweek_sheet else pd.DataFrame()
    swaps_df = pd.read_excel(xls, sheet_name=swaps_sheet) if swaps_sheet else pd.DataFrame()
    new_df = pd.read_excel(xls, sheet_name=new_sheet) if new_sheet else pd.DataFrame()

    # --- Staff
    name_c = pick_col(staff_df, ["Name","StaffName"])
    home_c = pick_col(staff_df, ["HomeSite","Site","BaseSite"], required=False)

    staff_df = staff_df.copy()
    staff_df["Name"] = staff_df[name_c].astype(str).str.strip()
    staff_df["HomeSite"] = staff_df[home_c].astype(str).str.strip().str.upper() if home_c else ""

    def bcol(cands, default=False):
        c = pick_col(staff_df, cands, required=False)
        if not c:
            return pd.Series([default]*len(staff_df))
        if staff_df[c].dtype == bool:
            return staff_df[c].fillna(default)
        return staff_df[c].apply(yn)

    staff_df["CanFrontDesk"] = bcol(["CanFrontDesk"])
    staff_df["CanTriage"] = bcol(["CanTriage"])
    staff_df["CanEmail"] = bcol(["CanEmail"])
    staff_df["CanPhones"] = bcol(["CanPhones"])
    staff_df["CanBookings"] = bcol(["CanBookings"])
    staff_df["CanEMIS"] = bcol(["CanEMIS"])
    staff_df["CanDocman"] = bcol(["CanDocman_PSA"]) | bcol(["CanDocman_AWAIT"]) | bcol(["CanDocman"])

    staff_df["FrontDeskOnly"] = bcol(["FrontDeskOnly"], default=False)
    staff_df["BreakRequired"] = bcol(["BreakRequired"], default=True)

    weight_cols = {
        "FrontDesk":"FrontDeskWeight",
        "Triage":"TriageWeight",
        "Email":"EmailWeight",
        "Phones":"PhonesWeight",
        "Bookings":"BookingsWeight",
        "EMIS":"EmisWeight",
        "Docman":"DocmanWeight",
        "Awaiting":"AwaitingWeight",
        "Misc":"MiscWeight",
    }

    staff_list: List[Staff] = []
    for _, r in staff_df.iterrows():
        weights: Dict[str,int] = {}
        for k, col in weight_cols.items():
            v = r.get(col, 3)
            try:
                if pd.isna(v) or v == "":
                    v = 3
                v = int(float(v))
            except Exception:
                v = 3
            weights[k] = max(0, min(5, v))
        staff_list.append(
            Staff(
                name=str(r["Name"]).strip(),
                home=str(r.get("HomeSite","")).strip().upper(),
                can_frontdesk=bool(r.get("CanFrontDesk", False)),
                can_triage=bool(r.get("CanTriage", False)),
                can_email=bool(r.get("CanEmail", False)),
                can_phones=bool(r.get("CanPhones", False)),
                can_bookings=bool(r.get("CanBookings", False)),
                can_emis=bool(r.get("CanEMIS", False)),
                can_docman=bool(r.get("CanDocman", False)),
                weights=weights,
                frontdesk_only=bool(r.get("FrontDeskOnly", False)),
                break_required=bool(r.get("BreakRequired", True)),
            )
        )

    # --- Working hours
    hours_df = hours_df.copy()
    hn = pick_col(hours_df, ["Name","StaffName"])
    hs = pick_col(hours_df, ["HomeSite","Site","BaseSite"], required=False)
    hours_df["Name"] = hours_df[hn].astype(str).str.strip()
    hours_df["HomeSite"] = hours_df[hs].astype(str).str.strip().str.upper() if hs else ""

    for dn in ["Mon","Tue","Wed","Thu","Fri"]:
        sc = pick_col(hours_df, [f"{dn}Start", f"{dn} Start", f"{dn}_Start"], required=False)
        ec = pick_col(hours_df, [f"{dn}End", f"{dn} End", f"{dn}_End"], required=False)
        hours_df[f"{dn}Start"] = hours_df[sc].apply(to_time) if sc else None
        hours_df[f"{dn}End"] = hours_df[ec].apply(to_time) if ec else None

    hours_map = {}
    for _, r in hours_df.iterrows():
        hours_map[str(r["Name"]).strip()] = {k: r.get(k) for k in hours_df.columns}

    # --- Align WorkingHours names to Staff names (common cause of 'blank working hours')
    # If a staff name in Staff sheet doesn't exactly match the WorkingHours sheet (extra spaces, case),
    # we map by normalized name so scheduling still finds their shift.
    norm_hours = {normalize(k): v for k, v in hours_map.items()}
    for st in staff_list:
        if st.name not in hours_map:
            nk = normalize(st.name)
            if nk in norm_hours:
                hours_map[st.name] = norm_hours[nk]
# --- Holidays ranges
    hols: List[Tuple[str,date,date,str]] = []
    if hols_df is not None and not hols_df.empty:
        ncol = pick_col(hols_df, ["Name","StaffName"], required=False) or hols_df.columns[0]
        sdcol = pick_col(hols_df, ["StartDate","Start"], required=False) or hols_df.columns[1]
        edcol = pick_col(hols_df, ["EndDate","End"], required=False) or hols_df.columns[2]
        notes_c = pick_col(hols_df, ["Notes","Note","Reason"], required=False)

        for _, r in hols_df.iterrows():
            nm = str(r.get(ncol,"")).strip()
            sd = to_date(r.get(sdcol))
            ed = to_date(r.get(edcol))
            note = "" if (not notes_c or pd.isna(r.get(notes_c))) else str(r.get(notes_c)).strip().lower()
            kind = "Holiday"
            if "sick" in note or "sickness" in note:
                kind = "Sick"
            elif "bank" in note:
                kind = "Bank Holiday"
            if nm and sd and ed:
                hols.append((nm, sd, ed, kind))

    # --- Targets
    def parse_hourly(df: pd.DataFrame) -> Dict[Tuple[str, time], int]:
        out: Dict[Tuple[str,time], int] = {}
        if df is None or df.empty:
            return out
        time_col = pick_col(df, ["Time"], required=False) or df.columns[0]
        ddf = df.copy()
        ddf["Time"] = ddf[time_col].apply(to_time)
        for dn in ["Mon","Tue","Wed","Thu","Fri"]:
            if dn not in ddf.columns:
                continue
            for _, r in ddf.iterrows():
                hh = r.get("Time")
                if not hh:
                    continue
                val = r.get(dn)
                if pd.isna(val) or val == "":
                    continue
                out[(dn, time(hh.hour, 0))] = int(float(val))
        return out

    phones_targets = parse_hourly(tph_df)
    bookings_targets = parse_hourly(tbk_df)

    weekly_targets = {"Bookings": 0.0, "EMIS": 0.0, "Docman": 0.0}
    if tweek_df is not None and not tweek_df.empty:
        task_c = pick_col(tweek_df, ["Task"], required=False) or tweek_df.columns[0]
        val_c = pick_col(tweek_df, ["WeekHoursTarget","Target","Hours"], required=False) or tweek_df.columns[1]
        for _, r in tweek_df.iterrows():
            tsk = str(r.get(task_c,"")).strip()
            val = r.get(val_c)
            if pd.isna(val) or val == "":
                continue
            if tsk in weekly_targets:
                weekly_targets[tsk] = float(val)

    # Swaps
    swaps: List[Tuple[date, str, str, Optional[time], Optional[time]]] = []
    if swaps_df is not None and not swaps_df.empty:
        dcol = pick_col(swaps_df, ["Date"], required=False) or swaps_df.columns[0]
        ncol = pick_col(swaps_df, ["Name"], required=False) or swaps_df.columns[1]
        swcol = pick_col(swaps_df, ["SwapWith"], required=False)
        nscol = pick_col(swaps_df, ["NewStart"], required=False)
        necol = pick_col(swaps_df, ["NewEnd"], required=False)
        for _, r in swaps_df.iterrows():
            dd = to_date(r.get(dcol))
            if not dd:
                continue
            nm = str(r.get(ncol,"")).strip()
            sw = str(r.get(swcol,"")).strip() if swcol else ""
            ns = to_time(r.get(nscol)) if nscol else None
            ne = to_time(r.get(necol)) if necol else None
            if nm:
                swaps.append((dd, nm, sw, ns, ne))

    # Buddies
    buddies: Dict[str,str] = {}
    if new_df is not None and not new_df.empty:
        nc = pick_col(new_df, ["NewStarterName","NewStarter","Starter"], required=False) or new_df.columns[0]
        bc = pick_col(new_df, ["BuddyName","Buddy"], required=False) or new_df.columns[1]
        for _, r in new_df.iterrows():
            n = str(r.get(nc,"")).strip()
            b = str(r.get(bc,"")).strip()
            if n and b:
                buddies[n] = b

    return TemplateData(
        staff=staff_list,
        hours_map=hours_map,
        hols=hols,
        call_handlers=callh_df if callh_df is not None else pd.DataFrame(),
        handler_leave=hleave_df if hleave_df is not None else pd.DataFrame(),
        phones_targets=phones_targets,
        bookings_targets=bookings_targets,
        weekly_targets=weekly_targets,
        swaps=swaps,
        buddies=buddies,
    )

# ---------- Availability helpers ----------
def holiday_kind(name: str, d: date, hols: List[Tuple[str,date,date,str]]) -> Optional[str]:
    for n, s, e, k in hols:
        if n.strip().lower() == name.strip().lower() and s and e and s <= d <= e:
            return k
    return None

def shift_window(hours_map: Dict[str, Dict[str, Optional[time]]], d: date, name: str) -> Tuple[Optional[time], Optional[time]]:
    dn = day_name(d)
    hr = hours_map.get(name)
    if hr is None:
        return None, None
    return hr.get(f"{dn}Start"), hr.get(f"{dn}End")

def is_working(hours_map: Dict[str, Dict[str, Optional[time]]], d: date, t: time, name: str) -> bool:
    stt, end = shift_window(hours_map, d, name)
    return bool(stt and end and (t >= stt) and (t < end))

# ---------- Swaps ----------
def apply_swaps(hours_map: Dict[str, Dict[str, Optional[time]]], swaps, week_dates: List[date]) -> Dict[str, Dict[str, Optional[time]]]:
    out = {k: dict(v) for k, v in hours_map.items()}
    in_week = set(week_dates)
    for dd, nm, sw, ns, ne in swaps:
        if dd not in in_week:
            continue
        dn = day_name(dd)
        if nm not in out:
            continue
        if sw and sw in out:
            a1, a2 = out[nm].get(f"{dn}Start"), out[nm].get(f"{dn}End")
            b1, b2 = out[sw].get(f"{dn}Start"), out[sw].get(f"{dn}End")
            out[nm][f"{dn}Start"], out[nm][f"{dn}End"] = b1, b2
            out[sw][f"{dn}Start"], out[sw][f"{dn}End"] = a1, a2
        elif ns and ne:
            out[nm][f"{dn}Start"], out[nm][f"{dn}End"] = ns, ne
    return out

# ---------- Call handler leave impact on phones ----------
def parse_handler_leave(df: pd.DataFrame) -> List[Tuple[str, date, date]]:
    if df is None or df.empty:
        return []
    ncol = pick_col(df, ["Name","HandlerName","CallHandler"], required=False) or df.columns[0]
    sdcol = pick_col(df, ["LeaveStartDate","LeaveStart","StartDate"], required=False) or df.columns[1]
    edcol = pick_col(df, ["LeaveEndDate","LeaveEnd","EndDate"], required=False) or df.columns[2]
    out = []
    for _, r in df.iterrows():
        nm = str(r.get(ncol,"")).strip()
        sd = to_date(r.get(sdcol))
        ed = to_date(r.get(edcol))
        if nm and sd and ed:
            out.append((nm, sd, ed))
    return out

def handler_working(callh_row: pd.Series, d: date, t: time) -> bool:
    dn = day_name(d)
    stt = to_time(callh_row.get(f"{dn}Start"))
    end = to_time(callh_row.get(f"{dn}End"))
    return bool(stt and end and (t >= stt) and (t < end))

def phones_required(tpl: TemplateData, d: date, t: time) -> int:
    dn = day_name(d)
    hour_key = time(t.hour, 0)
    base = int(tpl.phones_targets.get((dn, hour_key), 0) or 0)
    if base <= 0:
        return 0

    leave_ranges = parse_handler_leave(tpl.handler_leave)
    off = 0
    if tpl.call_handlers is not None and not tpl.call_handlers.empty:
        for _, r in tpl.call_handlers.iterrows():
            nm = str(r.get("Name","")).strip()
            if not nm:
                continue
            if not handler_working(r, d, t):
                continue
            for ln, sd, ed in leave_ranges:
                if ln.strip().lower() == nm.strip().lower() and sd <= d <= ed:
                    off += 1
                    break
    return base + off

# ---------- Site-of-day ----------
def awaiting_site_for_day(d: date) -> str:
    wd = d.weekday()
    if wd in (0,4):  # Mon/Fri
        return "SLGP"
    if wd in (1,3):  # Tue/Thu
        return "JEN"
    return "BGS"

def email_site_for_day(d: date) -> str:
    # Email preferred site pattern:
    # Mon = BGS, Tue/Wed/Thu = JEN, Fri = SLGP
    wd = d.weekday()
    if wd == 0:
        return "BGS"
    if wd in (1, 2, 3):
        return "JEN"
    return "SLGP"



# ---------- Break placement ----------
def pick_breaks_site_balanced(staff_list: List[Staff], hours_map, hols, week_dates: List[date], fixed_assignments: Set[Tuple[date,time,str]]) -> Dict[Tuple[date,time], Set[str]]:
    """
    Breaks only for staff with break_required AND shift >= 6h.
    Spread within site by balancing counts per time slot.
    Avoid placing break where staff is locked to fixed assignment (FD/Triage) in that slot.
    """
    breaks: Dict[Tuple[date,time], Set[str]] = {}
    break_load: Dict[Tuple[date,str,time], int] = {}

    slots = timeslots()

    for d in week_dates:
        for st in staff_list:
            if not st.break_required:
                continue
            if holiday_kind(st.name, d, hols):
                continue
            stt, end = shift_window(hours_map, d, st.name)
            if not stt or not end:
                continue
            dur = (dt_of(d, end) - dt_of(d, stt)).total_seconds() / 3600.0
            if dur < BREAK_THRESHOLD_HOURS:
                continue

            midpoint = dt_of(d, stt) + (dt_of(d, end) - dt_of(d, stt)) / 2

            best = None
            for bt in break_candidates_for_shift(stt, end):
                # Hard rule: breaks must never sit on fixed cover such as Front Desk or Triage.
                # If lunch falls during a fixed Front Desk band, move it before/after instead.
                if (d, bt, st.name) in fixed_assignments:
                    continue

                # avoid tiny fragments where possible
                before = (dt_of(d, bt) - dt_of(d, stt)).total_seconds() / 3600.0
                after = (dt_of(d, end) - dt_of(d, add_minutes(bt, 30))).total_seconds() / 3600.0
                frag_penalty = 0
                if before < 1.0:
                    frag_penalty += 10_000
                if after < 1.0:
                    frag_penalty += 10_000

                load = break_load.get((d, st.home, bt), 0)
                dist = abs((dt_of(d, bt) - midpoint).total_seconds())
                score = frag_penalty + (load * 3600) + dist
                if best is None or score < best[0]:
                    best = (score, bt)

            if best:
                bt = best[1]
                breaks.setdefault((d, bt), set()).add(st.name)
                break_load[(d, st.home, bt)] = break_load.get((d, st.home, bt), 0) + 1

    return breaks

# ---------- Scheduling ----------
def task_weight(st: Staff, task_key: str) -> int:
    # Default weight = 3 if not present
    return int(st.weights.get(task_key, 3) if st.weights is not None else 3)

def block_limits(task: str) -> Tuple[int,int]:
    if task == "Phones":
        return MIN_PHONES, MAX_PHONES
    if task == "Docman":
        return MIN_DOCMAN, MAX_DEFAULT
    # fixed tasks not here
    return MIN_DEFAULT, MAX_DEFAULT

def schedule_week(tpl: TemplateData, wk_start: date):
    slots = timeslots()
    dates = [wk_start + timedelta(days=i) for i in range(5)]

    # apply swaps
    hours_map = apply_swaps(tpl.hours_map, tpl.swaps, dates)

    staff_by_name = {s.name: s for s in tpl.staff}
    staff_names = [s.name for s in tpl.staff]

    # Assignments (d,t,name)->task
    a: Dict[Tuple[date,time,str], str] = {}
    gaps: List[Tuple[date,time,str,str]] = []

    # minutes tracking
    mins_task: Dict[Tuple[str,str], int] = {}           # (name, taskKey)->minutes
    mins_task_day: Dict[Tuple[date,str,str], int] = {}  # (d,name,taskKey)->minutes

    # Phones fairness: limit per-person phone time/bands per day (relaxed only if unavoidable)
    PHONE_DAY_MAX_MINS = 240  # 4h max on phones per person per day (unless unavoidable)
    PHONE_DAY_MAX_BANDS = 2   # max 2 phone blocks per day (unless unavoidable)
    phone_bands_day: Dict[Tuple[date, str], int] = {}  # (date,name)->count of started phone blocks

    def add_mins(d: date, nm: str, task_key: str, mins: int):
        mins_task[(nm, task_key)] = mins_task.get((nm, task_key), 0) + mins
        mins_task_day[(d, nm, task_key)] = mins_task_day.get((d, nm, task_key), 0) + mins

    def assigned(nm: str, d: date, t: time) -> Optional[str]:
        return a.get((d,t,nm))

    def is_free(nm: str, d: date, t: time) -> bool:
        return (d,t,nm) not in a

    # --- Fixed assignments: Front Desk + Triage bands
    fixed_slots: Set[Tuple[date,time,str]] = set()

    def can_cover_full_band(nm: str, d: date, bs: time, be: time) -> bool:
        """True if staff can cover the entire band AND does not already have locked/fixed work in that window."""
        stt, endt = shift_window(hours_map, d, nm)
        if not stt or not endt:
            return False
        if not ((stt <= bs) and (endt >= be)):
            return False
        # Prevent fixed-task overlap (e.g., don't let Triage overwrite Front Desk)
        for tt in slots:
            if tt < bs or tt >= be:
                continue
            if (d, tt, nm) in fixed_slots:
                return False
            if (d, tt, nm) in a:  # already assigned for any reason
                return False
        return True

    def pick_for_band(candidates: List[str], d: date, task_key: str, bs: time, be: time) -> Optional[str]:
        ok = []
        for nm in candidates:
            if holiday_kind(nm, d, tpl.hols):
                continue
            if not can_cover_full_band(nm, d, bs, be):
                continue
            ok.append(nm)
        if not ok:
            return None

        def score(nm: str):
            st = staff_by_name[nm]
            # FrontDeskOnly gets absolute precedence for FD bands
            fd_only = 1 if (task_key == "FrontDesk" and st.frontdesk_only) else 0
            # weight-first, then least used
            w = task_weight(st, task_key)
            used = mins_task.get((nm, task_key), 0)
            return (-fd_only, -w, used, nm.lower())

        ok.sort(key=score)
        return ok[0]

    # --------------------------------------------------
    # Front Desk — band-based enforcement (primary)
    # SLGP fixed bands: 08:00–11:00, 11:00–13:00, 13:00–16:00, 16:00–18:30.
    # JEN/BGS fixed bands: 08:00–10:30, 10:30–13:00, 13:00–16:00, 16:00–18:30.
    # Exactly 1 person per site per slot; no 30/60-min fragments.
    # --------------------------------------------------
    FD_BANDS_BY_SITE = {"SLGP": FD_BANDS, "JEN": FD_BANDS_FLEX, "BGS": FD_BANDS_FLEX}

    
    def assign_frontdesk_for_site_day(site: str, d: date) -> None:
        """Assign Front Desk in bands, with fairness:
        - Normally max 1 FD band per person per day
        - If impossible, allow a 2nd band but NEVER consecutive (must be different person than previous band)
        - Never exceed 2 FD bands per person per day
        """
        role = f"FrontDesk_{site}"
        cands_all = [s.name for s in tpl.staff if s.can_frontdesk and s.home == site]

        fd_count: Dict[str, int] = {nm: 0 for nm in cands_all}
        last_band_staff: Optional[str] = None

        def fd_score(nm: str):
            st = staff_by_name[nm]
            fd_only = 1 if st.frontdesk_only else 0
            w = task_weight(st, "FrontDesk")
            used = mins_task.get((nm, "FrontDesk"), 0)
            # deterministic: higher fd_only, higher weight, less used minutes, name
            return (-fd_only, -w, used, nm.lower())

        def pick_fd_for_band(bs: time, be: time) -> Optional[str]:
            """Pick staff for a single FD band with the fairness rules."""
            # Tier 1: people with 0 FD bands today, and not the last band staff
            tiers = [
                lambda nm: fd_count.get(nm, 0) == 0 and (last_band_staff is None or nm != last_band_staff),
                # Tier 2: allow a 2nd FD band, but still not consecutive
                lambda nm: fd_count.get(nm, 0) == 1 and (last_band_staff is None or nm != last_band_staff),
            ]
            for ok_rule in tiers:
                pool = []
                for nm in cands_all:
                    if not ok_rule(nm):
                        continue
                    if fd_count.get(nm, 0) >= 2:
                        continue
                    if holiday_kind(nm, d, tpl.hols):
                        continue
                    if not can_cover_full_band(nm, d, bs, be):
                        continue
                    pool.append(nm)
                if pool:
                    pool.sort(key=fd_score)
                    return pool[0]
            return None

        def try_bands(bands):
            nonlocal last_band_staff
            # Try to fill all bands without leaving partial state behind.
            orig_count = dict(fd_count)
            orig_last = last_band_staff

            tmp_count = dict(fd_count)
            tmp_last = last_band_staff
            chosen_by_band = []

            for bs, be in bands:
                # Make the picker see the in-progress state for this attempt.
                fd_count.clear(); fd_count.update(tmp_count)
                last_band_staff = tmp_last

                chosen = pick_fd_for_band(bs, be)
                if not chosen:
                    # restore and fail
                    fd_count.clear(); fd_count.update(orig_count)
                    last_band_staff = orig_last
                    return None

                chosen_by_band.append((bs, be, chosen))
                tmp_count[chosen] = tmp_count.get(chosen, 0) + 1
                tmp_last = chosen

            # commit the successful attempt state
            fd_count.clear(); fd_count.update(tmp_count)
            last_band_staff = tmp_last
            return chosen_by_band

        bands = FD_BANDS_BY_SITE.get(site, FD_BANDS)
        picked = try_bands(bands)

        if picked is None:
            for bs, _be in bands:
                gaps.append((d, bs, role, "No suitable staff for Front Desk band"))
            return

        for bs, be, chosen in picked:
            for tt in slots:
                if tt < bs or tt >= be:
                    continue
                a[(d, tt, chosen)] = role
                fixed_slots.add((d, tt, chosen))
                add_mins(d, chosen, "FrontDesk", SLOT_MIN)

    for d in dates:
        for site in ("SLGP", "JEN", "BGS"):
            assign_frontdesk_for_site_day(site, d)

    # Triage (SLGP/JEN)
    for d in dates:
        for site in ("SLGP","JEN","BGS"):
            role = f"Triage_Admin_{site}"
            cands = [s.name for s in tpl.staff if s.can_triage and s.home == site]
            for bs, be in TRIAGE_BANDS:
                chosen = pick_for_band(cands, d, "Triage", bs, be)
                if not chosen:
                    gaps.append((d, bs, role, "No suitable staff for full band"))
                    continue
                for tt in slots:
                    if tt < bs or tt >= be:
                        continue
                    a[(d, tt, chosen)] = role
                    fixed_slots.add((d, tt, chosen))
                    add_mins(d, chosen, "Triage", SLOT_MIN)

    # Optional: BGS Triage Admin early (08:00–10:00), flex by +30 mins if needed (08:30–10:30),
    # only if capacity allows (i.e., a suitable triage-capable BGS staff member can cover the full band).
    for d in dates:
        role = "Triage_Admin_BGS"
        band_options = [(time(8, 0), time(10, 0)), (time(8, 30), time(10, 30))]
        cands = [s.name for s in tpl.staff if s.can_triage and str(s.home).upper() == "BGS"]

        chosen = None
        chosen_band = None
        for bs, be in band_options:
            pick = pick_for_band(cands, d, "Triage", bs, be)
            if pick:
                chosen = pick
                chosen_band = (bs, be)
                break

        if chosen and chosen_band:
            bs, be = chosen_band
            for tt in slots:
                if tt < bs or tt >= be:
                    continue
                if (d, tt, chosen) in a:
                    continue
                a[(d, tt, chosen)] = role
                fixed_slots.add((d, tt, chosen))
                add_mins(d, chosen, "Triage", SLOT_MIN)


    
# --- Breaks (only >6h and BreakRequired)
    breaks = pick_breaks_site_balanced(tpl.staff, hours_map, tpl.hols, dates, fixed_slots)

    def on_break(nm: str, d: date, t: time) -> bool:
        return nm in breaks.get((d,t), set())

    # --- Active blocks for variable tasks
    active: Dict[Tuple[date,str], Tuple[str,int]] = {}  # (d,name)->(task, end_idx_excl)
    planned_day: Dict[Tuple[date,str], int] = {}  # (d, task) -> planned minutes for day (blocks started), used for caps

    def apply_active(nm: str, d: date, idx: int) -> bool:
        b = active.get((d,nm))
        if not b:
            return False
        task, end_idx = b
        if idx >= end_idx:
            del active[(d,nm)]
            return False
        t = slots[idx]
        if not is_free(nm,d,t):
            # someone already assigned (fixed or enforced), stop block
            del active[(d,nm)]
            return False
        a[(d,t,nm)] = task
        add_mins(d, nm, task_key_for_task(task), SLOT_MIN)
        return True

    def stop_block(nm: str, d: date):
        if (d,nm) in active:
            del active[(d,nm)]

    def task_key_for_task(task: str) -> str:
        if task.startswith("FrontDesk_"):
            return "FrontDesk"
        if task.startswith("Triage_Admin_"):
            return "Triage"
        if task == "Email_Box":
            return "Email"
        if task == "Awaiting_PSA_Admin":
            return "Awaiting"
        if task == "Phones":
            return "Phones"
        if task == "Bookings":
            return "Bookings"
        if task == "EMIS":
            return "EMIS"
        if task == "Docman":
            return "Docman"
        if task == "Misc_Tasks":
            return "Misc"
        return "Misc"

    def eligible(nm: str, task: str, d: date, t: time, allow_cross_site: bool=False) -> bool:
        st = staff_by_name[nm]
        if holiday_kind(nm, d, tpl.hols):
            return False
        if not is_working(hours_map, d, t, nm):
            return False
        if on_break(nm, d, t):
            return False
        # Hard rule: FrontDeskOnly staff must never be assigned to other tasks
        if getattr(st, "frontdesk_only", False):
            return False
        if task.startswith("FrontDesk_") or task.startswith("Triage_Admin_"):
            return False  # already handled fixed
        if task == "Email_Box":
            # Soft site preference (site-of-day first; allow cross-site if needed)
            if not st.can_email:
                return False
            if allow_cross_site:
                return True
            return st.home == email_site_for_day(d)
        if task == "Awaiting_PSA_Admin":
            if not st.can_docman:
                return False
            if allow_cross_site:
                return True
            return st.home == awaiting_site_for_day(d)
        if task == "Phones":
            return st.can_phones
        if task == "Bookings":
            if not st.can_bookings:
                return False
            if allow_cross_site:
                return True
            return st.home == "SLGP"
        if task == "EMIS":
            return st.can_emis
        if task == "Docman":
            return st.can_docman
        if task == "Misc_Tasks":
            return True
        return True

    def start_block(nm: str, task: str, d: date, start_idx: int, allow_short_end: bool=True) -> bool:
        mn, mx = block_limits(task)
        stt, end = shift_window(hours_map, d, nm)
        if not stt or not end:
            return False

        # compute end_idx by shift end or break or fixed assignment boundary
        end_idx = start_idx
        while end_idx < len(slots) and slots[end_idx] < end:
            # stop if a fixed assignment exists at this slot for nm
            if (d, slots[end_idx], nm) in fixed_slots:
                break
            # stop at break start
            if nm in breaks.get((d, slots[end_idx]), set()):
                break
            end_idx += 1

        remaining = end_idx - start_idx
        if remaining <= 0:
            return False

        # no floaters: must be >= mn unless it's end-of-shift remainder (allow_short_end)
        # Phones are a hard-coverage task: if we're short late in the day, allow shorter end blocks.
        if task == "Phones" and remaining < mn:
            allow_short_end = True
        if remaining < mn and not allow_short_end:
            return False

        L = remaining if remaining < mn else min(mx, remaining)
        # if we have enough remaining to meet min, enforce min
        if remaining >= mn:
            L = max(mn, L)

        # --- Per-day caps for Docman/EMIS (prevents one day swallowing the whole weekly target)
        if task in ("Docman","EMIS"):
            if task == "Docman" and target_doc > 0:
                day_cap = int(round((target_doc / 5.0) * 1.30))
                done_today = planned_day.get((d, "Docman"), 0)
            elif task == "EMIS" and target_emis > 0:
                day_cap = int(round((target_emis / 5.0) * 1.30))
                done_today = planned_day.get((d, "EMIS"), 0)
            else:
                day_cap = 0
                done_today = 0
            if day_cap > 0:
                cap_slots = max(0, int((day_cap - done_today) // SLOT_MIN))
                if cap_slots <= 0:
                    return False
                L = min(L, cap_slots)
                if L < mn and not allow_short_end:
                    return False

        active[(d,nm)] = (task, start_idx + L)
        if task in ("Docman","EMIS"):
            planned_day[(d, task)] = planned_day.get((d, task), 0) + (L * SLOT_MIN)
        return True

    def pick_candidates(task: str, d: date, t: time, allow_cross_site: bool=False, prefer_sites: Optional[List[str]]=None, phones_strict: bool=True) -> List[str]:
        cands = []
        for nm in staff_names:
            if not is_free(nm,d,t):
                continue
            if not eligible(nm, task, d, t, allow_cross_site=allow_cross_site):
                continue

            # Phones fairness guardrails (strict pass). Relax only if we cannot meet phones requirement.
            if task == "Phones" and phones_strict:
                bands = phone_bands_day.get((d, nm), 0)
                mins_today = mins_task_day.get((d, nm, "Phones"), 0)
                if bands >= PHONE_DAY_MAX_BANDS:
                    continue
                if mins_today >= PHONE_DAY_MAX_MINS:
                    continue

            cands.append(nm)

        # site preference for EMIS/Docman: JEN/BGS first
        if prefer_sites:
            cands.sort(key=lambda nm: (0 if staff_by_name[nm].home in prefer_sites else 1, nm.lower()))

        def score(nm: str):
            st = staff_by_name[nm]
            key = task_key_for_task(task)
            w = task_weight(st, key)
            used = mins_task.get((nm, key), 0)

            # Phones: spread load across staff (bands then minutes today), then weights.
            if task == "Phones":
                bands = phone_bands_day.get((d, nm), 0)
                mins_today = mins_task_day.get((d, nm, "Phones"), 0)
                return (bands, mins_today, -w, used, nm.lower())

            # Bookings: always prefer SLGP when cross-site is allowed (still deterministic).
            if task == "Bookings":
                pref = 0 if str(st.home).upper() == "SLGP" else 1
                return (pref, -w, used, nm.lower())

            return (-w, used, nm.lower())

        cands.sort(key=score)
        return cands

    def assign_block(nm: str, task: str, d: date, idx: int):
        """Assign a task starting at slot idx, respecting existing active blocks."""
        # If already active on the same task, just continue the block.
        b = active.get((d, nm))
        if b and b[0] == task:
            apply_active(nm, d, idx)
            return

        # If active on a different task, stop that block before starting a new one.
        if b and b[0] != task:
            stop_block(nm, d)

        # Start a new block
        ok = start_block(nm, task, d, idx, allow_short_end=(task == "Misc_Tasks"))
        if not ok:
            # fallback: if we cannot start a valid block, force Misc (true end-of-shift only)
            a[(d, slots[idx], nm)] = "Misc_Tasks"
            add_mins(d, nm, "Misc", SLOT_MIN)
            return

        # Track phone blocks (bands) only when a NEW Phones block starts
        if task == "Phones":
            phone_bands_day[(d, nm)] = phone_bands_day.get((d, nm), 0) + 1

        apply_active(nm, d, idx)

    # --- Weekly targets (minutes)
    target_book = int(round((tpl.weekly_targets.get("Bookings", 0.0) or 0.0) * 60))
    target_emis = int(round((tpl.weekly_targets.get("EMIS", 0.0) or 0.0) * 60))
    target_doc  = int(round((tpl.weekly_targets.get("Docman", 0.0) or 0.0) * 60))
    # Targets with buffer (avoid 30-minute micro-assignments and overshoot)
    BOOK_LOW = 0.90   # aim to hit at least 90% of target
    BOOK_HIGH = 1.15  # allow up to 15% overshoot
    EMIS_LOW = 0.90
    EMIS_HIGH = 1.15
    DOC_LOW = 0.90
    DOC_HIGH = 1.15

    def total_mins(task_key: str) -> int:
        return sum(v for (nm, tk), v in mins_task.items() if tk == task_key)

    # --- Daily caps for EMIS/Docman to prevent front-loading early in the week
    day_cap_emis = int(round((target_emis / 5.0) * 1.30)) if target_emis > 0 else 0
    day_cap_doc  = int(round((target_doc  / 5.0) * 1.30)) if target_doc  > 0 else 0
    day_mins: Dict[Tuple[str,str], int] = {}  # (date_label, task_key) -> minutes assigned that day

    # helper: compute dynamic bookings people-per-slot requirement to approach weekly target

    # helper: bookings requirement per slot until we reach buffered target (>=90% of target)
    def bookings_needed_this_slot(d: date, idx: int) -> int:
        if target_book <= 0:
            return 0
        t = slots[idx]
        done = total_mins("Bookings")
        low = int(round(target_book * BOOK_LOW))
        high = int(round(target_book * BOOK_HIGH))

        if done >= low:
            return 0  # target met (within buffer)

        remaining = max(0, low - done)

        # Remaining booking-eligible slots in the week (from now)
        rem_slots = 0
        for dd in dates:
            for tt in slots:
                if dd < d:
                    continue
                if dd == d and tt < t:
                    continue
                rem_slots += 1

        if rem_slots <= 0:
            return 0

        # People needed this slot to catch up smoothly; at least 1 if behind target
        ppl = math.ceil(remaining / (rem_slots * SLOT_MIN))
        return max(1, int(ppl))

    def enforce(task: str, need: int, d: date, idx: int, allow_cross_site: bool=False, prefer_sites: Optional[List[str]]=None, note_task_key: str=""):
        t = slots[idx]
        while True:
            current = len([nm for nm in staff_names if a.get((d,t,nm)) == task])
            if current >= need:
                return
            # Phones fairness: try strict selection first, relax only if we cannot meet requirement
            if task == "Phones":
                cands = pick_candidates(task, d, t, allow_cross_site=allow_cross_site, prefer_sites=prefer_sites, phones_strict=True)
                if not cands:
                    cands = pick_candidates(task, d, t, allow_cross_site=allow_cross_site, prefer_sites=prefer_sites, phones_strict=False)
            else:
                cands = pick_candidates(task, d, t, allow_cross_site=allow_cross_site, prefer_sites=prefer_sites, phones_strict=True)
            if not cands:
                # For Phones (hard requirement), allow stealing from non-fixed assignments.
                if task == "Phones":
                    steal = [
                        nm for nm in staff_names
                        if eligible(nm, task, d, t, allow_cross_site=True)
                        and (d, t, nm) in a
                        and not str(a.get((d, t, nm), "")).startswith(("FrontDesk_", "Triage_Admin_"))
                        and a.get((d, t, nm)) not in ("Email_Box", "Awaiting_PSA_Admin")
                    ]
                    if steal:
                        steal.sort(key=lambda nm: (-task_weight(staff_by_name[nm], "Phones"), mins_task.get((nm, "Phones"), 0), nm.lower()))
                        nm = steal[0]
                        # break any active block for this person so they don't snap back
                        stop_block(nm, d)
                        a[(d, t, nm)] = "Phones"
                        add_mins(d, nm, "Phones", SLOT_MIN)
                        continue
                gaps.append((d, t, task, f"Short by {need-current}"))
                return
            nm = cands[0]
            assign_block(nm, task, d, idx)

    # --- Main loop by slot
    for d in dates:
        # Continuity trackers (reduce job-hopping)
        fd_last: Dict[str, Optional[str]] = {s: None for s in SITES}
        fd_run: Dict[str, int] = {s: 0 for s in SITES}
        email_last: Optional[str] = None
        awaiting_last: Optional[str] = None

        # --------------------------------------------------------------
        # PSA Admin / Awaiting Response — stable block (min 2.5h)
        # Goal: cover 10:30–16:00 (later can be extended manually if needed).
        # Flex start by ±30 mins to avoid break fragmentation.
        # --------------------------------------------------------------
        psa_pref_site = awaiting_site_for_day(d)
        psa_window_start = time(10, 30)
        psa_window_end = time(16, 0)
        psa_start_options = [time(10, 0), time(10, 30), time(11, 0)]

        def _psa_score(nm: str) -> tuple:
            st = staff_by_name[nm]
            pref = 0 if st.home == psa_pref_site else 1
            w = task_weight(st, "Awaiting")
            used = mins_task.get((nm, "Awaiting"), 0)
            return (pref, -w, used, nm.lower())

        def _psa_assignable_slots(nm: str, st_time: time) -> int:
            st = staff_by_name[nm]
            if not st.can_docman:
                return 0
            if holiday_kind(nm, d, tpl.hols):
                return 0
            stt, endt = shift_window(hours_map, d, nm)
            if not stt or not endt:
                return 0
            # must intersect the window
            if endt <= psa_window_start or stt >= psa_window_end:
                return 0

            cnt = 0
            for tt in slots:
                if tt < psa_window_start or tt >= psa_window_end:
                    continue
                if tt < st_time:
                    continue
                if not is_working(hours_map, d, tt, nm):
                    continue
                if nm in breaks.get((d, tt), set()):
                    continue
                if (d, tt, nm) in fixed_slots:
                    continue
                cnt += 1
            return cnt

        # Pick a person + a start time (flex ±30) that yields at least 2.5h assignable time.
        psa_choice = None
        best = None
        for st_time in psa_start_options:
            cands = [nm for nm in staff_names if staff_by_name[nm].home == "BGS" and staff_by_name[nm].can_triage and a.get((dd, tt, nm)) == "Misc_Tasks"]
            if cands:
                chosen = sorted(cands, key=lambda nm: nm.lower())[0]
                a[(dd, tt, chosen)] = "Triage_Admin_BGS"


    return a, breaks, gaps, dates, slots, hours_map
# ---------- Excel output ----------
ROLE_COLORS = {
    "FrontDesk_SLGP": "FFF2CC",
    "FrontDesk_JEN":  "FFF2CC",
    "FrontDesk_BGS":  "FFF2CC",
    "Triage_Admin_SLGP": "D9EAD3",
    "Triage_Admin_JEN":  "D9EAD3",
    "Triage_Admin_BGS":  "D9EAD3",
    "Email_Box": "CFE2F3",
    "Phones": "C9DAF8",
    "Bookings": "FCE5CD",
    "EMIS": "EAD1DC",
    "Docman": "D0E0E3",
    "Awaiting_PSA_Admin": "D0E0E3",
    "Misc_Tasks": "FFFFFF",
    "Break": "DDDDDD",
    "Holiday": "F8CBAD",
    "Bank Holiday": "FFD966",
    "Sick": "F4CCCC",
    "": "DDDDDD",
}

def fill_for(value: str) -> PatternFill:
    return PatternFill("solid", fgColor=ROLE_COLORS.get(value, "FFFFFF"))

THICK = Side(style="thick")
THIN = Side(style="thin")
DAY_BORDER = Border(top=THICK)
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_workbook(tpl: TemplateData, start_monday: date, weeks: int) -> Workbook:
    """
    Creates an Excel workbook where Week*_MasterTimeline is the single source of truth.
    Site timelines are formula-linked to the Master so manual edits propagate.
    Adds:
      - Site timelines (SLGP/JEN/BGS)
      - Coverage_By_Slot (names per task per slot)
      - Coverage_Dashboard (at-a-glance metrics + target progress bars)
      - Clear day separators + wider columns + consistent colour fills
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.styles import Border

    wb = Workbook()
    wb.remove(wb.active)

    staff_names = [s.name for s in tpl.staff]
    staff_by_name = {s.name: s for s in tpl.staff}

    # Column widths
    DATE_W, TIME_W, STAFF_W = 14, 8, 18

    def is_day_start(val_time: str) -> bool:
        return val_time == DAY_START.strftime("%H:%M")

    def is_day_end(val_time: str) -> bool:
        # last slot begins at 18:00 when day ends 18:30
        return val_time == "18:00"

    def merged_border(base: Border, top=False, bottom=False) -> Border:
        return Border(
            left=base.left, right=base.right,
            top=THICK if top else base.top,
            bottom=THICK if bottom else base.bottom,
        )

    # Tasks for coverage sheets/dashboard
    COVER_TASKS = [
        "FrontDesk_SLGP","FrontDesk_JEN","FrontDesk_BGS",
        "Triage_Admin_SLGP","Triage_Admin_JEN","Triage_Admin_BGS",
        "Email_Box","Awaiting_PSA_Admin","Phones",
        "Bookings","EMIS","Docman","Misc_Tasks"
    ]

    for w in range(weeks):
        wk_start = start_monday + timedelta(days=7*w)
        a, breaks, gaps, dates, slots, hours_map = schedule_week(tpl, wk_start)

        # ----------------------------
        # 1) Master Timeline
        # ----------------------------
        ws = wb.create_sheet(f"Week{w+1}_MasterTimeline")
        ws.append(["Date","Time"] + staff_names)
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "C2"

        # Set widths
        ws.column_dimensions["A"].width = DATE_W
        ws.column_dimensions["B"].width = TIME_W
        for i in range(len(staff_names)):
            ws.column_dimensions[get_column_letter(3+i)].width = STAFF_W

        for d in dates:
            for t in slots:
                row = [d.strftime("%a %d-%b"), t.strftime("%H:%M")]
                for nm in staff_names:
                    hk = holiday_kind(nm, d, tpl.hols)
                    if hk:
                        val = hk
                    elif not is_working(hours_map, d, t, nm):
                        val = ""
                    elif a.get((d,t,nm), "").startswith(("FrontDesk_", "Triage_Admin_")):
                        val = a.get((d,t,nm), "")
                    elif nm in breaks.get((d,t), set()):
                        val = "Break"
                    else:
                        val = a.get((d,t,nm), "Misc_Tasks")
                    row.append(val)
                ws.append(row)

        # Style Master (fills + day borders)
        for rr in range(2, ws.max_row+1):
            tval = str(ws.cell(rr,2).value or "")
            day_start = is_day_start(tval)
            day_end = is_day_end(tval)
            for cc in range(1, ws.max_column+1):
                cell = ws.cell(rr,cc)
                b = CELL_BORDER
                if day_start:
                    b = merged_border(b, top=True)
                if day_end:
                    b = merged_border(b, bottom=True)
                cell.border = b

                if cc >= 3:
                    val = str(cell.value or "")
                    cell.fill = fill_for(val)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

        # ----------------------------
        # 2) Site Timelines (formula-linked)
        # ----------------------------
        def make_site_timeline(site: str):
            site_staff = [nm for nm in staff_names if str(staff_by_name[nm].home).upper() == site]
            if not site_staff:
                return
            ws_site = wb.create_sheet(f"Week{w+1}_{site}_Timeline")
            ws_site.append(["Date","Time"] + site_staff)
            for c in ws_site[1]:
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws_site.freeze_panes = "C2"

            ws_site.column_dimensions["A"].width = DATE_W
            ws_site.column_dimensions["B"].width = TIME_W
            for i in range(len(site_staff)):
                ws_site.column_dimensions[get_column_letter(3+i)].width = STAFF_W

            master_col = {nm: 3 + staff_names.index(nm) for nm in site_staff}

            for rr in range(2, ws.max_row+1):
                date_val = ws.cell(rr,1).value
                time_val = ws.cell(rr,2).value
                ws_site.append([date_val, time_val] + [""]*len(site_staff))
                site_rr = ws_site.max_row
                for i, nm in enumerate(site_staff):
                    mc = master_col[nm]
                    src = f"'{ws.title}'!{get_column_letter(mc)}{rr}"
                    ws_site.cell(site_rr, 3+i).value = f"={src}"

            # Styling: borders + fills (fills won't compute until Excel opens; apply conditional formatting instead)
            from openpyxl.formatting.rule import FormulaRule

            data_range = f"{get_column_letter(3)}2:{get_column_letter(ws_site.max_column)}{ws_site.max_row}"
            # Apply expression rules so formula-linked cells colour correctly.
            tl = f"{get_column_letter(3)}2"  # top-left cell in the formatted range
            for task, color in ROLE_COLORS.items():
                if not task:
                    continue
                fill = PatternFill("solid", fgColor=color)
                ws_site.conditional_formatting.add(
                    data_range,
                    FormulaRule(formula=[f'ISNUMBER(SEARCH("{task}",{tl}))'], fill=fill, stopIfTrue=False)
                )

            for rr in range(2, ws_site.max_row+1):
                tval = str(ws_site.cell(rr,2).value or "")
                day_start = is_day_start(tval)
                day_end = is_day_end(tval)
                for cc in range(1, ws_site.max_column+1):
                    cell = ws_site.cell(rr,cc)
                    b = CELL_BORDER
                    if day_start:
                        b = merged_border(b, top=True)
                    if day_end:
                        b = merged_border(b, bottom=True)
                    cell.border = b
                    if cc >= 3:
                        cell.alignment = Alignment(vertical="top", wrap_text=True)

        for site in ("SLGP","JEN","BGS"):
            make_site_timeline(site)

        # ----------------------------
        # 3) Coverage_By_Slot (names)
        # ----------------------------
        ws_cov = wb.create_sheet(f"Week{w+1}_Coverage_By_Slot")
        ws_cov.append(["Date","Time"] + COVER_TASKS)
        for c in ws_cov[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_cov.freeze_panes = "C2"

        ws_cov.column_dimensions["A"].width = DATE_W
        ws_cov.column_dimensions["B"].width = TIME_W
        for i in range(len(COVER_TASKS)):
            ws_cov.column_dimensions[get_column_letter(3+i)].width = 26

        for d in dates:
            for t in slots:
                row = [d.strftime("%a %d-%b"), t.strftime("%H:%M")]
                for task in COVER_TASKS:
                    names = [nm for nm in staff_names if a.get((d,t,nm)) == task]
                    row.append(", ".join(names))
                ws_cov.append(row)

        for rr in range(2, ws_cov.max_row+1):
            tval = str(ws_cov.cell(rr,2).value or "")
            day_start = is_day_start(tval)
            day_end = is_day_end(tval)
            for cc in range(1, ws_cov.max_column+1):
                cell = ws_cov.cell(rr,cc)
                b = CELL_BORDER
                if day_start:
                    b = merged_border(b, top=True)
                if day_end:
                    b = merged_border(b, bottom=True)
                cell.border = b
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Apply background colours to Coverage_By_Slot columns for readability
        for cc in range(3, ws_cov.max_column+1):
            task = str(ws_cov.cell(1, cc).value or "")
            f = fill_for(task)
            for rr in range(2, ws_cov.max_row+1):
                ws_cov.cell(rr, cc).fill = f

        # ----------------------------
        # 4) Totals
        # ----------------------------
        ws_tot = wb.create_sheet(f"Week{w+1}_Totals_Static")
        tasks_tot = [
            "FrontDesk_SLGP","FrontDesk_JEN","FrontDesk_BGS",
            "Triage_Admin_SLGP","Triage_Admin_JEN","Email_Box",
            "Phones","Awaiting_PSA_Admin","Bookings","EMIS","Docman",
            "Misc_Tasks","Break"
        ]
        ws_tot.append(["Name"] + tasks_tot + ["WeeklyTotalHours"])
        for c in ws_tot[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_tot.freeze_panes = "B2"
        ws_tot.column_dimensions["A"].width = 22
        for i in range(len(tasks_tot)):
            ws_tot.column_dimensions[get_column_letter(2+i)].width = 14

        hours = {(nm, task): 0.0 for nm in staff_names for task in tasks_tot}
        for d in dates:
            for t in slots:
                for nm in staff_names:
                    task = a.get((d,t,nm))
                    if task:
                        hours[(nm, task)] = hours.get((nm, task), 0.0) + 0.5
                for nm in breaks.get((d,t), set()):
                    hours[(nm, "Break")] = hours.get((nm, "Break"), 0.0) + 0.5

        for nm in staff_names:
            row = [nm]
            total = 0.0
            for task in tasks_tot:
                v = round(hours.get((nm, task), 0.0), 2)
                row.append(v)
                total += v
            row.append(round(total, 2))
            ws_tot.append(row)

        # ----------------------------
        # 5) Coverage Dashboard
        # ----------------------------
        ws_dash = wb.create_sheet(f"Week{w+1}_Coverage_Dashboard")
        ws_dash.column_dimensions["A"].width = 44
        ws_dash.column_dimensions["B"].width = 14
        ws_dash.column_dimensions["C"].width = 14
        ws_dash.column_dimensions["D"].width = 10

        ws_dash["A1"] = "Coverage Dashboard"
        ws_dash["A1"].font = Font(bold=True, size=14)

        total_slots = len(dates) * len(slots)

        phones_ok = 0
        for d in dates:
            for t in slots:
                req = phones_required(tpl, d, t)
                actual = sum(1 for nm in staff_names if a.get((d,t,nm)) == "Phones")
                if actual >= req:
                    phones_ok += 1
        phones_pct = phones_ok / total_slots if total_slots else 1.0

        fd_ok = 0
        fd_total = total_slots * 3
        for d in dates:
            for t in slots:
                for site in ("SLGP","JEN","BGS"):
                    role = f"FrontDesk_{site}"
                    actual = sum(1 for nm in staff_names if a.get((d,t,nm)) == role)
                    if actual == 1:
                        fd_ok += 1
        fd_pct = fd_ok / fd_total if fd_total else 1.0

        # Break compliance
        req_break = 0
        got_break = 0
        for d in dates:
            for nm in staff_names:
                stt, end = shift_window(hours_map, d, nm)
                if not stt or not end:
                    continue
                dur = (dt_of(d, end) - dt_of(d, stt)).total_seconds()/3600.0
                if dur >= BREAK_THRESHOLD_HOURS and staff_by_name[nm].break_required:
                    req_break += 1
                    had = any(nm in breaks.get((d, bt), set()) for bt in (time(12,0), time(12,30), time(13,0), time(13,30)))
                    if had:
                        got_break += 1
        break_pct = (got_break/req_break) if req_break else 1.0

        def achieved(task: str) -> float:
            return round(sum(0.5 for d in dates for t in slots for nm in staff_names if a.get((d,t,nm)) == task), 2)

        book_h = achieved("Bookings")
        emis_h = achieved("EMIS")
        doc_h = achieved("Docman")

        book_t = float(tpl.weekly_targets.get("Bookings", 0.0) or 0.0)
        emis_t = float(tpl.weekly_targets.get("EMIS", 0.0) or 0.0)
        doc_t = float(tpl.weekly_targets.get("Docman", 0.0) or 0.0)

        ws_dash.append(["Metric","Achieved","Target","%"])
        for c in ws_dash[2]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")

        rows = [
            ("Front Desk coverage (slots meeting rule)", fd_ok, fd_total, fd_pct),
            ("Phones coverage (slots meeting requirement)", phones_ok, total_slots, phones_pct),
            ("Break compliance (staff-days with break)", got_break, req_break, break_pct),
            ("Bookings hours", book_h, book_t, (book_h/book_t if book_t else 1.0)),
            ("EMIS hours", emis_h, emis_t, (emis_h/emis_t if emis_t else 1.0)),
            ("Docman hours", doc_h, doc_t, (doc_h/doc_t if doc_t else 1.0)),
        ]
        for r in rows:
            ws_dash.append([r[0], r[1], r[2], round(float(r[3]), 3)])

        # Data bars for %
        last_row = ws_dash.max_row
        ws_dash.conditional_formatting.add(f"D3:D{last_row}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="63C384"))

        for rr in range(2, ws_dash.max_row+1):
            for cc in range(1, 5):
                ws_dash.cell(rr, cc).border = CELL_BORDER
                ws_dash.cell(rr, cc).alignment = Alignment(vertical="center")

        # ----------------------------
        # 6) Notes/Gaps
        # ----------------------------
        ws_g = wb.create_sheet(f"Week{w+1}_NotesAndGaps")
        ws_g.append(["Date","Time","Task","Note"])
        for c in ws_g[1]:
            c.font = Font(bold=True)
        for d, t, task, note in gaps:
            ws_g.append([d.isoformat(), t.strftime("%H:%M") if t else "", task, note])


        # ---------------- Direct Site Timelines (values + fills) ----------------
        site_sheets = {}
        for site in ("SLGP","JEN","BGS"):
            ws_site2 = wb.create_sheet(f"Week{w+1}_{site}_Timeline")
            ws_site2.append(["Date","Time"] + staff_names)
            for c in ws_site2[1]:
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws_site2.freeze_panes = "C2"
            site_sheets[site] = ws_site2

        # Fill data rows
        for d in dates:
            for t in slots:
                date_label = d.strftime("%a %d-%b")
                time_label = t.strftime("%H:%M")
                for site, ws_site2 in site_sheets.items():
                    row = [date_label, time_label]
                    for nm in staff_names:
                        hk = holiday_kind(nm, d, tpl.hols)
                        if hk:
                            val = hk
                        elif not is_working(hours_map, d, t, nm):
                            val = ""
                        elif a.get((d, t, nm), "").startswith(("FrontDesk_", "Triage_Admin_")):
                            val = a.get((d, t, nm), "")
                        elif nm in breaks.get((d, t), set()):
                            val = "Break"
                        else:
                            val = a.get((d, t, nm), "")
                        row.append(val)

        # Apply fills + formatting to site sheets
        for site, ws_site2 in site_sheets.items():
            try:
                widen_columns(ws_site2, width=18)
            except Exception:
                pass
            ws_site2.column_dimensions["A"].width = 14
            ws_site2.column_dimensions["B"].width = 8
            for rr in range(2, ws_site2.max_row + 1):
                for cc in range(3, ws_site2.max_column + 1):
                    v = str(ws_site2.cell(rr, cc).value or "")
                    ws_site2.cell(rr, cc).fill = fill_for(v)
                    ws_site2.cell(rr, cc).alignment = Alignment(wrap_text=True, vertical="top")
            try:
                apply_day_borders(ws_site2)
            except Exception:
                pass

        # ---------------- Coverage By Slot By Site (names) ----------------
        ws_cov = wb.create_sheet(f"Week{w+1}_Coverage_By_Slot_By_Site")
        cov_cols = ["FD_SLGP","FD_JEN","FD_BGS","Triage_SLGP","Triage_JEN","Phones","Bookings","EMIS","Docman","Awaiting","Email","Misc"]
        ws_cov.append(["Date","Time"] + cov_cols)
        for c in ws_cov[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_cov.freeze_panes = "C2"
        try:
            widen_columns(ws_cov, width=22)
        except Exception:
            pass
        ws_cov.column_dimensions["A"].width = 14
        ws_cov.column_dimensions["B"].width = 8

        def names_for_prefix(prefix, d, t):
            return ", ".join([nm for nm in staff_names if str(a.get((d,t,nm),"")).startswith(prefix)])

        for d in dates:
            for t in slots:
                row = [d.strftime("%a %d-%b"), t.strftime("%H:%M")]
                row.append(names_for_prefix("FrontDesk_SLGP", d, t))
                row.append(names_for_prefix("FrontDesk_JEN", d, t))
                row.append(names_for_prefix("FrontDesk_BGS", d, t))
                row.append(names_for_prefix("Triage_Admin_SLGP", d, t))
                row.append(names_for_prefix("Triage_Admin_JEN", d, t))
                row.append(", ".join([nm for nm in staff_names if a.get((d,t,nm))=="Phones"]))
                row.append(", ".join([nm for nm in staff_names if a.get((d,t,nm))=="Bookings"]))
                row.append(", ".join([nm for nm in staff_names if a.get((d,t,nm))=="EMIS"]))
                row.append(", ".join([nm for nm in staff_names if a.get((d,t,nm))=="Docman"]))
                row.append(", ".join([nm for nm in staff_names if a.get((d,t,nm))=="Email_Box"]))
                row.append(", ".join([nm for nm in staff_names if a.get((d,t,nm))=="Awaiting_PSA_Admin"]))
                row.append(", ".join([nm for nm in staff_names if a.get((d,t,nm))=="Misc_Tasks"]))
                ws_cov.append(row)

        # Colour coverage columns by meaning
        col_fills = {
            "FD_": PatternFill("solid", fgColor=ROLE_COLORS.get("FrontDesk_SLGP","FFF2CC")),
            "Triage_": PatternFill("solid", fgColor=ROLE_COLORS.get("Triage_Admin_SLGP","D9EAD3")),
            "Phones": PatternFill("solid", fgColor=ROLE_COLORS.get("Phones","C9DAF8")),
            "Bookings": PatternFill("solid", fgColor=ROLE_COLORS.get("Bookings","FCE5CD")),
            "EMIS": PatternFill("solid", fgColor=ROLE_COLORS.get("EMIS","EAD1DC")),
            "Docman": PatternFill("solid", fgColor=ROLE_COLORS.get("Docman","D0E0E3")),
            "Awaiting": PatternFill("solid", fgColor=ROLE_COLORS.get("Awaiting_PSA_Admin","D0E0E3")),
            "Email": PatternFill("solid", fgColor=ROLE_COLORS.get("Email_Box","CFE2F3")),
            "Misc": PatternFill("solid", fgColor=ROLE_COLORS.get("Misc_Tasks","EFEFEF")),
        }
        for cc in range(3, ws_cov.max_column+1):
            header = str(ws_cov.cell(1,cc).value)
            fill = None
            for k,v in col_fills.items():
                if header.startswith(k) or header == k:
                    fill = v
                    break
            if fill:
                for rr in range(2, ws_cov.max_row+1):
                    ws_cov.cell(rr,cc).fill = fill
                    ws_cov.cell(rr,cc).alignment = Alignment(wrap_text=True, vertical="top")
        try:
            apply_day_borders(ws_cov)
        except Exception:
            pass

        except Exception:
            pass

        # ---------------- Dynamic Totals (based on Site sheets) ----------------
        task_keys = ["FrontDesk","Triage","Phones","Bookings","EMIS","Docman","Awaiting","Email","Misc","Break"]
        ws_dyn = write_dynamic_totals_from_site_sheets(wb, w+1, staff_names, task_keys)

    return wb



def apply_day_borders(ws):
    # Thick border between days (every 48 rows for 30-min slots * 24 hours approx 48 slots)
    for col in range(1, ws.max_column + 1):
        for row in range(1, ws.max_row + 1):
            if row > 1 and (row-2) % 48 == 0:
                cell = ws.cell(row=row, column=col)
                cell.border = Border(top=THICK)

def widen_columns(ws, width=18):
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = width

def add_progress_bars(ws, col_letter, max_value):
    rule = DataBarRule(start_type='num', start_value=0,
                       end_type='num', end_value=max_value,
                       color="63C384")
    ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)

# ===========================
# Site display mapping helpers (for Site Timelines output)
# ===========================
def home_site_of(name: str, staff_by_name: dict) -> str:
    try:
        return str(staff_by_name[name].get("HomeSite","")).strip().upper()
    except Exception:
        return ""

def display_site_for_assignment(role: str, d, name: str, staff_by_name: dict) -> str:
    """Decide which site sheet should display this assignment."""
    if not role:
        return ""
    if role.endswith("_SLGP"):
        return "SLGP"
    if role.endswith("_JEN"):
        return "JEN"
    if role.endswith("_BGS"):
        return "BGS"
    if role == "Bookings":
        return "SLGP"
    if role == "Awaiting_PSA_Admin":
        return awaiting_site_for_day(d)
    if role == "Email_Box":
        return home_site_of(name, staff_by_name)  # JEN/BGS
    # Phones / EMIS / Docman / Misc -> staff home site for display
    return home_site_of(name, staff_by_name)

# ===========================
# Dynamic Totals (updates when Site timelines edited)
# ===========================
def write_dynamic_totals_from_site_sheets(wb, week_num: int, staff_names, task_keys):
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(f"Week{week_num}_Totals")
    ws.append(["Name"] + task_keys + ["WeeklyTotal"])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r, nm in enumerate(staff_names, start=2):
        row = [nm]
        staff_col_idx = 3 + (r-2)  # timelines: Date(A), Time(B), staff start at C
        staff_col = get_column_letter(staff_col_idx)
        rng = f"{staff_col}$2:{staff_col}$500"

        for task in task_keys:
            if task == "FrontDesk":
                crit = "FrontDesk*"
            elif task == "Triage":
                crit = "Triage*"
            elif task == "Phones":
                crit = "Phones"
            elif task == "Bookings":
                crit = "Bookings"
            elif task == "EMIS":
                crit = "EMIS"
            elif task == "Docman":
                crit = "Docman"
            elif task == "Awaiting":
                crit = "Awaiting_PSA_Admin"
            elif task == "Email":
                crit = "Email_Box"
            elif task == "Misc":
                crit = "Misc_Tasks"
            elif task == "Break":
                crit = "Break"
            else:
                crit = task

            f = (
                f"=0.5*("
                f'COUNTIF(Week{week_num}_SLGP_Timeline!{rng},"{crit}")+'
                f'COUNTIF(Week{week_num}_JEN_Timeline!{rng},"{crit}")+'
                f'COUNTIF(Week{week_num}_BGS_Timeline!{rng},"{crit}")'
                f")"
            )
            row.append(f)

        start_letter = get_column_letter(2)
        end_letter = get_column_letter(1 + len(task_keys))
        row.append(f"=SUM({start_letter}{r}:{end_letter}{r})")
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 20
    for i in range(2, 2+len(task_keys)+1):
        ws.column_dimensions[get_column_letter(i)].width = 12
    return ws


# =========================================================
# v32 STABLE WORKBOOK WRITER + RECALC (no formulas, no CF)
# Source-of-truth = Site Timelines values
# =========================================================

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

THICK_SIDE = Side(style="thick")
THIN_SIDE = Side(style="thin")

def _border(top=False, bottom=False, left=False, right=False):
    return Border(
        top=THICK_SIDE if top else THIN_SIDE,
        bottom=THICK_SIDE if bottom else THIN_SIDE,
        left=THICK_SIDE if left else THIN_SIDE,
        right=THICK_SIDE if right else THIN_SIDE,
    )

def _time_str(t: time) -> str:
    return t.strftime("%H:%M")

def _day_sep_rows(slots_count: int) -> int:
    # One day = number of slots in a day (Mon..Fri separate by repeating header row)
    return slots_count

def _task_color(task: str) -> str:
    return ROLE_COLORS.get(task, "FFFFFF")

def _apply_fill_and_style(ws, start_row, start_col, end_row, end_col, repeat_header_every=None, header_row=1):
    """Apply cell fills based on value strings + basic borders; optionally repeat header row for printing."""
    for r in range(1, ws.max_row + 1):
        # repeat header row (copy values + styles) at boundaries
        if repeat_header_every and r > header_row and (r - header_row - 1) % repeat_header_every == 0:
            # this row is start of a new day block: already inserted by builder; style it as header
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            continue

    for r in range(start_row, end_row + 1):
        # thick top border at start of each day block (after repeated header row)
        if repeat_header_every:
            # day block starts right after each repeated header row
            # builder inserts header rows; detect by Time cell being 08:00
            time_val = str(ws.cell(r, 2).value or "")
            top_day = (time_val == DAY_START.strftime("%H:%M"))
        else:
            top_day = False

        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            val = str(cell.value or "")
            # Fill only for task columns (>=3); Date/Time columns left plain
            if c >= 3:
                cell.fill = PatternFill("solid", fgColor=_task_color(val))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            # Borders
            b = Border(
                top=THICK_SIDE if top_day else THIN_SIDE,
                bottom=THIN_SIDE,
                left=THIN_SIDE,
                right=THIN_SIDE,
            )
            cell.border = b

def _site_for_display(role: str, d: date, name: str, staff_by_name: dict) -> str:
    """Decide which site sheet should display this assignment.

    IMPORTANT: Site timelines are already split by staff home site (columns).
    So we should NOT blank tasks like Bookings/Awaiting/Email simply because they are
    'conceptually' owned by another site. If a staff member is assigned a task,
    it must display on their home-site timeline.
    """
    if not role:
        return ""
    if role.endswith("_SLGP"):
        return "SLGP"
    if role.endswith("_JEN"):
        return "JEN"
    if role.endswith("_BGS"):
        return "BGS"
    # Default: show on staff home site
    try:
        return str(staff_by_name[name].home).upper()
    except Exception:
        return ""

def _compute_totals_from_assignments(assignments, breaks, dates, slots, staff_names):
    totals = {}  # (name, task)->hours
    for nm in staff_names:
        totals[(nm, "WeeklyTotal")] = 0.0
    for d in dates:
        for t in slots:
            for nm in staff_names:
                task = assignments.get((d, t, nm))
                if task:
                    totals[(nm, task)] = totals.get((nm, task), 0.0) + 0.5
                    totals[(nm, "WeeklyTotal")] += 0.5
            for nm in breaks.get((d, t), set()):
                totals[(nm, "Break")] = totals.get((nm, "Break"), 0.0) + 0.5
                totals[(nm, "WeeklyTotal")] = totals.get((nm, "WeeklyTotal"), 0.0) + 0.5
    return totals

def _compute_totals_from_site_timelines(wb, week_num: int):
    """Read values from site timelines and compute per-staff totals."""
    # identify site sheets
    site_sheets = {}
    for site in ("SLGP", "JEN", "BGS"):
        nm = f"Week{week_num}_{site}_Timeline"
        if nm in wb.sheetnames:
            site_sheets[site] = wb[nm]
    if not site_sheets:
        raise ValueError(f"No site timeline sheets found for Week{week_num} (expected Week{week_num}_SLGP_Timeline etc.)")

    # staff headers are in row 1, columns C..end
    # all site sheets share same staff list (site-only, but headers present)
    # We'll take union in display order from each sheet.
    staff = []
    for ws in site_sheets.values():
        hdr = [ws.cell(1, c).value for c in range(3, ws.max_column + 1)]
        hdr = [str(x).strip() for x in hdr if x and str(x).strip()]
        for h in hdr:
            if h not in staff:
                staff.append(h)

    tasks = set()
    for ws in site_sheets.values():
        for r in range(2, ws.max_row + 1):
            # skip repeated header rows (Time cell is 'Time')
            if str(ws.cell(r,2).value or '').strip().lower() == 'time':
                continue
            for c in range(3, ws.max_column + 1):
                v = str(ws.cell(r,c).value or '').strip()
                if v:
                    tasks.add(v)
    # Normalize list of tasks of interest
    known = [
        "FrontDesk_SLGP","FrontDesk_JEN","FrontDesk_BGS",
        "Triage_Admin_SLGP","Triage_Admin_JEN","Triage_Admin_BGS",
        "Phones","Bookings","Email_Box","Awaiting_PSA_Admin",
        "EMIS","Docman","Misc_Tasks","Break","Holiday","Sick","Bank Holiday"
    ]
    tasks_order = [t for t in known if t in tasks]
    # include any unexpected tasks at end
    for t in sorted(tasks):
        if t not in tasks_order:
            tasks_order.append(t)

    totals = {(nm, t): 0.0 for nm in staff for t in tasks_order}
    totals.update({(nm, "WeeklyTotal"): 0.0 for nm in staff})

    for ws in site_sheets.values():
        # map staff col
        col_map = {}
        for c in range(3, ws.max_column + 1):
            h = ws.cell(1,c).value
            if h and str(h).strip() in staff:
                col_map[str(h).strip()] = c
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r,2).value or '').strip().lower() == 'time':
                continue
            for nm in staff:
                c = col_map.get(nm)
                if not c:
                    continue
                v = str(ws.cell(r,c).value or '').strip()
                if not v:
                    continue
                totals[(nm, v)] = totals.get((nm, v), 0.0) + 0.5
                totals[(nm, "WeeklyTotal")] = totals.get((nm, "WeeklyTotal"), 0.0) + 0.5
    return staff, tasks_order, totals

def _write_totals_sheet(wb, week_num: int, staff, tasks_order, totals):
    title = f"Week{week_num}_Totals"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws.append(["Name"] + tasks_order + ["WeeklyTotal"])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 22
    for i in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    for nm in staff:
        row = [nm]
        for t in tasks_order:
            row.append(round(totals.get((nm, t), 0.0), 2))
        row.append(round(totals.get((nm, "WeeklyTotal"), 0.0), 2))
        ws.append(row)
    return ws

def _write_coverage_sheet_from_site_timelines(wb, week_num: int):
    title = f"Week{week_num}_Coverage_By_Slot"
    if title in wb.sheetnames:
        del wb[title]
    ws_cov = wb.create_sheet(title)

    # load site sheets
    site_sheets = {site: wb[f"Week{week_num}_{site}_Timeline"] for site in ("SLGP","JEN","BGS") if f"Week{week_num}_{site}_Timeline" in wb.sheetnames}
    if not site_sheets:
        raise ValueError(f"No site timelines for Week{week_num}")

    # tasks columns
    tasks_cols = [
        "FrontDesk_SLGP","FrontDesk_JEN","FrontDesk_BGS",
        "Triage_Admin_SLGP","Triage_Admin_JEN","Triage_Admin_BGS",
        "Email_Box","Awaiting_PSA_Admin","Phones","Bookings","EMIS","Docman","Misc_Tasks"
    ]
    ws_cov.append(["Date","Time"] + tasks_cols)
    for c in ws_cov[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws_cov.freeze_panes = "C2"
    ws_cov.column_dimensions["A"].width = 14
    ws_cov.column_dimensions["B"].width = 8
    for i in range(len(tasks_cols)):
        ws_cov.column_dimensions[get_column_letter(3+i)].width = 28

    # Use SLGP sheet as row backbone (date/time)
    backbone = next(iter(site_sheets.values()))
    # collect staff headers per site
    headers = {}
    for site, ws in site_sheets.items():
        headers[site] = [str(ws.cell(1,c).value).strip() for c in range(3, ws.max_column+1) if ws.cell(1,c).value]

    for r in range(2, backbone.max_row + 1):
        # skip repeated headers
        if str(backbone.cell(r,2).value or '').strip().lower() == 'time':
            continue
        dval = backbone.cell(r,1).value
        tval = backbone.cell(r,2).value
        row = [dval, tval]
        # build lookup per task -> list names by scanning all site sheets
        per_task = {k: [] for k in tasks_cols}
        for site, ws in site_sheets.items():
            for ci, nm in enumerate(headers[site], start=3):
                v = str(ws.cell(r,ci).value or '').strip()
                if not v:
                    continue
                if v in per_task:
                    per_task[v].append(nm)
        for k in tasks_cols:
            row.append(", ".join(per_task[k]))
        ws_cov.append(row)

    # simple styling
    for rr in range(2, ws_cov.max_row+1):
        time_val = str(ws_cov.cell(rr,2).value or "")
        top_day = (time_val == DAY_START.strftime("%H:%M"))
        for cc in range(1, ws_cov.max_column+1):
            cell = ws_cov.cell(rr,cc)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(top=THICK_SIDE if top_day else THIN_SIDE, bottom=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE)
    
    # Apply background colours to Coverage_By_Slot columns for readability (recalc-safe)
    # Colour by task column header (not by names in cells).
    for cc in range(3, ws_cov.max_column + 1):
        task = str(ws_cov.cell(1, cc).value or "")
        f = fill_for(task)
        for rr in range(2, ws_cov.max_row + 1):
            ws_cov.cell(rr, cc).fill = f
    # Apply background colours to Coverage_By_Slot columns for readability
    for cc in range(3, ws_cov.max_column+1):
        task = str(ws_cov.cell(1, cc).value or "")
        f = fill_for(task)
        for rr in range(2, ws_cov.max_row+1):
            ws_cov.cell(rr, cc).fill = f


    return ws_cov

def recalc_workbook_from_site_timelines(xlsx_bytes: bytes) -> bytes:
    """Recalculate Coverage + Totals from edited site timelines. Returns updated workbook bytes."""
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    # detect weeks by sheet names
    week_nums = sorted({int(m.group(1)) for s in wb.sheetnames if (m:=re.match(r"Week(\d+)_SLGP_Timeline", s))})
    if not week_nums:
        raise ValueError("Could not find Week#_SLGP_Timeline sheets in uploaded workbook.")
    for w in week_nums:
        staff, tasks_order, totals = _compute_totals_from_site_timelines(wb, w)
        _write_totals_sheet(wb, w, staff, tasks_order, totals)
        _write_coverage_sheet_from_site_timelines(wb, w)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def build_workbook(tpl: TemplateData, start_monday: date, weeks: int) -> Workbook:
    """Stable build: write site timelines as VALUES + FILLS. Totals + Coverage are static (but can be recalculated by re-upload)."""
    wb = Workbook()
    wb.remove(wb.active)

    staff_names = [s.name for s in tpl.staff]
    staff_by_name = {s.name: s for s in tpl.staff}

    DATE_W, TIME_W, STAFF_W = 14, 8, 18

    for w in range(weeks):
        week_start = start_monday + timedelta(days=7*w)
        # schedule_week is defined earlier in this file and returns assignments, breaks, gaps, dates, slots, hours_map
        assignments, breaks, gaps, dates, slots, hours_map = schedule_week(tpl, week_start)

        # Create site sheets (site staff only)
        for site in ("SLGP","JEN","BGS"):
            site_staff = [nm for nm in staff_names if str(staff_by_name[nm].home).upper() == site]
            ws = wb.create_sheet(f"Week{w+1}_{site}_Timeline")
            ws.append(["Date","Time"] + site_staff)
            # style header
            for c in ws[1]:
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws.freeze_panes = "C2"

            ws.column_dimensions["A"].width = DATE_W
            ws.column_dimensions["B"].width = TIME_W
            for i in range(len(site_staff)):
                ws.column_dimensions[get_column_letter(3+i)].width = STAFF_W

            # write rows with repeated header per day for printing
            for d in dates:
                # repeat header row before each day (except first day already has top header)
                if d != dates[0]:
                    ws.append(["Date","Time"] + site_staff)
                    r = ws.max_row
                    for c in range(1, ws.max_column+1):
                        cell = ws.cell(r,c)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                for t in slots:
                    row = [d.strftime("%a %d-%b"), _time_str(t)]
                    for nm in site_staff:
                        hk = holiday_kind(nm, d, tpl.hols)
                        if hk:
                            val = hk
                        elif not is_working(hours_map, d, t, nm):
                            val = ""
                        elif nm in breaks.get((d,t), set()):
                            val = "Break"
                        else:
                            val = assignments.get((d,t,nm), "Misc_Tasks")
                        row.append(val)
                    ws.append(row)

            # apply fills + borders
            _apply_fill_and_style(ws, start_row=2, start_col=1, end_row=ws.max_row, end_col=ws.max_column, repeat_header_every=len(slots)+1)

        # Coverage + Totals (static from assignments for initial build)
        # Coverage
        ws_cov = wb.create_sheet(f"Week{w+1}_Coverage_By_Slot")
        tasks_cols = [
            "FrontDesk_SLGP","FrontDesk_JEN","FrontDesk_BGS",
            "Triage_Admin_SLGP","Triage_Admin_JEN","Triage_Admin_BGS",
            "Email_Box","Awaiting_PSA_Admin","Phones","Bookings","EMIS","Docman","Misc_Tasks"
        ]
        ws_cov.append(["Date","Time"] + tasks_cols)
        for c in ws_cov[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_cov.freeze_panes = "C2"
        ws_cov.column_dimensions["A"].width = 14
        ws_cov.column_dimensions["B"].width = 8
        for i in range(len(tasks_cols)):
            ws_cov.column_dimensions[get_column_letter(3+i)].width = 28

        for d in dates:
            for t in slots:
                row = [d.strftime("%a %d-%b"), _time_str(t)]
                per_task = {k: [] for k in tasks_cols}
                for nm in staff_names:
                    task = assignments.get((d,t,nm))
                    if task in per_task:
                        per_task[task].append(nm)
                for k in tasks_cols:
                    row.append(", ".join(per_task[k]))
                ws_cov.append(row)
        # style
        for rr in range(2, ws_cov.max_row+1):
            time_val = str(ws_cov.cell(rr,2).value or "")
            top_day = (time_val == DAY_START.strftime("%H:%M"))
            for cc in range(1, ws_cov.max_column+1):
                cell = ws_cov.cell(rr,cc)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Apply background colours to Coverage_By_Slot columns for readability
        for cc in range(3, ws_cov.max_column+1):
            task = str(ws_cov.cell(1, cc).value or "")
            f = fill_for(task)
            for rr in range(2, ws_cov.max_row+1):
                cell = ws_cov.cell(rr, cc)
                time_val = str(ws_cov.cell(rr,2).value or "")
                top_day = (time_val == DAY_START.strftime("%H:%M"))
                cell.fill = f
                cell.border = Border(top=THICK_SIDE if top_day else THIN_SIDE, bottom=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE)

        # Totals
        totals = _compute_totals_from_assignments(assignments, breaks, dates, slots, staff_names)
        ws_tot = wb.create_sheet(f"Week{w+1}_Totals")
        tasks_order = [
            "FrontDesk_SLGP","FrontDesk_JEN","FrontDesk_BGS",
            "Triage_Admin_SLGP","Triage_Admin_JEN","Triage_Admin_BGS",
            "Phones","Bookings","Email_Box","Awaiting_PSA_Admin",
            "EMIS","Docman","Misc_Tasks","Break"
        ]
        ws_tot.append(["Name"] + tasks_order + ["WeeklyTotal"])
        for c in ws_tot[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_tot.freeze_panes = "B2"
        ws_tot.column_dimensions["A"].width = 22
        for i in range(2, ws_tot.max_column+1):
            ws_tot.column_dimensions[get_column_letter(i)].width = 14
        for nm in staff_names:
            row = [nm] + [round(totals.get((nm,t), 0.0), 2) for t in tasks_order] + [round(totals.get((nm,"WeeklyTotal"), 0.0), 2)]
            ws_tot.append(row)


        # Target checks + heatmaps from template targets
        ws_tc = wb.create_sheet(f"Week{w+1}_Target_Checks")
        ws_tc.append(["Target check", "Actual", "Target", "%", "Status"])
        for c in ws_tc[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        def _actual_hours(task):
            return round(sum(0.5 for dd in dates for tt in slots for nm in staff_names if assignments.get((dd,tt,nm)) == task), 2)
        def _matrix_target_hours(target_map):
            return round(sum(float(v or 0) for v in target_map.values()), 2)
        def _matrix_actual_hours(task):
            return round(sum(0.5 for dd in dates for tt in slots for nm in staff_names if assignments.get((dd,tt,nm)) == task), 2)
        checks = [
            ("Phones hourly matrix hours", _matrix_actual_hours("Phones"), _matrix_target_hours(tpl.phones_targets)),
            ("Bookings hourly matrix hours", _matrix_actual_hours("Bookings"), _matrix_target_hours(tpl.bookings_targets)),
            ("Bookings weekly hours", _actual_hours("Bookings"), float(tpl.weekly_targets.get("Bookings", 0.0) or 0.0)),
            ("EMIS weekly hours", _actual_hours("EMIS"), float(tpl.weekly_targets.get("EMIS", 0.0) or 0.0)),
            ("Docman weekly hours", _actual_hours("Docman"), float(tpl.weekly_targets.get("Docman", 0.0) or 0.0)),
        ]
        for label, actual, target in checks:
            pct = (actual / target) if target else 1.0
            status = "Met" if pct >= 1 else ("Near" if pct >= 0.9 else "Short")
            ws_tc.append([label, actual, target, round(pct,3), status])
        for rr in range(2, ws_tc.max_row+1):
            status = str(ws_tc.cell(rr,5).value)
            fill = PatternFill("solid", fgColor="FFC6EFCE" if status == "Met" else ("FFFFEB9C" if status == "Near" else "FFFFC7CE"))
            for cc in range(1,6):
                ws_tc.cell(rr,cc).fill = fill
                ws_tc.cell(rr,cc).border = CELL_BORDER
        for cc,wid in {1:30,2:12,3:12,4:10,5:12}.items():
            ws_tc.column_dimensions[get_column_letter(cc)].width = wid

        ws_hm = wb.create_sheet(f"Week{w+1}_Heatmaps")
        ws_hm.append(["Hourly coverage heatmaps — actual / target from template"])
        ws_hm["A1"].font = Font(bold=True, size=14)
        def _hm_fill(actual, target):
            if target in (None, "") or float(target or 0) <= 0:
                return PatternFill("solid", fgColor="FFE7E6E6")
            ratio = float(actual or 0) / float(target or 1)
            if ratio >= 1:
                return PatternFill("solid", fgColor="FFC6EFCE")
            if ratio >= 0.9:
                return PatternFill("solid", fgColor="FFFFEB9C")
            return PatternFill("solid", fgColor="FFFFC7CE")
        def _write_hour_matrix(start_row, title, task_name, target_map):
            ws_hm.cell(start_row,1).value = title + " (half-hour actual headcount / hourly target)"
            ws_hm.cell(start_row,1).font = Font(bold=True)
            header_row = start_row + 1
            ws_hm.cell(header_row,1).value = "Time"
            for i, dn in enumerate(["Mon","Tue","Wed","Thu","Fri"], start=2):
                ws_hm.cell(header_row,i).value = dn
            for c in range(1,7):
                ws_hm.cell(header_row,c).font = Font(bold=True)
                ws_hm.cell(header_row,c).alignment = Alignment(horizontal="center")
            # Show every 30-minute slot as an integer headcount.
            # The target is the template's hourly people requirement, repeated for both half-hour slots.
            for r_i, tt in enumerate(slots, start=header_row+1):
                ws_hm.cell(r_i,1).value = tt.strftime("%H:%M")
                for c_i, dd in enumerate(dates, start=2):
                    dn = day_name(dd)
                    target = int(target_map.get((dn, time(tt.hour,0)), 0) or 0)
                    actual = int(sum(1 for nm in staff_names if assignments.get((dd,tt,nm)) == task_name))
                    cell = ws_hm.cell(r_i,c_i)
                    cell.value = f"{actual}/{target:g}"
                    cell.fill = _hm_fill(actual, target)
                    cell.alignment = Alignment(horizontal="center")
            for rr in range(start_row, header_row + len(slots) + 1):
                for cc in range(1,7):
                    ws_hm.cell(rr,cc).border = CELL_BORDER
            return header_row + len(slots) + 2
        nr = _write_hour_matrix(3, "Phones — actual / hourly target", "Phones", tpl.phones_targets)
        _write_hour_matrix(nr+1, "Bookings — actual / hourly target", "Bookings", tpl.bookings_targets)
        for cc in range(1,7):
            ws_hm.column_dimensions[get_column_letter(cc)].width = 16

        # Notes/gaps
        ws_g = wb.create_sheet(f"Week{w+1}_NotesAndGaps")
        ws_g.append(["Date","Time","Task","Note"])
        for c in ws_g[1]:
            c.font = Font(bold=True)
        for d,t,task,note in gaps:
            ws_g.append([d.isoformat(), _time_str(t) if t else "", task, note])

    return wb


# =========================================================
# v37_9 STABILISATION PATCH
# Implements:
# 1) Lock Front Desk bands first (fixed)
# 2) Lock Email owner shift next (fixed)
# 3) Move breaks ±30 mins to avoid splitting FD/PSA/Email
# 4) Enforce min 2.5h blocks for FD/PSA/Bookings (Phones only 1h min if needed)
# 5) Phone rotation limit (1 long or 2 short blocks/day)
# 6) Strong SLGP preference for Bookings
# =========================================================

def email_site_for_day(d: date) -> str:
    # As agreed:
    # Mon = BGS, Tue/Wed/Thu = JEN, Fri = SLGP
    wd = d.weekday()
    if wd == 0:
        return "BGS"
    if wd in (1,2,3):
        return "JEN"
    return "SLGP"

def schedule_week_v37_9(tpl: TemplateData, wk_start: date):
    slots = timeslots()
    dates = [wk_start + timedelta(days=i) for i in range(5)]
    hours_map = apply_swaps(tpl.hours_map, tpl.swaps, dates)

    staff_by_name = {s.name: s for s in tpl.staff}
    staff_names = [s.name for s in tpl.staff]

    a: Dict[Tuple[date,time,str], str] = {}
    gaps: List[Tuple[date,time,str,str]] = []
    fixed: Set[Tuple[date,time,str]] = set()

    # minutes tracking (needed for weighting / reporting)
    mins_task: Dict[Tuple[str,str], int] = {}
    mins_task_day: Dict[Tuple[date,str,str], int] = {}

    def add_mins(d: date, nm: str, task_key: str, mins: int):
        mins_task[(nm, task_key)] = mins_task.get((nm, task_key), 0) + mins
        mins_task_day[(d, nm, task_key)] = mins_task_day.get((d, nm, task_key), 0) + mins

    def task_key_for_task(task: str) -> str:
        if task.startswith("FrontDesk_"):
            return "FrontDesk"
        if task.startswith("Triage_Admin_"):
            return "Triage"
        if task == "Email_Box":
            return "Email"
        if task == "Awaiting_PSA_Admin":
            return "Awaiting"
        if task == "Phones":
            return "Phones"
        if task == "Bookings":
            return "Bookings"
        if task == "EMIS":
            return "EMIS"
        if task == "Docman":
            return "Docman"
        if task == "Misc_Tasks":
            return "Misc"
        if task == "Break":
            return "Break"
        return "Misc"

    def is_free(nm: str, d: date, t: time) -> bool:
        return (d,t,nm) not in a

    # ----------------------------
    # TRIAGE (keep existing fixed bands)
    # ----------------------------
    fixed_slots: Set[Tuple[date,time,str]] = set()
    def can_cover_full_band(nm: str, d: date, bs: time, be: time) -> bool:
        stt, endt = shift_window(hours_map, d, nm)
        if not stt or not endt:
            return False
        if not ((stt <= bs) and (endt >= be)):
            return False
        for tt in slots:
            if tt < bs or tt >= be:
                continue
            if (d, tt, nm) in fixed_slots:
                return False
            if (d, tt, nm) in a:
                return False
        return True

    def task_weight(st: Staff, task_key: str) -> int:
        return int(st.weights.get(task_key, 3) if st.weights is not None else 3)

    def pick_for_band(candidates: List[str], d: date, task_key: str, bs: time, be: time) -> Optional[str]:
        ok = []
        for nm in candidates:
            if holiday_kind(nm, d, tpl.hols):
                continue
            if not can_cover_full_band(nm, d, bs, be):
                continue
            ok.append(nm)
        if not ok:
            return None

        def score(nm: str):
            st = staff_by_name[nm]
            w = task_weight(st, task_key)
            used = mins_task.get((nm, task_key), 0)
            return (-w, used, nm.lower())

        ok.sort(key=score)
        return ok[0]

    for d in dates:
        for site in ("SLGP","JEN","BGS"):
            role = f"Triage_Admin_{site}"
            cands = [s.name for s in tpl.staff if s.can_triage and s.home == site]
            for bs, be in TRIAGE_BANDS:
                chosen = pick_for_band(cands, d, "Triage", bs, be)
                if not chosen:
                    gaps.append((d, bs, role, "No suitable staff for full band"))
                    continue
                for tt in slots:
                    if tt < bs or tt >= be:
                        continue
                    a[(d, tt, chosen)] = role
                    fixed_slots.add((d, tt, chosen))
                    fixed.add((d, tt, chosen))
                    add_mins(d, chosen, "Triage", SLOT_MIN)

    # ----------------------------
    # 1) FRONT DESK BANDS FIRST (hard fixed)
    # ----------------------------
    fd_bands_by_site = {"SLGP": FD_BANDS, "JEN": FD_BANDS_FLEX, "BGS": FD_BANDS_FLEX}

    fd_bands_count: Dict[Tuple[date,str], List[int]] = {}  # (d,nm)->band indices assigned

    def fd_candidate_ok(nm: str, d: date, site: str, bs: time, be: time) -> bool:
        st = staff_by_name[nm]
        if st.home != site or (not st.can_frontdesk):
            return False
        if holiday_kind(nm, d, tpl.hols):
            return False
        stt, endt = shift_window(hours_map, d, nm)
        if not stt or not endt:
            return False
        if not (stt <= bs and endt >= be):
            return False
        for tt in slots:
            if tt < bs or tt >= be:
                continue
            if (d, tt, nm) in fixed:
                return False
            if (d, tt, nm) in a:
                return False
        return True
    def fd_score_for(d: date, bs: time, be: time, nm: str) -> Tuple[int,int,int,int,str]:
        st = staff_by_name[nm]
        # frontdesk_only first
        fd_only = 0 if st.frontdesk_only else 1
        # avoid consuming email-capable staff in the main email window when there is a realistic alternative.
        # This keeps one stable email owner from 10:30 instead of fragmenting email/front desk.
        email_pen = 0
        if st.can_email and (bs < time(16,0) and be > time(10,30)):
            stt, endt = shift_window(hours_map, d, nm)
            if stt and endt and stt <= time(10,30) and endt >= time(15,30):
                email_pen = 3
                if st.home == email_site_for_day(d):
                    email_pen = 5
        w = -task_weight(st, "FrontDesk")
        used = mins_task.get((nm, "FrontDesk"), 0)
        return (fd_only, email_pen, w, used, nm.lower())

    for d in dates:
        for site in SITES:
            bands = fd_bands_by_site[site]
            for bi, (bs, be) in enumerate(bands):
                # Phase 1: strict (max 1 band/day)
                cands = [nm for nm in staff_names if fd_candidate_ok(nm, d, site, bs, be)]
                def strict_ok(nm: str) -> bool:
                    assigned = fd_bands_count.get((d,nm), [])
                    return len(assigned) == 0
                strict = [nm for nm in cands if strict_ok(nm)]
                strict.sort(key=lambda nm: fd_score_for(d, bs, be, nm))

                chosen = strict[0] if strict else None

                # Phase 2: allow 2nd band/day BUT never consecutive
                if not chosen:
                    def second_ok(nm: str) -> bool:
                        assigned = fd_bands_count.get((d,nm), [])
                        if len(assigned) >= 2:
                            return False
                        if len(assigned) == 0:
                            return True
                        # not consecutive
                        return all(abs(bi - x) > 1 for x in assigned)
                    second = [nm for nm in cands if second_ok(nm)]
                    second.sort(key=lambda nm: fd_score_for(d, bs, be, nm))
                    chosen = second[0] if second else None

                if not chosen:
                    gaps.append((d, bs, f"FrontDesk_{site}", "No suitable staff for FD band"))
                    continue

                for tt in slots:
                    if tt < bs or tt >= be:
                        continue
                    a[(d, tt, chosen)] = f"FrontDesk_{site}"
                    fixed.add((d, tt, chosen))
                    add_mins(d, chosen, "FrontDesk", SLOT_MIN)
                fd_bands_count.setdefault((d, chosen), []).append(bi)

    # ----------------------------
    # 2) EMAIL OWNER SHIFT NEXT (hard fixed; ignore break coverage)
    # ----------------------------
    email_owner: Dict[date, Optional[str]] = {}

    def can_hold_email(nm: str, d: date) -> bool:
        st = staff_by_name[nm]
        if not st.can_email:
            return False
        if holiday_kind(nm, d, tpl.hols):
            return False
        stt, endt = shift_window(hours_map, d, nm)
        if not stt or not endt:
            return False
        if not (stt <= time(10,30)):
            return False
        # must be able to cover to 16:00 minimum
        if endt < time(15,30):
            return False
        return True
    def email_window_end(nm: str, d: date) -> time:
        _, endt = shift_window(hours_map, d, nm)
        if not endt:
            return time(16,0)
        if endt >= time(18,30):
            return time(18,30)
        # otherwise cover until end of shift, but no later than 16:00 for standard shifts
        return endt if endt < time(16,0) else time(16,0)

    for d in dates:
        pref_site = email_site_for_day(d)
        cands = [nm for nm in staff_names if can_hold_email(nm, d)]
        if not cands:
            gaps.append((d, time(10,30), "Email_Box", "No eligible staff for Email owner"))
            email_owner[d] = None
            continue

        # Prefer: not already fixed in email window; then preferred site; then weight
        def email_score(nm: str):
            st = staff_by_name[nm]
            endt = email_window_end(nm, d)
            has_conflict = 0
            for tt in slots:
                if tt < time(10,30) or tt >= endt:
                    continue
                if (d, tt, nm) in fixed:
                    has_conflict = 1
                    break
            pref = 0 if st.home == pref_site else 1
            w = -task_weight(st, "Email")
            used = mins_task.get((nm, "Email"), 0)
            # Prefer a conflict-free owner who can hold email for longer.
            dur_slots = sum(1 for tt in slots if tt >= time(10,30) and tt < endt and is_working(hours_map, d, tt, nm) and (d, tt, nm) not in fixed)
            return (has_conflict, pref, -dur_slots, w, used, nm.lower())

        cands.sort(key=email_score)
        chosen = cands[0]
        email_owner[d] = chosen
        endt = email_window_end(chosen, d)

        for tt in slots:
            if tt < time(10,30) or tt >= endt:
                continue
            if not is_working(hours_map, d, tt, chosen):
                continue
            # If FD/Triage already fixed here, leave it (email will then be "not possible" but we try to avoid by scoring)
            if (d, tt, chosen) in fixed:
                continue
            a[(d, tt, chosen)] = "Email_Box"
            fixed.add((d, tt, chosen))
            add_mins(d, chosen, "Email", SLOT_MIN)

    # ----------------------------
    # 3) PSA ADMIN (Awaiting_PSA_Admin) — continuous 10:30–16:00 (min 2.5h) with ±30 break flexibility
    # ----------------------------
    psa_owner: Dict[date, Optional[str]] = {}

    def awaiting_site_for_day_local(d: date) -> str:
        wd = d.weekday()
        if wd in (0,4):
            return "SLGP"
        if wd in (1,3):
            return "JEN"
        return "BGS"

    def can_hold_psa(nm: str, d: date) -> bool:
        st = staff_by_name[nm]
        if not st.can_docman:
            return False
        if holiday_kind(nm, d, tpl.hols):
            return False
        stt, endt = shift_window(hours_map, d, nm)
        if not stt or not endt:
            return False
        if stt > time(10,30):
            return False
        if endt < time(16,0):
            return False
        return True

    for d in dates:
        pref_site = awaiting_site_for_day_local(d)
        cands = [nm for nm in staff_names if can_hold_psa(nm, d)]
        if not cands:
            gaps.append((d, time(10,30), "Awaiting_PSA_Admin", "No eligible staff for PSA Admin owner"))
            psa_owner[d] = None
            continue

        def psa_score(nm: str):
            st = staff_by_name[nm]
            conflict = 0
            for tt in slots:
                if tt < time(10,30) or tt >= time(16,0):
                    continue
                if (d, tt, nm) in fixed:
                    conflict = 1
                    break
            pref = 0 if st.home == pref_site else 1
            w = -task_weight(st, "Awaiting")
            used = mins_task.get((nm, "Awaiting"), 0)
            return (conflict, pref, w, used, nm.lower())

        cands.sort(key=psa_score)
        chosen = cands[0]
        psa_owner[d] = chosen

        for tt in slots:
            if tt < time(10,30) or tt >= time(16,0):
                continue
            if not is_working(hours_map, d, tt, chosen):
                continue
            if (d, tt, chosen) in fixed:
                continue
            a[(d, tt, chosen)] = "Awaiting_PSA_Admin"
            fixed.add((d, tt, chosen))
            add_mins(d, chosen, "Awaiting", SLOT_MIN)

    # ----------------------------
    # 4) BREAKS with ±30min move to avoid splitting FD/PSA/Email
    # ----------------------------
    breaks: Dict[Tuple[date,time], Set[str]] = {}

    for d in dates:
        for nm in staff_names:
            st = staff_by_name[nm]
            if not st.break_required:
                continue
            if holiday_kind(nm, d, tpl.hols):
                continue
            stt, endt = shift_window(hours_map, d, nm)
            if not stt or not endt:
                continue
            dur = (dt_of(d, endt) - dt_of(d, stt)).total_seconds()/3600.0
            if dur < BREAK_THRESHOLD_HOURS:
                continue

            # Candidate break starts. Late starters use later break options, so 12:00 starts are not missed.
            candidates = break_candidates_for_shift(stt, endt)

            if not candidates:
                continue

            def break_cost(bt: time) -> Tuple[int,int,int]:
                # Prefer not to collide with fixed work. If every option collides, the break still appears,
                # but it should be at the least disruptive point rather than forcing lots of task-swapping.
                collides_fixed = 0
                for tt in (bt,):
                    if (d, tt, nm) in fixed:
                        # Email can pause for 30 mins more safely than front desk/triage.
                        if a.get((d, tt, nm)) == "Email_Box":
                            collides_fixed += 1
                        else:
                            collides_fixed += 10
                midpoint = dt_of(d, stt) + (dt_of(d, endt)-dt_of(d, stt))/2
                dist = int(abs((dt_of(d, bt) - midpoint).total_seconds()))
                # Avoid 12:00 as the default if 12:30/13:00 is equally good; it creates fewer early handovers.
                early_penalty = 1 if bt == time(12,0) else 0
                return (collides_fixed, early_penalty, dist)

            candidates.sort(key=break_cost)
            bt = candidates[0]
            breaks.setdefault((d, bt), set()).add(nm)
            # overwrite assignment at break
            if is_working(hours_map, d, bt, nm) and not holiday_kind(nm, d, tpl.hols):
                a[(d, bt, nm)] = "Break"
                fixed.add((d, bt, nm))
                add_mins(d, nm, "Break", SLOT_MIN)

    def on_break(nm: str, d: date, t: time) -> bool:
        return nm in breaks.get((d,t), set())

    # ----------------------------
    # Helper: eligibility
    # ----------------------------
    def eligible(nm: str, task: str, d: date, t: time) -> bool:
        st = staff_by_name[nm]
        if holiday_kind(nm, d, tpl.hols):
            return False
        if not is_working(hours_map, d, t, nm):
            return False
        if on_break(nm, d, t):
            return False
        if (d, t, nm) in fixed:
            return False
        if task == "Phones":
            return st.can_phones
        if task == "Bookings":
            return st.can_bookings
        if task == "EMIS":
            return st.can_emis
        if task == "Docman":
            return st.can_docman
        if task == "Misc_Tasks":
            return True
        return True

    # ----------------------------
    # 5) PHONES enforcement with rotation limits
    # ----------------------------
    # Limits: 1 long (>=3h) OR 2 short blocks/day. Minimum 1h block unless unavoidable.
    phone_blocks: Dict[Tuple[date,str], List[Tuple[int,int]]] = {}  # (d,nm)->[(start_idx,end_idx)]
    phone_mins_day: Dict[Tuple[date,str], int] = {}

    def can_take_phones_block(nm: str, d: date, proposed_slots: int) -> bool:
        blks = phone_blocks.get((d,nm), [])
        if len(blks) == 0:
            return True
        if len(blks) >= 2:
            return False
        # If already has a long block, don't add another
        existing_long = any((e-s) >= 6 for s,e in blks)  # 6 slots = 3h
        if existing_long:
            return False
        # Second block must be short (<=2.5h)
        return proposed_slots <= 5

    def contiguous_free_slots(nm: str, d: date, start_idx: int) -> int:
        cnt = 0
        for j in range(start_idx, len(slots)):
            tt = slots[j]
            if not eligible(nm, "Phones", d, tt):
                break
            cnt += 1
        return cnt

    for di, d in enumerate(dates):
        for idx, t in enumerate(slots):
            req = phones_required(tpl, d, t)
            if req <= 0:
                continue
            cur = sum(1 for nm in staff_names if a.get((d,t,nm)) == "Phones")
            need = req - cur
            if need <= 0:
                continue

            # choose candidates: free + eligible; weighting; rotate
            cands = [nm for nm in staff_names if eligible(nm, "Phones", d, t)]
            # Prefer those not already on phones today, then weight
            def phones_score(nm: str):
                st = staff_by_name[nm]
                blks = phone_blocks.get((d,nm), [])
                already = 1 if blks else 0
                w = -task_weight(st, "Phones")
                used = mins_task_day.get((d,nm,"Phones"), 0)
                return (already, w, used, nm.lower())

            cands.sort(key=phones_score)

            for _ in range(need):
                chosen = None
                for nm in cands:
                    avail = contiguous_free_slots(nm, d, idx)
                    if avail <= 0:
                        continue
                    # propose a block length: prefer 3h (6 slots) first block, else 2h (4 slots)
                    prop = min(avail, 6) if not phone_blocks.get((d,nm), []) else min(avail, 4)
                    # minimum 1h (2 slots)
                    if prop < 2:
                        continue
                    if not can_take_phones_block(nm, d, prop):
                        continue
                    chosen = (nm, prop)
                    break

                if not chosen:
                    # unavoidable: allow 1h even if it breaks the rotation rules
                    for nm in cands:
                        avail = contiguous_free_slots(nm, d, idx)
                        prop = min(avail, 2)
                        if prop >= 1 and avail >= 1:
                            chosen = (nm, 1)
                            break

                if not chosen:
                    gaps.append((d, t, "Phones", f"Short by {need}"))
                    break

                nm, prop = chosen
                start = idx
                end = idx + prop
                for j in range(start, end):
                    tt = slots[j]
                    if eligible(nm, "Phones", d, tt):
                        a[(d, tt, nm)] = "Phones"
                        add_mins(d, nm, "Phones", SLOT_MIN)
                phone_blocks.setdefault((d,nm), []).append((start, end))
                phone_mins_day[(d,nm)] = phone_mins_day.get((d,nm), 0) + prop*SLOT_MIN
                # remove chosen from candidate list for this slot so we don't pick same person twice
                cands = [x for x in cands if x != nm]

    # ----------------------------
    # 6) BOOKINGS with strong SLGP preference, min 2.5h blocks
    # ----------------------------
    target_book = int(round((tpl.weekly_targets.get("Bookings", 0.0) or 0.0) * 60))
    BOOK_LOW = 0.90
    low_book = int(round(target_book * BOOK_LOW)) if target_book > 0 else 0

    def total_bookings_mins() -> int:
        return sum(v for (nm, tk), v in mins_task.items() if tk == "Bookings")

    def contiguous_free_for_task(nm: str, d: date, start_idx: int, task: str) -> int:
        cnt = 0
        for j in range(start_idx, len(slots)):
            tt = slots[j]
            if not eligible(nm, task, d, tt):
                break
            cnt += 1
        return cnt

    for d in dates:
        for idx, t in enumerate(slots):
            if t < time(10,30):
                continue
            if target_book <= 0 or total_bookings_mins() >= low_book:
                break
            # estimate remaining slots to distribute
            remaining = max(0, low_book - total_bookings_mins())
            # rough people needed now
            ppl = 1 if remaining > 0 else 0
            if ppl <= 0:
                continue

            # choose SLGP-first candidates
            cands = [nm for nm in staff_names if eligible(nm, "Bookings", d, t)]
            def book_score(nm: str):
                st = staff_by_name[nm]
                pref = 0 if st.home == "SLGP" else 1
                w = -task_weight(st, "Bookings")
                used = mins_task.get((nm, "Bookings"), 0)
                return (pref, w, used, nm.lower())
            cands.sort(key=book_score)

            if not cands:
                continue

            nm = cands[0]
            avail = contiguous_free_for_task(nm, d, idx, "Bookings")
            # min 2.5h block = 5 slots
            if avail < 5:
                continue
            prop = min(avail, 9)  # cap ~4.5h
            prop = max(5, prop)
            for j in range(idx, idx+prop):
                tt = slots[j]
                if eligible(nm, "Bookings", d, tt):
                    a[(d, tt, nm)] = "Bookings"
                    add_mins(d, nm, "Bookings", SLOT_MIN)

    # ----------------------------
    # Fill remaining: prefer Docman/EMIS until their weekly targets hit; else Misc
    # ----------------------------
    target_emis = int(round((tpl.weekly_targets.get("EMIS", 0.0) or 0.0) * 60))
    target_doc  = int(round((tpl.weekly_targets.get("Docman", 0.0) or 0.0) * 60))
    EMIS_LOW = 0.90
    DOC_LOW = 0.90
    low_emis = int(round(target_emis * EMIS_LOW)) if target_emis > 0 else 0
    low_doc  = int(round(target_doc  * DOC_LOW))  if target_doc  > 0 else 0

    # Daily caps to spread EMIS/Docman across the week (avoid front-loading Mon/Tue)
    day_cap_emis = int(round((target_emis / 5.0) * 1.30)) if target_emis > 0 else 0
    day_cap_doc  = int(round((target_doc  / 5.0) * 1.30)) if target_doc  > 0 else 0
    day_mins: Dict[Tuple[str,str], int] = {}  # (date_label, task_key) -> minutes assigned that day

    def total_mins(task_key: str) -> int:
        return sum(v for (nm, tk), v in mins_task.items() if tk == task_key)

    for d in dates:
        for idx, t in enumerate(slots):
            for nm in staff_names:
                if holiday_kind(nm, d, tpl.hols):
                    continue
                if not is_working(hours_map, d, t, nm):
                    continue
                if on_break(nm, d, t):
                    continue
                if (d, t, nm) in a:
                    continue

                # choose based on remaining targets
                chosen = None

                # Prefer Docman/EMIS on JEN/BGS (keep SLGP capacity for Bookings where possible)
                def _any_free_non_slgp_for(task_name: str) -> bool:
                    for other in staff_names:
                        if other == nm:
                            continue
                        if staff_by_name[other].home not in ("JEN","BGS"):
                            continue
                        if not is_free(other, d, t):
                            continue
                        if eligible(other, task_name, d, t):
                            return True
                    return False

                doc_day = day_mins.get((d, "Docman"), 0)
                emis_day = day_mins.get((d, "EMIS"), 0)

                # Dynamic daily caps: if we are behind, raise today's cap so we still hit weekly buffers
                days_left = max(1, len(dates) - di)
                rem_doc  = max(0, low_doc - total_mins("Docman")) if low_doc else 0
                rem_emis = max(0, low_emis - total_mins("EMIS")) if low_emis else 0
                cap_doc_today  = day_cap_doc  if day_cap_doc  else (int(round((rem_doc / days_left) * 1.50)) if rem_doc else 0)
                cap_emis_today = day_cap_emis if day_cap_emis else (int(round((rem_emis / days_left) * 1.50)) if rem_emis else 0)
                if rem_doc and day_cap_doc:
                    cap_doc_today = max(day_cap_doc, int(round((rem_doc / days_left) * 1.50)))
                if rem_emis and day_cap_emis:
                    cap_emis_today = max(day_cap_emis, int(round((rem_emis / days_left) * 1.50)))

                if low_doc and total_mins("Docman") < low_doc and (cap_doc_today == 0 or doc_day < cap_doc_today):
                    if staff_by_name[nm].home in ("JEN","BGS") or not _any_free_non_slgp_for("Docman"):
                        chosen = "Docman"

                if chosen is None and low_emis and total_mins("EMIS") < low_emis and (cap_emis_today == 0 or emis_day < cap_emis_today):
                    if staff_by_name[nm].home in ("JEN","BGS") or not _any_free_non_slgp_for("EMIS"):
                        chosen = "EMIS"

                if chosen is None:
                    chosen = "Misc_Tasks"

                a[(d,t,nm)] = chosen
                mins_task[(nm, task_key_for_task(chosen))] = mins_task.get((nm, task_key_for_task(chosen)), 0) + SLOT_MIN
                mins_task_day[(d,nm, task_key_for_task(chosen))] = mins_task_day.get((d,nm, task_key_for_task(chosen)), 0) + SLOT_MIN
                day_mins[(d, task_key_for_task(chosen))] = day_mins.get((d, task_key_for_task(chosen)), 0) + SLOT_MIN
                add_mins(d, nm, task_key_for_task(chosen), SLOT_MIN)

    # ----------------------------
    # FINAL HARD ENFORCEMENT: no working slot blank
    # ----------------------------
    for d in dates:
        for t in slots:
            for nm in staff_names:
                if holiday_kind(nm, d, tpl.hols):
                    continue
                if not is_working(hours_map, d, t, nm):
                    continue
                if on_break(nm, d, t):
                    continue
                if not a.get((d,t,nm)):
                    a[(d,t,nm)] = "Misc_Tasks"
                    add_mins(d, nm, "Misc", SLOT_MIN)

    # --------------------------------------------------
    # POST-OPTIMISE WEEKLY TARGETS (proportional smoothing)
    # Convert Misc_Tasks into Bookings / EMIS / Docman to
    # hit Targets_Weekly, while keeping:
    #   - breaks/holidays untouched
    #   - minimum block length = 2.5h (5 slots)
    # Priorities:
    #   - Bookings: SLGP staff only (cross-site ONLY if impossible)
    #   - EMIS/Docman: prefer JEN/BGS so SLGP can keep bookings
    # --------------------------------------------------
    MIN_BLOCK_SLOTS = 5  # 2.5h

    def _count_slots(task_name: str) -> int:
        return sum(1 for dd in dates for tt in slots for nm in staff_names if a.get((dd, tt, nm)) == task_name)

    def _contig_misc_runs(dd, nm: str):
        runs = []
        run = []
        for tt in slots:
            if (dd, tt, nm) not in a:
                if run:
                    runs.append(run)
                    run = []
                continue
            if a.get((dd, tt, nm)) == "Misc_Tasks":
                run.append(tt)
            else:
                if run:
                    runs.append(run)
                    run = []
        if run:
            runs.append(run)
        return runs

    def _assign_from_runs(task_name: str, need_slots: int, candidates: list[str]) -> int:
        assigned_slots = 0
        if need_slots <= 0 or not candidates:
            return 0

        per_day = max(1, math.ceil(need_slots / len(dates))) if dates else need_slots

        for dd in dates:
            day_need = min(per_day, need_slots - assigned_slots)
            if day_need <= 0:
                break

            def day_task_used(nm: str) -> int:
                return sum(1 for tt in slots if a.get((dd, tt, nm)) == task_name)

            ordered = sorted(candidates, key=lambda nm: (day_task_used(nm), nm.lower()))

            for nm in ordered:
                if day_need <= 0:
                    break
                for run in _contig_misc_runs(dd, nm):
                    if len(run) < MIN_BLOCK_SLOTS:
                        continue
                    take = min(len(run), day_need)
                    if take < MIN_BLOCK_SLOTS:
                        continue
                    for tt in run[:take]:
                        a[(dd, tt, nm)] = task_name
                    assigned_slots += take
                    day_need -= take
                    if assigned_slots >= need_slots:
                        return assigned_slots
        return assigned_slots

    # Targets are in HOURS; convert to SLOTS (0.5h each)
    target_book_slots = int(round(float(tpl.weekly_targets.get("Bookings", 0.0) or 0.0) * 2))
    target_emis_slots = int(round(float(tpl.weekly_targets.get("EMIS", 0.0) or 0.0) * 2))
    target_doc_slots  = int(round(float(tpl.weekly_targets.get("Docman", 0.0) or 0.0) * 2))

    cur_book = _count_slots("Bookings")
    cur_emis = _count_slots("EMIS")
    cur_doc  = _count_slots("Docman")

    need_book = max(0, target_book_slots - cur_book)
    need_emis = max(0, target_emis_slots - cur_emis)
    need_doc  = max(0, target_doc_slots  - cur_doc)

    slgp_bookers = [nm for nm in staff_names if staff_by_name[nm].home == "SLGP" and staff_by_name[nm].can_bookings]
    jenbgs_emis  = [nm for nm in staff_names if staff_by_name[nm].home in ("JEN","BGS") and staff_by_name[nm].can_emis]
    jenbgs_doc   = [nm for nm in staff_names if staff_by_name[nm].home in ("JEN","BGS") and staff_by_name[nm].can_docman]

    _assign_from_runs("Docman", need_doc, jenbgs_doc)
    _assign_from_runs("EMIS", need_emis, jenbgs_emis)
    got_book = _assign_from_runs("Bookings", need_book, slgp_bookers)

    if got_book < need_book:
        any_bookers = [nm for nm in staff_names if staff_by_name[nm].can_bookings]
        _assign_from_runs("Bookings", need_book - got_book, any_bookers)


    # --------------------------------------------------
    # FINAL TARGET ENFORCEMENT FROM TEMPLATE MATRICES
    # Phones and Bookings hourly targets are checked from the template first;
    # EMIS/Docman weekly targets are then filled from remaining non-fixed capacity.
    # This avoids leaving people on Misc_Tasks while template targets are short.
    # --------------------------------------------------
    PROTECTED_TASKS = {"Break", "Email_Box", "Awaiting_PSA_Admin"}
    def _is_fixed_role(task: str) -> bool:
        return str(task or "").startswith("FrontDesk_") or str(task or "").startswith("Triage_Admin_") or task in PROTECTED_TASKS

    def _can_do_task(nm: str, task: str, dd: date, tt: time) -> bool:
        st = staff_by_name[nm]
        if holiday_kind(nm, dd, tpl.hols):
            return False
        if not is_working(hours_map, dd, tt, nm):
            return False
        if on_break(nm, dd, tt):
            return False
        current = a.get((dd, tt, nm))
        if _is_fixed_role(current):
            return False
        # Once phones have been allocated, do not let bookings/admin targets steal them back.
        if current == "Phones" and task != "Phones":
            return False
        if task == "Phones":
            return bool(st.can_phones)
        if task == "Bookings":
            return bool(st.can_bookings)
        if task == "EMIS":
            return bool(st.can_emis)
        if task == "Docman":
            return bool(st.can_docman)
        return True

    def _task_actual(dd: date, tt: time, task: str) -> int:
        return sum(1 for nm in staff_names if a.get((dd, tt, nm)) == task)

    def _choose_reassign_candidate(dd: date, tt: time, task: str, prefer_sites: Optional[List[str]]=None) -> Optional[str]:
        cands = [nm for nm in staff_names if _can_do_task(nm, task, dd, tt) and a.get((dd, tt, nm)) != task]
        if not cands:
            return None
        def score(nm: str):
            st = staff_by_name[nm]
            cur = a.get((dd, tt, nm))
            cur_rank = {"Misc_Tasks": 0, "EMIS": 1, "Docman": 1, "Bookings": 2, "Phones": 3}.get(cur, 1)
            pref = 0 if (prefer_sites and st.home in prefer_sites) else 1
            w = -task_weight(st, task)
            return (cur_rank, pref, w, nm.lower())
        cands.sort(key=score)
        return cands[0]

    # Phones hourly target
    for dd in dates:
        for tt in slots:
            req = phones_required(tpl, dd, tt)
            while req and _task_actual(dd, tt, "Phones") < req:
                nm = _choose_reassign_candidate(dd, tt, "Phones")
                if not nm:
                    gaps.append((dd, tt, "Phones", f"Target short after optimisation: {req - _task_actual(dd, tt, 'Phones')}"))
                    break
                a[(dd, tt, nm)] = "Phones"

    # Bookings hourly target
    for dd in dates:
        dn = day_name(dd)
        for tt in slots:
            hour_key = time(tt.hour, 0)
            req = int(tpl.bookings_targets.get((dn, hour_key), 0) or 0)
            while req and _task_actual(dd, tt, "Bookings") < req:
                nm = _choose_reassign_candidate(dd, tt, "Bookings", prefer_sites=["SLGP"])
                if not nm:
                    gaps.append((dd, tt, "Bookings", f"Target short after optimisation: {req - _task_actual(dd, tt, 'Bookings')}"))
                    break
                a[(dd, tt, nm)] = "Bookings"

    # Weekly EMIS / Docman targets from remaining capacity
    def _count_task_hours(task: str) -> float:
        return sum(0.5 for dd in dates for tt in slots for nm in staff_names if a.get((dd,tt,nm)) == task)

    for task, pref_sites in (("Docman", ["JEN","BGS"]), ("EMIS", ["JEN","BGS"])):
        target_hours = float(tpl.weekly_targets.get(task, 0.0) or 0.0)
        while target_hours and _count_task_hours(task) < target_hours:
            chosen = None
            for dd in dates:
                for tt in slots:
                    nm = _choose_reassign_candidate(dd, tt, task, prefer_sites=pref_sites)
                    if nm and a.get((dd,tt,nm)) == "Misc_Tasks":
                        chosen = (dd, tt, nm)
                        break
                if chosen:
                    break
            if not chosen:
                gaps.append((dates[-1], slots[-1], task, f"Weekly target short after optimisation: {round(target_hours - _count_task_hours(task), 2)} hours"))
                break
            dd, tt, nm = chosen
            a[(dd, tt, nm)] = task
    return a, breaks, gaps, dates, slots, hours_map

# --------------------------------------------------
# v37_19 SAFE PATCH — stability, target and rota-rules wrapper
# Keeps the last working scheduler but applies the operational rules requested:
#   * Bookings must be in blocks of at least 1 hour (2 x 30-min slots).
#   * EMIS/Docman may still be shorter than 1 hour.
#   * Phones must not exceed 3 hours continuously; a break does not reset phones.
#     There must be at least 1 hour off phones before another phones block.
#   * Front desk weights should win where a higher-weight FD person is free/misc,
#     especially where moving the current FD person helps phone targets.
# --------------------------------------------------
_base_schedule_week_v37_9 = schedule_week_v37_9

def schedule_week(tpl: TemplateData, wk_start: date):
    a, breaks, gaps, dates, slots, hours_map = _base_schedule_week_v37_9(tpl, wk_start)
    staff_by_name = {s.name: s for s in tpl.staff}
    staff_names = [s.name for s in tpl.staff]

    def _on_break(nm: str, dd: date, tt: time) -> bool:
        return nm in breaks.get((dd, tt), set())

    def _protected(task: str) -> bool:
        task = str(task or "")
        return task.startswith("FrontDesk_") or task.startswith("Triage_Admin_") or task in {"Email_Box", "Awaiting_PSA_Admin", "Break", "Holiday", "Sick", "Bank Holiday"}

    def _can_work(nm: str, dd: date, tt: time) -> bool:
        return (not holiday_kind(nm, dd, tpl.hols)) and is_working(hours_map, dd, tt, nm) and not _on_break(nm, dd, tt)

    def _capable(nm: str, task: str) -> bool:
        st = staff_by_name[nm]
        if task == "Phones":
            return bool(st.can_phones)
        if task == "Bookings":
            return bool(st.can_bookings)
        if task == "EMIS":
            return bool(st.can_emis)
        if task == "Docman":
            return bool(st.can_docman)
        if task.startswith("FrontDesk_"):
            return bool(st.can_frontdesk)
        return True

    def _can_overwrite_for(nm: str, dd: date, tt: time, task: str, allow_booking_steal: bool=False) -> bool:
        if not _can_work(nm, dd, tt) or not _capable(nm, task):
            return False
        cur = a.get((dd, tt, nm), "")
        if _protected(cur):
            return False
        if task != "Phones" and cur == "Phones":
            return False
        if task == "Phones" and cur == "Bookings" and not allow_booking_steal:
            return False
        return True

    def _phone_indices(nm: str, dd: date, extra_idx: Optional[int]=None, remove_idx: Optional[int]=None) -> List[int]:
        out = []
        for i, tt in enumerate(slots):
            is_phone = a.get((dd, tt, nm)) == "Phones"
            if extra_idx is not None and i == extra_idx:
                is_phone = True
            if remove_idx is not None and i == remove_idx:
                is_phone = False
            if is_phone:
                out.append(i)
        return out

    def _phone_ok_if(nm: str, dd: date, idx: int) -> bool:
        inds = _phone_indices(nm, dd, extra_idx=idx)
        if not inds:
            return True
        cluster = [inds[0]]
        clusters = []
        for x in inds[1:]:
            # A gap of only one non-phone slot (e.g. a 30-min break) does not reset phones.
            # Need at least two non-phone slots = 1 hour away from phones.
            if x - cluster[-1] <= 2:
                cluster.append(x)
            else:
                clusters.append(cluster)
                cluster = [x]
        clusters.append(cluster)
        return all(len(c) <= 6 for c in clusters)  # 6 half-hours = 3 hours

    def _phone_actual(dd: date, tt: time) -> int:
        return sum(1 for nm in staff_names if a.get((dd, tt, nm)) == "Phones")

    def _phone_required(dd: date, tt: time) -> int:
        return int(phones_required(tpl, dd, tt) or 0)

    # 1) Prefer higher weighted front desk staff in a stable band-level swap.
    #    This catches cases like: Mandy on Misc, Christine on FD, phones short -> Mandy FD, Christine Phones.
    fd_bands_by_site = {"SLGP": FD_BANDS, "JEN": FD_BANDS_FLEX, "BGS": FD_BANDS_FLEX}
    for dd in dates:
        for site in SITES:
            role = f"FrontDesk_{site}"
            for bs, be in fd_bands_by_site.get(site, FD_BANDS):
                band_slots = [tt for tt in slots if bs <= tt < be]
                if not band_slots:
                    continue
                holders = [nm for nm in staff_names if sum(1 for tt in band_slots if a.get((dd,tt,nm)) == role) >= max(1, len(band_slots)//2)]
                if not holders:
                    continue
                current = holders[0]
                cur_w = int(staff_by_name[current].weights.get("FrontDesk", 3))

                def fd_candidate_ok(nm: str) -> bool:
                    if nm == current:
                        return False
                    st = staff_by_name[nm]
                    if st.home != site or not st.can_frontdesk:
                        return False
                    if int(st.weights.get("FrontDesk", 3)) <= cur_w:
                        return False
                    # Candidate must be available for the whole band and not already holding protected work.
                    for tt in band_slots:
                        if not _can_work(nm, dd, tt):
                            return False
                        if _protected(a.get((dd,tt,nm), "")):
                            return False
                    return True

                cands = [nm for nm in staff_names if fd_candidate_ok(nm)]
                if not cands:
                    continue

                def fd_cand_score(nm: str):
                    # prefer someone sitting on Misc for most of the band, then highest FD weight
                    misc_count = sum(1 for tt in band_slots if a.get((dd,tt,nm)) in ("", "Misc_Tasks"))
                    fdw = -int(staff_by_name[nm].weights.get("FrontDesk", 3))
                    return (-misc_count, fdw, nm.lower())

                new_fd = sorted(cands, key=fd_cand_score)[0]
                # Only swap where it improves phone cover or removes avoidable misc.
                phone_help = any(_phone_actual(dd, tt) < _phone_required(dd, tt) for tt in band_slots)
                new_fd_misc = sum(1 for tt in band_slots if a.get((dd,tt,new_fd), "") in ("", "Misc_Tasks")) >= max(1, len(band_slots)//2)
                if not (phone_help or new_fd_misc):
                    continue
                for tt in band_slots:
                    old_task = a.get((dd, tt, new_fd), "Misc_Tasks") or "Misc_Tasks"
                    a[(dd, tt, new_fd)] = role
                    if phone_help and _capable(current, "Phones") and _can_work(current, dd, tt) and _phone_actual(dd, tt) < _phone_required(dd, tt):
                        idx = slots.index(tt)
                        if _phone_ok_if(current, dd, idx):
                            a[(dd, tt, current)] = "Phones"
                        else:
                            a[(dd, tt, current)] = old_task if not _protected(old_task) else "Misc_Tasks"
                    else:
                        a[(dd, tt, current)] = old_task if not _protected(old_task) else "Misc_Tasks"

    # 2) Remove phone allocations that breach max 3 continuous hours / <1h gap.
    for dd in dates:
        for nm in staff_names:
            inds = _phone_indices(nm, dd)
            if not inds:
                continue
            clusters = []
            cluster = [inds[0]]
            for x in inds[1:]:
                if x - cluster[-1] <= 2:
                    cluster.append(x)
                else:
                    clusters.append(cluster)
                    cluster = [x]
            clusters.append(cluster)
            for cluster in clusters:
                if len(cluster) > 6:
                    for idx in cluster[6:]:
                        tt = slots[idx]
                        if a.get((dd,tt,nm)) == "Phones":
                            a[(dd,tt,nm)] = "Misc_Tasks"

    # 3) Refill phone hourly targets without breaching phone rotation, preferring not to steal bookings.
    for dd in dates:
        for idx, tt in enumerate(slots):
            req = _phone_required(dd, tt)
            while req and _phone_actual(dd, tt) < req:
                cands = [nm for nm in staff_names if _can_overwrite_for(nm, dd, tt, "Phones", allow_booking_steal=False) and a.get((dd,tt,nm)) != "Phones" and _phone_ok_if(nm, dd, idx)]
                if not cands:
                    cands = [nm for nm in staff_names if _can_overwrite_for(nm, dd, tt, "Phones", allow_booking_steal=True) and a.get((dd,tt,nm)) != "Phones" and _phone_ok_if(nm, dd, idx)]
                if not cands:
                    gaps.append((dd, tt, "Phones", f"Target short after phone-rotation rule: {req - _phone_actual(dd, tt)}"))
                    break
                def phone_score(nm: str):
                    st = staff_by_name[nm]
                    cur = a.get((dd,tt,nm), "")
                    cur_rank = {"Misc_Tasks":0, "EMIS":1, "Docman":1, "Bookings":4}.get(cur, 2)
                    used_today = sum(1 for x in slots if a.get((dd,x,nm)) == "Phones")
                    return (cur_rank, used_today, -int(st.weights.get("Phones",3)), nm.lower())
                chosen = sorted(cands, key=phone_score)[0]
                a[(dd, tt, chosen)] = "Phones"

    # 4) Bookings must not appear as isolated 30-minute allocations.
    def _booking_runs(dd: date, nm: str):
        runs = []
        cur = []
        for i, tt in enumerate(slots):
            if a.get((dd,tt,nm)) == "Bookings":
                cur.append(i)
            else:
                if cur:
                    runs.append(cur); cur=[]
        if cur:
            runs.append(cur)
        return runs

    def _can_booking_slot(nm: str, dd: date, idx: int) -> bool:
        tt = slots[idx]
        return _can_overwrite_for(nm, dd, tt, "Bookings", allow_booking_steal=False) and a.get((dd,tt,nm)) != "Phones"

    for dd in dates:
        for nm in staff_names:
            for run in list(_booking_runs(dd, nm)):
                if len(run) >= 2:
                    continue
                idx = run[0]
                extended = False
                for adj in (idx-1, idx+1):
                    if 0 <= adj < len(slots) and _can_booking_slot(nm, dd, adj):
                        a[(dd, slots[adj], nm)] = "Bookings"
                        extended = True
                        break
                if not extended:
                    a[(dd, slots[idx], nm)] = "Misc_Tasks"

    # 5) Refill bookings hourly targets using 1-hour blocks only.
    def _booking_actual(dd: date, tt: time) -> int:
        return sum(1 for nm in staff_names if a.get((dd, tt, nm)) == "Bookings")

    for dd in dates:
        dn = day_name(dd)
        for idx, tt in enumerate(slots):
            req = int(tpl.bookings_targets.get((dn, time(tt.hour,0)), 0) or 0)
            while req and _booking_actual(dd, tt) < req:
                block_options = []
                if idx + 1 < len(slots):
                    block_options.append([idx, idx+1])
                if idx - 1 >= 0:
                    block_options.append([idx-1, idx])
                candidates = []
                for nm in staff_names:
                    if not _capable(nm, "Bookings"):
                        continue
                    # Must add a NEW bookings person to the target half-hour slot;
                    # otherwise the while loop would not progress.
                    if a.get((dd, tt, nm)) == "Bookings":
                        continue
                    for block in block_options:
                        if all(_can_booking_slot(nm, dd, j) or a.get((dd,slots[j],nm)) == "Bookings" for j in block):
                            candidates.append((nm, block))
                            break
                if not candidates:
                    gaps.append((dd, tt, "Bookings", f"Target short after 1-hour booking-block rule: {req - _booking_actual(dd, tt)}"))
                    break
                def book_score(item):
                    nm, block = item
                    st = staff_by_name[nm]
                    pref = 0 if st.home == "SLGP" else 1
                    overwrite_cost = sum({"Misc_Tasks":0, "EMIS":1, "Docman":1, "":0}.get(a.get((dd,slots[j],nm), ""), 2) for j in block)
                    used = sum(1 for x in slots if a.get((dd,x,nm)) == "Bookings")
                    return (pref, overwrite_cost, used, -int(st.weights.get("Bookings",3)), nm.lower())
                nm, block = sorted(candidates, key=book_score)[0]
                for j in block:
                    a[(dd, slots[j], nm)] = "Bookings"

    # 6) Final pass: if fixing bookings created a new isolated block somehow, remove it rather than leaving 30 mins.
    for dd in dates:
        for nm in staff_names:
            for run in list(_booking_runs(dd, nm)):
                if len(run) < 2:
                    for idx in run:
                        a[(dd, slots[idx], nm)] = "Misc_Tasks"

    # 7) Front desk weight clean-up, slot by slot.
    # If a higher-weight FD person is sitting on Misc at their own site, put them on FD
    # and release the lower-weight FD holder. This is deliberately after the main
    # rota build so it catches cases like Mandy left on Misc on Friday.
    for dd in dates:
        for tt in slots:
            for site in SITES:
                role = f"FrontDesk_{site}"
                holders = [nm for nm in staff_names if a.get((dd, tt, nm)) == role]
                if not holders:
                    continue
                current = holders[0]
                cur_w = int(staff_by_name[current].weights.get("FrontDesk", 3))
                cands = []
                for nm in staff_names:
                    if nm == current:
                        continue
                    st = staff_by_name[nm]
                    if st.home != site or not st.can_frontdesk:
                        continue
                    # Equal FD weight is still worth swapping when the candidate is otherwise on Misc,
                    # because the released FD holder may be able to cover Phones/EMIS/Docman.
                    if int(st.weights.get("FrontDesk", 3)) < cur_w:
                        continue
                    if not _can_work(nm, dd, tt):
                        continue
                    if a.get((dd, tt, nm), "") != "Misc_Tasks":
                        continue
                    cands.append(nm)
                if not cands:
                    continue
                def _fd_slot_score(nm: str):
                    st = staff_by_name[nm]
                    # Prefer the highest FD weight, then the person with the fewest other useful task skills
                    # so multi-skilled staff are released to Phones/EMIS/Docman/Bookings.
                    other_skills = int(bool(st.can_phones)) + int(bool(st.can_bookings)) + int(bool(st.can_emis)) + int(bool(st.can_docman))
                    return (-int(st.weights.get("FrontDesk", 3)), other_skills, nm.lower())
                new_fd = sorted(cands, key=_fd_slot_score)[0]
                a[(dd, tt, new_fd)] = role
                idx = slots.index(tt)
                # Use the released person for phones if the template target is short and the phone rule allows it;
                # otherwise leave them as Misc for the admin-target clean-up below.
                if _capable(current, "Phones") and _can_work(current, dd, tt) and _phone_actual(dd, tt) < _phone_required(dd, tt) and _phone_ok_if(current, dd, idx):
                    a[(dd, tt, current)] = "Phones"
                else:
                    a[(dd, tt, current)] = "Misc_Tasks"

    # 8) Misc must be a last resort. Convert remaining Misc to EMIS/Docman while weekly
    # targets are still short. EMIS/Docman are allowed to be shorter than 1 hour, so this
    # may use single 30-minute gaps that genuinely do not fit anything else.
    def _task_slots(task_name: str) -> int:
        return sum(1 for dd in dates for tt in slots for nm in staff_names if a.get((dd, tt, nm)) == task_name)

    def _weekly_target_slots(task_name: str) -> int:
        return int(round(float(tpl.weekly_targets.get(task_name, 0.0) or 0.0) * 2))

    def _admin_need(task_name: str) -> int:
        return max(0, _weekly_target_slots(task_name) - _task_slots(task_name))

    def _misc_admin_candidates(task_name: str):
        out = []
        for dd in dates:
            for tt in slots:
                for nm in staff_names:
                    if a.get((dd, tt, nm)) != "Misc_Tasks":
                        continue
                    if not _can_overwrite_for(nm, dd, tt, task_name, allow_booking_steal=False):
                        continue
                    st = staff_by_name[nm]
                    # Prefer JEN/BGS for Docman/EMIS, but allow SLGP if it is the only spare capacity.
                    site_pref = 0 if st.home in ("JEN", "BGS") else 1
                    weight = -int(st.weights.get(task_name, 3))
                    out.append((site_pref, weight, dd, tt, nm))
        out.sort(key=lambda x: (x[0], x[1], x[2], x[3], x[4].lower()))
        return out

    # Allocate to whichever target has the largest remaining shortfall, repeating until no useful Misc remains.
    safety = 0
    while safety < 10000 and (_admin_need("Docman") > 0 or _admin_need("EMIS") > 0):
        safety += 1
        doc_need = _admin_need("Docman")
        emis_need = _admin_need("EMIS")
        task = "Docman" if doc_need >= emis_need and doc_need > 0 else "EMIS"
        cands = _misc_admin_candidates(task)
        if not cands and task == "Docman" and emis_need > 0:
            task = "EMIS"
            cands = _misc_admin_candidates(task)
        elif not cands and task == "EMIS" and doc_need > 0:
            task = "Docman"
            cands = _misc_admin_candidates(task)
        if not cands:
            break
        _, _, dd, tt, nm = cands[0]
        a[(dd, tt, nm)] = task

    # 9) Record an explicit gap note if Misc remains only because no eligible target task fits.
    for task in ("Docman", "EMIS"):
        short = _admin_need(task)
        if short > 0:
            gaps.append((dates[-1], slots[-1], task, f"Weekly target still short after converting all eligible Misc: {short * 0.5:.1f} hours"))

    return a, breaks, gaps, dates, slots, hours_map

# --------------------------------------------------
# v37_19 SAFE PATCH - CONTINUITY / ADMIN PRIORITY TIDY
# --------------------------------------------------
# This keeps the stable safe engine, then tidies avoidable choppiness:
# * Misc remains last resort.
# * EMIS/Docman are made continuous where possible.
# * Breaks can move out of the middle of an admin block where a flexible slot exists.
# * If EMIS/Docman are materially short, FD can be rebalanced so a FD-suitable,
#   non-admin person covers front desk and the admin-capable FD holder covers EMIS/Docman.
# --------------------------------------------------
_schedule_week_before_continuity_tidy = schedule_week


def schedule_week(tpl: TemplateData, wk_start: date):
    a, breaks, gaps, dates, slots, hours_map = _schedule_week_before_continuity_tidy(tpl, wk_start)
    staff_by_name = {s.name: s for s in tpl.staff}
    staff_names = [s.name for s in tpl.staff]
    ADMIN_TASKS = ("EMIS", "Docman")

    def _on_break(nm, dd, tt):
        return nm in breaks.get((dd, tt), set())

    def _can_work(nm, dd, tt):
        return (not holiday_kind(nm, dd, tpl.hols)) and is_working(hours_map, dd, tt, nm) and not _on_break(nm, dd, tt)

    def _hard(task):
        task = str(task or "")
        return task.startswith("Triage_Admin_") or task in {"Awaiting_PSA_Admin", "Phones", "Break", "Holiday", "Sick", "Bank Holiday"}

    def _very_hard(task):
        task = str(task or "")
        return task.startswith("Triage_Admin_") or task in {"Awaiting_PSA_Admin", "Phones", "Break", "Holiday", "Sick", "Bank Holiday"}

    def _capable(nm, task):
        st = staff_by_name[nm]
        if task == "EMIS": return bool(st.can_emis)
        if task == "Docman": return bool(st.can_docman)
        if task == "Bookings": return bool(st.can_bookings)
        if task == "Phones": return bool(st.can_phones)
        if task.startswith("FrontDesk_"): return bool(st.can_frontdesk)
        return True

    def _target_slots(task):
        return int(round(float(tpl.weekly_targets.get(task, 0.0) or 0.0) * 2))

    def _actual_slots(task):
        return sum(1 for dd in dates for tt in slots for nm in staff_names if a.get((dd, tt, nm)) == task)

    def _need(task):
        return max(0, _target_slots(task) - _actual_slots(task))

    def _admin_task_to_use(nm=None):
        # Choose the task with the larger remaining shortfall, constrained by capability if nm supplied.
        candidates = []
        for task in ADMIN_TASKS:
            if nm is not None and not _capable(nm, task):
                continue
            candidates.append((_need(task), task))
        candidates.sort(reverse=True)
        return candidates[0][1] if candidates and candidates[0][0] > 0 else None

    # 1) Admin-priority front desk rebalance by stable bands.
    # Example this is intended to catch: Mandy can cover FD but not EMIS; Madison can do EMIS and is holding FD.
    fd_bands_by_site = {"SLGP": FD_BANDS, "JEN": FD_BANDS_FLEX, "BGS": FD_BANDS_FLEX}
    for dd in dates:
        for site in SITES:
            role = f"FrontDesk_{site}"
            for bs, be in fd_bands_by_site.get(site, FD_BANDS):
                band = [tt for tt in slots if bs <= tt < be]
                if not band:
                    continue
                holders = []
                for nm in staff_names:
                    c = sum(1 for tt in band if a.get((dd, tt, nm)) == role)
                    if c:
                        holders.append((c, nm))
                if not holders:
                    continue
                holders.sort(reverse=True)
                current = holders[0][1]
                task_for_current = _admin_task_to_use(current)
                if not task_for_current:
                    continue
                cur_fd_w = int(staff_by_name[current].weights.get("FrontDesk", 3))

                alts = []
                for nm in staff_names:
                    if nm == current:
                        continue
                    st = staff_by_name[nm]
                    if st.home != site or not st.can_frontdesk:
                        continue
                    if int(st.weights.get("FrontDesk", 3)) < cur_fd_w:
                        continue
                    # Prefer FD-capable staff who cannot do the short admin task.
                    if _capable(nm, task_for_current):
                        continue
                    ok = 0
                    release_cost = 0
                    for tt in band:
                        cur = a.get((dd, tt, nm), "") or ""
                        if _can_work(nm, dd, tt) and not _very_hard(cur) and not str(cur).startswith("FrontDesk_"):
                            ok += 1
                            # Email is allowed to be interrupted only where admin targets are short, but costs more.
                            release_cost += {"":0, "Misc_Tasks":0, "EMIS":1, "Docman":1, "Bookings":2, "Email_Box":3}.get(cur, 4)
                    if ok >= max(1, len(band)-1):
                        alts.append((release_cost, -int(st.weights.get("FrontDesk",3)), nm))
                if not alts:
                    continue
                alts.sort(key=lambda x: (x[0], x[1], x[2].lower()))
                new_fd = alts[0][2]
                for tt in band:
                    if not _can_work(new_fd, dd, tt) or not _can_work(current, dd, tt):
                        continue
                    cur_new = a.get((dd, tt, new_fd), "") or ""
                    if _very_hard(cur_new) or str(cur_new).startswith("FrontDesk_"):
                        continue
                    a[(dd, tt, new_fd)] = role
                    # Release current into the short admin task; if target fills, keep their previous flexible task.
                    task_for_current = _admin_task_to_use(current)
                    if task_for_current:
                        a[(dd, tt, current)] = task_for_current
                    else:
                        a[(dd, tt, current)] = cur_new if cur_new in ("EMIS", "Docman", "Bookings") else "Misc_Tasks"

    # Helpers for continuity passes.
    def _flex_for_admin(nm, dd, tt, task):
        if not _can_work(nm, dd, tt) or not _capable(nm, task):
            return False
        cur = a.get((dd, tt, nm), "") or ""
        if _hard(cur) or str(cur).startswith("FrontDesk_"):
            return False
        # Do not casually steal bookings here. FD rebalance above is the controlled way to release admin time.
        return cur in ("", "Misc_Tasks", "EMIS", "Docman")

    # 2) Fill a one-slot flexible gap between same admin task: EMIS/Misc/EMIS -> EMIS/EMIS/EMIS.
    for dd in dates:
        for nm in staff_names:
            for i in range(1, len(slots)-1):
                left = a.get((dd, slots[i-1], nm))
                right = a.get((dd, slots[i+1], nm))
                cur = a.get((dd, slots[i], nm), "") or ""
                if left in ADMIN_TASKS and left == right and cur in ("", "Misc_Tasks") and _flex_for_admin(nm, dd, slots[i], left):
                    a[(dd, slots[i], nm)] = left

    # 3) If an admin fragment is isolated, try to extend it by 30 minutes next to it rather than leaving a lone slot.
    for dd in dates:
        for nm in staff_names:
            for task in ADMIN_TASKS:
                inds = [i for i, tt in enumerate(slots) if a.get((dd, tt, nm)) == task]
                for i in list(inds):
                    prev_same = i > 0 and a.get((dd, slots[i-1], nm)) == task
                    next_same = i + 1 < len(slots) and a.get((dd, slots[i+1], nm)) == task
                    if prev_same or next_same:
                        continue
                    # Prefer extending forward, then backward, but only into genuine flexible time.
                    for j in (i+1, i-1):
                        if 0 <= j < len(slots) and _flex_for_admin(nm, dd, slots[j], task):
                            if a.get((dd, slots[j], nm), "") in ("", "Misc_Tasks"):
                                a[(dd, slots[j], nm)] = task
                                break

    # 4) Allocate remaining Misc/blank to EMIS/Docman using continuous runs first.
    def _candidate_runs(task):
        runs = []
        for dd in dates:
            for nm in staff_names:
                if not _capable(nm, task):
                    continue
                cur_run = []
                for i, tt in enumerate(slots):
                    if _flex_for_admin(nm, dd, tt, task) and (a.get((dd, tt, nm), "") in ("", "Misc_Tasks")):
                        cur_run.append(i)
                    else:
                        if cur_run:
                            runs.append((dd, nm, cur_run)); cur_run=[]
                if cur_run:
                    runs.append((dd, nm, cur_run))
        return runs

    safety = 0
    while safety < 500 and (_need("EMIS") > 0 or _need("Docman") > 0):
        safety += 1
        task = "EMIS" if _need("EMIS") >= _need("Docman") else "Docman"
        if _need(task) <= 0:
            task = "Docman" if task == "EMIS" else "EMIS"
        if _need(task) <= 0:
            break
        runs = _candidate_runs(task)
        if not runs:
            other = "Docman" if task == "EMIS" else "EMIS"
            runs = _candidate_runs(other)
            if not runs or _need(other) <= 0:
                break
            task = other
        scored = []
        for dd, nm, run in runs:
            st = staff_by_name[nm]
            # Prefer longer runs and runs adjacent to same task so they become continuous.
            before = run[0] > 0 and a.get((dd, slots[run[0]-1], nm)) == task
            after = run[-1] + 1 < len(slots) and a.get((dd, slots[run[-1]+1], nm)) == task
            adj = -int(before or after)
            site_pref = 0 if st.home in ("JEN", "BGS") else 1
            scored.append((-len(run), adj, site_pref, -int(st.weights.get(task,3)), dd, run[0], nm, run))
        scored.sort(key=lambda x: (x[0], x[1], x[2], x[3], x[4], x[5], x[6].lower()))
        _, _, _, _, dd, _start, nm, run = scored[0]
        need = _need(task)
        chunk_len = min(len(run), need)
        if len(run) >= 2 and chunk_len == 1 and need > 1:
            chunk_len = 2
        if chunk_len > 6 and need > 6:
            chunk_len = 6
        for i in run[:chunk_len]:
            a[(dd, slots[i], nm)] = task

    # 5) Move breaks if they split the same admin task and a flexible slot exists nearby.
    def _move_break(dd, nm, old_tt, new_tt):
        if (dd, old_tt) in breaks and nm in breaks[(dd, old_tt)]:
            breaks[(dd, old_tt)].discard(nm)
            if not breaks[(dd, old_tt)]:
                breaks.pop((dd, old_tt), None)
        breaks.setdefault((dd, new_tt), set()).add(nm)

    for dd in dates:
        for nm in staff_names:
            bts = [tt for (bd, tt), names in list(breaks.items()) if bd == dd and nm in names]
            for bt in bts:
                if bt not in slots:
                    continue
                i = slots.index(bt)
                if i <= 0 or i >= len(slots)-1:
                    continue
                left = a.get((dd, slots[i-1], nm))
                right = a.get((dd, slots[i+1], nm))
                if left not in ADMIN_TASKS or left != right:
                    continue
                choices = []
                for j, nt in enumerate(slots):
                    if j == i:
                        continue
                    cur = a.get((dd, nt, nm), "") or ""
                    if cur not in ("", "Misc_Tasks"):
                        continue
                    if not is_working(hours_map, dd, nt, nm) or holiday_kind(nm, dd, tpl.hols):
                        continue
                    stt, _endt = shift_window(hours_map, dd, nm)
                    if stt and stt >= time(11,30):
                        if not (time(14,0) <= nt <= time(16,30)):
                            continue
                    else:
                        if not (time(12,0) <= nt <= time(15,30)):
                            continue
                    choices.append((abs(j-i), j, nt))
                if choices:
                    choices.sort()
                    _, _j, nt = choices[0]
                    _move_break(dd, nm, bt, nt)
                    a[(dd, bt, nm)] = left

    # 6) Final Misc last-resort note if targets are still short.
    for task in ADMIN_TASKS:
        if _need(task) > 0:
            gaps.append((dates[-1], slots[-1], task, f"Weekly target still short after continuity/admin-priority tidy: {_need(task) * 0.5:.1f} hours"))

    return a, breaks, gaps, dates, slots, hours_map


# --------------------------------------------------
# v37_19 SAFE PATCH - CONTINUITY / FLEXIBLE BREAKS FINAL TIDY
# --------------------------------------------------
_prev_schedule_week_continuity = schedule_week

def schedule_week(tpl: TemplateData, wk_start: date):
    a, breaks, gaps, dates, slots, hours_map = _prev_schedule_week_continuity(tpl, wk_start)
    staff_by_name = {s.name: s for s in tpl.staff}
    staff_names = [s.name for s in tpl.staff]
    ADMIN_TASKS = ("EMIS", "Docman")

    def _task(nm, dd, tt):
        return str(a.get((dd, tt, nm), "") or "")

    def _is_holiday(nm, dd):
        return bool(holiday_kind(nm, dd, tpl.hols))

    def _break_at(nm, dd):
        return [tt for (bd, tt), names in list(breaks.items()) if bd == dd and nm in names]

    def _remove_break(nm, dd, tt):
        if (dd, tt) in breaks and nm in breaks[(dd, tt)]:
            breaks[(dd, tt)].discard(nm)
            if not breaks[(dd, tt)]:
                breaks.pop((dd, tt), None)

    def _set_single_break(nm, dd, new_tt):
        for old in _break_at(nm, dd):
            _remove_break(nm, dd, old)
        breaks.setdefault((dd, new_tt), set()).add(nm)

    def _dt(dd, tt):
        return dt_of(dd, tt)

    def _middle_third_candidates(nm, dd):
        stt, endt = shift_window(hours_map, dd, nm)
        if not stt or not endt:
            return []
        dur = (_dt(dd, endt) - _dt(dd, stt)).total_seconds() / 3600.0
        if dur < 6.0 or _is_holiday(nm, dd):
            return []
        start_dt = _dt(dd, stt)
        end_dt = _dt(dd, endt)
        lo = start_dt + (end_dt - start_dt) / 3
        hi = start_dt + (end_dt - start_dt) * 2 / 3
        out = []
        for tt in slots:
            tt_dt = _dt(dd, tt)
            # The break must be fully inside the shift and start broadly in the middle third.
            if tt < stt or add_minutes(tt, 30) > endt:
                continue
            if lo <= tt_dt <= hi:
                out.append(tt)
        # If a short/odd shift gives no exact half-hour in the third, use any internal slot away from edges.
        if not out:
            for tt in slots:
                tt_dt = _dt(dd, tt)
                if tt < stt or add_minutes(tt, 30) > endt:
                    continue
                before = (tt_dt - start_dt).total_seconds() / 3600.0
                after = (end_dt - (tt_dt + timedelta(minutes=30))).total_seconds() / 3600.0
                if before >= 1.5 and after >= 1.5:
                    out.append(tt)
        return out

    def _break_score(nm, dd, tt):
        cur = _task(nm, dd, tt)
        i = slots.index(tt) if tt in slots else -1
        left = _task(nm, dd, slots[i-1]) if i > 0 else ""
        right = _task(nm, dd, slots[i+1]) if 0 <= i < len(slots)-1 else ""
        # Breaks can move flexibly, but avoid breaking protected cover first.
        if cur.startswith("FrontDesk_") or cur.startswith("Triage_Admin_") or cur == "Awaiting_PSA_Admin":
            base = 10000
        elif cur == "Phones":
            base = 2000
        elif cur == "Email_Box":
            base = 300
        elif cur == "Bookings":
            base = 100
        elif cur in ADMIN_TASKS:
            base = 600
        elif cur in ("", "Misc_Tasks"):
            base = 0
        else:
            base = 50
        # Strongly avoid splitting a continuous admin block.
        if left in ADMIN_TASKS and left == right:
            base += 2500
        # Prefer breaks not right at the edge of an existing admin/booking block.
        if cur in ADMIN_TASKS and (left == cur or right == cur):
            base += 500
        return base

    # A) Re-place breaks in the middle third of the day, choosing the least disruptive slot.
    # This makes admin blocks less choppy and allows 12:00 starters to still receive a break.
    for dd in dates:
        for st in tpl.staff:
            nm = st.name
            cands = _middle_third_candidates(nm, dd)
            if not cands:
                # Remove accidental breaks for sub-6h/non-working days.
                for bt in _break_at(nm, dd):
                    _remove_break(nm, dd, bt)
                continue
            existing = _break_at(nm, dd)
            best = sorted(cands, key=lambda tt: (_break_score(nm, dd, tt), abs(slots.index(tt) - (len(slots)//2)), tt))[0]
            if not existing or existing[0] not in cands or _break_score(nm, dd, best) + 50 < _break_score(nm, dd, existing[0]):
                _set_single_break(nm, dd, best)

    def _on_break(nm, dd, tt):
        return nm in breaks.get((dd, tt), set())

    def _can_work(nm, dd, tt):
        return (not _is_holiday(nm, dd)) and is_working(hours_map, dd, tt, nm) and not _on_break(nm, dd, tt)

    def _capable(nm, task):
        st = staff_by_name[nm]
        if task == "EMIS": return bool(st.can_emis)
        if task == "Docman": return bool(st.can_docman)
        if task == "Bookings": return bool(st.can_bookings)
        if task == "Phones": return bool(st.can_phones)
        if task.startswith("FrontDesk_"): return bool(st.can_frontdesk)
        return True

    def _hard(cur):
        cur = str(cur or "")
        return cur.startswith("FrontDesk_") or cur.startswith("Triage_Admin_") or cur in {"Awaiting_PSA_Admin", "Email_Box", "Phones", "Break", "Holiday", "Sick", "Bank Holiday"}

    def _target_slots(task):
        return int(round(float(tpl.weekly_targets.get(task, 0.0) or 0.0) * 2))

    def _actual_slots(task):
        return sum(1 for dd in dates for tt in slots for nm in staff_names if a.get((dd, tt, nm)) == task)

    def _need(task):
        return max(0, _target_slots(task) - _actual_slots(task))

    def _booking_surplus():
        return max(0, _actual_slots("Bookings") - _target_slots("Bookings"))

    def _flex_for_admin(nm, dd, tt, task, allow_booking=False):
        if not _can_work(nm, dd, tt) or not _capable(nm, task):
            return False
        cur = _task(nm, dd, tt)
        if _hard(cur):
            return False
        if cur in ("", "Misc_Tasks", "EMIS", "Docman"):
            return True
        if allow_booking and cur == "Bookings" and _booking_surplus() > 0:
            return True
        return False

    # B) If a 30-min admin island cannot be extended, release it back to flexible time and re-use that
    # capacity in longer runs. This avoids Will-style EMIS/Misc/EMIS or one random EMIS slot later.
    def _admin_runs(dd, nm, task):
        runs = []
        cur = []
        for i, tt in enumerate(slots):
            if a.get((dd, tt, nm)) == task:
                cur.append(i)
            else:
                if cur:
                    runs.append(cur); cur = []
        if cur:
            runs.append(cur)
        return runs

    # First join one-slot gaps between the same admin task, if the gap is flexible.
    for dd in dates:
        for nm in staff_names:
            for i in range(1, len(slots)-1):
                left = a.get((dd, slots[i-1], nm))
                right = a.get((dd, slots[i+1], nm))
                if left in ADMIN_TASKS and left == right and _flex_for_admin(nm, dd, slots[i], left, allow_booking=True):
                    a[(dd, slots[i], nm)] = left

    # Then remove genuinely isolated single admin slots where there is alternative flexible capacity.
    for dd in dates:
        for nm in staff_names:
            for task in ADMIN_TASKS:
                for run in list(_admin_runs(dd, nm, task)):
                    if len(run) != 1:
                        continue
                    i = run[0]
                    prev_same = i > 0 and a.get((dd, slots[i-1], nm)) == task
                    next_same = i + 1 < len(slots) and a.get((dd, slots[i+1], nm)) == task
                    if prev_same or next_same:
                        continue
                    # Try to extend first.
                    extended = False
                    for j in (i+1, i-1):
                        if 0 <= j < len(slots) and _flex_for_admin(nm, dd, slots[j], task, allow_booking=True):
                            a[(dd, slots[j], nm)] = task
                            extended = True
                            break
                    if not extended:
                        a[(dd, slots[i], nm)] = "Misc_Tasks"

    # C) Reallocate admin shortfall into the longest continuous flexible runs first.
    # Allow surplus bookings to convert to EMIS/Docman where admin targets are short.
    def _candidate_runs(task):
        runs = []
        allow_booking = True  # admin targets take priority over bookings when EMIS/Docman are short
        for dd in dates:
            for nm in staff_names:
                if not _capable(nm, task):
                    continue
                cur_run = []
                for i, tt in enumerate(slots):
                    cur_val = _task(nm, dd, tt)
                    # For new allocation, only take genuine flexible time or bookings.
                    # Do not steal the other admin task, or EMIS/Docman can oscillate.
                    if cur_val in ("", "Misc_Tasks", "Bookings") and _flex_for_admin(nm, dd, tt, task, allow_booking=allow_booking):
                        cur_run.append(i)
                    else:
                        if cur_run:
                            runs.append((dd, nm, cur_run)); cur_run = []
                if cur_run:
                    runs.append((dd, nm, cur_run))
        return runs

    safety = 0
    while safety < 1000 and (_need("EMIS") > 0 or _need("Docman") > 0):
        safety += 1
        # Prioritise the larger shortfall, but use whichever task has a useful run.
        tasks = sorted(ADMIN_TASKS, key=lambda t: _need(t), reverse=True)
        chosen_task = None
        chosen_run = None
        for task in tasks:
            if _need(task) <= 0:
                continue
            runs = _candidate_runs(task)
            if not runs:
                continue
            scored = []
            for dd, nm, run in runs:
                st = staff_by_name[nm]
                before = run[0] > 0 and a.get((dd, slots[run[0]-1], nm)) == task
                after = run[-1] + 1 < len(slots) and a.get((dd, slots[run[-1]+1], nm)) == task
                # Prefer Madison/other high weight, longer blocks, and adjacency to same task.
                site_pref = 0 if st.home in ("JEN", "BGS") else 1
                scored.append((-len(run), -int(before or after), site_pref, -int(st.weights.get(task, 3)), dd, run[0], nm, run))
            scored.sort(key=lambda x: (x[0], x[1], x[2], x[3], x[4], x[5], x[6].lower()))
            chosen_task = task
            chosen_run = scored[0]
            break
        if not chosen_task or not chosen_run:
            break
        _, _, _, _, dd, _start, nm, run = chosen_run
        need = _need(chosen_task)
        # Use a block, not a scatter: minimum 1h where possible, max 3h per chunk to avoid huge blocks.
        take = min(len(run), need)
        if len(run) >= 2 and take == 1 and need > 1:
            take = 2
        take = min(take, 6)
        for i in run[:take]:
            a[(dd, slots[i], nm)] = chosen_task

    # D) After admin allocation, join avoidable Misc gaps one more time.
    for dd in dates:
        for nm in staff_names:
            for i in range(1, len(slots)-1):
                left = a.get((dd, slots[i-1], nm))
                right = a.get((dd, slots[i+1], nm))
                cur = _task(nm, dd, slots[i])
                if left in ADMIN_TASKS and left == right and cur in ("", "Misc_Tasks") and _flex_for_admin(nm, dd, slots[i], left, allow_booking=True):
                    a[(dd, slots[i], nm)] = left

    # E) Notes if still short after the continuity tidy.
    for task in ADMIN_TASKS:
        if _need(task) > 0:
            gaps.append((dates[-1], slots[-1], task, f"Weekly target still short after flexible-break continuity tidy: {_need(task) * 0.5:.1f} hours"))

    return a, breaks, gaps, dates, slots, hours_map

# =========================================================
# PATCH — Email required daily + protected continuous cover
# =========================================================
_prev_schedule_week_email_daily = schedule_week

def schedule_week(tpl: TemplateData, wk_start: date):
    """Final wrapper: ensure Email_Box is covered every weekday as a daily priority task.

    Rule added for Jo:
      - Email must be covered every day, ideally 10:30–15:30 as one continuous block.
      - Prefer the email site-of-day, but use any email-capable staff if needed.
      - Do not steal Front Desk, Triage, Awaiting PSA, Phones, Holiday/Sick/Bank Holiday, or Break.
      - If no full 10:30–15:30 block is possible, cover the longest available continuous block and record the gap.
    """
    a, breaks, gaps, dates, slots, hours_map = _prev_schedule_week_email_daily(tpl, wk_start)
    staff_by_name = {s.name: s for s in tpl.staff}
    staff_names = [s.name for s in tpl.staff]

    EMAIL_START = time(10, 30)
    EMAIL_END = time(15, 30)
    email_slots = [tt for tt in slots if EMAIL_START <= tt < EMAIL_END]
    HARD_FOR_EMAIL = {"Awaiting_PSA_Admin", "Phones", "Break", "Holiday", "Sick", "Bank Holiday"}
    ADMIN_OR_FLEX = {"", "Misc_Tasks", "EMIS", "Docman", "Bookings"}

    def _on_break(nm, dd, tt):
        return nm in breaks.get((dd, tt), set())

    def _cur(nm, dd, tt):
        return str(a.get((dd, tt, nm), "") or "")

    def _is_hard(cur):
        cur = str(cur or "")
        return cur.startswith("FrontDesk_") or cur.startswith("Triage_Admin_") or cur in HARD_FOR_EMAIL

    def _can_email_slot(nm, dd, tt):
        st = staff_by_name[nm]
        if not getattr(st, "can_email", False):
            return False
        if holiday_kind(nm, dd, tpl.hols):
            return False
        if not is_working(hours_map, dd, tt, nm):
            return False
        if _on_break(nm, dd, tt):
            return False
        cur = _cur(nm, dd, tt)
        if _is_hard(cur):
            return False
        return cur in ADMIN_OR_FLEX or cur == "Email_Box"

    def _covered(dd, tt):
        return any(a.get((dd, tt, nm)) == "Email_Box" for nm in staff_names)

    def _contiguous_runs(indices):
        runs = []
        cur = []
        prev = None
        for idx in indices:
            if prev is None or idx == prev + 1:
                cur.append(idx)
            else:
                if cur:
                    runs.append(cur)
                cur = [idx]
            prev = idx
        if cur:
            runs.append(cur)
        return runs

    def _candidate_runs(dd, nm):
        idxs = [slots.index(tt) for tt in email_slots if not _covered(dd, tt) and _can_email_slot(nm, dd, tt)]
        return _contiguous_runs(idxs)

    def _score_run(dd, nm, run):
        st = staff_by_name[nm]
        pref_site = email_site_for_day(dd)
        site_score = 0 if str(st.home).upper() == str(pref_site).upper() else 1
        # Prefer taking Misc/blank first; avoid breaking admin/booking unless needed.
        release_cost = 0
        for i in run:
            cur = _cur(nm, dd, slots[i])
            release_cost += {"": 0, "Misc_Tasks": 0, "Bookings": 1, "EMIS": 3, "Docman": 3, "Email_Box": 0}.get(cur, 5)
        weight = int(st.weights.get("Email", 3)) if getattr(st, "weights", None) else 3
        return (-len(run), site_score, release_cost, -weight, nm.lower(), run[0])

    # First, if a day already has partial email, extend it using the same person where possible.
    for dd in dates:
        for tt in email_slots:
            if _covered(dd, tt):
                continue
            current_email_people = [nm for nm in staff_names if any(a.get((dd, t2, nm)) == "Email_Box" for t2 in email_slots)]
            patched = False
            for nm in sorted(current_email_people, key=lambda n: (0 if staff_by_name[n].home == email_site_for_day(dd) else 1, n.lower())):
                if _can_email_slot(nm, dd, tt):
                    a[(dd, tt, nm)] = "Email_Box"
                    patched = True
                    break
            if patched:
                continue

        # Then allocate uncovered email slots as a continuous block, preferring the full 10:30–15:30 window.
        safety = 0
        while safety < 20 and any(not _covered(dd, tt) for tt in email_slots):
            safety += 1
            scored = []
            for nm in staff_names:
                if not getattr(staff_by_name[nm], "can_email", False):
                    continue
                for run in _candidate_runs(dd, nm):
                    scored.append((*_score_run(dd, nm, run), dd, nm, run))
            if not scored:
                break
            scored.sort(key=lambda x: x[:6])
            *_, dd2, nm, run = scored[0]
            for i in run:
                a[(dd2, slots[i], nm)] = "Email_Box"

        missing = [tt.strftime("%H:%M") for tt in email_slots if not _covered(dd, tt)]
        if missing:
            gaps.append((dd, EMAIL_START, "Email_Box", "Daily email cover short at: " + ", ".join(missing)))

    return a, breaks, gaps, dates, slots, hours_map

# =========================================================
# PATCH — Email daily priority v2 (can displace phones, then refill phones)
# =========================================================
_prev_schedule_week_email_daily_v2 = schedule_week

def schedule_week(tpl: TemplateData, wk_start: date):
    a, breaks, gaps, dates, slots, hours_map = _prev_schedule_week_email_daily_v2(tpl, wk_start)
    staff_by_name = {s.name: s for s in tpl.staff}
    staff_names = [s.name for s in tpl.staff]
    EMAIL_START = time(10, 30)
    EMAIL_END = time(15, 30)
    email_slots = [tt for tt in slots if EMAIL_START <= tt < EMAIL_END]

    def _on_break(nm, dd, tt):
        return nm in breaks.get((dd, tt), set())
    def _cur(nm, dd, tt):
        return str(a.get((dd, tt, nm), "") or "")
    def _covered(dd, tt):
        return any(a.get((dd, tt, nm)) == "Email_Box" for nm in staff_names)
    def _hard_for_email(cur):
        cur = str(cur or "")
        return cur.startswith("FrontDesk_") or cur.startswith("Triage_Admin_") or cur in {"Awaiting_PSA_Admin", "Break", "Holiday", "Sick", "Bank Holiday"}
    def _can_email(nm, dd, tt, allow_phones=True):
        st = staff_by_name[nm]
        if not getattr(st, "can_email", False): return False
        if holiday_kind(nm, dd, tpl.hols): return False
        if not is_working(hours_map, dd, tt, nm): return False
        if _on_break(nm, dd, tt): return False
        cur = _cur(nm, dd, tt)
        if _hard_for_email(cur): return False
        if cur in ("", "Misc_Tasks", "EMIS", "Docman", "Bookings", "Email_Box"): return True
        if allow_phones and cur == "Phones": return True
        return False
    def _email_score(nm, dd, tt):
        st = staff_by_name[nm]
        cur = _cur(nm, dd, tt)
        cost = {"":0, "Misc_Tasks":0, "Bookings":1, "EMIS":3, "Docman":3, "Phones":6, "Email_Box":0}.get(cur, 9)
        site = 0 if str(st.home).upper() == str(email_site_for_day(dd)).upper() else 1
        weight = int(st.weights.get("Email", 3)) if getattr(st, "weights", None) else 3
        # prefer continuing existing email blocks either side
        idx = slots.index(tt)
        continuity = 0
        if idx > 0 and a.get((dd, slots[idx-1], nm)) == "Email_Box": continuity -= 2
        if idx + 1 < len(slots) and a.get((dd, slots[idx+1], nm)) == "Email_Box": continuity -= 2
        return (cost, site, continuity, -weight, nm.lower())

    # force-cover each email slot where any safe email-capable person exists
    for dd in dates:
        for tt in email_slots:
            if _covered(dd, tt):
                continue
            cands = [nm for nm in staff_names if _can_email(nm, dd, tt, allow_phones=True)]
            if cands:
                chosen = sorted(cands, key=lambda nm: _email_score(nm, dd, tt))[0]
                a[(dd, tt, chosen)] = "Email_Box"

    # refill phone shortfalls created by email priority from genuinely flexible capacity
    def _phones_required(dd, tt):
        try:
            return int(phones_required(tpl, dd, tt) or 0)
        except Exception:
            return 0
    def _phones_actual(dd, tt):
        return sum(1 for nm in staff_names if a.get((dd, tt, nm)) == "Phones")
    def _can_phone(nm, dd, tt):
        st = staff_by_name[nm]
        if not getattr(st, "can_phones", False): return False
        if holiday_kind(nm, dd, tpl.hols): return False
        if not is_working(hours_map, dd, tt, nm): return False
        if _on_break(nm, dd, tt): return False
        cur = _cur(nm, dd, tt)
        if cur.startswith("FrontDesk_") or cur.startswith("Triage_Admin_") or cur in {"Email_Box", "Awaiting_PSA_Admin", "Break", "Holiday", "Sick", "Bank Holiday", "Phones"}:
            return False
        return cur in ("", "Misc_Tasks", "Bookings", "EMIS", "Docman")
    def _phone_score(nm, dd, tt):
        st = staff_by_name[nm]
        cur = _cur(nm, dd, tt)
        cost = {"":0, "Misc_Tasks":0, "Bookings":1, "EMIS":4, "Docman":4}.get(cur, 8)
        weight = int(st.weights.get("Phones", 3)) if getattr(st, "weights", None) else 3
        return (cost, -weight, nm.lower())

    for dd in dates:
        for tt in slots:
            need = _phones_required(dd, tt)
            while need > 0 and _phones_actual(dd, tt) < need:
                cands = [nm for nm in staff_names if _can_phone(nm, dd, tt)]
                if not cands:
                    break
                chosen = sorted(cands, key=lambda nm: _phone_score(nm, dd, tt))[0]
                a[(dd, tt, chosen)] = "Phones"

    # final daily email gap note, with clear impossibility reason rather than silently dropping it
    for dd in dates:
        missing = [tt.strftime("%H:%M") for tt in email_slots if not _covered(dd, tt)]
        if missing:
            gaps.append((dd, EMAIL_START, "Email_Box", "Daily email cover still not possible at: " + ", ".join(missing) + " (email-capable staff are on fixed cover, break, not working, or leave)"))
    return a, breaks, gaps, dates, slots, hours_map

# =========================================================
# PATCH — tidy duplicate email gap notes
# =========================================================
_prev_schedule_week_tidy_email_notes = schedule_week

def schedule_week(tpl: TemplateData, wk_start: date):
    a, breaks, gaps, dates, slots, hours_map = _prev_schedule_week_tidy_email_notes(tpl, wk_start)
    # Remove obsolete notes from earlier email patch layers, keeping only the final clearer message.
    clean = []
    seen = set()
    for dd, tt, task, note in gaps:
        text = str(note or "")
        if task == "Email_Box" and text.startswith("Daily email cover short at:"):
            continue
        key = (dd, tt, task, text)
        if key not in seen:
            clean.append((dd, tt, task, note)); seen.add(key)
    return a, breaks, clean, dates, slots, hours_map

# =========================================================
# PATCH — Email can be split by a normal lunch break
# =========================================================
_prev_schedule_week_email_break_ok = schedule_week

def schedule_week(tpl: TemplateData, wk_start: date):
    """Allow the nominated Email_Box person to take their break in the middle of
    the email window without treating that 30-minute break as an email failure.

    This keeps the visible Break on the rota, keeps email blocks either side as
    continuous as possible, and only reports genuine email gaps where no email is
    covered and the missing slot is not simply the email owner's break.
    """
    a, breaks, gaps, dates, slots, hours_map = _prev_schedule_week_email_break_ok(tpl, wk_start)
    staff_names = [s.name for s in tpl.staff]
    EMAIL_START = time(10, 30)
    EMAIL_END = time(15, 30)
    email_slots = [tt for tt in slots if EMAIL_START <= tt < EMAIL_END]

    def _has_email(dd, tt):
        return any(a.get((dd, tt, nm)) == "Email_Box" for nm in staff_names)

    def _is_email_break_gap(dd, tt):
        """True if tt is uncovered only because the email owner is on Break,
        with Email_Box before and after the break on the same person/day.
        """
        idx = slots.index(tt)
        for nm in staff_names:
            if nm not in breaks.get((dd, tt), set()):
                continue
            # Look backwards/forwards within the email window, allowing adjacent
            # break slots but requiring actual Email_Box on both sides.
            before = False
            j = idx - 1
            while j >= 0 and EMAIL_START <= slots[j] < EMAIL_END:
                if nm in breaks.get((dd, slots[j]), set()):
                    j -= 1
                    continue
                before = (a.get((dd, slots[j], nm)) == "Email_Box")
                break
            after = False
            j = idx + 1
            while j < len(slots) and EMAIL_START <= slots[j] < EMAIL_END:
                if nm in breaks.get((dd, slots[j]), set()):
                    j += 1
                    continue
                after = (a.get((dd, slots[j], nm)) == "Email_Box")
                break
            if before and after:
                return True
        return False

    # Rewrite/remove email gap notes so lunch-break interruptions are acceptable.
    cleaned = []
    seen = set()
    for dd, tt, task, note in gaps:
        text = str(note or "")
        if task == "Email_Box" and "Daily email cover" in text:
            # Recalculate the genuine missing slots for that day rather than relying
            # on older wrapper notes which treated Break as impossible.
            real_missing = []
            for es in email_slots:
                if _has_email(dd, es):
                    continue
                if _is_email_break_gap(dd, es):
                    continue
                real_missing.append(es.strftime("%H:%M"))
            if not real_missing:
                continue
            note = "Daily email cover still not possible at: " + ", ".join(real_missing) + " (email-capable staff are on fixed cover, not working, leave, or another protected task)"
        key = (dd, tt, task, str(note))
        if key not in seen:
            cleaned.append((dd, tt, task, note)); seen.add(key)
    return a, breaks, cleaned, dates, slots, hours_map

# =========================================================
# PATCH — Front Desk is fixed cover; breaks must move around it
# =========================================================
_prev_schedule_week_frontdesk_break_fixed = schedule_week

def schedule_week(tpl: TemplateData, wk_start: date):
    """Final safety pass: front desk bands are hard/set cover.

    Lunch breaks may move before/after a front-desk band, but must not hide or
    remove front-desk cover in the visible rota. This pass relocates any break
    that has landed on a FrontDesk_* slot and then repairs any remaining
    front-desk gaps using the highest-weight available front-desk person.
    """
    a, breaks, gaps, dates, slots, hours_map = _prev_schedule_week_frontdesk_break_fixed(tpl, wk_start)
    staff_names = [s.name for s in tpl.staff]
    staff_by_name = {s.name: s for s in tpl.staff}

    # Use the same band pattern as the safe v37 engine currently uses.
    fd_bands_by_site = {"SLGP": FD_BANDS, "JEN": FD_BANDS_FLEX, "BGS": FD_BANDS_FLEX}

    def _shift(nm, dd):
        return shift_window(hours_map, dd, nm)

    def _is_holiday(nm, dd):
        return bool(holiday_kind(nm, dd, tpl.hols))

    def _duration_hours(nm, dd):
        stt, endt = _shift(nm, dd)
        if not stt or not endt:
            return 0.0
        return (dt_of(dd, endt) - dt_of(dd, stt)).total_seconds() / 3600.0

    def _on_break(nm, dd, tt):
        return nm in breaks.get((dd, tt), set())

    def _remove_break(nm, dd, tt):
        if (dd, tt) in breaks and nm in breaks[(dd, tt)]:
            breaks[(dd, tt)].remove(nm)
            if not breaks[(dd, tt)]:
                del breaks[(dd, tt)]

    def _set_break(nm, dd, tt):
        # One break per day: clear any old visible break first.
        for old in list(slots):
            _remove_break(nm, dd, old)
        breaks.setdefault((dd, tt), set()).add(nm)

    def _current(nm, dd, tt):
        return str(a.get((dd, tt, nm), "") or "")

    def _is_fd(nm, dd, tt):
        return _current(nm, dd, tt).startswith("FrontDesk_")

    def _candidate_breaks(nm, dd):
        stt, endt = _shift(nm, dd)
        if not stt or not endt or _is_holiday(nm, dd):
            return []
        if _duration_hours(nm, dd) < BREAK_THRESHOLD_HOURS:
            return []
        start_dt, end_dt = dt_of(dd, stt), dt_of(dd, endt)
        mid = start_dt + (end_dt - start_dt) / 2
        lo = start_dt + (end_dt - start_dt) / 3
        hi = start_dt + (end_dt - start_dt) * 2 / 3
        cands = []
        for tt in slots:
            if tt < stt or add_minutes(tt, 30) > endt:
                continue
            if _is_fd(nm, dd, tt):
                continue
            cur = _current(nm, dd, tt)
            # Do not move a break onto another hard fixed role if avoidable.
            hard = cur.startswith("Triage_Admin_") or cur == "Awaiting_PSA_Admin"
            # Middle third preferred, but if FD occupies the middle, before/after is fine.
            tt_dt = dt_of(dd, tt)
            outside_mid = 0 if lo <= tt_dt <= hi else 1
            # Prefer genuinely flexible work, then email/admin, then phones/bookings, then other fixed.
            if cur in ("", "Misc_Tasks"):
                cost = 0
            elif cur == "Email_Box":
                cost = 40
            elif cur in ("EMIS", "Docman"):
                cost = 120
            elif cur == "Bookings":
                cost = 180
            elif cur == "Phones":
                cost = 300
            elif hard:
                cost = 5000
            else:
                cost = 500
            dist = abs((tt_dt - mid).total_seconds()) / 60
            cands.append((outside_mid, cost, dist, tt))
        cands.sort()
        return [x[-1] for x in cands]

    # 1) Move any break that overlaps Front Desk. FD must remain visible.
    for dd in dates:
        for nm in staff_names:
            for tt in list(slots):
                if _on_break(nm, dd, tt) and _is_fd(nm, dd, tt):
                    cands = _candidate_breaks(nm, dd)
                    if cands:
                        _set_break(nm, dd, cands[0])
                    else:
                        # Last resort: remove the break rather than breaking FD cover.
                        _remove_break(nm, dd, tt)
                        gaps.append((dd, tt, "Break", f"Break could not be placed without disrupting Front Desk for {nm}"))

    def _visible_fd_staff(dd, tt, site):
        role = f"FrontDesk_{site}"
        return [nm for nm in staff_names if a.get((dd, tt, nm)) == role and not _on_break(nm, dd, tt)]

    def _can_take_fd(nm, dd, tt, site):
        st = staff_by_name[nm]
        if not st.can_frontdesk or st.home != site or _is_holiday(nm, dd):
            return False
        if _on_break(nm, dd, tt):
            # Try moving the break away from this fixed slot.
            cands = [bt for bt in _candidate_breaks(nm, dd) if bt != tt]
            if cands:
                _set_break(nm, dd, cands[0])
            else:
                return False
        stt, endt = _shift(nm, dd)
        return bool(stt and endt and stt <= tt < endt)

    def _fd_score(nm, dd, tt):
        st = staff_by_name[nm]
        cur = _current(nm, dd, tt)
        # Prefer front-desk-only/high-weight staff, and avoid stealing scarce admin/email/phones.
        fd_only = 0 if st.frontdesk_only else 1
        weight = -int(st.weights.get("FrontDesk", 3)) if getattr(st, "weights", None) else -3
        steal_cost = {"":0, "Misc_Tasks":0, "Bookings":1, "EMIS":5, "Docman":5, "Email_Box":8, "Phones":9}.get(cur, 20)
        return (fd_only, weight, steal_cost, nm.lower())

    # 2) Repair any remaining visible FD gaps, and remove false FD gap notes after repair.
    for dd in dates:
        for site, bands in fd_bands_by_site.items():
            role = f"FrontDesk_{site}"
            for bs, be in bands:
                for tt in [x for x in slots if bs <= x < be]:
                    if _visible_fd_staff(dd, tt, site):
                        continue
                    cands = [nm for nm in staff_names if _can_take_fd(nm, dd, tt, site)]
                    if cands:
                        chosen = sorted(cands, key=lambda nm: _fd_score(nm, dd, tt))[0]
                        a[(dd, tt, chosen)] = role
                    else:
                        gaps.append((dd, tt, role, "Front Desk cover still not possible after moving breaks"))

    # 3) De-duplicate/clean old FD notes where cover now exists.
    cleaned = []
    seen = set()
    for dd, tt, task, note in gaps:
        if str(task).startswith("FrontDesk_"):
            site = str(task).replace("FrontDesk_", "")
            if tt and _visible_fd_staff(dd, tt, site):
                continue
        key = (dd, tt, task, str(note))
        if key not in seen:
            cleaned.append((dd, tt, task, note)); seen.add(key)
    return a, breaks, cleaned, dates, slots, hours_map

# =========================================================
# PATCH: Break de-duplication / one visible break per staff-day
# =========================================================
_prev_schedule_week_break_dedupe = schedule_week

def schedule_week(tpl: TemplateData, wk_start: date):
    """Final wrapper: enforce exactly one 30-minute break per eligible staff-day.

    Previous repair layers could leave a visible two-slot break (e.g. 13:00 and
    13:30) because one layer tracked breaks in the `breaks` structure while
    another also wrote 'Break' directly into assignments. This pass normalises
    both sources into one visible 30-minute break and then repairs Front Desk
    cover again so fixed FD bands remain protected.
    """
    a, breaks, gaps, dates, slots, hours_map = _prev_schedule_week_break_dedupe(tpl, wk_start)
    staff_names = [s.name for s in tpl.staff]
    staff_by_name = {s.name: s for s in tpl.staff}

    def _shift(nm, dd):
        return shift_window(hours_map, dd, nm)

    def _is_holiday(nm, dd):
        return bool(holiday_kind(nm, dd, tpl.hols))

    def _working(nm, dd, tt):
        return is_working(hours_map, dd, tt, nm) and not _is_holiday(nm, dd)

    def _dur_hours(nm, dd):
        stt, endt = _shift(nm, dd)
        if not stt or not endt:
            return 0.0
        return (dt_of(dd, endt) - dt_of(dd, stt)).total_seconds() / 3600.0

    def _current(nm, dd, tt):
        return str(a.get((dd, tt, nm), "") or "")

    def _clear_break_membership(nm, dd):
        for tt in list(slots):
            if (dd, tt) in breaks and nm in breaks[(dd, tt)]:
                breaks[(dd, tt)].remove(nm)
                if not breaks[(dd, tt)]:
                    del breaks[(dd, tt)]

    def _is_frontdesk_slot(nm, dd, tt):
        return _current(nm, dd, tt).startswith("FrontDesk_")

    def _break_candidates(nm, dd):
        stt, endt = _shift(nm, dd)
        if not stt or not endt or _dur_hours(nm, dd) < BREAK_THRESHOLD_HOURS:
            return []
        start_dt, end_dt = dt_of(dd, stt), dt_of(dd, endt)
        mid = start_dt + (end_dt - start_dt) / 2
        lo = start_dt + (end_dt - start_dt) / 3
        hi = start_dt + (end_dt - start_dt) * 2 / 3
        out = []
        for tt in slots:
            if tt < stt or add_minutes(tt, 30) > endt:
                continue
            if _is_frontdesk_slot(nm, dd, tt):
                continue
            cur = _current(nm, dd, tt)
            # Avoid hard fixed cover. Email can have a break in the middle.
            if cur.startswith("Triage_Admin_") or cur == "Awaiting_PSA_Admin":
                task_cost = 5000
            elif cur in ("", "Misc_Tasks"):
                task_cost = 0
            elif cur == "Email_Box":
                task_cost = 20
            elif cur in ("EMIS", "Docman"):
                task_cost = 100
            elif cur == "Bookings":
                task_cost = 150
            elif cur == "Phones":
                task_cost = 250
            elif cur == "Break":
                task_cost = -10  # if already a break, prefer keeping it if sensible
            else:
                task_cost = 400
            tt_dt = dt_of(dd, tt)
            outside_mid = 0 if lo <= tt_dt <= hi else 1
            dist = abs((tt_dt - mid).total_seconds()) / 60
            out.append((outside_mid, task_cost, dist, tt))
        out.sort()
        return [x[-1] for x in out]

    # Normalise break visibility: one 30-minute break only.
    for dd in dates:
        for nm in staff_names:
            st = staff_by_name[nm]
            # collect all break slots from both structures
            break_slots = set()
            for tt in slots:
                if nm in breaks.get((dd, tt), set()) or _current(nm, dd, tt) == "Break":
                    break_slots.add(tt)

            # First clear all existing break markers/cells for that person-day.
            _clear_break_membership(nm, dd)
            for tt in break_slots:
                if a.get((dd, tt, nm)) == "Break":
                    # Restore duplicate/old breaks to misc only if the person is actually working.
                    if _working(nm, dd, tt):
                        a[(dd, tt, nm)] = "Misc_Tasks"
                    else:
                        a.pop((dd, tt, nm), None)

            if not st.break_required or _is_holiday(nm, dd) or _dur_hours(nm, dd) < BREAK_THRESHOLD_HOURS:
                continue

            cands = _break_candidates(nm, dd)
            if not cands:
                gaps.append((dd, None, "Break", f"Break could not be placed for {nm} without disrupting fixed cover"))
                continue

            # If one of the old break slots is valid and well placed, keep the best old one.
            old_valid = [tt for tt in break_slots if tt in cands and not _is_frontdesk_slot(nm, dd, tt)]
            chosen = old_valid[0] if old_valid else cands[0]
            # Use assignment cell only as the break marker; keep breaks dict in sync for on-break style checks.
            a[(dd, chosen, nm)] = "Break"
            breaks.setdefault((dd, chosen), set()).add(nm)

    # Repair front desk after the de-dupe in case an old duplicate break had hidden cover.
    fd_bands_by_site = {"SLGP": FD_BANDS, "JEN": FD_BANDS_FLEX, "BGS": FD_BANDS_FLEX}

    def _on_break(nm, dd, tt):
        return nm in breaks.get((dd, tt), set()) or a.get((dd, tt, nm)) == "Break"

    def _visible_fd_staff(dd, tt, site):
        role = f"FrontDesk_{site}"
        return [nm for nm in staff_names if a.get((dd, tt, nm)) == role and not _on_break(nm, dd, tt)]

    def _can_fd(nm, dd, tt, site):
        st = staff_by_name[nm]
        if not st.can_frontdesk or st.home != site or _is_holiday(nm, dd):
            return False
        stt, endt = _shift(nm, dd)
        return bool(stt and endt and stt <= tt < endt and not _on_break(nm, dd, tt))

    def _fd_score(nm, dd, tt):
        st = staff_by_name[nm]
        cur = _current(nm, dd, tt)
        fd_only = 0 if getattr(st, "frontdesk_only", False) else 1
        w = -int(st.weights.get("FrontDesk", 3)) if getattr(st, "weights", None) else -3
        steal_cost = {"":0, "Misc_Tasks":0, "Bookings":1, "EMIS":5, "Docman":5, "Email_Box":8, "Phones":9}.get(cur, 20)
        return (fd_only, w, steal_cost, nm.lower())

    for dd in dates:
        for site, bands in fd_bands_by_site.items():
            role = f"FrontDesk_{site}"
            for bs, be in bands:
                for tt in [x for x in slots if bs <= x < be]:
                    if _visible_fd_staff(dd, tt, site):
                        continue
                    cands = [nm for nm in staff_names if _can_fd(nm, dd, tt, site)]
                    if cands:
                        chosen = sorted(cands, key=lambda x: _fd_score(x, dd, tt))[0]
                        a[(dd, tt, chosen)] = role
                    else:
                        gaps.append((dd, tt, role, "Front Desk cover still not possible after break de-duplication"))

    # Clean duplicate gap rows.
    cleaned, seen = [], set()
    for item in gaps:
        key = tuple(str(x) for x in item)
        if key not in seen:
            cleaned.append(item); seen.add(key)
    return a, breaks, cleaned, dates, slots, hours_map

# =========================================================
# PATCH: Front Desk is first rule, fixed bands, no per-slot drift
# =========================================================
_prev_schedule_week_fd_band_lock = schedule_week

def schedule_week(tpl: TemplateData, wk_start: date):
    """Final safety wrapper: lock Front Desk as fixed band cover.

    Front Desk is the first rule. For every weekday/site/band there must be one
    visible FrontDesk_* assignment across the whole fixed band. Breaks move
    around those bands; they must not split or replace Front Desk cover.
    """
    a, breaks, gaps, dates, slots, hours_map = _prev_schedule_week_fd_band_lock(tpl, wk_start)
    staff_names = [s.name for s in tpl.staff]
    staff_by_name = {s.name: s for s in tpl.staff}

    # Front Desk fixed bands are site-specific; see module-level FD_FIXED_BANDS_BY_SITE.

    HARD_FIXED_PREFIXES = ("Triage_Admin_",)
    HARD_FIXED_TASKS = {"Awaiting_PSA_Admin", "Holiday", "Sick", "Bank Holiday"}

    def _shift(nm, dd):
        return shift_window(hours_map, dd, nm)

    def _is_holiday(nm, dd):
        return bool(holiday_kind(nm, dd, tpl.hols))

    def _working(nm, dd, tt):
        return is_working(hours_map, dd, tt, nm) and not _is_holiday(nm, dd)

    def _duration_hours(nm, dd):
        stt, endt = _shift(nm, dd)
        if not stt or not endt:
            return 0.0
        return (dt_of(dd, endt) - dt_of(dd, stt)).total_seconds() / 3600.0

    def _current(nm, dd, tt):
        return str(a.get((dd, tt, nm), "") or "")

    def _on_break(nm, dd, tt):
        return nm in breaks.get((dd, tt), set()) or _current(nm, dd, tt) == "Break"

    def _clear_break(nm, dd):
        for tt in list(slots):
            if (dd, tt) in breaks and nm in breaks[(dd, tt)]:
                breaks[(dd, tt)].remove(nm)
                if not breaks[(dd, tt)]:
                    del breaks[(dd, tt)]
            if a.get((dd, tt, nm)) == "Break":
                a[(dd, tt, nm)] = "Misc_Tasks" if _working(nm, dd, tt) else ""

    def _fd_slots(bs, be):
        return [tt for tt in slots if bs <= tt < be]

    def _person_fd_bands(nm, dd):
        """All FD slots the person is already locked to after repair."""
        return {tt for tt in slots if str(a.get((dd, tt, nm), "") or "").startswith("FrontDesk_")}

    def _break_candidates(nm, dd):
        stt, endt = _shift(nm, dd)
        if not stt or not endt or _duration_hours(nm, dd) < BREAK_THRESHOLD_HOURS:
            return []
        start_dt, end_dt = dt_of(dd, stt), dt_of(dd, endt)
        lo = start_dt + (end_dt - start_dt) / 3
        hi = start_dt + (end_dt - start_dt) * 2 / 3
        mid = start_dt + (end_dt - start_dt) / 2
        fd_taken = _person_fd_bands(nm, dd)
        cands = []
        for tt in slots:
            if tt < stt or add_minutes(tt, 30) > endt:
                continue
            if tt in fd_taken:
                continue
            cur = _current(nm, dd, tt)
            if cur.startswith(HARD_FIXED_PREFIXES) or cur in HARD_FIXED_TASKS:
                continue
            # Email can be interrupted by break; admin less ideal but allowed; phones/bookings less ideal.
            if cur in ("", "Misc_Tasks", "Break"):
                cost = 0
            elif cur == "Email_Box":
                cost = 20
            elif cur in ("EMIS", "Docman"):
                cost = 80
            elif cur == "Bookings":
                cost = 130
            elif cur == "Phones":
                cost = 250
            else:
                cost = 300
            tt_dt = dt_of(dd, tt)
            outside_mid = 0 if lo <= tt_dt <= hi else 1
            dist = abs((tt_dt - mid).total_seconds()) / 60
            cands.append((outside_mid, cost, dist, tt))
        cands.sort()
        return [x[-1] for x in cands]

    def _place_one_break(nm, dd):
        st = staff_by_name[nm]
        _clear_break(nm, dd)
        if not st.break_required or _is_holiday(nm, dd) or _duration_hours(nm, dd) < BREAK_THRESHOLD_HOURS:
            return True
        cands = _break_candidates(nm, dd)
        if not cands:
            gaps.append((dd, None, "Break", f"Break could not be placed for {nm} without disrupting fixed Front Desk/Triage cover"))
            return False
        bt = cands[0]
        a[(dd, bt, nm)] = "Break"
        breaks.setdefault((dd, bt), set()).add(nm)
        return True

    def _can_cover_band(nm, dd, site, bs, be):
        st = staff_by_name[nm]
        if not st.can_frontdesk or str(st.home).upper() != site or _is_holiday(nm, dd):
            return False
        stt, endt = _shift(nm, dd)
        if not stt or not endt or not (stt <= bs and endt >= be):
            return False
        # Do not steal someone from hard fixed cover inside the band.
        for tt in _fd_slots(bs, be):
            cur = _current(nm, dd, tt)
            if cur.startswith(HARD_FIXED_PREFIXES) or cur in HARD_FIXED_TASKS:
                return False
        return True

    def _band_score(nm, dd, site, bs, be):
        st = staff_by_name[nm]
        fd_only = 0 if getattr(st, "frontdesk_only", False) else 1
        fd_weight = -int(st.weights.get("FrontDesk", 3)) if getattr(st, "weights", None) else -3
        # Prefer people already on this FD band, then people on misc/low-priority work.
        already_fd = 0 if all(a.get((dd, tt, nm)) == f"FrontDesk_{site}" for tt in _fd_slots(bs, be)) else 1
        steal_cost = 0
        for tt in _fd_slots(bs, be):
            cur = _current(nm, dd, tt)
            steal_cost += {
                "": 0, "Misc_Tasks": 0, "Break": 0,
                "Bookings": 2, "EMIS": 8, "Docman": 8,
                "Email_Box": 15, "Phones": 20,
            }.get(cur, 50)
        return (already_fd, fd_only, fd_weight, steal_cost, nm.lower())

    # Lock FD band by band. Clear all FD drift in the band first, then write one person continuously.
    for dd in dates:
        for site in ("SLGP", "JEN", "BGS"):
            role = f"FrontDesk_{site}"
            for bs, be in FD_FIXED_BANDS_BY_SITE.get(site, FD_BANDS):
                band_slots = _fd_slots(bs, be)
                # Candidate picked before clearing, so existing continuous cover is preserved if sensible.
                cands = [nm for nm in staff_names if _can_cover_band(nm, dd, site, bs, be)]
                if not cands:
                    gaps.append((dd, bs, role, "No suitable staff for fixed Front Desk band"))
                    continue
                chosen = sorted(cands, key=lambda nm: _band_score(nm, dd, site, bs, be))[0]

                # Remove all front desk labels for this site/band from everyone else.
                for tt in band_slots:
                    for nm in staff_names:
                        if a.get((dd, tt, nm)) == role and nm != chosen:
                            a[(dd, tt, nm)] = "Misc_Tasks" if _working(nm, dd, tt) else ""

                # Put chosen on the whole band; FD is visible and continuous.
                for tt in band_slots:
                    if _on_break(chosen, dd, tt):
                        _clear_break(chosen, dd)
                    a[(dd, tt, chosen)] = role

                # Re-place break around FD if they need one.
                _place_one_break(chosen, dd)

    # After FD locking, make sure every eligible person has no more than one break and never on FD.
    for dd in dates:
        for nm in staff_names:
            break_slots = [tt for tt in slots if _on_break(nm, dd, tt)]
            # Clear duplicates and FD-overlap breaks, then re-place once if required.
            if len(break_slots) > 1 or any(str(a.get((dd, tt, nm), "") or "").startswith("FrontDesk_") for tt in break_slots):
                _place_one_break(nm, dd)

    # Convert flexible Misc into shortfall admin where possible (Misc remains last resort after FD repair).
    def _task_total(task):
        return sum(0.5 for dd in dates for tt in slots for nm in staff_names if a.get((dd, tt, nm)) == task)

    target_emis = float(tpl.weekly_targets.get("EMIS", 0.0) or 0.0)
    target_doc = float(tpl.weekly_targets.get("Docman", 0.0) or 0.0)

    def _assign_misc_to(task, can_attr, target):
        if target <= 0:
            return
        # Greedy continuous blocks: process by person/day so Will-type patterns stay together.
        for dd in dates:
            for nm in staff_names:
                st = staff_by_name[nm]
                if not getattr(st, can_attr):
                    continue
                if _task_total(task) >= target:
                    return
                run = []
                for tt in slots:
                    if a.get((dd, tt, nm)) == "Misc_Tasks" and _working(nm, dd, tt):
                        run.append(tt)
                    else:
                        if run:
                            for rt in run:
                                if _task_total(task) < target:
                                    a[(dd, rt, nm)] = task
                            run = []
                if run:
                    for rt in run:
                        if _task_total(task) < target:
                            a[(dd, rt, nm)] = task

    # Fill Docman/EMIS shortfalls; these are the only tasks allowed to be under 1h if needed.
    _assign_misc_to("EMIS", "can_emis", target_emis)
    _assign_misc_to("Docman", "can_docman", target_doc)

    # Final FD verification and clean old/false gap notes.
    cleaned, seen = [], set()
    for dd in dates:
        for site in ("SLGP", "JEN", "BGS"):
            role = f"FrontDesk_{site}"
            for bs, be in FD_FIXED_BANDS_BY_SITE.get(site, FD_BANDS):
                for tt in _fd_slots(bs, be):
                    visible = [nm for nm in staff_names if a.get((dd, tt, nm)) == role and not _on_break(nm, dd, tt)]
                    if len(visible) != 1:
                        gaps.append((dd, tt, role, f"Front Desk fixed cover error: {len(visible)} visible staff"))
    for dd, tt, task, note in gaps:
        if str(task).startswith("FrontDesk_") and tt:
            site = str(task).replace("FrontDesk_", "")
            role = f"FrontDesk_{site}"
            visible = [nm for nm in staff_names if a.get((dd, tt, nm)) == role and not _on_break(nm, dd, tt)]
            if len(visible) == 1:
                continue
        key = (dd, tt, task, str(note))
        if key not in seen:
            cleaned.append((dd, tt, task, note)); seen.add(key)
    return a, breaks, cleaned, dates, slots, hours_map

# =========================================================
# PATCH: Absolute Front Desk band lock (FD overrides triage/admin if needed)
# =========================================================
_prev_schedule_week_fd_absolute = schedule_week

def schedule_week(tpl: TemplateData, wk_start: date):
    """Absolute final pass: front desk is the first rule.

    If the only person able to cover a whole fixed Front Desk band is currently
    on another task (including triage/admin/phones), move them to Front Desk.
    Breaks are then moved outside their FD slots. This prevents per-slot FD
    swapping and visible FD gaps.
    """
    a, breaks, gaps, dates, slots, hours_map = _prev_schedule_week_fd_absolute(tpl, wk_start)
    staff_names = [s.name for s in tpl.staff]
    staff_by_name = {s.name: s for s in tpl.staff}
    # Front Desk bands are site-specific:
    #   SLGP: Carol/SLGP pattern 08:00-11:00, 11:00-13:00, 13:00-16:00, 16:00-18:30
    #   JEN/BGS: 08:00-10:30, 10:30-13:00, 13:00-16:00, 16:00-18:30
    FD_FIXED_BANDS_BY_SITE = {
        "SLGP": [(time(8,0), time(11,0)), (time(11,0), time(13,0)), (time(13,0), time(16,0)), (time(16,0), time(18,30))],
        "JEN":  [(time(8,0), time(10,30)), (time(10,30), time(13,0)), (time(13,0), time(16,0)), (time(16,0), time(18,30))],
        "BGS":  [(time(8,0), time(10,30)), (time(10,30), time(13,0)), (time(13,0), time(16,0)), (time(16,0), time(18,30))],
    }

    def _shift(nm, dd):
        return shift_window(hours_map, dd, nm)
    def _holiday(nm, dd):
        return bool(holiday_kind(nm, dd, tpl.hols))
    def _working(nm, dd, tt):
        stt, endt = _shift(nm, dd)
        return bool(stt and endt and stt <= tt < endt and not _holiday(nm, dd))
    def _dur(nm, dd):
        stt, endt = _shift(nm, dd)
        return 0.0 if not stt or not endt else (dt_of(dd, endt) - dt_of(dd, stt)).total_seconds()/3600.0
    def _cur(nm, dd, tt):
        return str(a.get((dd, tt, nm), "") or "")
    def _fd_slots(bs, be):
        return [tt for tt in slots if bs <= tt < be]
    def _clear_breaks(nm, dd):
        for tt in slots:
            if (dd, tt) in breaks and nm in breaks[(dd, tt)]:
                breaks[(dd, tt)].remove(nm)
                if not breaks[(dd, tt)]:
                    del breaks[(dd, tt)]
            if a.get((dd, tt, nm)) == "Break":
                a[(dd, tt, nm)] = "Misc_Tasks" if _working(nm, dd, tt) else ""
    def _on_fd(nm, dd, tt):
        return _cur(nm, dd, tt).startswith("FrontDesk_")
    def _break_candidates(nm, dd):
        stt, endt = _shift(nm, dd)
        if not stt or not endt or _dur(nm, dd) < BREAK_THRESHOLD_HOURS:
            return []
        start_dt, end_dt = dt_of(dd, stt), dt_of(dd, endt)
        lo = start_dt + (end_dt-start_dt)/3
        hi = start_dt + (end_dt-start_dt)*2/3
        mid = start_dt + (end_dt-start_dt)/2
        out=[]
        for tt in slots:
            if tt < stt or add_minutes(tt, 30) > endt or _on_fd(nm, dd, tt):
                continue
            cur = _cur(nm, dd, tt)
            if cur in {"Holiday", "Sick", "Bank Holiday"}:
                continue
            cost = {"":0,"Misc_Tasks":0,"Break":0,"Email_Box":20,"EMIS":80,"Docman":80,"Bookings":130,"Phones":250,"Awaiting_PSA_Admin":400}.get(cur, 300)
            if cur.startswith("Triage_Admin_"):
                cost = 350
            tt_dt = dt_of(dd, tt)
            out.append((0 if lo <= tt_dt <= hi else 1, cost, abs((tt_dt-mid).total_seconds())/60, tt))
        out.sort()
        return [x[-1] for x in out]
    def _place_break(nm, dd):
        _clear_breaks(nm, dd)
        st = staff_by_name[nm]
        if not st.break_required or _holiday(nm, dd) or _dur(nm, dd) < BREAK_THRESHOLD_HOURS:
            return
        cands = _break_candidates(nm, dd)
        if cands:
            bt = cands[0]
            a[(dd, bt, nm)] = "Break"
            breaks.setdefault((dd, bt), set()).add(nm)
        else:
            gaps.append((dd, None, "Break", f"Break could not be placed for {nm} after Front Desk was locked"))
    def _can_fd_band(nm, dd, site, bs, be):
        st = staff_by_name[nm]
        if not st.can_frontdesk or str(st.home).upper() != site or _holiday(nm, dd):
            return False
        stt, endt = _shift(nm, dd)
        return bool(stt and endt and stt <= bs and endt >= be)
    def _score(nm, dd, site, bs, be):
        st = staff_by_name[nm]
        already = 0 if all(a.get((dd, tt, nm)) == f"FrontDesk_{site}" for tt in _fd_slots(bs, be)) else 1
        fd_only = 0 if getattr(st, "frontdesk_only", False) else 1
        weight = -int(st.weights.get("FrontDesk", 3)) if getattr(st, "weights", None) else -3
        steal = 0
        for tt in _fd_slots(bs, be):
            cur=_cur(nm,dd,tt)
            steal += {"":0,"Misc_Tasks":0,"Break":0,"Bookings":2,"EMIS":8,"Docman":8,"Email_Box":14,"Phones":18,"Awaiting_PSA_Admin":30}.get(cur, 25)
            if cur.startswith("Triage_Admin_"):
                steal += 12
        return (already, fd_only, weight, steal, nm.lower())

    for dd in dates:
        for site in ("SLGP", "JEN", "BGS"):
            role = f"FrontDesk_{site}"
            for bs, be in FD_FIXED_BANDS_BY_SITE.get(site, FD_BANDS):
                band = _fd_slots(bs, be)
                cands = [nm for nm in staff_names if _can_fd_band(nm, dd, site, bs, be)]
                if not cands:
                    gaps.append((dd, bs, role, "No one works the whole fixed Front Desk band"))
                    continue
                chosen = sorted(cands, key=lambda nm: _score(nm, dd, site, bs, be))[0]
                # Remove front desk labels in this band from everyone else.
                for tt in band:
                    for nm in staff_names:
                        if a.get((dd, tt, nm)) == role and nm != chosen:
                            a[(dd, tt, nm)] = "Misc_Tasks" if _working(nm, dd, tt) else ""
                # Lock the whole band to chosen.
                for tt in band:
                    if a.get((dd, tt, chosen)) == "Break" or chosen in breaks.get((dd, tt), set()):
                        _clear_breaks(chosen, dd)
                    a[(dd, tt, chosen)] = role
                _place_break(chosen, dd)

    # Final one-break clean-up: no duplicate breaks, no breaks on FD.
    for dd in dates:
        for nm in staff_names:
            bs=[tt for tt in slots if a.get((dd,tt,nm)) == "Break" or nm in breaks.get((dd,tt), set())]
            if len(bs)>1 or any(_on_fd(nm, dd, tt) for tt in bs):
                _place_break(nm, dd)

    # Verify FD and remove stale FD gap rows where now fixed.
    cleaned=[]; seen=set()
    for dd in dates:
        for site in ("SLGP","JEN","BGS"):
            role=f"FrontDesk_{site}"
            for bs,be in FD_FIXED_BANDS_BY_SITE.get(site, FD_BANDS):
                last=None; switched=False
                for tt in _fd_slots(bs,be):
                    vis=[nm for nm in staff_names if a.get((dd,tt,nm))==role and a.get((dd,tt,nm)) != "Break" and nm not in breaks.get((dd,tt),set())]
                    if len(vis)!=1:
                        gaps.append((dd,tt,role,f"Front Desk fixed cover error: {len(vis)} visible staff"))
                    elif last is None:
                        last=vis[0]
                    elif vis[0] != last:
                        switched=True
                if switched:
                    gaps.append((dd,bs,role,"Front Desk band switches staff within fixed band"))
    for dd,tt,task,note in gaps:
        if str(task).startswith("FrontDesk_") and tt:
            site=str(task).replace("FrontDesk_",""); role=f"FrontDesk_{site}"
            vis=[nm for nm in staff_names if a.get((dd,tt,nm))==role and nm not in breaks.get((dd,tt),set())]
            if len(vis)==1 and "switches" not in str(note):
                continue
        key=(dd,tt,task,str(note))
        if key not in seen:
            cleaned.append((dd,tt,task,note)); seen.add(key)
    return a, breaks, cleaned, dates, slots, hours_map

# =========================================================
# PATCH: Carol fixed SLGP morning FD + Mandy email priority
# =========================================================
_prev_schedule_week_carol_mandy = schedule_week

def schedule_week(tpl: TemplateData, wk_start: date):
    """Final business-rule pass.

    Rules added after the site-specific FD lock:
      * Carol is the fixed SLGP Front Desk cover 08:00-11:00 on every working day,
        unless she is on leave/not working that full band.
      * Mandy is prioritised for Email_Box from 10:30 onwards on her working days.
        To support this, if Mandy has been used for Front Desk after 10:30 and another
        full-band FD-capable person is available, Front Desk is switched to the
        alternative and Mandy is moved to Email_Box.
    """
    a, breaks, gaps, dates, slots, hours_map = _prev_schedule_week_carol_mandy(tpl, wk_start)
    staff_names = [s.name for s in tpl.staff]
    staff_by_name = {s.name: s for s in tpl.staff}

    FD_FIXED_BANDS_BY_SITE = {
        "SLGP": [(time(8,0), time(11,0)), (time(11,0), time(13,0)), (time(13,0), time(16,0)), (time(16,0), time(18,30))],
        "JEN":  [(time(8,0), time(10,30)), (time(10,30), time(13,0)), (time(13,0), time(16,0)), (time(16,0), time(18,30))],
        "BGS":  [(time(8,0), time(10,30)), (time(10,30), time(13,0)), (time(13,0), time(16,0)), (time(16,0), time(18,30))],
    }
    EMAIL_START, EMAIL_END = time(10,30), time(15,30)

    def _find_name(label):
        nl = normalize(label)
        for nm in staff_names:
            if normalize(nm) == nl or normalize(nm).startswith(nl) or nl in normalize(nm):
                return nm
        return None

    carol = _find_name("Carol")
    mandy = _find_name("Mandy")

    def _shift(nm, dd):
        return shift_window(hours_map, dd, nm)
    def _holiday(nm, dd):
        return bool(holiday_kind(nm, dd, tpl.hols))
    def _working(nm, dd, tt):
        stt, endt = _shift(nm, dd)
        return bool(stt and endt and stt <= tt < endt and not _holiday(nm, dd))
    def _works_band(nm, dd, bs, be):
        stt, endt = _shift(nm, dd)
        return bool(stt and endt and stt <= bs and endt >= be and not _holiday(nm, dd))
    def _cur(nm, dd, tt):
        return str(a.get((dd, tt, nm), "") or "")
    def _dur(nm, dd):
        stt, endt = _shift(nm, dd)
        return 0.0 if not stt or not endt else (dt_of(dd, endt)-dt_of(dd, stt)).total_seconds()/3600.0
    def _band_slots(bs, be):
        return [tt for tt in slots if bs <= tt < be]
    def _is_fd_task(task):
        return str(task or "").startswith("FrontDesk_")
    def _is_fixed_hard(task):
        task = str(task or "")
        return task.startswith("FrontDesk_") or task.startswith("Triage_Admin_") or task in {"Awaiting_PSA_Admin", "Holiday", "Sick", "Bank Holiday"}

    def _remove_breaks(nm, dd):
        for tt in list(slots):
            if nm in breaks.get((dd, tt), set()):
                breaks[(dd, tt)].discard(nm)
                if not breaks[(dd, tt)]:
                    breaks.pop((dd, tt), None)
            if a.get((dd, tt, nm)) == "Break":
                a[(dd, tt, nm)] = "Misc_Tasks" if _working(nm, dd, tt) else ""

    def _place_one_break(nm, dd):
        _remove_breaks(nm, dd)
        st = staff_by_name.get(nm)
        if not st or not st.break_required or _holiday(nm, dd) or _dur(nm, dd) < BREAK_THRESHOLD_HOURS:
            return
        stt, endt = _shift(nm, dd)
        if not stt or not endt:
            return
        start_dt, end_dt = dt_of(dd, stt), dt_of(dd, endt)
        lo = start_dt + (end_dt - start_dt) / 3
        hi = start_dt + (end_dt - start_dt) * 2 / 3
        mid = start_dt + (end_dt - start_dt) / 2
        cands = []
        for tt in slots:
            if tt < stt or add_minutes(tt, 30) > endt:
                continue
            cur = _cur(nm, dd, tt)
            # Never break Front Desk; email break is allowed, as requested.
            if cur.startswith("FrontDesk_") or cur.startswith("Triage_Admin_") or cur in {"Awaiting_PSA_Admin", "Holiday", "Sick", "Bank Holiday"}:
                continue
            cost = {"":0, "Misc_Tasks":0, "Break":0, "Email_Box":5, "EMIS":20, "Docman":20, "Bookings":50, "Phones":70}.get(cur, 80)
            tt_dt = dt_of(dd, tt)
            cands.append((0 if lo <= tt_dt <= hi else 1, cost, abs((tt_dt-mid).total_seconds())/60, tt))
        if cands:
            cands.sort()
            bt = cands[0][-1]
            a[(dd, bt, nm)] = "Break"
            breaks.setdefault((dd, bt), set()).add(nm)
        else:
            gaps.append((dd, None, "Break", f"Break could not be placed for {nm} without disrupting fixed cover"))

    def _clear_role_in_band(role, chosen, dd, band):
        for tt in band:
            for nm in staff_names:
                if nm != chosen and a.get((dd, tt, nm)) == role:
                    a[(dd, tt, nm)] = "Misc_Tasks" if _working(nm, dd, tt) else ""

    # 1) Carol fixed on SLGP 08:00-11:00 on each working day unless leave/not working.
    if carol:
        for dd in dates:
            bs, be = time(8,0), time(11,0)
            role = "FrontDesk_SLGP"
            band = _band_slots(bs, be)
            if _works_band(carol, dd, bs, be):
                _remove_breaks(carol, dd)
                _clear_role_in_band(role, carol, dd, band)
                for tt in band:
                    a[(dd, tt, carol)] = role
                _place_one_break(carol, dd)
            else:
                gaps.append((dd, bs, role, "Carol unavailable for fixed SLGP 08:00-11:00 Front Desk"))

    # 2) If Mandy is on FD after 10:30 and another full-band FD person can cover, release Mandy for email.
    if mandy:
        for dd in dates:
            if _holiday(mandy, dd):
                continue
            for site, bands in FD_FIXED_BANDS_BY_SITE.items():
                for bs, be in bands:
                    if be <= EMAIL_START:
                        continue
                    role = f"FrontDesk_{site}"
                    band = _band_slots(bs, be)
                    if not any(a.get((dd, tt, mandy)) == role for tt in band):
                        continue
                    alts = []
                    for nm in staff_names:
                        if nm == mandy:
                            continue
                        st = staff_by_name[nm]
                        if not st.can_frontdesk or str(st.home).upper() != site or not _works_band(nm, dd, bs, be):
                            continue
                        # Do not pull someone out of harder fixed cover to release Mandy unless unavoidable.
                        conflict = 0
                        for tt in band:
                            cur = _cur(nm, dd, tt)
                            if cur.startswith("FrontDesk_") and cur != role:
                                conflict += 999
                            elif cur.startswith("Triage_Admin_") or cur == "Awaiting_PSA_Admin":
                                conflict += 200
                            else:
                                conflict += {"":0, "Misc_Tasks":0, "Break":0, "EMIS":8, "Docman":8, "Bookings":12, "Phones":18, "Email_Box":25}.get(cur, 30)
                        fd_weight = -int(st.weights.get("FrontDesk", 3)) if getattr(st, "weights", None) else -3
                        alts.append((conflict, fd_weight, nm.lower(), nm))
                    if alts:
                        alts.sort()
                        chosen = alts[0][-1]
                        _remove_breaks(chosen, dd)
                        _clear_role_in_band(role, chosen, dd, band)
                        for tt in band:
                            a[(dd, tt, chosen)] = role
                            if a.get((dd, tt, mandy)) == role:
                                a[(dd, tt, mandy)] = "Misc_Tasks" if _working(mandy, dd, tt) else ""
                        _place_one_break(chosen, dd)

    # 3) Mandy email priority from 10:30 onwards on working days. Preserve fixed cover and her one break.
    if mandy:
        for dd in dates:
            if _holiday(mandy, dd):
                continue
            for tt in slots:
                if not (EMAIL_START <= tt < EMAIL_END):
                    continue
                if not _working(mandy, dd, tt):
                    continue
                cur = _cur(mandy, dd, tt)
                if cur == "Break" or mandy in breaks.get((dd, tt), set()):
                    continue
                if _is_fixed_hard(cur):
                    continue
                a[(dd, tt, mandy)] = "Email_Box"
            _place_one_break(mandy, dd)

    # 4) Final FD verification; FD remains first rule.
    for dd in dates:
        for site, bands in FD_FIXED_BANDS_BY_SITE.items():
            role = f"FrontDesk_{site}"
            for bs, be in bands:
                band = _band_slots(bs, be)
                # Re-assert Carol for SLGP first band if available.
                if site == "SLGP" and bs == time(8,0) and be == time(11,0) and carol and _works_band(carol, dd, bs, be):
                    _clear_role_in_band(role, carol, dd, band)
                    for tt in band:
                        a[(dd, tt, carol)] = role
                    _place_one_break(carol, dd)
                for tt in band:
                    vis = [nm for nm in staff_names if a.get((dd, tt, nm)) == role and nm not in breaks.get((dd, tt), set())]
                    if len(vis) != 1:
                        gaps.append((dd, tt, role, f"Front Desk fixed cover error after Carol/Mandy pass: {len(vis)} visible staff"))

    # 5) Clean stale FD gap notes if now fixed.
    cleaned = []
    seen = set()
    for dd, tt, task, note in gaps:
        if str(task).startswith("FrontDesk_") and tt:
            role = str(task)
            vis = [nm for nm in staff_names if a.get((dd, tt, nm)) == role and nm not in breaks.get((dd, tt), set())]
            if len(vis) == 1 and "Carol unavailable" not in str(note):
                continue
        key = (dd, tt, task, str(note))
        if key not in seen:
            cleaned.append((dd, tt, task, note)); seen.add(key)
    return a, breaks, cleaned, dates, slots, hours_map
