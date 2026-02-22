from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import io
import re
from dataclasses import dataclass
from datetime import datetime, date, time, timedelta
from typing import Dict, List, Tuple, Optional, Set

import math

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# =========================================================
# Rota Generator Engine — v15 (Weighted, Block-Stable, No Floaters)
# =========================================================

DAY_START = time(8, 0)
DAY_END = time(18, 30)
SLOT_MIN = 30

SITES = ["SLGP", "JEN", "BGS"]

FD_BANDS = [
    (time(8, 0), time(11, 0)),
    (time(11, 0), time(13, 0)),
    (time(13, 0), time(16, 0)),
    (time(16, 0), time(18, 30)),
]
TRIAGE_BANDS = [
    (time(8, 0), time(10, 30)),
    (time(10, 30), time(13, 0)),
    (time(13, 30), time(16, 0)),
]

BREAK_WINDOW = (time(12, 0), time(14, 0))
BREAK_CANDIDATES = [time(12, 0), time(12, 30), time(13, 0), time(13, 30)]
BREAK_THRESHOLD_HOURS = 6.0

# Block mins (slots)
MIN_PHONES = 2          # 1.5h
MAX_PHONES = 8          # 4h
MIN_DEFAULT = 5         # 2.5h
MAX_DEFAULT = 9         # 4.5h
MIN_DOCMAN = 6          # 3h

def timeslots() -> List[time]:
    cur = datetime(2000, 1, 1, DAY_START.hour, DAY_START.minute)
    end = datetime(2000, 1, 1, DAY_END.hour, DAY_END.minute)
    out: List[time] = []
    while cur < end:
        out.append(cur.time())
        cur += timedelta(minutes=SLOT_MIN)
    return out

def ensure_monday(d: date) -> date:
    return d - timedelta(days=d.weekday())

def day_name(d: date) -> str:
    return ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"][d.weekday()]

def t_in_range(t: time, a: time, b: time) -> bool:
    return (t >= a) and (t < b)

def normalize(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s).strip().lower())

def find_sheet(xls: pd.ExcelFile, candidates: List[str]) -> Optional[str]:
    names = {normalize(n): n for n in xls.sheet_names}
    for c in candidates:
        k = normalize(c)
        if k in names:
            return names[k]
    for n in xls.sheet_names:
        nn = normalize(n)
        for c in candidates:
            if normalize(c) in nn:
                return n
    return None

def pick_col(df: pd.DataFrame, candidates: List[str], required: bool=True) -> Optional[str]:
    cols = {normalize(c): c for c in df.columns}
    for cand in candidates:
        k = normalize(cand)
        if k in cols:
            return cols[k]
    for c in df.columns:
        nc = normalize(c)
        for cand in candidates:
            if normalize(cand) in nc:
                return c
    if required:
        raise KeyError(f"Missing required column among {candidates}. Available: {list(df.columns)}")
    return None

def to_time(x):
    if pd.isna(x) or x == "":
        return None
    if isinstance(x, time):
        return x
    if isinstance(x, datetime):
        return x.time()
    if isinstance(x, (float, int)):
        seconds = int(round(float(x) * 86400))
        return (datetime(2000,1,1) + timedelta(seconds=seconds)).time()
    return pd.to_datetime(str(x)).time()

def to_date(x):
    if pd.isna(x) or x == "":
        return None
    if isinstance(x, (datetime, pd.Timestamp)):
        return x.date()
    if isinstance(x, date):
        return x
    return pd.to_datetime(x).date()

def dt_of(d: date, t: time) -> datetime:
    return datetime(d.year, d.month, d.day, t.hour, t.minute)

def add_minutes(t: time, mins: int) -> time:
    return (datetime(2000,1,1,t.hour,t.minute) + timedelta(minutes=mins)).time()

def yn(v) -> bool:
    if pd.isna(v):
        return False
    s = str(v).strip().lower()
    return s in {"y","yes","true","1","t"}

@dataclass
class Staff:
    name: str
    home: str
    can_frontdesk: bool
    can_triage: bool
    can_email: bool
    can_phones: bool
    can_bookings: bool
    can_emis: bool
    can_docman: bool
    weights: Dict[str, int]
    frontdesk_only: bool
    break_required: bool

@dataclass
class TemplateData:
    staff: List[Staff]
    hours_map: Dict[str, Dict[str, Optional[time]]]
    hols: List[Tuple[str, date, date, str]]
    call_handlers: pd.DataFrame
    handler_leave: pd.DataFrame
    phones_targets: Dict[Tuple[str, time], int]
    bookings_targets: Dict[Tuple[str, time], int]
    weekly_targets: Dict[str, float]
    swaps: List[Tuple[date, str, str, Optional[time], Optional[time]]]
    buddies: Dict[str, str]

def read_template(uploaded_bytes: bytes) -> TemplateData:
    xls = pd.ExcelFile(io.BytesIO(uploaded_bytes))

    staff_sheet = find_sheet(xls, ["Staff"])
    hours_sheet = find_sheet(xls, ["WorkingHours", "Hours"])
    hols_sheet = find_sheet(xls, ["Holidays", "Leave", "Absence"])
    callh_sheet = find_sheet(xls, ["CallHandlers", "Call Handlers"])
    hleave_sheet = find_sheet(xls, ["Handler_Leave", "Handler Leave", "CallHandler_Leave"])
    tph_sheet = find_sheet(xls, ["Targets_Phones_Hourly", "PhonesTargets", "Targets Phones Hourly"])
    tbk_sheet = find_sheet(xls, ["Targets_Bookings_Hourly", "BookingsTargets", "Targets Bookings Hourly"])
    tweek_sheet = find_sheet(xls, ["Targets_Weekly", "Targets Weekly"])
    swaps_sheet = find_sheet(xls, ["Swaps"])
    new_sheet = find_sheet(xls, ["NewStarters", "New Starters"])

    if not staff_sheet or not hours_sheet:
        raise ValueError(f"Missing Staff/WorkingHours sheets. Found: {xls.sheet_names}")

    staff_df = pd.read_excel(xls, sheet_name=staff_sheet)
    hours_df = pd.read_excel(xls, sheet_name=hours_sheet)
    hols_df = pd.read_excel(xls, sheet_name=hols_sheet) if hols_sheet else pd.DataFrame()
    callh_df = pd.read_excel(xls, sheet_name=callh_sheet) if callh_sheet else pd.DataFrame()
    hleave_df = pd.read_excel(xls, sheet_name=hleave_sheet) if hleave_sheet else pd.DataFrame()
    tph_df = pd.read_excel(xls, sheet_name=tph_sheet) if tph_sheet else pd.DataFrame()
    tbk_df = pd.read_excel(xls, sheet_name=tbk_sheet) if tbk_sheet else pd.DataFrame()
    tweek_df = pd.read_excel(xls, sheet_name=tweek_sheet) if tweek_sheet else pd.DataFrame()
    swaps_df = pd.read_excel(xls, sheet_name=swaps_sheet) if swaps_sheet else pd.DataFrame()
    new_df = pd.read_excel(xls, sheet_name=new_sheet) if new_sheet else pd.DataFrame()

    # --- Staff
    name_c = pick_col(staff_df, ["Name","StaffName"])
    home_c = pick_col(staff_df, ["HomeSite","Site","BaseSite"], required=False)

    staff_df = staff_df.copy()
    staff_df["Name"] = staff_df[name_c].astype(str).str.strip()
    staff_df["HomeSite"] = staff_df[home_c].astype(str).str.strip().str.upper() if home_c else ""

    def bcol(cands, default=False):
        c = pick_col(staff_df, cands, required=False)
        if not c:
            return pd.Series([default]*len(staff_df))
        if staff_df[c].dtype == bool:
            return staff_df[c].fillna(default)
        return staff_df[c].apply(yn)

    staff_df["CanFrontDesk"] = bcol(["CanFrontDesk"])
    staff_df["CanTriage"] = bcol(["CanTriage"])
    staff_df["CanEmail"] = bcol(["CanEmail"])
    staff_df["CanPhones"] = bcol(["CanPhones"])
    staff_df["CanBookings"] = bcol(["CanBookings"])
    staff_df["CanEMIS"] = bcol(["CanEMIS"])
    staff_df["CanDocman"] = bcol(["CanDocman_PSA"]) | bcol(["CanDocman_AWAIT"]) | bcol(["CanDocman"])

    staff_df["FrontDeskOnly"] = bcol(["FrontDeskOnly"], default=False)
    staff_df["BreakRequired"] = bcol(["BreakRequired"], default=True)

    weight_cols = {
        "FrontDesk":"FrontDeskWeight",
        "Triage":"TriageWeight",
        "Email":"EmailWeight",
        "Phones":"PhonesWeight",
        "Bookings":"BookingsWeight",
        "EMIS":"EmisWeight",
        "Docman":"DocmanWeight",
        "Awaiting":"AwaitingWeight",
        "Misc":"MiscWeight",
    }

    staff_list: List[Staff] = []
    for _, r in staff_df.iterrows():
        weights: Dict[str,int] = {}
        for k, col in weight_cols.items():
            v = r.get(col, 3)
            try:
                if pd.isna(v) or v == "":
                    v = 3
                v = int(float(v))
            except Exception:
                v = 3
            weights[k] = max(0, min(5, v))
        staff_list.append(
            Staff(
                name=str(r["Name"]).strip(),
                home=str(r.get("HomeSite","")).strip().upper(),
                can_frontdesk=bool(r.get("CanFrontDesk", False)),
                can_triage=bool(r.get("CanTriage", False)),
                can_email=bool(r.get("CanEmail", False)),
                can_phones=bool(r.get("CanPhones", False)),
                can_bookings=bool(r.get("CanBookings", False)),
                can_emis=bool(r.get("CanEMIS", False)),
                can_docman=bool(r.get("CanDocman", False)),
                weights=weights,
                frontdesk_only=bool(r.get("FrontDeskOnly", False)),
                break_required=bool(r.get("BreakRequired", True)),
            )
        )

    # --- Working hours
    hours_df = hours_df.copy()
    hn = pick_col(hours_df, ["Name","StaffName"])
    hs = pick_col(hours_df, ["HomeSite","Site","BaseSite"], required=False)
    hours_df["Name"] = hours_df[hn].astype(str).str.strip()
    hours_df["HomeSite"] = hours_df[hs].astype(str).str.strip().str.upper() if hs else ""

    for dn in ["Mon","Tue","Wed","Thu","Fri"]:
        sc = pick_col(hours_df, [f"{dn}Start", f"{dn} Start", f"{dn}_Start"], required=False)
        ec = pick_col(hours_df, [f"{dn}End", f"{dn} End", f"{dn}_End"], required=False)
        hours_df[f"{dn}Start"] = hours_df[sc].apply(to_time) if sc else None
        hours_df[f"{dn}End"] = hours_df[ec].apply(to_time) if ec else None

    hours_map = {}
    for _, r in hours_df.iterrows():
        hours_map[str(r["Name"]).strip()] = {k: r.get(k) for k in hours_df.columns}

    # --- Align WorkingHours names to Staff names (common cause of 'blank working hours')
    # If a staff name in Staff sheet doesn't exactly match the WorkingHours sheet (extra spaces, case),
    # we map by normalized name so scheduling still finds their shift.
    norm_hours = {normalize(k): v for k, v in hours_map.items()}
    for st in staff_list:
        if st.name not in hours_map:
            nk = normalize(st.name)
            if nk in norm_hours:
                hours_map[st.name] = norm_hours[nk]
# --- Holidays ranges
    hols: List[Tuple[str,date,date,str]] = []
    if hols_df is not None and not hols_df.empty:
        ncol = pick_col(hols_df, ["Name","StaffName"], required=False) or hols_df.columns[0]
        sdcol = pick_col(hols_df, ["StartDate","Start"], required=False) or hols_df.columns[1]
        edcol = pick_col(hols_df, ["EndDate","End"], required=False) or hols_df.columns[2]
        notes_c = pick_col(hols_df, ["Notes","Note","Reason"], required=False)

        for _, r in hols_df.iterrows():
            nm = str(r.get(ncol,"")).strip()
            sd = to_date(r.get(sdcol))
            ed = to_date(r.get(edcol))
            note = "" if (not notes_c or pd.isna(r.get(notes_c))) else str(r.get(notes_c)).strip().lower()
            kind = "Holiday"
            if "sick" in note or "sickness" in note:
                kind = "Sick"
            elif "bank" in note:
                kind = "Bank Holiday"
            if nm and sd and ed:
                hols.append((nm, sd, ed, kind))

    # --- Targets
    def parse_hourly(df: pd.DataFrame) -> Dict[Tuple[str, time], int]:
        out: Dict[Tuple[str,time], int] = {}
        if df is None or df.empty:
            return out
        time_col = pick_col(df, ["Time"], required=False) or df.columns[0]
        ddf = df.copy()
        ddf["Time"] = ddf[time_col].apply(to_time)
        for dn in ["Mon","Tue","Wed","Thu","Fri"]:
            if dn not in ddf.columns:
                continue
            for _, r in ddf.iterrows():
                hh = r.get("Time")
                if not hh:
                    continue
                val = r.get(dn)
                if pd.isna(val) or val == "":
                    continue
                out[(dn, time(hh.hour, 0))] = int(float(val))
        return out

    phones_targets = parse_hourly(tph_df)
    bookings_targets = parse_hourly(tbk_df)

    weekly_targets = {"Bookings": 0.0, "EMIS": 0.0, "Docman": 0.0}
    if tweek_df is not None and not tweek_df.empty:
        task_c = pick_col(tweek_df, ["Task"], required=False) or tweek_df.columns[0]
        val_c = pick_col(tweek_df, ["WeekHoursTarget","Target","Hours"], required=False) or tweek_df.columns[1]
        for _, r in tweek_df.iterrows():
            tsk = str(r.get(task_c,"")).strip()
            val = r.get(val_c)
            if pd.isna(val) or val == "":
                continue
            if tsk in weekly_targets:
                weekly_targets[tsk] = float(val)

    # Swaps
    swaps: List[Tuple[date, str, str, Optional[time], Optional[time]]] = []
    if swaps_df is not None and not swaps_df.empty:
        dcol = pick_col(swaps_df, ["Date"], required=False) or swaps_df.columns[0]
        ncol = pick_col(swaps_df, ["Name"], required=False) or swaps_df.columns[1]
        swcol = pick_col(swaps_df, ["SwapWith"], required=False)
        nscol = pick_col(swaps_df, ["NewStart"], required=False)
        necol = pick_col(swaps_df, ["NewEnd"], required=False)
        for _, r in swaps_df.iterrows():
            dd = to_date(r.get(dcol))
            if not dd:
                continue
            nm = str(r.get(ncol,"")).strip()
            sw = str(r.get(swcol,"")).strip() if swcol else ""
            ns = to_time(r.get(nscol)) if nscol else None
            ne = to_time(r.get(necol)) if necol else None
            if nm:
                swaps.append((dd, nm, sw, ns, ne))

    # Buddies
    buddies: Dict[str,str] = {}
    if new_df is not None and not new_df.empty:
        nc = pick_col(new_df, ["NewStarterName","NewStarter","Starter"], required=False) or new_df.columns[0]
        bc = pick_col(new_df, ["BuddyName","Buddy"], required=False) or new_df.columns[1]
        for _, r in new_df.iterrows():
            n = str(r.get(nc,"")).strip()
            b = str(r.get(bc,"")).strip()
            if n and b:
                buddies[n] = b

    return TemplateData(
        staff=staff_list,
        hours_map=hours_map,
        hols=hols,
        call_handlers=callh_df if callh_df is not None else pd.DataFrame(),
        handler_leave=hleave_df if hleave_df is not None else pd.DataFrame(),
        phones_targets=phones_targets,
        bookings_targets=bookings_targets,
        weekly_targets=weekly_targets,
        swaps=swaps,
        buddies=buddies,
    )

# ---------- Availability helpers ----------
def holiday_kind(name: str, d: date, hols: List[Tuple[str,date,date,str]]) -> Optional[str]:
    for n, s, e, k in hols:
        if n.strip().lower() == name.strip().lower() and s and e and s <= d <= e:
            return k
    return None

def shift_window(hours_map: Dict[str, Dict[str, Optional[time]]], d: date, name: str) -> Tuple[Optional[time], Optional[time]]:
    dn = day_name(d)
    hr = hours_map.get(name)
    if hr is None:
        return None, None
    return hr.get(f"{dn}Start"), hr.get(f"{dn}End")

def is_working(hours_map: Dict[str, Dict[str, Optional[time]]], d: date, t: time, name: str) -> bool:
    stt, end = shift_window(hours_map, d, name)
    return bool(stt and end and (t >= stt) and (t < end))

# ---------- Swaps ----------
def apply_swaps(hours_map: Dict[str, Dict[str, Optional[time]]], swaps, week_dates: List[date]) -> Dict[str, Dict[str, Optional[time]]]:
    out = {k: dict(v) for k, v in hours_map.items()}
    in_week = set(week_dates)
    for dd, nm, sw, ns, ne in swaps:
        if dd not in in_week:
            continue
        dn = day_name(dd)
        if nm not in out:
            continue
        if sw and sw in out:
            a1, a2 = out[nm].get(f"{dn}Start"), out[nm].get(f"{dn}End")
            b1, b2 = out[sw].get(f"{dn}Start"), out[sw].get(f"{dn}End")
            out[nm][f"{dn}Start"], out[nm][f"{dn}End"] = b1, b2
            out[sw][f"{dn}Start"], out[sw][f"{dn}End"] = a1, a2
        elif ns and ne:
            out[nm][f"{dn}Start"], out[nm][f"{dn}End"] = ns, ne
    return out

# ---------- Call handler leave impact on phones ----------
def parse_handler_leave(df: pd.DataFrame) -> List[Tuple[str, date, date]]:
    if df is None or df.empty:
        return []
    ncol = pick_col(df, ["Name","HandlerName","CallHandler"], required=False) or df.columns[0]
    sdcol = pick_col(df, ["LeaveStartDate","LeaveStart","StartDate"], required=False) or df.columns[1]
    edcol = pick_col(df, ["LeaveEndDate","LeaveEnd","EndDate"], required=False) or df.columns[2]
    out = []
    for _, r in df.iterrows():
        nm = str(r.get(ncol,"")).strip()
        sd = to_date(r.get(sdcol))
        ed = to_date(r.get(edcol))
        if nm and sd and ed:
            out.append((nm, sd, ed))
    return out

def handler_working(callh_row: pd.Series, d: date, t: time) -> bool:
    dn = day_name(d)
    stt = to_time(callh_row.get(f"{dn}Start"))
    end = to_time(callh_row.get(f"{dn}End"))
    return bool(stt and end and (t >= stt) and (t < end))

def phones_required(tpl: TemplateData, d: date, t: time) -> int:
    dn = day_name(d)
    hour_key = time(t.hour, 0)
    base = int(tpl.phones_targets.get((dn, hour_key), 0) or 0)
    if base <= 0:
        return 0

    leave_ranges = parse_handler_leave(tpl.handler_leave)
    off = 0
    if tpl.call_handlers is not None and not tpl.call_handlers.empty:
        for _, r in tpl.call_handlers.iterrows():
            nm = str(r.get("Name","")).strip()
            if not nm:
                continue
            if not handler_working(r, d, t):
                continue
            for ln, sd, ed in leave_ranges:
                if ln.strip().lower() == nm.strip().lower() and sd <= d <= ed:
                    off += 1
                    break
    return base + off

# ---------- Site-of-day ----------
def awaiting_site_for_day(d: date) -> str:
    wd = d.weekday()
    if wd in (0,4):  # Mon/Fri
        return "SLGP"
    if wd in (1,3):  # Tue/Thu
        return "JEN"
    return "BGS"

def email_site_for_day(d: date) -> str:
    # Email preferred site pattern:
    # Mon = BGS, Tue/Wed/Thu = JEN, Fri = SLGP
    wd = d.weekday()
    if wd == 0:
        return "BGS"
    if wd in (1, 2, 3):
        return "JEN"
    return "SLGP"



# ---------- Break placement ----------
def pick_breaks_site_balanced(staff_list: List[Staff], hours_map, hols, week_dates: List[date], fixed_assignments: Set[Tuple[date,time,str]]) -> Dict[Tuple[date,time], Set[str]]:
    """
    Breaks only for staff with break_required AND shift > 6h.
    Spread within site by balancing counts per time slot.
    Avoid placing break where staff is locked to fixed assignment (FD/Triage) in that slot.
    """
    breaks: Dict[Tuple[date,time], Set[str]] = {}
    break_load: Dict[Tuple[date,str,time], int] = {}

    slots = timeslots()

    for d in week_dates:
        for st in staff_list:
            if not st.break_required:
                continue
            if holiday_kind(st.name, d, hols):
                continue
            stt, end = shift_window(hours_map, d, st.name)
            if not stt or not end:
                continue
            dur = (dt_of(d, end) - dt_of(d, stt)).total_seconds() / 3600.0
            if dur <= BREAK_THRESHOLD_HOURS:
                continue

            midpoint = dt_of(d, stt) + (dt_of(d, end) - dt_of(d, stt)) / 2

            best = None
            for bt in BREAK_CANDIDATES:
                if bt < stt or add_minutes(bt, 30) > end:
                    continue
                if not t_in_range(bt, BREAK_WINDOW[0], BREAK_WINDOW[1]):
                    continue
                # avoid if fixed assigned at that slot
                if (d, bt, st.name) in fixed_assignments:
                    continue

                # avoid tiny fragments
                before = (dt_of(d, bt) - dt_of(d, stt)).total_seconds() / 3600.0
                after = (dt_of(d, end) - dt_of(d, add_minutes(bt, 30))).total_seconds() / 3600.0
                frag_penalty = 0
                if before < 1.0:
                    frag_penalty += 10_000
                if after < 1.0:
                    frag_penalty += 10_000

                load = break_load.get((d, st.home, bt), 0)
                dist = abs((dt_of(d, bt) - midpoint).total_seconds())
                score = frag_penalty + (load * 3600) + dist
                if best is None or score < best[0]:
                    best = (score, bt)

            if best:
                bt = best[1]
                breaks.setdefault((d, bt), set()).add(st.name)
                break_load[(d, st.home, bt)] = break_load.get((d, st.home, bt), 0) + 1

    return breaks

# ---------- Scheduling ----------
def task_weight(st: Staff, task_key: str) -> int:
    # Default weight = 3 if not present
    return int(st.weights.get(task_key, 3) if st.weights is not None else 3)

def block_limits(task: str) -> Tuple[int,int]:
    if task == "Phones":
        return MIN_PHONES, MAX_PHONES
    if task == "Docman":
        return MIN_DOCMAN, MAX_DEFAULT
    # fixed tasks not here
    return MIN_DEFAULT, MAX_DEFAULT

def schedule_week(tpl: TemplateData, wk_start: date):
    slots = timeslots()
    dates = [wk_start + timedelta(days=i) for i in range(5)]

    # apply swaps
    hours_map = apply_swaps(tpl.hours_map, tpl.swaps, dates)

    staff_by_name = {s.name: s for s in tpl.staff}
    staff_names = [s.name for s in tpl.staff]

    # Assignments (d,t,name)->task
    a: Dict[Tuple[date,time,str], str] = {}
    gaps: List[Tuple[date,time,str,str]] = []

    # minutes tracking
    mins_task: Dict[Tuple[str,str], int] = {}           # (name, taskKey)->minutes
    mins_task_day: Dict[Tuple[date,str,str], int] = {}  # (d,name,taskKey)->minutes

    # Phones fairness: limit per-person phone time/bands per day (relaxed only if unavoidable)
    PHONE_DAY_MAX_MINS = 240  # 4h max on phones per person per day (unless unavoidable)
    PHONE_DAY_MAX_BANDS = 2   # max 2 phone blocks per day (unless unavoidable)
    phone_bands_day: Dict[Tuple[date, str], int] = {}  # (date,name)->count of started phone blocks

    def add_mins(d: date, nm: str, task_key: str, mins: int):
        mins_task[(nm, task_key)] = mins_task.get((nm, task_key), 0) + mins
        mins_task_day[(d, nm, task_key)] = mins_task_day.get((d, nm, task_key), 0) + mins

    def assigned(nm: str, d: date, t: time) -> Optional[str]:
        return a.get((d,t,nm))

    def is_free(nm: str, d: date, t: time) -> bool:
        return (d,t,nm) not in a

    # --- Fixed assignments: Front Desk + Triage bands
    fixed_slots: Set[Tuple[date,time,str]] = set()

    def can_cover_full_band(nm: str, d: date, bs: time, be: time) -> bool:
        """True if staff can cover the entire band AND does not already have locked/fixed work in that window."""
        stt, endt = shift_window(hours_map, d, nm)
        if not stt or not endt:
            return False
        if not ((stt <= bs) and (endt >= be)):
            return False
        # Prevent fixed-task overlap (e.g., don't let Triage overwrite Front Desk)
        for tt in slots:
            if tt < bs or tt >= be:
                continue
            if (d, tt, nm) in fixed_slots:
                return False
            if (d, tt, nm) in a:  # already assigned for any reason
                return False
        return True

    def pick_for_band(candidates: List[str], d: date, task_key: str, bs: time, be: time) -> Optional[str]:
        ok = []
        for nm in candidates:
            if holiday_kind(nm, d, tpl.hols):
                continue
            if not can_cover_full_band(nm, d, bs, be):
                continue
            ok.append(nm)
        if not ok:
            return None

        def score(nm: str):
            st = staff_by_name[nm]
            # FrontDeskOnly gets absolute precedence for FD bands
            fd_only = 1 if (task_key == "FrontDesk" and st.frontdesk_only) else 0
            # weight-first, then least used
            w = task_weight(st, task_key)
            used = mins_task.get((nm, task_key), 0)
            return (-fd_only, -w, used, nm.lower())

        ok.sort(key=score)
        return ok[0]

    # --------------------------------------------------
    # Front Desk — band-based enforcement (primary)
    # Bands: 08:00–11:00, 11:00–13:00, 13:00–16:00, 16:00–18:30
    # JEN/BGS can fall back to 30-min flex pattern if STRICT bands cannot be filled.
    # Exactly 1 person per site per slot; no 30/60-min fragments.
    # --------------------------------------------------
    FD_BANDS_STRICT = FD_BANDS
    FD_BANDS_FLEX = [
        (time(8, 0), time(11, 0)),
        (time(11, 0), time(13, 30)),
        (time(13, 30), time(16, 0)),
        (time(16, 0), time(18, 30)),
    ]

    
    def assign_frontdesk_for_site_day(site: str, d: date) -> None:
        """Assign Front Desk in bands, with fairness:
        - Normally max 1 FD band per person per day
        - If impossible, allow a 2nd band but NEVER consecutive (must be different person than previous band)
        - Never exceed 2 FD bands per person per day
        """
        role = f"FrontDesk_{site}"
        cands_all = [s.name for s in tpl.staff if s.can_frontdesk and s.home == site]

        fd_count: Dict[str, int] = {nm: 0 for nm in cands_all}
        last_band_staff: Optional[str] = None

        def fd_score(nm: str):
            st = staff_by_name[nm]
            fd_only = 1 if st.frontdesk_only else 0
            w = task_weight(st, "FrontDesk")
            used = mins_task.get((nm, "FrontDesk"), 0)
            # deterministic: higher fd_only, higher weight, less used minutes, name
            return (-fd_only, -w, used, nm.lower())

        def pick_fd_for_band(bs: time, be: time) -> Optional[str]:
            """Pick staff for a single FD band with the fairness rules."""
            # Tier 1: people with 0 FD bands today, and not the last band staff
            tiers = [
                lambda nm: fd_count.get(nm, 0) == 0 and (last_band_staff is None or nm != last_band_staff),
                # Tier 2: allow a 2nd FD band, but still not consecutive
                lambda nm: fd_count.get(nm, 0) == 1 and (last_band_staff is None or nm != last_band_staff),
            ]
            for ok_rule in tiers:
                pool = []
                for nm in cands_all:
                    if not ok_rule(nm):
                        continue
                    if fd_count.get(nm, 0) >= 2:
                        continue
                    if holiday_kind(nm, d, tpl.hols):
                        continue
                    if not can_cover_full_band(nm, d, bs, be):
                        continue
                    pool.append(nm)
                if pool:
                    pool.sort(key=fd_score)
                    return pool[0]
            return None

        def try_bands(bands):
            nonlocal last_band_staff
            # Try to fill all bands without leaving partial state behind.
            orig_count = dict(fd_count)
            orig_last = last_band_staff

            tmp_count = dict(fd_count)
            tmp_last = last_band_staff
            chosen_by_band = []

            for bs, be in bands:
                # Make the picker see the in-progress state for this attempt.
                fd_count.clear(); fd_count.update(tmp_count)
                last_band_staff = tmp_last

                chosen = pick_fd_for_band(bs, be)
                if not chosen:
                    # restore and fail
                    fd_count.clear(); fd_count.update(orig_count)
                    last_band_staff = orig_last
                    return None

                chosen_by_band.append((bs, be, chosen))
                tmp_count[chosen] = tmp_count.get(chosen, 0) + 1
                tmp_last = chosen

            # commit the successful attempt state
            fd_count.clear(); fd_count.update(tmp_count)
            last_band_staff = tmp_last
            return chosen_by_band

        bands = FD_BANDS_STRICT
        picked = try_bands(bands)
        if picked is None and site in ("JEN", "BGS"):
            bands = FD_BANDS_FLEX
            picked = try_bands(bands)

        if picked is None:
            for bs, _be in (FD_BANDS_FLEX if site in ("JEN","BGS") else FD_BANDS_STRICT):
                gaps.append((d, bs, role, "No suitable staff for Front Desk band"))
            return

        for bs, be, chosen in picked:
            for tt in slots:
                if tt < bs or tt >= be:
                    continue
                a[(d, tt, chosen)] = role
                fixed_slots.add((d, tt, chosen))
                add_mins(d, chosen, "FrontDesk", SLOT_MIN)

    for d in dates:
        for site in ("SLGP", "JEN", "BGS"):
            assign_frontdesk_for_site_day(site, d)

    # Triage (SLGP/JEN)
    for d in dates:
        for site in ("SLGP","JEN"):
            role = f"Triage_Admin_{site}"
            cands = [s.name for s in tpl.staff if s.can_triage and s.home == site]
            for bs, be in TRIAGE_BANDS:
                chosen = pick_for_band(cands, d, "Triage", bs, be)
                if not chosen:
                    gaps.append((d, bs, role, "No suitable staff for full band"))
                    continue
                for tt in slots:
                    if tt < bs or tt >= be:
                        continue
                    a[(d, tt, chosen)] = role
                    fixed_slots.add((d, tt, chosen))
                    add_mins(d, chosen, "Triage", SLOT_MIN)

    # --- Breaks (only >6h and BreakRequired)
    breaks = pick_breaks_site_balanced(tpl.staff, hours_map, tpl.hols, dates, fixed_slots)

    def on_break(nm: str, d: date, t: time) -> bool:
        return nm in breaks.get((d,t), set())

    # --- Active blocks for variable tasks
    active: Dict[Tuple[date,str], Tuple[str,int]] = {}  # (d,name)->(task, end_idx_excl)

    def apply_active(nm: str, d: date, idx: int) -> bool:
        b = active.get((d,nm))
        if not b:
            return False
        task, end_idx = b
        if idx >= end_idx:
            del active[(d,nm)]
            return False
        t = slots[idx]
        if not is_free(nm,d,t):
            # someone already assigned (fixed or enforced), stop block
            del active[(d,nm)]
            return False
        a[(d,t,nm)] = task
        add_mins(d, nm, task_key_for_task(task), SLOT_MIN)
        return True

    def stop_block(nm: str, d: date):
        if (d,nm) in active:
            del active[(d,nm)]

    def task_key_for_task(task: str) -> str:
        if task.startswith("FrontDesk_"):
            return "FrontDesk"
        if task.startswith("Triage_Admin_"):
            return "Triage"
        if task == "Email_Box":
            return "Email"
        if task == "Awaiting_PSA_Admin":
            return "Awaiting"
        if task == "Phones":
            return "Phones"
        if task == "Bookings":
            return "Bookings"
        if task == "EMIS":
            return "EMIS"
        if task == "Docman":
            return "Docman"
        if task == "Misc_Tasks":
            return "Misc"
        return "Misc"

    def eligible(nm: str, task: str, d: date, t: time, allow_cross_site: bool=False) -> bool:
        st = staff_by_name[nm]
        if holiday_kind(nm, d, tpl.hols):
            return False
        if not is_working(hours_map, d, t, nm):
            return False
        if on_break(nm, d, t):
            return False
        # Hard rule: FrontDeskOnly staff must never be assigned to other tasks
        if getattr(st, "frontdesk_only", False):
            return False
        if task.startswith("FrontDesk_") or task.startswith("Triage_Admin_"):
            return False  # already handled fixed
        if task == "Email_Box":
            # Soft site preference (site-of-day first; allow cross-site if needed)
            if not st.can_email:
                return False
            if allow_cross_site:
                return True
            return st.home == email_site_for_day(d)
        if task == "Awaiting_PSA_Admin":
            if not st.can_docman:
                return False
            if allow_cross_site:
                return True
            return st.home == awaiting_site_for_day(d)
        if task == "Phones":
            return st.can_phones
        if task == "Bookings":
            if not st.can_bookings:
                return False
            if allow_cross_site:
                return True
            return st.home == "SLGP"
        if task == "EMIS":
            return st.can_emis
        if task == "Docman":
            return st.can_docman
        if task == "Misc_Tasks":
            return True
        return True

    def start_block(nm: str, task: str, d: date, start_idx: int, allow_short_end: bool=True) -> bool:
        mn, mx = block_limits(task)
        stt, end = shift_window(hours_map, d, nm)
        if not stt or not end:
            return False

        # compute end_idx by shift end or break or fixed assignment boundary
        end_idx = start_idx
        while end_idx < len(slots) and slots[end_idx] < end:
            # stop if a fixed assignment exists at this slot for nm
            if (d, slots[end_idx], nm) in fixed_slots:
                break
            # stop at break start
            if nm in breaks.get((d, slots[end_idx]), set()):
                break
            end_idx += 1

        remaining = end_idx - start_idx
        if remaining <= 0:
            return False

        # no floaters: must be >= mn unless it's end-of-shift remainder (allow_short_end)
        # Phones are a hard-coverage task: if we're short late in the day, allow shorter end blocks.
        if task == "Phones" and remaining < mn:
            allow_short_end = True
        if remaining < mn and not allow_short_end:
            return False

        L = remaining if remaining < mn else min(mx, remaining)
        # if we have enough remaining to meet min, enforce min
        if remaining >= mn:
            L = max(mn, L)

        active[(d,nm)] = (task, start_idx + L)
        return True

    def pick_candidates(task: str, d: date, t: time, allow_cross_site: bool=False, prefer_sites: Optional[List[str]]=None, phones_strict: bool=True) -> List[str]:
        cands = []
        for nm in staff_names:
            if not is_free(nm,d,t):
                continue
            if not eligible(nm, task, d, t, allow_cross_site=allow_cross_site):
                continue

            # Phones fairness guardrails (strict pass). Relax only if we cannot meet phones requirement.
            if task == "Phones" and phones_strict:
                bands = phone_bands_day.get((d, nm), 0)
                mins_today = mins_task_day.get((d, nm, "Phones"), 0)
                if bands >= PHONE_DAY_MAX_BANDS:
                    continue
                if mins_today >= PHONE_DAY_MAX_MINS:
                    continue

            cands.append(nm)

        # site preference for EMIS/Docman: JEN/BGS first
        if prefer_sites:
            cands.sort(key=lambda nm: (0 if staff_by_name[nm].home in prefer_sites else 1, nm.lower()))

        def score(nm: str):
            st = staff_by_name[nm]
            key = task_key_for_task(task)
            w = task_weight(st, key)
            used = mins_task.get((nm, key), 0)

            # Phones: spread load across staff (bands then minutes today), then weights.
            if task == "Phones":
                bands = phone_bands_day.get((d, nm), 0)
                mins_today = mins_task_day.get((d, nm, "Phones"), 0)
                return (bands, mins_today, -w, used, nm.lower())

            # Bookings: always prefer SLGP when cross-site is allowed (still deterministic).
            if task == "Bookings":
                pref = 0 if str(st.home).upper() == "SLGP" else 1
                return (pref, -w, used, nm.lower())

            return (-w, used, nm.lower())

        cands.sort(key=score)
        return cands

    def assign_block(nm: str, task: str, d: date, idx: int):
        """Assign a task starting at slot idx, respecting existing active blocks."""
        # If already active on the same task, just continue the block.
        b = active.get((d, nm))
        if b and b[0] == task:
            apply_active(nm, d, idx)
            return

        # If active on a different task, stop that block before starting a new one.
        if b and b[0] != task:
            stop_block(nm, d)

        # Start a new block
        ok = start_block(nm, task, d, idx, allow_short_end=(task == "Misc_Tasks"))
        if not ok:
            # fallback: if we cannot start a valid block, force Misc (true end-of-shift only)
            a[(d, slots[idx], nm)] = "Misc_Tasks"
            add_mins(d, nm, "Misc", SLOT_MIN)
            return

        # Track phone blocks (bands) only when a NEW Phones block starts
        if task == "Phones":
            phone_bands_day[(d, nm)] = phone_bands_day.get((d, nm), 0) + 1

        apply_active(nm, d, idx)

    # --- Weekly targets (minutes)
    target_book = int(round((tpl.weekly_targets.get("Bookings", 0.0) or 0.0) * 60))
    target_emis = int(round((tpl.weekly_targets.get("EMIS", 0.0) or 0.0) * 60))
    target_doc  = int(round((tpl.weekly_targets.get("Docman", 0.0) or 0.0) * 60))
    # Targets with buffer (avoid 30-minute micro-assignments and overshoot)
    BOOK_LOW = 0.90   # aim to hit at least 90% of target
    BOOK_HIGH = 1.15  # allow up to 15% overshoot
    EMIS_LOW = 0.90
    EMIS_HIGH = 1.15
    DOC_LOW = 0.90
    DOC_HIGH = 1.15

    def total_mins(task_key: str) -> int:
        return sum(v for (nm, tk), v in mins_task.items() if tk == task_key)

    # helper: compute dynamic bookings people-per-slot requirement to approach weekly target

    # helper: bookings requirement per slot until we reach buffered target (>=90% of target)
    def bookings_needed_this_slot(d: date, idx: int) -> int:
        if target_book <= 0:
            return 0
        t = slots[idx]
        if t < time(10,30):
            return 0

        done = total_mins("Bookings")
        low = int(round(target_book * BOOK_LOW))
        high = int(round(target_book * BOOK_HIGH))

        if done >= low:
            return 0  # target met (within buffer)

        remaining = max(0, low - done)

        # Remaining booking-eligible slots in the week (from now)
        rem_slots = 0
        for dd in dates:
            for tt in slots:
                if dd < d:
                    continue
                if dd == d and tt < t:
                    continue
                if tt >= time(10,30):
                    rem_slots += 1

        if rem_slots <= 0:
            return 0

        # People needed this slot to catch up smoothly; at least 1 if behind target
        ppl = math.ceil(remaining / (rem_slots * SLOT_MIN))
        return max(1, int(ppl))

    def enforce(task: str, need: int, d: date, idx: int, allow_cross_site: bool=False, prefer_sites: Optional[List[str]]=None, note_task_key: str=""):
        t = slots[idx]
        while True:
            current = len([nm for nm in staff_names if a.get((d,t,nm)) == task])
            if current >= need:
                return
            # Phones fairness: try strict selection first, relax only if we cannot meet requirement
            if task == "Phones":
                cands = pick_candidates(task, d, t, allow_cross_site=allow_cross_site, prefer_sites=prefer_sites, phones_strict=True)
                if not cands:
                    cands = pick_candidates(task, d, t, allow_cross_site=allow_cross_site, prefer_sites=prefer_sites, phones_strict=False)
            else:
                cands = pick_candidates(task, d, t, allow_cross_site=allow_cross_site, prefer_sites=prefer_sites, phones_strict=True)
            if not cands:
                # For Phones (hard requirement), allow stealing from non-fixed assignments.
                if task == "Phones":
                    steal = [
                        nm for nm in staff_names
                        if eligible(nm, task, d, t, allow_cross_site=True)
                        and (d, t, nm) in a
                        and not str(a.get((d, t, nm), "")).startswith(("FrontDesk_", "Triage_Admin_"))
                        and a.get((d, t, nm)) not in ("Email_Box", "Awaiting_PSA_Admin")
                    ]
                    if steal:
                        steal.sort(key=lambda nm: (-task_weight(staff_by_name[nm], "Phones"), mins_task.get((nm, "Phones"), 0), nm.lower()))
                        nm = steal[0]
                        # break any active block for this person so they don't snap back
                        stop_block(nm, d)
                        a[(d, t, nm)] = "Phones"
                        add_mins(d, nm, "Phones", SLOT_MIN)
                        continue
                gaps.append((d, t, task, f"Short by {need-current}"))
                return
            nm = cands[0]
            assign_block(nm, task, d, idx)

    # --- Main loop by slot
    for d in dates:
        # Continuity trackers (reduce job-hopping)
        fd_last: Dict[str, Optional[str]] = {s: None for s in SITES}
        fd_run: Dict[str, int] = {s: 0 for s in SITES}
        email_last: Optional[str] = None
        awaiting_last: Optional[str] = None

        # --------------------------------------------------------------
        # PSA Admin / Awaiting Response — stable block (min 2.5h)
        # Goal: cover 10:30–16:00 (later can be extended manually if needed).
        # Flex start by ±30 mins to avoid break fragmentation.
        # --------------------------------------------------------------
        psa_pref_site = awaiting_site_for_day(d)
        psa_window_start = time(10, 30)
        psa_window_end = time(16, 0)
        psa_start_options = [time(10, 0), time(10, 30), time(11, 0)]

        def _psa_score(nm: str) -> tuple:
            st = staff_by_name[nm]
            pref = 0 if st.home == psa_pref_site else 1
            w = task_weight(st, "Awaiting")
            used = mins_task.get((nm, "Awaiting"), 0)
            return (pref, -w, used, nm.lower())

        def _psa_assignable_slots(nm: str, st_time: time) -> int:
            st = staff_by_name[nm]
            if not st.can_docman:
                return 0
            if holiday_kind(nm, d, tpl.hols):
                return 0
            stt, endt = shift_window(hours_map, d, nm)
            if not stt or not endt:
                return 0
            # must intersect the window
            if endt <= psa_window_start or stt >= psa_window_end:
                return 0

            cnt = 0
            for tt in slots:
                if tt < psa_window_start or tt >= psa_window_end:
                    continue
                if tt < st_time:
                    continue
                if not is_working(hours_map, d, tt, nm):
                    continue
                if nm in breaks.get((d, tt), set()):
                    continue
                if (d, tt, nm) in fixed_slots:
                    continue
                cnt += 1
            return cnt

        # Pick a person + a start time (flex ±30) that yields at least 2.5h assignable time.
        psa_choice = None
        best = None
        for st_time in psa_start_options:
            cands = [nm for nm in staff_names if _psa_assignable_slots(nm, st_time) >= 5]
            cands.sort(key=_psa_score)
            if not cands:
                continue
            nm = cands[0]
            sc = (_psa_score(nm), st_time)
            if best is None or sc < best:
                best = sc
                psa_choice = (nm, st_time)

        if psa_choice:
            psa_nm, psa_st = psa_choice
            for tt in slots:
                if tt < psa_window_start or tt >= psa_window_end:
                    continue
                if tt < psa_st:
                    continue
                if not is_working(hours_map, d, tt, psa_nm):
                    continue
                if psa_nm in breaks.get((d, tt), set()):
                    continue
                if (d, tt, psa_nm) in fixed_slots:
                    continue
                if (d, tt, psa_nm) in a:
                    continue
                a[(d, tt, psa_nm)] = "Awaiting_PSA_Admin"
                fixed_slots.add((d, tt, psa_nm))
                add_mins(d, psa_nm, "Awaiting", SLOT_MIN)
        else:
            # Log gaps for each slot in window (useful for dashboard)
            for tt in slots:
                if tt < psa_window_start or tt >= psa_window_end:
                    continue
                gaps.append((d, tt, "Awaiting_PSA_Admin", "No suitable staff for PSA Admin window"))

        # --------------------------------------------------------------
        # Email — single named owner per day (10:30–16:00, or to 18:30 if late shift)
        # Prefer site-of-day: Mon=BGS, Tue/Wed/Thu=JEN, Fri=SLGP. If nobody suitable, allow cross-site.
        # Do NOT cover email break; the owner simply pauses during their break.
        # --------------------------------------------------------------
        email_pref_site = email_site_for_day(d)
        email_owner: Optional[str] = None
        email_end_target = time(16, 0)

        def _email_score(nm: str) -> tuple:
            st = staff_by_name[nm]
            pref = 0 if st.home == email_pref_site else 1
            w = task_weight(st, "Email")
            used = mins_task.get((nm, "Email"), 0)
            return (pref, -w, used, nm.lower())

        email_cands = []
        for nm in staff_names:
            st = staff_by_name[nm]
            if not st.can_email:
                continue
            if holiday_kind(nm, d, tpl.hols):
                continue
            stt, endt = shift_window(hours_map, d, nm)
            if not stt or not endt:
                continue
            # must start by 10:30 and run at least to 16:00
            if not (stt <= time(10, 30) and endt >= time(16, 0)):
                continue
            # must be free (not fixed) for the majority of the window; we'll enforce per-slot later
            email_cands.append(nm)

        email_cands.sort(key=_email_score)
        if email_cands:
            email_owner = email_cands[0]
            # If late shift, keep email to end-of-shift (up to 18:30)
            _stt, _endt = shift_window(hours_map, d, email_owner)
            if _endt and _endt > time(16, 0):
                email_end_target = min(_endt, DAY_END)

        for idx, t in enumerate(slots):
            # Skip non-working: handled per eligible check

            # -----------------------------------------------------------------
            # -----------------------------------------------------------------

            # ---------------- Email (stable owner) ----------------
            if email_owner and t_in_range(t, time(10,30), email_end_target):
                nm = email_owner
                if (
                    not holiday_kind(nm, d, tpl.hols)
                    and is_working(hours_map, d, t, nm)
                    and (not on_break(nm, d, t))
                    and (d, t, nm) not in fixed_slots
                ):
                    # Only assign if not already fixed to something else at this slot
                    a[(d, t, nm)] = "Email_Box"
                    fixed_slots.add((d, t, nm))
                    add_mins(d, nm, "Email", SLOT_MIN)

            # Apply active blocks after fixed assignments (FD/Triage/Email/Awaiting)
            for nm in staff_names:
                if (d, t, nm) in a:
                    continue
                if on_break(nm, d, t):
                    continue
                apply_active(nm, d, idx)
            # ---------------- Email coverage (hard shift 10:30–16:00) ----------------
            # Must be covered for every slot in the window.
            # Prefer the site-of-day, but allow cross-site fallback so it is always covered.
            # We allocate once at 10:30 by filling the whole window with as few staff changes as possible.
            
            # ---------------- Email Fixed Shift (handled by stable owner above) ----------------

# Awaiting/PSA Admin is allocated as a stable daily block (see start of day).


            # Phones (hard) all day per matrix — FAST per-slot allocator (no infinite loops)
            req_p = phones_required(tpl, d, t)
            if req_p > 0:
                # Prefer continuity: keep people already on Phones in the previous slot.
                prev_t = slots[idx-1] if idx > 0 else None
                keep = []
                if prev_t:
                    for nm in staff_names:
                        if a.get((d, prev_t, nm)) == "Phones" and eligible(nm, "Phones", d, t, allow_cross_site=True) and (d, t, nm) not in fixed_slots:
                            keep.append(nm)

                chosen = []
                for nm in keep:
                    if is_free(nm, d, t):
                        chosen.append(nm)
                        if len(chosen) >= req_p:
                            break

                def can_take(nm: str) -> bool:
                    return eligible(nm, "Phones", d, t, allow_cross_site=True) and (d, t, nm) not in fixed_slots

                # 1) fill from free candidates
                if len(chosen) < req_p:
                    free_cands = [nm for nm in staff_names if is_free(nm, d, t) and can_take(nm)]
                    free_cands.sort(key=lambda nm: (-task_weight(staff_by_name[nm], "Phones"), mins_task.get((nm, "Phones"), 0), nm.lower()))
                    for nm in free_cands:
                        if nm in chosen:
                            continue
                        chosen.append(nm)
                        if len(chosen) >= req_p:
                            break

                # 2) if still short, steal from non-fixed, non-essential tasks (never steal FD/Triage/Email/Awaiting)
                if len(chosen) < req_p:
                    stealable = []
                    for nm in staff_names:
                        if nm in chosen:
                            continue
                        if not can_take(nm):
                            continue
                        cur = a.get((d, t, nm))
                        if not cur:
                            continue
                        if str(cur).startswith(("FrontDesk_", "Triage_Admin_")):
                            continue
                        if cur in ("Email_Box", "Awaiting_PSA_Admin"):
                            continue
                        stealable.append(nm)
                    stealable.sort(key=lambda nm: (-task_weight(staff_by_name[nm], "Phones"), mins_task.get((nm, "Phones"), 0), nm.lower()))
                    for nm in stealable:
                        chosen.append(nm)
                        if len(chosen) >= req_p:
                            break

                # Apply phones for this slot
                if len(chosen) < req_p:
                    gaps.append((d, t, "Phones", f"Short by {req_p-len(chosen)}"))
                for nm in chosen[:req_p]:
                    # stop any running block so we don't auto-overwrite later
                    stop_block(nm, d)
                    a[(d, t, nm)] = "Phones"
                    add_mins(d, nm, "Phones", SLOT_MIN)

            # Bookings: dynamic weekly pressure, SLGP first; can use other sites if still short
            need_b = bookings_needed_this_slot(d, idx)
            if need_b > 0:
                # enforce on SLGP first
                enforce("Bookings", need_b, d, idx, allow_cross_site=False)
                cur = len([nm for nm in staff_names if a.get((d,t,nm)) == "Bookings"])
                if cur < need_b:
                    # top-up cross-site (but NEVER steal from phones below req - enforce uses only free staff)
                    enforce("Bookings", need_b, d, idx, allow_cross_site=True)

            # Fill remaining staff with EMIS/Docman until targets met, pref JEN/BGS

            # Fill remaining staff with Bookings/EMIS/Docman until buffered targets hit.
            # No Misc unless all targets are within buffer OR staff is not eligible for remaining target tasks.
            emis_done = total_mins("EMIS")
            doc_done  = total_mins("Docman")
            book_done = total_mins("Bookings")

            low_book = int(round(target_book * BOOK_LOW)) if target_book > 0 else 0
            low_emis = int(round(target_emis * EMIS_LOW)) if target_emis > 0 else 0
            low_doc  = int(round(target_doc  * DOC_LOW))  if target_doc  > 0 else 0

            filler_tasks: List[str] = []
            if target_book > 0 and book_done < low_book and t >= time(10,30):
                filler_tasks.append("Bookings")
            if target_doc > 0 and doc_done < low_doc:
                filler_tasks.append("Docman")
            if target_emis > 0 and emis_done < low_emis:
                filler_tasks.append("EMIS")

            if not filler_tasks:
                filler_tasks = ["Misc_Tasks"]

            # assign fillers to any remaining free staff, pref site logic
            for nm in staff_names:
                if not is_working(hours_map, d, t, nm):
                    continue
                if holiday_kind(nm, d, tpl.hols):
                    continue
                if on_break(nm, d, t):
                    continue
                if not is_free(nm, d, t):
                    continue

                # choose filler respecting site preference
                chosen = None
                for ft in filler_tasks:
                    if ft in ("EMIS","Docman"):
                        # JEN/BGS first; SLGP only if still needed (handled by prefer_sites ranking)
                        pass
                    if eligible(nm, ft, d, t, allow_cross_site=True):
                        chosen = ft
                        break
                if not chosen:
                    chosen = "Misc_Tasks"

                # for EMIS/Docman, prefer JEN/BGS by candidate picker rather than per-person; but we're here per-person.
                # We'll implement: if task is EMIS/Docman and nm is SLGP and there exists any free JEN/BGS who can do it, give SLGP misc instead.
                if chosen in ("EMIS","Docman") and staff_by_name[nm].home == "SLGP":
                    # look for any other free JEN/BGS candidate for same task
                    any_other = False
                    for other in staff_names:
                        if other == nm:
                            continue
                        if staff_by_name[other].home not in ("JEN","BGS"):
                            continue
                        if not is_free(other, d, t):
                            continue
                        if eligible(other, chosen, d, t, allow_cross_site=True):
                            any_other = True
                            break
                    if any_other:
                        chosen = "Misc_Tasks"

                assign_block(nm, chosen, d, idx)

    # Post-pass: enforce FD exactly one per site per slot (safety)
    for d in dates:
        for t in slots:
            for site in SITES:
                role = f"FrontDesk_{site}"
                assigned_staff = [nm for nm in staff_names if a.get((d,t,nm)) == role]
                if len(assigned_staff) > 1:
                    # keep the highest weight (FrontDesk), then least used
                    def fd_score(nm):
                        st = staff_by_name[nm]
                        w = task_weight(st, "FrontDesk")
                        used = mins_task.get((nm, "FrontDesk"), 0)
                        return (-w, used, nm.lower())
                    assigned_staff.sort(key=fd_score)
                    keep = assigned_staff[0]
                    for extra in assigned_staff[1:]:
                        a[(d,t,extra)] = "Misc_Tasks"
                        gaps.append((d,t,role,f"Removed duplicate FD assignment for {extra}; kept {keep}"))


    # ----------------------------
    # No 30-minute fragments: if a staff member has a single-slot task between two other blocks,
    # continue the previous task into that slot (fallback: next task if no previous).
    # Does not change fixed bands (FrontDesk_*, Triage_Admin_*) and does not override Holidays/Break.
    # ----------------------------
    def smooth_single_slot_blocks():
        FIXED_PREFIXES = ("FrontDesk_", "Triage_Admin_")
        SPECIAL = ("Break", "Holiday", "Bank Holiday", "Sick")
        for d in dates:
            for nm in staff_names:
                seq = [a.get((d, t, nm)) for t in slots]
                i = 0
                while i < len(slots):
                    task = seq[i]
                    if not task:
                        i += 1
                        continue
                    j = i + 1
                    while j < len(slots) and seq[j] == task:
                        j += 1
                    block_len = j - i
                    if block_len == 1:
                        tsk = str(task)
                        if tsk.startswith(FIXED_PREFIXES) or tsk in SPECIAL:
                            i = j
                            continue
                        prev = seq[i-1] if i-1 >= 0 else None
                        nxt = seq[j] if j < len(slots) else None
                        chosen = None
                        if prev and (not str(prev).startswith(FIXED_PREFIXES)) and str(prev) not in SPECIAL:
                            chosen = prev
                        elif nxt and (not str(nxt).startswith(FIXED_PREFIXES)) and str(nxt) not in SPECIAL:
                            chosen = nxt
                        # Do not auto-extend FrontDesk into other people's single-slot gaps (prevents duplicates)
                        if chosen:
                            a[(d, slots[i], nm)] = chosen
                            seq[i] = chosen
                    i = j

    # Safety: ensure nobody who is working (and not on break/holiday) is left blank
    for d in dates:
        for t in slots:
            for nm in staff_names:

                if not is_working(hours_map, d, t, nm):
                    continue

                if holiday_kind(nm, d, tpl.hols):
                    continue

                if nm in breaks.get((d, t), set()):
                    continue

                if not a.get((d, t, nm)):
                    a[(d, t, nm)] = "Misc_Tasks"
                    add_mins(d, nm, "Misc", SLOT_MIN)

    return a, breaks, gaps, dates, slots, hours_map
# ---------- Excel output ----------
ROLE_COLORS = {
    "FrontDesk_SLGP": "FFF2CC",
    "FrontDesk_JEN":  "FFF2CC",
    "FrontDesk_BGS":  "FFF2CC",
    "Triage_Admin_SLGP": "D9EAD3",
    "Triage_Admin_JEN":  "D9EAD3",
    "Email_Box": "CFE2F3",
    "Phones": "C9DAF8",
    "Bookings": "FCE5CD",
    "EMIS": "EAD1DC",
    "Docman": "D0E0E3",
    "Awaiting_PSA_Admin": "D0E0E3",
    "Misc_Tasks": "FFFFFF",
    "Break": "DDDDDD",
    "Holiday": "F8CBAD",
    "Bank Holiday": "FFD966",
    "Sick": "F4CCCC",
    "": "DDDDDD",
}

def fill_for(value: str) -> PatternFill:
    return PatternFill("solid", fgColor=ROLE_COLORS.get(value, "FFFFFF"))

THICK = Side(style="thick")
THIN = Side(style="thin")
DAY_BORDER = Border(top=THICK)
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_workbook(tpl: TemplateData, start_monday: date, weeks: int) -> Workbook:
    """
    Creates an Excel workbook where Week*_MasterTimeline is the single source of truth.
    Site timelines are formula-linked to the Master so manual edits propagate.
    Adds:
      - Site timelines (SLGP/JEN/BGS)
      - Coverage_By_Slot (names per task per slot)
      - Coverage_Dashboard (at-a-glance metrics + target progress bars)
      - Clear day separators + wider columns + consistent colour fills
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.styles import Border

    wb = Workbook()
    wb.remove(wb.active)

    staff_names = [s.name for s in tpl.staff]
    staff_by_name = {s.name: s for s in tpl.staff}

    # Column widths
    DATE_W, TIME_W, STAFF_W = 14, 8, 18

    def is_day_start(val_time: str) -> bool:
        return val_time == DAY_START.strftime("%H:%M")

    def is_day_end(val_time: str) -> bool:
        # last slot begins at 18:00 when day ends 18:30
        return val_time == "18:00"

    def merged_border(base: Border, top=False, bottom=False) -> Border:
        return Border(
            left=base.left, right=base.right,
            top=THICK if top else base.top,
            bottom=THICK if bottom else base.bottom,
        )

    # Tasks for coverage sheets/dashboard
    COVER_TASKS = [
        "FrontDesk_SLGP","FrontDesk_JEN","FrontDesk_BGS",
        "Triage_Admin_SLGP","Triage_Admin_JEN",
        "Email_Box","Awaiting_PSA_Admin","Phones",
        "Bookings","EMIS","Docman","Misc_Tasks"
    ]

    for w in range(weeks):
        wk_start = start_monday + timedelta(days=7*w)
        a, breaks, gaps, dates, slots, hours_map = schedule_week(tpl, wk_start)

        # ----------------------------
        # 1) Master Timeline
        # ----------------------------
        ws = wb.create_sheet(f"Week{w+1}_MasterTimeline")
        ws.append(["Date","Time"] + staff_names)
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "C2"

        # Set widths
        ws.column_dimensions["A"].width = DATE_W
        ws.column_dimensions["B"].width = TIME_W
        for i in range(len(staff_names)):
            ws.column_dimensions[get_column_letter(3+i)].width = STAFF_W

        for d in dates:
            for t in slots:
                row = [d.strftime("%a %d-%b"), t.strftime("%H:%M")]
                for nm in staff_names:
                    hk = holiday_kind(nm, d, tpl.hols)
                    if hk:
                        val = hk
                    elif not is_working(hours_map, d, t, nm):
                        val = ""
                    elif nm in breaks.get((d,t), set()):
                        val = "Break"
                    else:
                        val = a.get((d,t,nm), "Misc_Tasks")
                    row.append(val)
                ws.append(row)

        # Style Master (fills + day borders)
        for rr in range(2, ws.max_row+1):
            tval = str(ws.cell(rr,2).value or "")
            day_start = is_day_start(tval)
            day_end = is_day_end(tval)
            for cc in range(1, ws.max_column+1):
                cell = ws.cell(rr,cc)
                b = CELL_BORDER
                if day_start:
                    b = merged_border(b, top=True)
                if day_end:
                    b = merged_border(b, bottom=True)
                cell.border = b

                if cc >= 3:
                    val = str(cell.value or "")
                    cell.fill = fill_for(val)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

        # ----------------------------
        # 2) Site Timelines (formula-linked)
        # ----------------------------
        def make_site_timeline(site: str):
            site_staff = [nm for nm in staff_names if str(staff_by_name[nm].home).upper() == site]
            if not site_staff:
                return
            ws_site = wb.create_sheet(f"Week{w+1}_{site}_Timeline")
            ws_site.append(["Date","Time"] + site_staff)
            for c in ws_site[1]:
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws_site.freeze_panes = "C2"

            ws_site.column_dimensions["A"].width = DATE_W
            ws_site.column_dimensions["B"].width = TIME_W
            for i in range(len(site_staff)):
                ws_site.column_dimensions[get_column_letter(3+i)].width = STAFF_W

            master_col = {nm: 3 + staff_names.index(nm) for nm in site_staff}

            for rr in range(2, ws.max_row+1):
                date_val = ws.cell(rr,1).value
                time_val = ws.cell(rr,2).value
                ws_site.append([date_val, time_val] + [""]*len(site_staff))
                site_rr = ws_site.max_row
                for i, nm in enumerate(site_staff):
                    mc = master_col[nm]
                    src = f"'{ws.title}'!{get_column_letter(mc)}{rr}"
                    ws_site.cell(site_rr, 3+i).value = f"={src}"

            # Styling: borders + fills (fills won't compute until Excel opens; apply conditional formatting instead)
            from openpyxl.formatting.rule import FormulaRule

            data_range = f"{get_column_letter(3)}2:{get_column_letter(ws_site.max_column)}{ws_site.max_row}"
            # Apply expression rules so formula-linked cells colour correctly.
            tl = f"{get_column_letter(3)}2"  # top-left cell in the formatted range
            for task, color in ROLE_COLORS.items():
                if not task:
                    continue
                fill = PatternFill("solid", fgColor=color)
                ws_site.conditional_formatting.add(
                    data_range,
                    FormulaRule(formula=[f'ISNUMBER(SEARCH("{task}",{tl}))'], fill=fill, stopIfTrue=False)
                )

            for rr in range(2, ws_site.max_row+1):
                tval = str(ws_site.cell(rr,2).value or "")
                day_start = is_day_start(tval)
                day_end = is_day_end(tval)
                for cc in range(1, ws_site.max_column+1):
                    cell = ws_site.cell(rr,cc)
                    b = CELL_BORDER
                    if day_start:
                        b = merged_border(b, top=True)
                    if day_end:
                        b = merged_border(b, bottom=True)
                    cell.border = b
                    if cc >= 3:
                        cell.alignment = Alignment(vertical="top", wrap_text=True)

        for site in ("SLGP","JEN","BGS"):
            make_site_timeline(site)

        # ----------------------------
        # 3) Coverage_By_Slot (names)
        # ----------------------------
        ws_cov = wb.create_sheet(f"Week{w+1}_Coverage_By_Slot")
        ws_cov.append(["Date","Time"] + COVER_TASKS)
        for c in ws_cov[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_cov.freeze_panes = "C2"

        ws_cov.column_dimensions["A"].width = DATE_W
        ws_cov.column_dimensions["B"].width = TIME_W
        for i in range(len(COVER_TASKS)):
            ws_cov.column_dimensions[get_column_letter(3+i)].width = 26

        for d in dates:
            for t in slots:
                row = [d.strftime("%a %d-%b"), t.strftime("%H:%M")]
                for task in COVER_TASKS:
                    names = [nm for nm in staff_names if a.get((d,t,nm)) == task]
                    row.append(", ".join(names))
                ws_cov.append(row)

        for rr in range(2, ws_cov.max_row+1):
            tval = str(ws_cov.cell(rr,2).value or "")
            day_start = is_day_start(tval)
            day_end = is_day_end(tval)
            for cc in range(1, ws_cov.max_column+1):
                cell = ws_cov.cell(rr,cc)
                b = CELL_BORDER
                if day_start:
                    b = merged_border(b, top=True)
                if day_end:
                    b = merged_border(b, bottom=True)
                cell.border = b
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Apply background colours to Coverage_By_Slot columns for readability
        for cc in range(3, ws_cov.max_column+1):
            task = str(ws_cov.cell(1, cc).value or "")
            f = fill_for(task)
            for rr in range(2, ws_cov.max_row+1):
                ws_cov.cell(rr, cc).fill = f

        # ----------------------------
        # 4) Totals
        # ----------------------------
        ws_tot = wb.create_sheet(f"Week{w+1}_Totals_Static")
        tasks_tot = [
            "FrontDesk_SLGP","FrontDesk_JEN","FrontDesk_BGS",
            "Triage_Admin_SLGP","Triage_Admin_JEN","Email_Box",
            "Phones","Awaiting_PSA_Admin","Bookings","EMIS","Docman",
            "Misc_Tasks","Break"
        ]
        ws_tot.append(["Name"] + tasks_tot + ["WeeklyTotalHours"])
        for c in ws_tot[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_tot.freeze_panes = "B2"
        ws_tot.column_dimensions["A"].width = 22
        for i in range(len(tasks_tot)):
            ws_tot.column_dimensions[get_column_letter(2+i)].width = 14

        hours = {(nm, task): 0.0 for nm in staff_names for task in tasks_tot}
        for d in dates:
            for t in slots:
                for nm in staff_names:
                    task = a.get((d,t,nm))
                    if task:
                        hours[(nm, task)] = hours.get((nm, task), 0.0) + 0.5
                for nm in breaks.get((d,t), set()):
                    hours[(nm, "Break")] = hours.get((nm, "Break"), 0.0) + 0.5

        for nm in staff_names:
            row = [nm]
            total = 0.0
            for task in tasks_tot:
                v = round(hours.get((nm, task), 0.0), 2)
                row.append(v)
                total += v
            row.append(round(total, 2))
            ws_tot.append(row)

        # ----------------------------
        # 5) Coverage Dashboard
        # ----------------------------
        ws_dash = wb.create_sheet(f"Week{w+1}_Coverage_Dashboard")
        ws_dash.column_dimensions["A"].width = 44
        ws_dash.column_dimensions["B"].width = 14
        ws_dash.column_dimensions["C"].width = 14
        ws_dash.column_dimensions["D"].width = 10

        ws_dash["A1"] = "Coverage Dashboard"
        ws_dash["A1"].font = Font(bold=True, size=14)

        total_slots = len(dates) * len(slots)

        phones_ok = 0
        for d in dates:
            for t in slots:
                req = phones_required(tpl, d, t)
                actual = sum(1 for nm in staff_names if a.get((d,t,nm)) == "Phones")
                if actual >= req:
                    phones_ok += 1
        phones_pct = phones_ok / total_slots if total_slots else 1.0

        fd_ok = 0
        fd_total = total_slots * 3
        for d in dates:
            for t in slots:
                for site in ("SLGP","JEN","BGS"):
                    role = f"FrontDesk_{site}"
                    actual = sum(1 for nm in staff_names if a.get((d,t,nm)) == role)
                    if actual == 1:
                        fd_ok += 1
        fd_pct = fd_ok / fd_total if fd_total else 1.0

        # Break compliance
        req_break = 0
        got_break = 0
        for d in dates:
            for nm in staff_names:
                stt, end = shift_window(hours_map, d, nm)
                if not stt or not end:
                    continue
                dur = (dt_of(d, end) - dt_of(d, stt)).total_seconds()/3600.0
                if dur > BREAK_THRESHOLD_HOURS and staff_by_name[nm].break_required:
                    req_break += 1
                    had = any(nm in breaks.get((d, bt), set()) for bt in (time(12,0), time(12,30), time(13,0), time(13,30)))
                    if had:
                        got_break += 1
        break_pct = (got_break/req_break) if req_break else 1.0

        def achieved(task: str) -> float:
            return round(sum(0.5 for d in dates for t in slots for nm in staff_names if a.get((d,t,nm)) == task), 2)

        book_h = achieved("Bookings")
        emis_h = achieved("EMIS")
        doc_h = achieved("Docman")

        book_t = float(tpl.weekly_targets.get("Bookings", 0.0) or 0.0)
        emis_t = float(tpl.weekly_targets.get("EMIS", 0.0) or 0.0)
        doc_t = float(tpl.weekly_targets.get("Docman", 0.0) or 0.0)

        ws_dash.append(["Metric","Achieved","Target","%"])
        for c in ws_dash[2]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")

        rows = [
            ("Front Desk coverage (slots meeting rule)", fd_ok, fd_total, fd_pct),
            ("Phones coverage (slots meeting requirement)", phones_ok, total_slots, phones_pct),
            ("Break compliance (staff-days with break)", got_break, req_break, break_pct),
            ("Bookings hours", book_h, book_t, (book_h/book_t if book_t else 1.0)),
            ("EMIS hours", emis_h, emis_t, (emis_h/emis_t if emis_t else 1.0)),
            ("Docman hours", doc_h, doc_t, (doc_h/doc_t if doc_t else 1.0)),
        ]
        for r in rows:
            ws_dash.append([r[0], r[1], r[2], round(float(r[3]), 3)])

        # Data bars for %
        last_row = ws_dash.max_row
        ws_dash.conditional_formatting.add(f"D3:D{last_row}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="63C384"))

        for rr in range(2, ws_dash.max_row+1):
            for cc in range(1, 5):
                ws_dash.cell(rr, cc).border = CELL_BORDER
                ws_dash.cell(rr, cc).alignment = Alignment(vertical="center")

        # ----------------------------
        # 6) Notes/Gaps
        # ----------------------------
        ws_g = wb.create_sheet(f"Week{w+1}_NotesAndGaps")
        ws_g.append(["Date","Time","Task","Note"])
        for c in ws_g[1]:
            c.font = Font(bold=True)
        for d, t, task, note in gaps:
            ws_g.append([d.isoformat(), t.strftime("%H:%M") if t else "", task, note])


        # ---------------- Direct Site Timelines (values + fills) ----------------
        site_sheets = {}
        for site in ("SLGP","JEN","BGS"):
            ws_site2 = wb.create_sheet(f"Week{w+1}_{site}_Timeline")
            ws_site2.append(["Date","Time"] + staff_names)
            for c in ws_site2[1]:
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws_site2.freeze_panes = "C2"
            site_sheets[site] = ws_site2

        # Fill data rows
        for d in dates:
            for t in slots:
                date_label = d.strftime("%a %d-%b")
                time_label = t.strftime("%H:%M")
                for site, ws_site2 in site_sheets.items():
                    row = [date_label, time_label]
                    for nm in staff_names:
                        hk = holiday_kind(nm, d, tpl.hols)
                        if hk:
                            val = hk
                        elif not is_working(hours_map, d, t, nm):
                            val = ""
                        elif nm in breaks.get((d, t), set()):
                            val = "Break"
                        else:
                            val = a.get((d, t, nm), "")
                            row.append(val)

        # Apply fills + formatting to site sheets
        for site, ws_site2 in site_sheets.items():
            try:
                widen_columns(ws_site2, width=18)
            except Exception:
                pass
            ws_site2.column_dimensions["A"].width = 14
            ws_site2.column_dimensions["B"].width = 8
            for rr in range(2, ws_site2.max_row + 1):
                for cc in range(3, ws_site2.max_column + 1):
                    v = str(ws_site2.cell(rr, cc).value or "")
                    ws_site2.cell(rr, cc).fill = fill_for(v)
                    ws_site2.cell(rr, cc).alignment = Alignment(wrap_text=True, vertical="top")
            try:
                apply_day_borders(ws_site2)
            except Exception:
                pass

        # ---------------- Coverage By Slot By Site (names) ----------------
        ws_cov = wb.create_sheet(f"Week{w+1}_Coverage_By_Slot_By_Site")
        cov_cols = ["FD_SLGP","FD_JEN","FD_BGS","Triage_SLGP","Triage_JEN","Phones","Bookings","EMIS","Docman","Awaiting","Email","Misc"]
        ws_cov.append(["Date","Time"] + cov_cols)
        for c in ws_cov[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_cov.freeze_panes = "C2"
        try:
            widen_columns(ws_cov, width=22)
        except Exception:
            pass
        ws_cov.column_dimensions["A"].width = 14
        ws_cov.column_dimensions["B"].width = 8

        def names_for_prefix(prefix, d, t):
            return ", ".join([nm for nm in staff_names if str(a.get((d,t,nm),"")).startswith(prefix)])

        for d in dates:
            for t in slots:
                row = [d.strftime("%a %d-%b"), t.strftime("%H:%M")]
                row.append(names_for_prefix("FrontDesk_SLGP", d, t))
                row.append(names_for_prefix("FrontDesk_JEN", d, t))
                row.append(names_for_prefix("FrontDesk_BGS", d, t))
                row.append(names_for_prefix("Triage_Admin_SLGP", d, t))
                row.append(names_for_prefix("Triage_Admin_JEN", d, t))
                row.append(", ".join([nm for nm in staff_names if a.get((d,t,nm))=="Phones"]))
                row.append(", ".join([nm for nm in staff_names if a.get((d,t,nm))=="Bookings"]))
                row.append(", ".join([nm for nm in staff_names if a.get((d,t,nm))=="EMIS"]))
                row.append(", ".join([nm for nm in staff_names if a.get((d,t,nm))=="Docman"]))
                row.append(", ".join([nm for nm in staff_names if a.get((d,t,nm))=="Email_Box"]))
                row.append(", ".join([nm for nm in staff_names if a.get((d,t,nm))=="Awaiting_PSA_Admin"]))
                row.append(", ".join([nm for nm in staff_names if a.get((d,t,nm))=="Misc_Tasks"]))
                ws_cov.append(row)

        # Colour coverage columns by meaning
        col_fills = {
            "FD_": PatternFill("solid", fgColor=ROLE_COLORS.get("FrontDesk_SLGP","FFF2CC")),
            "Triage_": PatternFill("solid", fgColor=ROLE_COLORS.get("Triage_Admin_SLGP","D9EAD3")),
            "Phones": PatternFill("solid", fgColor=ROLE_COLORS.get("Phones","C9DAF8")),
            "Bookings": PatternFill("solid", fgColor=ROLE_COLORS.get("Bookings","FCE5CD")),
            "EMIS": PatternFill("solid", fgColor=ROLE_COLORS.get("EMIS","EAD1DC")),
            "Docman": PatternFill("solid", fgColor=ROLE_COLORS.get("Docman","D0E0E3")),
            "Awaiting": PatternFill("solid", fgColor=ROLE_COLORS.get("Awaiting_PSA_Admin","D0E0E3")),
            "Email": PatternFill("solid", fgColor=ROLE_COLORS.get("Email_Box","CFE2F3")),
            "Misc": PatternFill("solid", fgColor=ROLE_COLORS.get("Misc_Tasks","EFEFEF")),
        }
        for cc in range(3, ws_cov.max_column+1):
            header = str(ws_cov.cell(1,cc).value)
            fill = None
            for k,v in col_fills.items():
                if header.startswith(k) or header == k:
                    fill = v
                    break
            if fill:
                for rr in range(2, ws_cov.max_row+1):
                    ws_cov.cell(rr,cc).fill = fill
                    ws_cov.cell(rr,cc).alignment = Alignment(wrap_text=True, vertical="top")
        try:
            apply_day_borders(ws_cov)
        except Exception:
            pass

        except Exception:
            pass

        # ---------------- Dynamic Totals (based on Site sheets) ----------------
        task_keys = ["FrontDesk","Triage","Phones","Bookings","EMIS","Docman","Awaiting","Email","Misc","Break"]
        ws_dyn = write_dynamic_totals_from_site_sheets(wb, w+1, staff_names, task_keys)

    return wb



def apply_day_borders(ws):
    # Thick border between days (every 48 rows for 30-min slots * 24 hours approx 48 slots)
    for col in range(1, ws.max_column + 1):
        for row in range(1, ws.max_row + 1):
            if row > 1 and (row-2) % 48 == 0:
                cell = ws.cell(row=row, column=col)
                cell.border = Border(top=THICK)

def widen_columns(ws, width=18):
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = width

def add_progress_bars(ws, col_letter, max_value):
    rule = DataBarRule(start_type='num', start_value=0,
                       end_type='num', end_value=max_value,
                       color="63C384")
    ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)

# ===========================
# Site display mapping helpers (for Site Timelines output)
# ===========================
def home_site_of(name: str, staff_by_name: dict) -> str:
    try:
        return str(staff_by_name[name].get("HomeSite","")).strip().upper()
    except Exception:
        return ""

def display_site_for_assignment(role: str, d, name: str, staff_by_name: dict) -> str:
    """Decide which site sheet should display this assignment."""
    if not role:
        return ""
    if role.endswith("_SLGP"):
        return "SLGP"
    if role.endswith("_JEN"):
        return "JEN"
    if role.endswith("_BGS"):
        return "BGS"
    if role == "Bookings":
        return "SLGP"
    if role == "Awaiting_PSA_Admin":
        return awaiting_site_for_day(d)
    if role == "Email_Box":
        return home_site_of(name, staff_by_name)  # JEN/BGS
    # Phones / EMIS / Docman / Misc -> staff home site for display
    return home_site_of(name, staff_by_name)

# ===========================
# Dynamic Totals (updates when Site timelines edited)
# ===========================
def write_dynamic_totals_from_site_sheets(wb, week_num: int, staff_names, task_keys):
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(f"Week{week_num}_Totals")
    ws.append(["Name"] + task_keys + ["WeeklyTotal"])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r, nm in enumerate(staff_names, start=2):
        row = [nm]
        staff_col_idx = 3 + (r-2)  # timelines: Date(A), Time(B), staff start at C
        staff_col = get_column_letter(staff_col_idx)
        rng = f"{staff_col}$2:{staff_col}$500"

        for task in task_keys:
            if task == "FrontDesk":
                crit = "FrontDesk*"
            elif task == "Triage":
                crit = "Triage*"
            elif task == "Phones":
                crit = "Phones"
            elif task == "Bookings":
                crit = "Bookings"
            elif task == "EMIS":
                crit = "EMIS"
            elif task == "Docman":
                crit = "Docman"
            elif task == "Awaiting":
                crit = "Awaiting_PSA_Admin"
            elif task == "Email":
                crit = "Email_Box"
            elif task == "Misc":
                crit = "Misc_Tasks"
            elif task == "Break":
                crit = "Break"
            else:
                crit = task

            f = (
                f"=0.5*("
                f'COUNTIF(Week{week_num}_SLGP_Timeline!{rng},"{crit}")+'
                f'COUNTIF(Week{week_num}_JEN_Timeline!{rng},"{crit}")+'
                f'COUNTIF(Week{week_num}_BGS_Timeline!{rng},"{crit}")'
                f")"
            )
            row.append(f)

        start_letter = get_column_letter(2)
        end_letter = get_column_letter(1 + len(task_keys))
        row.append(f"=SUM({start_letter}{r}:{end_letter}{r})")
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 20
    for i in range(2, 2+len(task_keys)+1):
        ws.column_dimensions[get_column_letter(i)].width = 12
    return ws


# =========================================================
# v32 STABLE WORKBOOK WRITER + RECALC (no formulas, no CF)
# Source-of-truth = Site Timelines values
# =========================================================

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

THICK_SIDE = Side(style="thick")
THIN_SIDE = Side(style="thin")

def _border(top=False, bottom=False, left=False, right=False):
    return Border(
        top=THICK_SIDE if top else THIN_SIDE,
        bottom=THICK_SIDE if bottom else THIN_SIDE,
        left=THICK_SIDE if left else THIN_SIDE,
        right=THICK_SIDE if right else THIN_SIDE,
    )

def _time_str(t: time) -> str:
    return t.strftime("%H:%M")

def _day_sep_rows(slots_count: int) -> int:
    # One day = number of slots in a day (Mon..Fri separate by repeating header row)
    return slots_count

def _task_color(task: str) -> str:
    return ROLE_COLORS.get(task, "FFFFFF")

def _apply_fill_and_style(ws, start_row, start_col, end_row, end_col, repeat_header_every=None, header_row=1):
    """Apply cell fills based on value strings + basic borders; optionally repeat header row for printing."""
    for r in range(1, ws.max_row + 1):
        # repeat header row (copy values + styles) at boundaries
        if repeat_header_every and r > header_row and (r - header_row - 1) % repeat_header_every == 0:
            # this row is start of a new day block: already inserted by builder; style it as header
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            continue

    for r in range(start_row, end_row + 1):
        # thick top border at start of each day block (after repeated header row)
        if repeat_header_every:
            # day block starts right after each repeated header row
            # builder inserts header rows; detect by Time cell being 08:00
            time_val = str(ws.cell(r, 2).value or "")
            top_day = (time_val == DAY_START.strftime("%H:%M"))
        else:
            top_day = False

        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            val = str(cell.value or "")
            # Fill only for task columns (>=3); Date/Time columns left plain
            if c >= 3:
                cell.fill = PatternFill("solid", fgColor=_task_color(val))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            # Borders
            b = Border(
                top=THICK_SIDE if top_day else THIN_SIDE,
                bottom=THIN_SIDE,
                left=THIN_SIDE,
                right=THIN_SIDE,
            )
            cell.border = b

def _site_for_display(role: str, d: date, name: str, staff_by_name: dict) -> str:
    """Decide which site sheet should display this assignment.

    IMPORTANT: Site timelines are already split by staff home site (columns).
    So we should NOT blank tasks like Bookings/Awaiting/Email simply because they are
    'conceptually' owned by another site. If a staff member is assigned a task,
    it must display on their home-site timeline.
    """
    if not role:
        return ""
    if role.endswith("_SLGP"):
        return "SLGP"
    if role.endswith("_JEN"):
        return "JEN"
    if role.endswith("_BGS"):
        return "BGS"
    # Default: show on staff home site
    try:
        return str(staff_by_name[name].home).upper()
    except Exception:
        return ""

def _compute_totals_from_assignments(assignments, breaks, dates, slots, staff_names):
    totals = {}  # (name, task)->hours
    for nm in staff_names:
        totals[(nm, "WeeklyTotal")] = 0.0
    for d in dates:
        for t in slots:
            for nm in staff_names:
                task = assignments.get((d, t, nm))
                if task:
                    totals[(nm, task)] = totals.get((nm, task), 0.0) + 0.5
                    totals[(nm, "WeeklyTotal")] += 0.5
            for nm in breaks.get((d, t), set()):
                totals[(nm, "Break")] = totals.get((nm, "Break"), 0.0) + 0.5
                totals[(nm, "WeeklyTotal")] = totals.get((nm, "WeeklyTotal"), 0.0) + 0.5
    return totals

def _compute_totals_from_site_timelines(wb, week_num: int):
    """Read values from site timelines and compute per-staff totals."""
    # identify site sheets
    site_sheets = {}
    for site in ("SLGP", "JEN", "BGS"):
        nm = f"Week{week_num}_{site}_Timeline"
        if nm in wb.sheetnames:
            site_sheets[site] = wb[nm]
    if not site_sheets:
        raise ValueError(f"No site timeline sheets found for Week{week_num} (expected Week{week_num}_SLGP_Timeline etc.)")

    # staff headers are in row 1, columns C..end
    # all site sheets share same staff list (site-only, but headers present)
    # We'll take union in display order from each sheet.
    staff = []
    for ws in site_sheets.values():
        hdr = [ws.cell(1, c).value for c in range(3, ws.max_column + 1)]
        hdr = [str(x).strip() for x in hdr if x and str(x).strip()]
        for h in hdr:
            if h not in staff:
                staff.append(h)

    tasks = set()
    for ws in site_sheets.values():
        for r in range(2, ws.max_row + 1):
            # skip repeated header rows (Time cell is 'Time')
            if str(ws.cell(r,2).value or '').strip().lower() == 'time':
                continue
            for c in range(3, ws.max_column + 1):
                v = str(ws.cell(r,c).value or '').strip()
                if v:
                    tasks.add(v)
    # Normalize list of tasks of interest
    known = [
        "FrontDesk_SLGP","FrontDesk_JEN","FrontDesk_BGS",
        "Triage_Admin_SLGP","Triage_Admin_JEN",
        "Phones","Bookings","Email_Box","Awaiting_PSA_Admin",
        "EMIS","Docman","Misc_Tasks","Break","Holiday","Sick","Bank Holiday"
    ]
    tasks_order = [t for t in known if t in tasks]
    # include any unexpected tasks at end
    for t in sorted(tasks):
        if t not in tasks_order:
            tasks_order.append(t)

    totals = {(nm, t): 0.0 for nm in staff for t in tasks_order}
    totals.update({(nm, "WeeklyTotal"): 0.0 for nm in staff})

    for ws in site_sheets.values():
        # map staff col
        col_map = {}
        for c in range(3, ws.max_column + 1):
            h = ws.cell(1,c).value
            if h and str(h).strip() in staff:
                col_map[str(h).strip()] = c
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r,2).value or '').strip().lower() == 'time':
                continue
            for nm in staff:
                c = col_map.get(nm)
                if not c:
                    continue
                v = str(ws.cell(r,c).value or '').strip()
                if not v:
                    continue
                totals[(nm, v)] = totals.get((nm, v), 0.0) + 0.5
                totals[(nm, "WeeklyTotal")] = totals.get((nm, "WeeklyTotal"), 0.0) + 0.5
    return staff, tasks_order, totals

def _write_totals_sheet(wb, week_num: int, staff, tasks_order, totals):
    title = f"Week{week_num}_Totals"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws.append(["Name"] + tasks_order + ["WeeklyTotal"])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 22
    for i in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    for nm in staff:
        row = [nm]
        for t in tasks_order:
            row.append(round(totals.get((nm, t), 0.0), 2))
        row.append(round(totals.get((nm, "WeeklyTotal"), 0.0), 2))
        ws.append(row)
    return ws

def _write_coverage_sheet_from_site_timelines(wb, week_num: int):
    title = f"Week{week_num}_Coverage_By_Slot"
    if title in wb.sheetnames:
        del wb[title]
    ws_cov = wb.create_sheet(title)

    # load site sheets
    site_sheets = {site: wb[f"Week{week_num}_{site}_Timeline"] for site in ("SLGP","JEN","BGS") if f"Week{week_num}_{site}_Timeline" in wb.sheetnames}
    if not site_sheets:
        raise ValueError(f"No site timelines for Week{week_num}")

    # tasks columns
    tasks_cols = [
        "FrontDesk_SLGP","FrontDesk_JEN","FrontDesk_BGS",
        "Triage_Admin_SLGP","Triage_Admin_JEN",
        "Email_Box","Awaiting_PSA_Admin","Phones","Bookings","EMIS","Docman","Misc_Tasks"
    ]
    ws_cov.append(["Date","Time"] + tasks_cols)
    for c in ws_cov[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws_cov.freeze_panes = "C2"
    ws_cov.column_dimensions["A"].width = 14
    ws_cov.column_dimensions["B"].width = 8
    for i in range(len(tasks_cols)):
        ws_cov.column_dimensions[get_column_letter(3+i)].width = 28

    # Use SLGP sheet as row backbone (date/time)
    backbone = next(iter(site_sheets.values()))
    # collect staff headers per site
    headers = {}
    for site, ws in site_sheets.items():
        headers[site] = [str(ws.cell(1,c).value).strip() for c in range(3, ws.max_column+1) if ws.cell(1,c).value]

    for r in range(2, backbone.max_row + 1):
        # skip repeated headers
        if str(backbone.cell(r,2).value or '').strip().lower() == 'time':
            continue
        dval = backbone.cell(r,1).value
        tval = backbone.cell(r,2).value
        row = [dval, tval]
        # build lookup per task -> list names by scanning all site sheets
        per_task = {k: [] for k in tasks_cols}
        for site, ws in site_sheets.items():
            for ci, nm in enumerate(headers[site], start=3):
                v = str(ws.cell(r,ci).value or '').strip()
                if not v:
                    continue
                if v in per_task:
                    per_task[v].append(nm)
        for k in tasks_cols:
            row.append(", ".join(per_task[k]))
        ws_cov.append(row)

    # simple styling
    for rr in range(2, ws_cov.max_row+1):
        time_val = str(ws_cov.cell(rr,2).value or "")
        top_day = (time_val == DAY_START.strftime("%H:%M"))
        for cc in range(1, ws_cov.max_column+1):
            cell = ws_cov.cell(rr,cc)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(top=THICK_SIDE if top_day else THIN_SIDE, bottom=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE)
    
    # Apply background colours to Coverage_By_Slot columns for readability (recalc-safe)
    # Colour by task column header (not by names in cells).
    for cc in range(3, ws_cov.max_column + 1):
        task = str(ws_cov.cell(1, cc).value or "")
        f = fill_for(task)
        for rr in range(2, ws_cov.max_row + 1):
            ws_cov.cell(rr, cc).fill = f


    # Apply background colours to Coverage_By_Slot columns for readability (so recalc keeps colours)
    for cc in range(3, ws_cov.max_column + 1):
        task = str(ws_cov.cell(1, cc).value or "")
        f = fill_for(task)
        for rr in range(2, ws_cov.max_row + 1):
            ws_cov.cell(rr, cc).fill = f

    # Colour the coverage columns so it stays coloured after recalc (no formulas needed)
    for cc in range(3, ws_cov.max_column + 1):
        task = str(ws_cov.cell(1, cc).value or "")
        f = fill_for(task)
        for rr in range(2, ws_cov.max_row + 1):
            ws_cov.cell(rr, cc).fill = f
    return ws_cov

def recalc_workbook_from_site_timelines(xlsx_bytes: bytes) -> bytes:
    """Recalculate Coverage + Totals from edited site timelines. Returns updated workbook bytes."""
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    # detect weeks by sheet names
    week_nums = sorted({int(m.group(1)) for s in wb.sheetnames if (m:=re.match(r"Week(\d+)_SLGP_Timeline", s))})
    if not week_nums:
        raise ValueError("Could not find Week#_SLGP_Timeline sheets in uploaded workbook.")
    for w in week_nums:
        staff, tasks_order, totals = _compute_totals_from_site_timelines(wb, w)
        _write_totals_sheet(wb, w, staff, tasks_order, totals)
        _write_coverage_sheet_from_site_timelines(wb, w)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def build_workbook(tpl: TemplateData, start_monday: date, weeks: int) -> Workbook:
    """Stable build: write site timelines as VALUES + FILLS. Totals + Coverage are static (but can be recalculated by re-upload)."""
    wb = Workbook()
    wb.remove(wb.active)

    staff_names = [s.name for s in tpl.staff]
    staff_by_name = {s.name: s for s in tpl.staff}

    DATE_W, TIME_W, STAFF_W = 14, 8, 18

    for w in range(weeks):
        week_start = start_monday + timedelta(days=7*w)
        # schedule_week is defined earlier in this file and returns assignments, breaks, gaps, dates, slots, hours_map
        assignments, breaks, gaps, dates, slots, hours_map = schedule_week(tpl, week_start)

        # Create site sheets (site staff only)
        for site in ("SLGP","JEN","BGS"):
            site_staff = [nm for nm in staff_names if str(staff_by_name[nm].home).upper() == site]
            ws = wb.create_sheet(f"Week{w+1}_{site}_Timeline")
            ws.append(["Date","Time"] + site_staff)
            # style header
            for c in ws[1]:
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws.freeze_panes = "C2"

            ws.column_dimensions["A"].width = DATE_W
            ws.column_dimensions["B"].width = TIME_W
            for i in range(len(site_staff)):
                ws.column_dimensions[get_column_letter(3+i)].width = STAFF_W

            # write rows with repeated header per day for printing
            for d in dates:
                # repeat header row before each day (except first day already has top header)
                if d != dates[0]:
                    ws.append(["Date","Time"] + site_staff)
                    r = ws.max_row
                    for c in range(1, ws.max_column+1):
                        cell = ws.cell(r,c)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                for t in slots:
                    row = [d.strftime("%a %d-%b"), _time_str(t)]
                    for nm in site_staff:
                        hk = holiday_kind(nm, d, tpl.hols)
                        if hk:
                            val = hk
                        elif not is_working(hours_map, d, t, nm):
                            val = ""
                        elif nm in breaks.get((d,t), set()):
                            val = "Break"
                        else:
                            val = assignments.get((d,t,nm), "Misc_Tasks")
                        row.append(val)
                    ws.append(row)

            # apply fills + borders
            _apply_fill_and_style(ws, start_row=2, start_col=1, end_row=ws.max_row, end_col=ws.max_column, repeat_header_every=len(slots)+1)

        # Coverage + Totals (static from assignments for initial build)
        # Coverage
        ws_cov = wb.create_sheet(f"Week{w+1}_Coverage_By_Slot")
        tasks_cols = [
            "FrontDesk_SLGP","FrontDesk_JEN","FrontDesk_BGS",
            "Triage_Admin_SLGP","Triage_Admin_JEN",
            "Email_Box","Awaiting_PSA_Admin","Phones","Bookings","EMIS","Docman","Misc_Tasks"
        ]
        ws_cov.append(["Date","Time"] + tasks_cols)
        for c in ws_cov[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_cov.freeze_panes = "C2"
        ws_cov.column_dimensions["A"].width = 14
        ws_cov.column_dimensions["B"].width = 8
        for i in range(len(tasks_cols)):
            ws_cov.column_dimensions[get_column_letter(3+i)].width = 28

        for d in dates:
            for t in slots:
                row = [d.strftime("%a %d-%b"), _time_str(t)]
                per_task = {k: [] for k in tasks_cols}
                for nm in staff_names:
                    task = assignments.get((d,t,nm))
                    if task in per_task:
                        per_task[task].append(nm)
                for k in tasks_cols:
                    row.append(", ".join(per_task[k]))
                ws_cov.append(row)
        # style
        for rr in range(2, ws_cov.max_row+1):
            time_val = str(ws_cov.cell(rr,2).value or "")
            top_day = (time_val == DAY_START.strftime("%H:%M"))
            for cc in range(1, ws_cov.max_column+1):
                cell = ws_cov.cell(rr,cc)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Apply background colours to Coverage_By_Slot columns for readability
        for cc in range(3, ws_cov.max_column+1):
            task = str(ws_cov.cell(1, cc).value or "")
            f = fill_for(task)
            for rr in range(2, ws_cov.max_row+1):
                ws_cov.cell(rr, cc).fill = f
                cell.border = Border(top=THICK_SIDE if top_day else THIN_SIDE, bottom=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE)

        # Totals
        totals = _compute_totals_from_assignments(assignments, breaks, dates, slots, staff_names)
        ws_tot = wb.create_sheet(f"Week{w+1}_Totals")
        tasks_order = [
            "FrontDesk_SLGP","FrontDesk_JEN","FrontDesk_BGS",
            "Triage_Admin_SLGP","Triage_Admin_JEN",
            "Phones","Bookings","Email_Box","Awaiting_PSA_Admin",
            "EMIS","Docman","Misc_Tasks","Break"
        ]
        ws_tot.append(["Name"] + tasks_order + ["WeeklyTotal"])
        for c in ws_tot[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_tot.freeze_panes = "B2"
        ws_tot.column_dimensions["A"].width = 22
        for i in range(2, ws_tot.max_column+1):
            ws_tot.column_dimensions[get_column_letter(i)].width = 14
        for nm in staff_names:
            row = [nm] + [round(totals.get((nm,t), 0.0), 2) for t in tasks_order] + [round(totals.get((nm,"WeeklyTotal"), 0.0), 2)]
            ws_tot.append(row)

        # Notes/gaps
        ws_g = wb.create_sheet(f"Week{w+1}_NotesAndGaps")
        ws_g.append(["Date","Time","Task","Note"])
        for c in ws_g[1]:
            c.font = Font(bold=True)
        for d,t,task,note in gaps:
            ws_g.append([d.isoformat(), _time_str(t) if t else "", task, note])

    return wb


# =========================================================
# v37_9 STABILISATION PATCH
# Implements:
# 1) Lock Front Desk bands first (fixed)
# 2) Lock Email owner shift next (fixed)
# 3) Move breaks ±30 mins to avoid splitting FD/PSA/Email
# 4) Enforce min 2.5h blocks for FD/PSA/Bookings (Phones only 1h min if needed)
# 5) Phone rotation limit (1 long or 2 short blocks/day)
# 6) Strong SLGP preference for Bookings
# =========================================================

def email_site_for_day(d: date) -> str:
    # As agreed:
    # Mon = BGS, Tue/Wed/Thu = JEN, Fri = SLGP
    wd = d.weekday()
    if wd == 0:
        return "BGS"
    if wd in (1,2,3):
        return "JEN"
    return "SLGP"

def schedule_week_v37_9(tpl: TemplateData, wk_start: date):
    slots = timeslots()
    dates = [wk_start + timedelta(days=i) for i in range(5)]
    hours_map = apply_swaps(tpl.hours_map, tpl.swaps, dates)

    staff_by_name = {s.name: s for s in tpl.staff}
    staff_names = [s.name for s in tpl.staff]

    a: Dict[Tuple[date,time,str], str] = {}
    gaps: List[Tuple[date,time,str,str]] = []
    fixed: Set[Tuple[date,time,str]] = set()

    # minutes tracking (needed for weighting / reporting)
    mins_task: Dict[Tuple[str,str], int] = {}
    mins_task_day: Dict[Tuple[date,str,str], int] = {}

    def add_mins(d: date, nm: str, task_key: str, mins: int):
        mins_task[(nm, task_key)] = mins_task.get((nm, task_key), 0) + mins
        mins_task_day[(d, nm, task_key)] = mins_task_day.get((d, nm, task_key), 0) + mins

    def task_key_for_task(task: str) -> str:
        if task.startswith("FrontDesk_"):
            return "FrontDesk"
        if task.startswith("Triage_Admin_"):
            return "Triage"
        if task == "Email_Box":
            return "Email"
        if task == "Awaiting_PSA_Admin":
            return "Awaiting"
        if task == "Phones":
            return "Phones"
        if task == "Bookings":
            return "Bookings"
        if task == "EMIS":
            return "EMIS"
        if task == "Docman":
            return "Docman"
        if task == "Misc_Tasks":
            return "Misc"
        if task == "Break":
            return "Break"
        return "Misc"

    def is_free(nm: str, d: date, t: time) -> bool:
        return (d,t,nm) not in a

    # ----------------------------
    # TRIAGE (keep existing fixed bands)
    # ----------------------------
    fixed_slots: Set[Tuple[date,time,str]] = set()
    def can_cover_full_band(nm: str, d: date, bs: time, be: time) -> bool:
        stt, endt = shift_window(hours_map, d, nm)
        if not stt or not endt:
            return False
        if not ((stt <= bs) and (endt >= be)):
            return False
        for tt in slots:
            if tt < bs or tt >= be:
                continue
            if (d, tt, nm) in fixed_slots:
                return False
            if (d, tt, nm) in a:
                return False
        return True

    def task_weight(st: Staff, task_key: str) -> int:
        return int(st.weights.get(task_key, 3) if st.weights is not None else 3)

    def pick_for_band(candidates: List[str], d: date, task_key: str, bs: time, be: time) -> Optional[str]:
        ok = []
        for nm in candidates:
            if holiday_kind(nm, d, tpl.hols):
                continue
            if not can_cover_full_band(nm, d, bs, be):
                continue
            ok.append(nm)
        if not ok:
            return None

        def score(nm: str):
            st = staff_by_name[nm]
            w = task_weight(st, task_key)
            used = mins_task.get((nm, task_key), 0)
            return (-w, used, nm.lower())

        ok.sort(key=score)
        return ok[0]

    for d in dates:
        for site in ("SLGP","JEN"):
            role = f"Triage_Admin_{site}"
            cands = [s.name for s in tpl.staff if s.can_triage and s.home == site]
            for bs, be in TRIAGE_BANDS:
                chosen = pick_for_band(cands, d, "Triage", bs, be)
                if not chosen:
                    gaps.append((d, bs, role, "No suitable staff for full band"))
                    continue
                for tt in slots:
                    if tt < bs or tt >= be:
                        continue
                    a[(d, tt, chosen)] = role
                    fixed_slots.add((d, tt, chosen))
                    fixed.add((d, tt, chosen))
                    add_mins(d, chosen, "Triage", SLOT_MIN)

    # ----------------------------
    # 1) FRONT DESK BANDS FIRST (hard fixed)
    # ----------------------------
    fd_bands_by_site = {s: FD_BANDS[:] for s in SITES}

    fd_bands_count: Dict[Tuple[date,str], List[int]] = {}  # (d,nm)->band indices assigned

    def fd_candidate_ok(nm: str, d: date, site: str, bs: time, be: time) -> bool:
        st = staff_by_name[nm]
        if st.home != site or (not st.can_frontdesk):
            return False
        if holiday_kind(nm, d, tpl.hols):
            return False
        stt, endt = shift_window(hours_map, d, nm)
        if not stt or not endt:
            return False
        if not (stt <= bs and endt >= be):
            return False
        for tt in slots:
            if tt < bs or tt >= be:
                continue
            if (d, tt, nm) in fixed:
                return False
            if (d, tt, nm) in a:
                return False
        return True
    def fd_score_for(d: date, bs: time, be: time, nm: str) -> Tuple[int,int,int,int,str]:
        st = staff_by_name[nm]
        # frontdesk_only first
        fd_only = 0 if st.frontdesk_only else 1
        # avoid consuming Email-capable staff on bands overlapping 10:30
        email_pen = 0
        if (bs <= time(10,30) < be) and st.can_email:
            stt, endt = shift_window(hours_map, d, nm)
            if stt and endt and stt <= time(10,30) and endt >= time(16,0):
                email_pen = 1
        w = -task_weight(st, "FrontDesk")
        used = mins_task.get((nm, "FrontDesk"), 0)
        return (fd_only, email_pen, w, used, nm.lower())

    for d in dates:
        for site in SITES:
            bands = fd_bands_by_site[site]
            for bi, (bs, be) in enumerate(bands):
                # Phase 1: strict (max 1 band/day)
                cands = [nm for nm in staff_names if fd_candidate_ok(nm, d, site, bs, be)]
                def strict_ok(nm: str) -> bool:
                    assigned = fd_bands_count.get((d,nm), [])
                    return len(assigned) == 0
                strict = [nm for nm in cands if strict_ok(nm)]
                strict.sort(key=lambda nm: fd_score_for(d, bs, be, nm))

                chosen = strict[0] if strict else None

                # Phase 2: allow 2nd band/day BUT never consecutive
                if not chosen:
                    def second_ok(nm: str) -> bool:
                        assigned = fd_bands_count.get((d,nm), [])
                        if len(assigned) >= 2:
                            return False
                        if len(assigned) == 0:
                            return True
                        # not consecutive
                        return all(abs(bi - x) > 1 for x in assigned)
                    second = [nm for nm in cands if second_ok(nm)]
                    second.sort(key=lambda nm: fd_score_for(d, bs, be, nm))
                    chosen = second[0] if second else None

                if not chosen:
                    gaps.append((d, bs, f"FrontDesk_{site}", "No suitable staff for FD band"))
                    continue

                for tt in slots:
                    if tt < bs or tt >= be:
                        continue
                    a[(d, tt, chosen)] = f"FrontDesk_{site}"
                    fixed.add((d, tt, chosen))
                    add_mins(d, chosen, "FrontDesk", SLOT_MIN)
                fd_bands_count.setdefault((d, chosen), []).append(bi)

    # ----------------------------
    # 2) EMAIL OWNER SHIFT NEXT (hard fixed; ignore break coverage)
    # ----------------------------
    email_owner: Dict[date, Optional[str]] = {}

    def can_hold_email(nm: str, d: date) -> bool:
        st = staff_by_name[nm]
        if not st.can_email:
            return False
        if holiday_kind(nm, d, tpl.hols):
            return False
        stt, endt = shift_window(hours_map, d, nm)
        if not stt or not endt:
            return False
        if not (stt <= time(10,30)):
            return False
        # must be able to cover to 16:00 minimum
        if endt < time(15,30):
            return False
        return True
    def email_window_end(nm: str, d: date) -> time:
        _, endt = shift_window(hours_map, d, nm)
        if not endt:
            return time(16,0)
        if endt >= time(18,30):
            return time(18,30)
        # otherwise cover until end of shift, but no later than 16:00 for standard shifts
        return endt if endt < time(16,0) else time(16,0)

    for d in dates:
        pref_site = email_site_for_day(d)
        cands = [nm for nm in staff_names if can_hold_email(nm, d)]
        if not cands:
            gaps.append((d, time(10,30), "Email_Box", "No eligible staff for Email owner"))
            email_owner[d] = None
            continue

        # Prefer: not already fixed in email window; then preferred site; then weight
        def email_score(nm: str):
            st = staff_by_name[nm]
            endt = email_window_end(nm, d)
            has_conflict = 0
            for tt in slots:
                if tt < time(10,30) or tt >= endt:
                    continue
                if (d, tt, nm) in fixed:
                    has_conflict = 1
                    break
            pref = 0 if st.home == pref_site else 1
            w = -task_weight(st, "Email")
            used = mins_task.get((nm, "Email"), 0)
            return (has_conflict, pref, w, used, nm.lower())

        cands.sort(key=email_score)
        chosen = cands[0]
        email_owner[d] = chosen
        endt = email_window_end(chosen, d)

        for tt in slots:
            if tt < time(10,30) or tt >= endt:
                continue
            if not is_working(hours_map, d, tt, chosen):
                continue
            # If FD/Triage already fixed here, leave it (email will then be "not possible" but we try to avoid by scoring)
            if (d, tt, chosen) in fixed:
                continue
            a[(d, tt, chosen)] = "Email_Box"
            fixed.add((d, tt, chosen))
            add_mins(d, chosen, "Email", SLOT_MIN)

    # ----------------------------
    # 3) PSA ADMIN (Awaiting_PSA_Admin) — continuous 10:30–16:00 (min 2.5h) with ±30 break flexibility
    # ----------------------------
    psa_owner: Dict[date, Optional[str]] = {}

    def awaiting_site_for_day_local(d: date) -> str:
        wd = d.weekday()
        if wd in (0,4):
            return "SLGP"
        if wd in (1,3):
            return "JEN"
        return "BGS"

    def can_hold_psa(nm: str, d: date) -> bool:
        st = staff_by_name[nm]
        if not st.can_docman:
            return False
        if holiday_kind(nm, d, tpl.hols):
            return False
        stt, endt = shift_window(hours_map, d, nm)
        if not stt or not endt:
            return False
        if stt > time(10,30):
            return False
        if endt < time(16,0):
            return False
        return True

    for d in dates:
        pref_site = awaiting_site_for_day_local(d)
        cands = [nm for nm in staff_names if can_hold_psa(nm, d)]
        if not cands:
            gaps.append((d, time(10,30), "Awaiting_PSA_Admin", "No eligible staff for PSA Admin owner"))
            psa_owner[d] = None
            continue

        def psa_score(nm: str):
            st = staff_by_name[nm]
            conflict = 0
            for tt in slots:
                if tt < time(10,30) or tt >= time(16,0):
                    continue
                if (d, tt, nm) in fixed:
                    conflict = 1
                    break
            pref = 0 if st.home == pref_site else 1
            w = -task_weight(st, "Awaiting")
            used = mins_task.get((nm, "Awaiting"), 0)
            return (conflict, pref, w, used, nm.lower())

        cands.sort(key=psa_score)
        chosen = cands[0]
        psa_owner[d] = chosen

        for tt in slots:
            if tt < time(10,30) or tt >= time(16,0):
                continue
            if not is_working(hours_map, d, tt, chosen):
                continue
            if (d, tt, chosen) in fixed:
                continue
            a[(d, tt, chosen)] = "Awaiting_PSA_Admin"
            fixed.add((d, tt, chosen))
            add_mins(d, chosen, "Awaiting", SLOT_MIN)

    # ----------------------------
    # 4) BREAKS with ±30min move to avoid splitting FD/PSA/Email
    # ----------------------------
    breaks: Dict[Tuple[date,time], Set[str]] = {}

    for d in dates:
        for nm in staff_names:
            st = staff_by_name[nm]
            if not st.break_required:
                continue
            if holiday_kind(nm, d, tpl.hols):
                continue
            stt, endt = shift_window(hours_map, d, nm)
            if not stt or not endt:
                continue
            dur = (dt_of(d, endt) - dt_of(d, stt)).total_seconds()/3600.0
            if dur <= BREAK_THRESHOLD_HOURS:
                continue

            # Candidate break starts (and allow ±30 flexibility inherently via list)
            candidates = [bt for bt in BREAK_CANDIDATES if (bt >= stt and add_minutes(bt, 30) <= endt)]

            if not candidates:
                continue

            def break_cost(bt: time) -> Tuple[int,int]:
                # Prefer not to collide with fixed, except it's OK to collide with Email (email doesn't need cover at break)
                collides_fixed = 0
                for tt in (bt,):
                    if (d, tt, nm) in fixed:
                        # allow if the fixed task is Email_Box
                        if a.get((d, tt, nm)) == "Email_Box":
                            collides_fixed += 0
                        else:
                            collides_fixed += 1
                # Prefer midpoint-ish
                midpoint = dt_of(d, stt) + (dt_of(d, endt)-dt_of(d, stt))/2
                dist = int(abs((dt_of(d, bt) - midpoint).total_seconds()))
                return (collides_fixed, dist)

            candidates.sort(key=break_cost)
            bt = candidates[0]
            breaks.setdefault((d, bt), set()).add(nm)
            # overwrite assignment at break
            if is_working(hours_map, d, bt, nm) and not holiday_kind(nm, d, tpl.hols):
                a[(d, bt, nm)] = "Break"
                fixed.add((d, bt, nm))
                add_mins(d, nm, "Break", SLOT_MIN)

    def on_break(nm: str, d: date, t: time) -> bool:
        return nm in breaks.get((d,t), set())

    # ----------------------------
    # Helper: eligibility
    # ----------------------------
    def eligible(nm: str, task: str, d: date, t: time) -> bool:
        st = staff_by_name[nm]
        if holiday_kind(nm, d, tpl.hols):
            return False
        if not is_working(hours_map, d, t, nm):
            return False
        if on_break(nm, d, t):
            return False
        if (d, t, nm) in fixed:
            return False
        if task == "Phones":
            return st.can_phones
        if task == "Bookings":
            return st.can_bookings
        if task == "EMIS":
            return st.can_emis
        if task == "Docman":
            return st.can_docman
        if task == "Misc_Tasks":
            return True
        return True

    # ----------------------------
    # 5) PHONES enforcement with rotation limits
    # ----------------------------
    # Limits: 1 long (>=3h) OR 2 short blocks/day. Minimum 1h block unless unavoidable.
    phone_blocks: Dict[Tuple[date,str], List[Tuple[int,int]]] = {}  # (d,nm)->[(start_idx,end_idx)]
    phone_mins_day: Dict[Tuple[date,str], int] = {}

    def can_take_phones_block(nm: str, d: date, proposed_slots: int) -> bool:
        blks = phone_blocks.get((d,nm), [])
        if len(blks) == 0:
            return True
        if len(blks) >= 2:
            return False
        # If already has a long block, don't add another
        existing_long = any((e-s) >= 6 for s,e in blks)  # 6 slots = 3h
        if existing_long:
            return False
        # Second block must be short (<=2.5h)
        return proposed_slots <= 5

    def contiguous_free_slots(nm: str, d: date, start_idx: int) -> int:
        cnt = 0
        for j in range(start_idx, len(slots)):
            tt = slots[j]
            if not eligible(nm, "Phones", d, tt):
                break
            cnt += 1
        return cnt

    for d in dates:
        for idx, t in enumerate(slots):
            req = phones_required(tpl, d, t)
            if req <= 0:
                continue
            cur = sum(1 for nm in staff_names if a.get((d,t,nm)) == "Phones")
            need = req - cur
            if need <= 0:
                continue

            # choose candidates: free + eligible; weighting; rotate
            cands = [nm for nm in staff_names if eligible(nm, "Phones", d, t)]
            # Prefer those not already on phones today, then weight
            def phones_score(nm: str):
                st = staff_by_name[nm]
                blks = phone_blocks.get((d,nm), [])
                already = 1 if blks else 0
                w = -task_weight(st, "Phones")
                used = mins_task_day.get((d,nm,"Phones"), 0)
                return (already, w, used, nm.lower())

            cands.sort(key=phones_score)

            for _ in range(need):
                chosen = None
                for nm in cands:
                    avail = contiguous_free_slots(nm, d, idx)
                    if avail <= 0:
                        continue
                    # propose a block length: prefer 3h (6 slots) first block, else 2h (4 slots)
                    prop = min(avail, 6) if not phone_blocks.get((d,nm), []) else min(avail, 4)
                    # minimum 1h (2 slots)
                    if prop < 2:
                        continue
                    if not can_take_phones_block(nm, d, prop):
                        continue
                    chosen = (nm, prop)
                    break

                if not chosen:
                    # unavoidable: allow 1h even if it breaks the rotation rules
                    for nm in cands:
                        avail = contiguous_free_slots(nm, d, idx)
                        prop = min(avail, 2)
                        if prop >= 1 and avail >= 1:
                            chosen = (nm, 1)
                            break

                if not chosen:
                    gaps.append((d, t, "Phones", f"Short by {need}"))
                    break

                nm, prop = chosen
                start = idx
                end = idx + prop
                for j in range(start, end):
                    tt = slots[j]
                    if eligible(nm, "Phones", d, tt):
                        a[(d, tt, nm)] = "Phones"
                        add_mins(d, nm, "Phones", SLOT_MIN)
                phone_blocks.setdefault((d,nm), []).append((start, end))
                phone_mins_day[(d,nm)] = phone_mins_day.get((d,nm), 0) + prop*SLOT_MIN
                # remove chosen from candidate list for this slot so we don't pick same person twice
                cands = [x for x in cands if x != nm]

    # ----------------------------
    # 6) BOOKINGS with strong SLGP preference, min 2.5h blocks
    # ----------------------------
    target_book = int(round((tpl.weekly_targets.get("Bookings", 0.0) or 0.0) * 60))
    BOOK_LOW = 0.90
    low_book = int(round(target_book * BOOK_LOW)) if target_book > 0 else 0

    def total_bookings_mins() -> int:
        return sum(v for (nm, tk), v in mins_task.items() if tk == "Bookings")

    def contiguous_free_for_task(nm: str, d: date, start_idx: int, task: str) -> int:
        cnt = 0
        for j in range(start_idx, len(slots)):
            tt = slots[j]
            if not eligible(nm, task, d, tt):
                break
            cnt += 1
        return cnt

    for d in dates:
        for idx, t in enumerate(slots):
            if t < time(10,30):
                continue
            if target_book <= 0 or total_bookings_mins() >= low_book:
                break
            # estimate remaining slots to distribute
            remaining = max(0, low_book - total_bookings_mins())
            # rough people needed now
            ppl = 1 if remaining > 0 else 0
            if ppl <= 0:
                continue

            # choose SLGP-first candidates
            cands = [nm for nm in staff_names if eligible(nm, "Bookings", d, t)]
            def book_score(nm: str):
                st = staff_by_name[nm]
                pref = 0 if st.home == "SLGP" else 1
                w = -task_weight(st, "Bookings")
                used = mins_task.get((nm, "Bookings"), 0)
                return (pref, w, used, nm.lower())
            cands.sort(key=book_score)

            if not cands:
                continue

            nm = cands[0]
            avail = contiguous_free_for_task(nm, d, idx, "Bookings")
            # min 2.5h block = 5 slots
            if avail < 5:
                continue
            prop = min(avail, 9)  # cap ~4.5h
            prop = max(5, prop)
            for j in range(idx, idx+prop):
                tt = slots[j]
                if eligible(nm, "Bookings", d, tt):
                    a[(d, tt, nm)] = "Bookings"
                    add_mins(d, nm, "Bookings", SLOT_MIN)

    # ----------------------------
    # Fill remaining: prefer Docman/EMIS until their weekly targets hit; else Misc
    # ----------------------------
    target_emis = int(round((tpl.weekly_targets.get("EMIS", 0.0) or 0.0) * 60))
    target_doc  = int(round((tpl.weekly_targets.get("Docman", 0.0) or 0.0) * 60))
    EMIS_LOW = 0.90
    DOC_LOW = 0.90
    low_emis = int(round(target_emis * EMIS_LOW)) if target_emis > 0 else 0
    low_doc  = int(round(target_doc  * DOC_LOW))  if target_doc  > 0 else 0

    def total_mins(task_key: str) -> int:
        return sum(v for (nm, tk), v in mins_task.items() if tk == task_key)

    for d in dates:
        for idx, t in enumerate(slots):
            for nm in staff_names:
                if holiday_kind(nm, d, tpl.hols):
                    continue
                if not is_working(hours_map, d, t, nm):
                    continue
                if on_break(nm, d, t):
                    continue
                if (d, t, nm) in a:
                    continue

                # choose based on remaining targets
                chosen = None
                if low_doc > 0 and total_mins("Docman") < low_doc and eligible(nm, "Docman", d, t):
                    chosen = "Docman"
                elif low_emis > 0 and total_mins("EMIS") < low_emis and eligible(nm, "EMIS", d, t):
                    chosen = "EMIS"
                elif eligible(nm, "Misc_Tasks", d, t):
                    chosen = "Misc_Tasks"

                if chosen:
                    a[(d, t, nm)] = chosen
                    add_mins(d, nm, task_key_for_task(chosen), SLOT_MIN)

    # ----------------------------
    # FINAL HARD ENFORCEMENT: no working slot blank
    # ----------------------------
    for d in dates:
        for t in slots:
            for nm in staff_names:
                if holiday_kind(nm, d, tpl.hols):
                    continue
                if not is_working(hours_map, d, t, nm):
                    continue
                if on_break(nm, d, t):
                    continue
                if not a.get((d,t,nm)):
                    a[(d,t,nm)] = "Misc_Tasks"
                    add_mins(d, nm, "Misc", SLOT_MIN)

    return a, breaks, gaps, dates, slots, hours_map

# Override the scheduler used by build_workbook
schedule_week = schedule_week_v37_9