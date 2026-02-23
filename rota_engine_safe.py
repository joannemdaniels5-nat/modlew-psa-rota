
import io
from dataclasses import dataclass
from datetime import date, time, datetime, timedelta
from typing import Dict, List, Tuple, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- Helpers ----------
DAY_START = time(8, 0)
DAY_END = time(18, 30)
SLOT_MIN = 30

@dataclass
class TemplateData:
    staff_names: List[str]
    home_sites: Dict[str, str]  # name -> site (SLGP/JEN/BGS)
    hours_map: Dict[str, Dict[str, Optional[time]]]  # name -> {MonStart, MonEnd, ...}


def ensure_monday(d: date) -> date:
    return d - timedelta(days=d.weekday())


def _normalize(x):
    return str(x).strip().upper()


def _to_time(x):
    if pd.isna(x) or x == "":
        return None
    if isinstance(x, time):
        return x
    if isinstance(x, datetime):
        return x.time()
    # excel numeric time
    if isinstance(x, (float, int)):
        seconds = int(round(float(x) * 86400))
        return (datetime(2000,1,1) + timedelta(seconds=seconds)).time()
    return pd.to_datetime(str(x)).time()


def read_template(uploaded_bytes: bytes) -> TemplateData:
    xls = pd.ExcelFile(io.BytesIO(uploaded_bytes))
    # Try to find Staff and WorkingHours sheets
    def find_sheet(candidates):
        names = {str(n).strip().lower(): n for n in xls.sheet_names}
        for c in candidates:
            k = str(c).strip().lower()
            if k in names:
                return names[k]
        # substring match
        for n in xls.sheet_names:
            nn = str(n).strip().lower()
            if any(str(c).strip().lower() in nn for c in candidates):
                return n
        return None

    staff_sheet = find_sheet(["Staff"]) or xls.sheet_names[0]
    hours_sheet = find_sheet(["WorkingHours", "Hours"]) or xls.sheet_names[0]

    staff_df = pd.read_excel(xls, sheet_name=staff_sheet)
    hours_df = pd.read_excel(xls, sheet_name=hours_sheet)

    # Columns
    def pick_col(df, candidates, default=None):
        cols = {str(c).strip().lower(): c for c in df.columns}
        for cand in candidates:
            k = str(cand).strip().lower()
            if k in cols:
                return cols[k]
        # contains
        for c in df.columns:
            cl = str(c).strip().lower()
            if any(str(k).strip().lower() in cl for k in candidates):
                return c
        return default

    name_c1 = pick_col(staff_df, ["Name", "StaffName"]) or staff_df.columns[0]
    site_c1 = pick_col(staff_df, ["HomeSite", "Site", "BaseSite"])  # may be None

    names = [str(n).strip() for n in staff_df[name_c1].dropna().tolist() if str(n).strip()]
    sites = {n: (str(staff_df.loc[i, site_c1]).strip().upper() if site_c1 else "SLGP")
             for i, n in enumerate(staff_df[name_c1]) if str(n).strip()}

    # Working hours
    name_c2 = pick_col(hours_df, ["Name", "StaffName"]) or hours_df.columns[0]
    hours_df = hours_df.copy()
    hours_df["Name"] = hours_df[name_c2].astype(str).str.strip()
    for dn in ["Mon","Tue","Wed","Thu","Fri"]:
        s_col = pick_col(hours_df, [f"{dn}Start", f"{dn} Start", f"{dn}_Start"]) or None
        e_col = pick_col(hours_df, [f"{dn}End", f"{dn} End", f"{dn}_End"]) or None
        hours_df[f"{dn}Start"] = hours_df[s_col].apply(_to_time) if s_col else None
        hours_df[f"{dn}End"] = hours_df[e_col].apply(_to_time) if e_col else None

    hours_map = {}
    for _, r in hours_df.iterrows():
        nm = str(r["Name"]).strip()
        if not nm:
            continue
        hours_map[nm] = {k: r.get(k) for k in [
            "MonStart","MonEnd","TueStart","TueEnd","WedStart","WedEnd",
            "ThuStart","ThuEnd","FriStart","FriEnd"
        ]}

    return TemplateData(staff_names=names, home_sites=sites, hours_map=hours_map)


# ---------- Scheduling (minimal: fill working slots with Misc_Tasks) ----------

def _timeslots():
    cur = datetime(2000,1,1, DAY_START.hour, DAY_START.minute)
    end = datetime(2000,1,1, DAY_END.hour, DAY_END.minute)
    out = []
    while cur < end:
        out.append(cur.time())
        cur += timedelta(minutes=SLOT_MIN)
    return out

ROLE_COLORS = {
    "Misc_Tasks": "FFFFFF",
    "Break": "DDDDDD",
}

THIN = Side(style="thin")
THICK = Side(style="thick")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build_workbook(tpl: TemplateData, start_monday: date, weeks: int) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    names = tpl.staff_names
    slots = _timeslots()

    def day_name(d: date) -> str:
        return ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"][d.weekday()]

    def is_working(nm: str, d: date, t: time) -> bool:
        hm = tpl.hours_map.get(nm, {})
        dn = day_name(d)
        st = hm.get(f"{dn}Start")
        en = hm.get(f"{dn}End")
        return bool(st and en and (t >= st) and (t < en))

    for w in range(weeks):
        dates = [start_monday + timedelta(days=7*w + i) for i in range(5)]
        # Site sheets by home site
        sites = {s for s in (tpl.home_sites.get(n, 'SLGP') for n in names)}
        for site in sorted(sites):
            site_staff = [n for n in names if (tpl.home_sites.get(n, 'SLGP') == site)]
            if not site_staff:
                continue
            ws = wb.create_sheet(f"Week{w+1}_{site}_Timeline")
            ws.append(["Date","Time"] + site_staff)
            for c in ws[1]:
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws.freeze_panes = "C2"
            ws.column_dimensions['A'].width = 14
            ws.column_dimensions['B'].width = 8
            for i in range(len(site_staff)):
                ws.column_dimensions[get_column_letter(3+i)].width = 18

            for d in dates:
                for t in slots:
                    row = [d.strftime("%a %d-%b"), t.strftime("%H:%M")]
                    for nm in site_staff:
                        row.append("Misc_Tasks" if is_working(nm, d, t) else "")
                    ws.append(row)

            # style
            for rr in range(2, ws.max_row+1):
                for cc in range(1, ws.max_column+1):
                    cell = ws.cell(rr, cc)
                    cell.border = CELL_BORDER
                    if cc >= 3:
                        val = str(cell.value or "")
                        color = ROLE_COLORS.get(val, "FFFFFF")
                        cell.fill = PatternFill("solid", fgColor=color)
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Totals
        ws_tot = wb.create_sheet(f"Week{w+1}_Totals")
        tasks = ["Misc_Tasks", "Break"]
        ws_tot.append(["Name"] + tasks + ["WeeklyTotal"])
        for c in ws_tot[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_tot.freeze_panes = "B2"
        ws_tot.column_dimensions['A'].width = 22
        for i in range(2, 2+len(tasks)+1):
            ws_tot.column_dimensions[get_column_letter(i)].width = 12

        hours = {(nm, task): 0.0 for nm in names for task in tasks}
        for d in dates:
            for t in slots:
                for nm in names:
                    if is_working(nm, d, t):
                        hours[(nm, "Misc_Tasks")] = hours.get((nm, "Misc_Tasks"), 0.0) + 0.5
        for nm in names:
            total = sum(hours.get((nm, task), 0.0) for task in tasks)
            ws_tot.append([nm] + [round(hours.get((nm, task), 0.0), 2) for task in tasks] + [round(total, 2)])

        # Coverage by slot (names)
        ws_cov = wb.create_sheet(f"Week{w+1}_Coverage_By_Slot")
        ws_cov.append(["Date","Time","Misc_Tasks"])
        for c in ws_cov[1]:
            c.font = Font(bold=True)
        ws_cov.freeze_panes = "C2"
        ws_cov.column_dimensions['A'].width = 14
        ws_cov.column_dimensions['B'].width = 8
        ws_cov.column_dimensions['C'].width = 28
        for d in dates:
            for t in slots:
                names_on = [nm for nm in names if is_working(nm, d, t)]
                ws_cov.append([d.strftime("%a %d-%b"), t.strftime("%H:%M"), ", ".join(names_on)])

    return wb


# ---------- Recalc from edited site timelines (rebuild totals + coverage) ----------

def recalc_workbook_from_site_timelines(xlsx_bytes: bytes) -> bytes:
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    # detect week numbers
    import re as _re
    week_nums = sorted({int(m.group(1)) for s in wb.sheetnames if (m:=_re.match(r"Week(\d+)_\w+_Timeline", s))})
    if not week_nums:
        raise ValueError("Could not find any Week#_SITE_Timeline sheets.")

    for w in week_nums:
        # Collect site sheets
        site_sheets = {s: wb[s] for s in wb.sheetnames if s.startswith(f"Week{w}_") and s.endswith("_Timeline")}
        # Staff list (union in display order)
        staff = []
        for ws in site_sheets.values():
            for c in range(3, ws.max_column+1):
                h = ws.cell(1,c).value
                if h and h not in staff:
                    staff.append(h)
        # Recreate Totals
        title = f"Week{w}_Totals"
        if title in wb.sheetnames:
            del wb[title]
        ws_tot = wb.create_sheet(title)
        tasks = ["Misc_Tasks", "Break"]
        ws_tot.append(["Name"] + tasks + ["WeeklyTotal"])
        for c in ws_tot[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_tot.freeze_panes = "B2"
        ws_tot.column_dimensions['A'].width = 22
        for i in range(2, 2+len(tasks)+1):
            ws_tot.column_dimensions[get_column_letter(i)].width = 12
        tot = {(nm, t): 0.0 for nm in staff for t in tasks}
        for ws in site_sheets.values():
            for r in range(2, ws.max_row+1):
                # skip repeated headers if any
                if str(ws.cell(r,2).value).strip().lower() == 'time':
                    continue
                for c in range(3, ws.max_column+1):
                    nm = ws.cell(1,c).value
                    if not nm:
                        continue
                    v = str(ws.cell(r,c).value or '')
                    if v in tasks:
                        tot[(nm, v)] = tot.get((nm, v), 0.0) + 0.5
        for nm in staff:
            total = sum(tot.get((nm, t), 0.0) for t in tasks)
            ws_tot.append([nm] + [round(tot.get((nm, t), 0.0), 2) for t in tasks] + [round(total, 2)])

        # Recreate Coverage_By_Slot
        cov_title = f"Week{w}_Coverage_By_Slot"
        if cov_title in wb.sheetnames:
            del wb[cov_title]
        ws_cov = wb.create_sheet(cov_title)
        ws_cov.append(["Date","Time","Misc_Tasks"])
        for c in ws_cov[1]:
            c.font = Font(bold=True)
        ws_cov.freeze_panes = "C2"
        ws_cov.column_dimensions['A'].width = 14
        ws_cov.column_dimensions['B'].width = 8
        ws_cov.column_dimensions['C'].width = 28

        # choose a backbone sheet for date/time
        backbone = next(iter(site_sheets.values()))
        # Build per-row names
        for r in range(2, backbone.max_row+1):
            if str(backbone.cell(r,2).value).strip().lower() == 'time':
                continue
            dval = backbone.cell(r,1).value
            tval = backbone.cell(r,2).value
            names_on = []
            for ws in site_sheets.values():
                for c in range(3, ws.max_column+1):
                    v = str(ws.cell(r,c).value or '')
                    if v == 'Misc_Tasks':
                        nm = ws.cell(1,c).value
                        if nm and nm not in names_on:
                            names_on.append(nm)
            ws_cov.append([dval, tval, ", ".join(names_on)])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
